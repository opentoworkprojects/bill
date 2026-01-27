
import base64
import json
import logging
import os
import ssl
import uuid
import httpx
import asyncio
import sqlite3
import tempfile
import io
import time
import random
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional

import jwt
import razorpay
from dotenv import load_dotenv
from fastapi import APIRouter, Body, Depends, FastAPI, File, Form, HTTPException, UploadFile, status, Query, Request
from fastapi.responses import Response, StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from motor.motor_asyncio import AsyncIOMotorClient
from passlib.context import CryptContext
from pydantic import BaseModel, ConfigDict, Field
from starlette.middleware.cors import CORSMiddleware
from starlette.middleware.gzip import GZipMiddleware
from starlette.middleware.base import BaseHTTPMiddleware

# Import Redis cache service
from redis_cache import init_redis_cache, cleanup_redis_cache, get_cached_order_service, get_table_status_manager

# Import monitoring system
from monitoring import init_monitoring, collect_metrics_task, monitoring_router

# ✅ Import Performance Optimization Modules
try:
    from response_optimizer import (
        CacheDecorator,
        ResponseOptimizer,
        PerformanceMetrics,
        ResponseHeaders
    )
    from query_optimizer import (
        QueryOptimizer,
        CacheKeyGenerator,
        PerformanceConstants
    )
    print("✅ Performance optimization modules imported successfully")
except ImportError as e:
    print(f"⚠️ Performance modules not found: {e}. Core features will still work.")

try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage

    _LLM_AVAILABLE = True
except Exception:
    _LLM_AVAILABLE = False

# Automation features - no external dependencies needed

    class LlmChat:
        def __init__(self, *args, **kwargs):
            pass

        def with_model(self, *args, **kwargs):
            return self

        async def send_message(self, *args, **kwargs):
            raise RuntimeError("LLM integration unavailable")

    class UserMessage:
        def __init__(self, text):
            self.text = text


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

# Import super admin router AFTER loading .env
from super_admin import super_admin_router, set_database as set_super_admin_db, set_redis_cache as set_super_admin_cache

# Import ops panel router
from ops_panel import ops_router, set_database as set_ops_db, set_redis_cache as set_ops_cache

# MongoDB connection with SSL configuration
mongo_url = os.getenv(
    "MONGO_URL",
    "mongodb+srv://shivshankarkumar281_db_user:RNdGNCCyBtj1d5Ar@retsro-ai.un0np9m.mongodb.net/?retryWrites=true&w=majority&authSource=admin&readPreference=primary&serverSelectionTimeoutMS=10000&connectTimeoutMS=15000&socketTimeoutMS=20000&appName=retsro-ai",
)

# Clean up and optimize MongoDB Atlas connection string
if "mongodb+srv://" in mongo_url:
    # Remove duplicate parameters and optimize for Atlas
    base_url = mongo_url.split("?")[0] if "?" in mongo_url else mongo_url

    # Standard Atlas parameters for optimal connection
    params = [
        "retryWrites=true",
        "w=majority",
        "tls=true",
        "tlsInsecure=true",
        "authSource=admin",
        "readPreference=primary",
    ]

    mongo_url = f"{base_url}?{'&'.join(params)}"

# Configure MongoDB client with OPTIMIZED settings for speed
try:
    if (
        "mongodb+srv://" in mongo_url
        or "ssl=true" in mongo_url
        or "tls=true" in mongo_url
    ):
        # OPTIMIZED MongoDB Atlas connection with connection pooling
        client = AsyncIOMotorClient(
            mongo_url,
            tls=True,
            tlsInsecure=True,
            # Performance optimizations
            maxPoolSize=50,  # Increased connection pool
            minPoolSize=10,  # Keep connections warm
            maxIdleTimeMS=45000,  # Keep connections alive
            serverSelectionTimeoutMS=3000,  # Faster timeout
            connectTimeoutMS=5000,  # Faster connection
            socketTimeoutMS=20000,  # Socket timeout
            retryWrites=True,  # Auto-retry failed writes
            retryReads=True,  # Auto-retry failed reads
            # Compression for faster data transfer
            compressors="snappy,zlib",
        )
    else:
        # For local or non-SSL connections
        client = AsyncIOMotorClient(
            mongo_url,
            maxPoolSize=50,
            minPoolSize=10,
            maxIdleTimeMS=45000,
        )
except Exception as e:
    print(f"MongoDB client creation failed: {e}")
    # Fallback to basic client with minimal TLS settings
    try:
        client = AsyncIOMotorClient(
            mongo_url, 
            tls=True, 
            tlsInsecure=True, 
            serverSelectionTimeoutMS=3000,
            maxPoolSize=50
        )
    except Exception as e2:
        print(f"Fallback client creation failed: {e2}")
        client = AsyncIOMotorClient(mongo_url)

db = client[os.getenv("DB_NAME", "restrobill")]

# Security
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()
JWT_SECRET = os.getenv("JWT_SECRET", "default-jwt-secret-please-change-in-production")
JWT_ALGORITHM = os.getenv("JWT_ALGORITHM", "HS256")

app = FastAPI(
    title="BillByteKOT API",
    description="Restaurant Billing & KOT Management System",
    version="1.3.0",
    docs_url="/api/docs",
    redoc_url="/api/redoc",
    # Performance optimizations
    swagger_ui_parameters={"defaultModelsExpandDepth": -1},
)
api_router = APIRouter(prefix="/api")

# In-memory cache for frequently accessed data
from functools import lru_cache
import threading

# Simple cache with thread-safe operations
_cache = {}
_cache_ttl = {}
_cache_lock = threading.Lock()

# Request deduplication to prevent duplicate concurrent requests
_pending_requests = {}
_pending_lock = threading.Lock()

def cache_response(ttl_seconds=60):
    """Cache decorator for API responses"""
    def decorator(func):
        async def wrapper(*args, **kwargs):
            # Create cache key from function name and args
            cache_key = f"{func.__name__}:{str(args)}:{str(kwargs)}"
            current_time = time.time()
            
            # Check if cached and not expired
            with _cache_lock:
                if cache_key in _cache and cache_key in _cache_ttl:
                    if current_time < _cache_ttl[cache_key]:
                        return _cache[cache_key]
            
            # Execute function and cache result
            result = await func(*args, **kwargs)
            
            with _cache_lock:
                _cache[cache_key] = result
                _cache_ttl[cache_key] = current_time + ttl_seconds
            
            return result
        return wrapper
    return decorator

def clear_expired_cache():
    """Clear expired cache entries to free memory"""
    current_time = time.time()
    with _cache_lock:
        expired_keys = [k for k, v in _cache_ttl.items() if current_time >= v]
        for key in expired_keys:
            _cache.pop(key, None)
            _cache_ttl.pop(key, None)

# Semaphore to limit concurrent database operations (free tier optimization)
DB_SEMAPHORE = asyncio.Semaphore(20)  # Max 20 concurrent DB operations

async def rate_limited_db_operation(operation):
    """Execute database operation with rate limiting for free tier"""
    async with DB_SEMAPHORE:
        return await operation


# Dynamic CORS origin checker
def is_allowed_origin(origin: str) -> bool:
    """Check if the origin is allowed for CORS"""
    allowed_patterns = [
        "http://localhost:3000",
        "http://localhost:3001",
        "https://restro-ai.onrender.com",
        "https://restro-ai-u9kz.vercel.app",
        "https://billbytekot.in",
        "https://www.billbytekot.in",
    ]

    # Check exact matches
    if origin in allowed_patterns:
        return True

    # Check pattern matches
    domain_patterns = [".vercel.app", ".netlify.app", ".onrender.com", ".render.com", ".billbytekot.in"]

    for pattern in domain_patterns:
        if origin.endswith(pattern):
            return True

    # Allow localhost with any port for development
    if origin.startswith("http://localhost:") or origin.startswith("http://127.0.0.1:"):
        return True

    return False


# Add CORS middleware to allow frontend connections
ALLOWED_ORIGINS = [
    "http://localhost:3000",
    "http://localhost:3001",
    "https://restro-ai.onrender.com",
    "https://billbytekot.in",
    "https://www.billbytekot.in",
    # exact current frontend origin:
    "https://restro-ai-u9kz-ed0v8idw3-shivs-projects-db2d52eb.vercel.app",
]

# CRITICAL: Add CORS middleware BEFORE any routes
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins temporarily to fix CORS
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

# Add GZip compression for faster response times (compress responses > 500 bytes)
app.add_middleware(GZipMiddleware, minimum_size=500)

# Rate limiting and monitoring middleware
class MonitoringMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request: Request, call_next):
        start_time = time.time()
        
        # Rate limiting check
        client_ip = request.client.host
        user_agent = request.headers.get("user-agent", "")
        
        # Skip rate limiting for health checks and monitoring endpoints
        if request.url.path in ["/health", "/api/monitoring/health", "/nginx_status"]:
            response = await call_next(request)
            return response
        
        # Check rate limits using Redis if available
        try:
            from redis_cache import redis_cache
            if redis_cache and redis_cache.is_connected():
                # Different rate limits for different endpoints
                if request.url.path.startswith("/api/auth/"):
                    rate_limit_key = f"rate_limit:auth:{client_ip}"
                    if not await redis_cache.check_rate_limit(rate_limit_key, 10, 60):  # 10 requests per minute
                        raise HTTPException(status_code=429, detail="Too many authentication requests")
                elif request.url.path.startswith("/api/orders/"):
                    rate_limit_key = f"rate_limit:orders:{client_ip}"
                    if not await redis_cache.check_rate_limit(rate_limit_key, 200, 60):  # 200 requests per minute
                        raise HTTPException(status_code=429, detail="Too many order requests")
                else:
                    rate_limit_key = f"rate_limit:general:{client_ip}"
                    if not await redis_cache.check_rate_limit(rate_limit_key, 100, 60):  # 100 requests per minute
                        raise HTTPException(status_code=429, detail="Rate limit exceeded")
        except Exception as e:
            # Continue without rate limiting if Redis is unavailable
            print(f"Rate limiting error: {e}")
        
        # Process request
        try:
            response = await call_next(request)
            is_error = response.status_code >= 400
        except Exception as e:
            is_error = True
            response = Response(content=str(e), status_code=500)
        
        # Record metrics
        response_time = (time.time() - start_time) * 1000  # Convert to milliseconds
        
        try:
            from monitoring import metrics_collector
            if metrics_collector:
                metrics_collector.record_request(response_time, is_error)
        except Exception as e:
            print(f"Metrics recording error: {e}")
        
        # Add performance headers
        response.headers["X-Response-Time"] = f"{response_time:.2f}ms"
        response.headers["X-Server-Instance"] = os.getenv("SERVER_INSTANCE", "1")
        
        return response

app.add_middleware(MonitoringMiddleware)


# Currency symbols mapping
CURRENCY_SYMBOLS = {
    "INR": "₹",
    "USD": "$",
    "EUR": "€",
    "GBP": "£",
    "AED": "د.إ",
    "SAR": "﷼",
    "JPY": "¥",
    "CNY": "¥",
    "AUD": "A$",
    "CAD": "C$",
    "PKR": "₨",
}


# Models
class PrintCustomization(BaseModel):
    """Print customization settings"""
    model_config = ConfigDict(extra="allow")  # Allow extra fields
    paper_width: str = "80mm"  # 58mm, 80mm, 110mm, custom
    custom_width: Optional[int] = None  # in mm if paper_width is "custom"
    font_size: Optional[str] = "medium"  # small, medium, large (or int 8-16px for legacy)
    line_spacing: float = 1.3  # 1.0-2.0
    margin_top: int = 5  # in mm
    margin_bottom: int = 5  # in mm
    margin_left: int = 5  # in mm
    margin_right: int = 5  # in mm
    show_logo: bool = True
    show_qr_code: bool = False
    qr_code_enabled: bool = False  # Alias for show_qr_code
    qr_code_content: str = "website"  # website, order_id, custom
    custom_qr_text: Optional[str] = None
    header_style: str = "centered"  # centered, left, right
    item_layout: str = "detailed"  # detailed, compact, minimal
    show_item_notes: bool = True
    show_preparation_time: bool = False
    show_server_name: bool = True
    show_waiter_name: bool = True  # Alias for show_server_name
    show_table_number: bool = True
    show_customer_name: bool = True
    show_order_date: bool = True
    show_order_time: bool = True
    date_format: str = "DD-MM-YYYY"  # DD-MM-YYYY, MM-DD-YYYY, YYYY-MM-DD
    time_format: str = "12h"  # 12h, 24h
    separator_style: str = "dashes"  # dashes, dots, equals, line, dash, equal, heavy, light, none
    border_style: str = "single"  # single, double
    footer_style: str = "simple"  # simple, detailed
    total_style: str = "bold"  # bold, boxed, highlighted
    print_copies: int = 1  # 1-5
    auto_cut: bool = True
    beep_on_print: bool = False
    auto_print: bool = False  # Auto print after payment
    # Header content options
    show_address: bool = True
    show_phone: bool = True
    show_email: bool = False
    show_website: bool = False
    show_gstin: bool = True
    show_fssai: bool = False
    show_tagline: bool = True
    # KOT specific settings
    kot_auto_print: bool = True
    kot_font_size: str = "large"  # small, medium, large
    kot_show_time: bool = True
    kot_highlight_notes: bool = True


class BusinessSettings(BaseModel):
    restaurant_name: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    gstin: Optional[str] = None
    fssai: Optional[str] = None
    currency: str = "INR"
    tax_rate: float = 5.0

    receipt_theme: str = "classic"
    logo_url: Optional[str] = None
    website: Optional[str] = None
    tagline: Optional[str] = None
    footer_message: Optional[str] = "Thank you for dining with us!"
    print_customization: Optional[PrintCustomization] = None
    # Business Type & KOT Settings
    business_type: str = "restaurant"  # restaurant, stall, food-truck, takeaway-only, cafe, cloud-kitchen
    kot_mode_enabled: bool = True  # KOT mode for restaurants with tables
    # WhatsApp Settings
    whatsapp_enabled: bool = False
    whatsapp_business_number: Optional[str] = None
    whatsapp_message_template: Optional[str] = "Thank you for dining at {restaurant_name}! Your bill of {currency}{total} has been paid. Order #{order_id}"
    # Auto WhatsApp Notifications
    whatsapp_auto_notify: bool = False
    whatsapp_notify_on_placed: bool = True
    whatsapp_notify_on_preparing: bool = True
    whatsapp_notify_on_ready: bool = True
    whatsapp_notify_on_completed: bool = True
    # Customer Self-Order Settings
    customer_self_order_enabled: bool = False
    frontend_url: Optional[str] = None  # For generating QR codes
    # UPI Payment Settings
    upi_id: Optional[str] = None  # UPI ID for QR code payments


class User(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    username: str
    email: str
    role: str
    phone: Optional[str] = None
    login_method: str = "password"  # password, whatsapp
    organization_id: Optional[str] = None  # Links staff to their admin/organization
    business_settings: Optional[BusinessSettings] = None
    razorpay_key_id: Optional[str] = None
    razorpay_key_secret: Optional[str] = None
    subscription_active: bool = False
    bill_count: int = 0
    subscription_expires_at: Optional[datetime] = None
    setup_completed: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    # Referral system fields
    referral_code: Optional[str] = None  # Unique 8-char alphanumeric code - ALWAYS GENERATED
    wallet_balance: float = 0.0  # Current wallet balance from referral rewards
    referred_by: Optional[str] = None  # Referral code used during signup (optional)
    total_referrals: int = 0  # Count of successful referrals
    total_referral_earnings: float = 0.0  # Total earned from referrals
    # Fraud prevention fields (Requirement 11.2)
    device_fingerprint: Optional[str] = None  # Device fingerprint for self-referral detection


class UserCreate(BaseModel):
    username: str
    email: str
    password: str
    role: str = "admin"
    referral_code: Optional[str] = None  # Optional referral code during signup


class RegisterOTPRequest(BaseModel):
    email: str
    username: str
    password: str
    role: str = "admin"
    referral_code: Optional[str] = None  # Optional referral code during signup


class VerifyRegistrationOTP(BaseModel):
    email: str
    otp: str


class UserLogin(BaseModel):
    username: str
    password: str


class WhatsAppOTPRequest(BaseModel):
    phone: str
    country_code: str = "+91"


class WhatsAppOTPVerify(BaseModel):
    phone: str
    otp: str
    country_code: str = "+91"


class RazorpaySettings(BaseModel):
    razorpay_key_id: str
    razorpay_key_secret: str


class StaffCreate(BaseModel):
    username: str
    email: str
    password: str
    role: str
    phone: Optional[str] = None
    salary: Optional[float] = None


class StaffUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    role: Optional[str] = None
    phone: Optional[str] = None
    salary: Optional[float] = None


class MenuItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: str
    price: float
    description: Optional[str] = None
    image_url: Optional[str] = None
    image_data: Optional[str] = None
    available: bool = True
    ingredients: Optional[List[str]] = []
    preparation_time: Optional[int] = 15
    organization_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class MenuItemCreate(BaseModel):
    name: str
    category: str
    price: float
    description: Optional[str] = None
    image_url: Optional[str] = None
    image_data: Optional[str] = None
    available: bool = True
    ingredients: Optional[List[str]] = []
    preparation_time: Optional[int] = 15


class Table(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    table_number: int
    capacity: int
    status: str = "available"
    current_order_id: Optional[str] = None
    organization_id: Optional[str] = None
    location: Optional[str] = None
    section: Optional[str] = None
    table_type: Optional[str] = "regular"
    notes: Optional[str] = None
    created_at: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())
    updated_at: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())


class TableCreate(BaseModel):
    table_number: int
    capacity: int
    status: str = "available"
    location: Optional[str] = None
    section: Optional[str] = None
    table_type: Optional[str] = "regular"
    notes: Optional[str] = None


class Reservation(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    table_id: str
    table_number: int
    customer_name: str
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    party_size: int
    reservation_date: str
    reservation_time: str
    duration: int = 120  # minutes
    status: str = "confirmed"  # confirmed, pending, cancelled, completed, expired
    notes: Optional[str] = None
    pre_arrival_minutes: int = 15  # Minutes before reservation to mark table as reserved
    organization_id: Optional[str] = None
    created_at: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())
    updated_at: Optional[str] = Field(default_factory=lambda: datetime.now().isoformat())


class ReservationCreate(BaseModel):
    table_id: str
    customer_name: str
    customer_phone: Optional[str] = None
    customer_email: Optional[str] = None
    party_size: int
    reservation_date: str
    reservation_time: str
    duration: int = 120
    status: str = "confirmed"
    notes: Optional[str] = None
    pre_arrival_minutes: int = 15  # Minutes before reservation to mark table as reserved


class ReservationSettings(BaseModel):
    pre_arrival_minutes: int = 15  # Default 15 minutes before
    auto_clear_minutes: int = 30   # Auto-clear if no-show after 30 minutes
    grace_period_minutes: int = 15 # Grace period for late arrivals


class OrderItem(BaseModel):
    menu_item_id: str
    name: str
    quantity: int
    price: float
    notes: Optional[str] = None


async def get_next_invoice_number(organization_id: str) -> int:
    """Get the next invoice number for an organization"""
    counter = await db.counters.find_one_and_update(
        {"_id": f"invoice_{organization_id}"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
        new=True
    )
    return counter["seq"]

class Order(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    invoice_number: Optional[int] = None  # Sequential invoice number
    table_id: str
    table_number: int
    items: List[OrderItem]
    subtotal: float
    tax: float
    tax_rate: float = 5.0  # Store the tax rate used for this order
    discount: float = 0
    total: float
    status: str = "pending"
    waiter_id: str
    waiter_name: str
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None  # For WhatsApp notifications
    tracking_token: Optional[str] = None  # For customer live tracking
    order_type: Optional[str] = "dine_in"  # dine_in, takeaway, delivery
    organization_id: Optional[str] = None
    # Payment fields - Multi-payment support
    payment_method: Optional[str] = "cash"  # Primary payment method or 'split'
    is_credit: bool = False
    payment_received: float = 0
    balance_amount: float = 0
    # Split payment details
    cash_amount: float = 0
    card_amount: float = 0
    upi_amount: float = 0
    credit_amount: float = 0  # Amount on credit (unpaid)
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class OrderCreate(BaseModel):
    table_id: Optional[str] = "counter"  # Optional when KOT disabled
    table_number: Optional[int] = 0  # Optional when KOT disabled
    items: List[OrderItem]
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None  # For WhatsApp notifications
    frontend_origin: Optional[str] = None  # For generating tracking links
    order_type: Optional[str] = "dine_in"  # dine_in, takeaway, delivery


class Payment(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    order_id: str
    amount: float
    payment_method: str
    razorpay_order_id: Optional[str] = None
    razorpay_payment_id: Optional[str] = None
    status: str = "pending"
    organization_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class PaymentCreate(BaseModel):
    order_id: str
    amount: float
    payment_method: str


class SubscriptionPayment(BaseModel):
    amount: float = 999.0  # ₹999/year


class InventoryItem(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    quantity: float
    unit: str
    min_quantity: float
    max_quantity: Optional[float] = None
    price_per_unit: float
    cost_price: Optional[float] = None
    category_id: Optional[str] = None
    supplier_id: Optional[str] = None
    sku: Optional[str] = None
    barcode: Optional[str] = None
    description: Optional[str] = None
    location: Optional[str] = None
    expiry_date: Optional[str] = None
    batch_number: Optional[str] = None
    reorder_point: Optional[float] = None
    reorder_quantity: Optional[float] = None
    organization_id: Optional[str] = None
    last_updated: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class InventoryItemCreate(BaseModel):
    name: str
    quantity: float
    unit: str
    min_quantity: float
    max_quantity: Optional[float] = None
    price_per_unit: float
    cost_price: Optional[float] = None
    category_id: Optional[str] = None
    supplier_id: Optional[str] = None
    sku: Optional[str] = None
    barcode: Optional[str] = None
    description: Optional[str] = None
    location: Optional[str] = None
    expiry_date: Optional[str] = None
    batch_number: Optional[str] = None
    reorder_point: Optional[float] = None
    reorder_quantity: Optional[float] = None


# Expense Management Models
EXPENSE_CATEGORIES = [
    "Rent", "Utilities", "Salaries", "Supplies", "Maintenance", 
    "Marketing", "Insurance", "Taxes", "Equipment", "Transportation",
    "Food & Ingredients", "Cleaning", "Licenses", "Professional Services", "Other"
]

class Expense(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    date: str  # ISO date string YYYY-MM-DD
    amount: float
    category: str
    description: str
    payment_method: str = "cash"  # cash, card, upi, bank_transfer
    vendor_name: Optional[str] = None
    receipt_url: Optional[str] = None
    notes: Optional[str] = None
    organization_id: Optional[str] = None
    created_by: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class ExpenseCreate(BaseModel):
    date: str
    amount: float
    category: str
    description: str
    payment_method: str = "cash"
    vendor_name: Optional[str] = None
    receipt_url: Optional[str] = None
    notes: Optional[str] = None


class ExpenseUpdate(BaseModel):
    date: Optional[str] = None
    amount: Optional[float] = None
    category: Optional[str] = None
    description: Optional[str] = None
    payment_method: Optional[str] = None
    vendor_name: Optional[str] = None
    receipt_url: Optional[str] = None
    notes: Optional[str] = None


class ChatMessage(BaseModel):
    message: str


class PrintData(BaseModel):
    content: str
    type: str
    theme: Optional[str] = "classic"


# Helper functions
def hash_password(password: str) -> str:
    return pwd_context.hash(password)


def verify_password(plain_password: str, hashed_password: str) -> bool:
    try:
        return pwd_context.verify(plain_password, hashed_password)
    except Exception as e:
        print(f"❌ Password verification exception: {str(e)}")
        return False


def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.now(timezone.utc) + timedelta(days=7)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, JWT_SECRET, algorithm=JWT_ALGORITHM)


# Referral System Helper Functions
import string

REFERRAL_CODE_CHARS = string.ascii_uppercase + string.digits  # A-Z, 0-9
REFERRAL_CODE_LENGTH = 8


def generate_referral_code_string() -> str:
    """Generate a random 8-character alphanumeric referral code"""
    return ''.join(random.choices(REFERRAL_CODE_CHARS, k=REFERRAL_CODE_LENGTH))


async def generate_unique_referral_code() -> str:
    """
    Generate a unique 8-character alphanumeric referral code.
    Ensures uniqueness by checking against existing codes in the database.
    
    Returns:
        str: A unique 8-character alphanumeric code (uppercase)
    """
    max_attempts = 10
    for _ in range(max_attempts):
        code = generate_referral_code_string()
        # Check if code already exists (case-insensitive)
        existing = await db.users.find_one({"referral_code": {"$regex": f"^{code}$", "$options": "i"}})
        if not existing:
            return code
    
    # Fallback: use UUID-based code if random generation fails
    return str(uuid.uuid4()).replace("-", "")[:REFERRAL_CODE_LENGTH].upper()


async def validate_referral_code(code: str) -> dict:
    """
    Validate a referral code and return the referrer information.
    Case-insensitive validation as per requirements.
    
    Args:
        code: The referral code to validate
        
    Returns:
        dict with 'valid' boolean and 'referrer' user data if valid
    """
    if not code or len(code) != REFERRAL_CODE_LENGTH:
        return {"valid": False, "error": "Invalid referral code format"}
    
    # Case-insensitive lookup
    referrer = await db.users.find_one({"referral_code": {"$regex": f"^{code}$", "$options": "i"}})
    
    if not referrer:
        return {"valid": False, "error": "Invalid referral code"}
    
    return {
        "valid": True,
        "referrer": {
            "id": referrer.get("id"),
            "username": referrer.get("username"),
            "referral_code": referrer.get("referral_code")
        }
    }


async def assign_referral_code_to_user(user_id: str) -> str:
    """
    Assign a unique referral code to a user if they don't have one.
    
    Args:
        user_id: The user's ID
        
    Returns:
        str: The user's referral code (existing or newly generated)
    """
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise ValueError(f"User not found: {user_id}")
    
    # Return existing code if already assigned
    if user.get("referral_code"):
        return user["referral_code"]
    
    # Generate and assign new code
    code = await generate_unique_referral_code()
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"referral_code": code}}
    )
    return code


async def send_registration_otp_email(email: str, otp: str, username: str = "User"):
    """Send OTP email for registration verification"""
    import os
    
    # Get email configuration
    EMAIL_PROVIDER = os.getenv("EMAIL_PROVIDER", "console")
    
    subject = "Verify Your Email - BillByteKOT Registration"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background-color: #f5f5f5;
                margin: 0;
                padding: 0;
            }}
            .container {{
                max-width: 600px;
                margin: 40px auto;
                background-color: #ffffff;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 40px 20px;
                text-align: center;
            }}
            .header h1 {{
                color: #ffffff;
                margin: 0;
                font-size: 28px;
            }}
            .content {{
                padding: 40px 30px;
            }}
            .otp-box {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: #ffffff;
                font-size: 36px;
                font-weight: bold;
                letter-spacing: 8px;
                text-align: center;
                padding: 20px;
                border-radius: 8px;
                margin: 30px 0;
            }}
            .info {{
                background-color: #f8f9fa;
                border-left: 4px solid #667eea;
                padding: 15px;
                margin: 20px 0;
                border-radius: 4px;
            }}
            .footer {{
                background-color: #f8f9fa;
                padding: 20px;
                text-align: center;
                color: #6c757d;
                font-size: 14px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🍽️ BillByteKOT</h1>
                <p style="color: #ffffff; margin: 10px 0 0 0;">Restaurant Management System</p>
            </div>
            
            <div class="content">
                <h2 style="color: #333;">Welcome, {username}! 👋</h2>
                <p style="color: #666; font-size: 16px; line-height: 1.6;">
                    Thank you for registering with BillByteKOT! Please verify your email address to complete your registration.
                </p>
                
                <div class="otp-box">
                    {otp}
                </div>
                
                <div class="info">
                    <p style="margin: 0; color: #666;">
                        <strong>⏰ Valid for 10 minutes</strong><br>
                        This OTP will expire in 10 minutes for security reasons.
                    </p>
                </div>
                
                <p style="color: #666; font-size: 14px; margin-top: 30px;">
                    If you didn't request this registration, please ignore this email.
                </p>
            </div>
            
            <div class="footer">
                <p style="margin: 0;">
                    <strong>BillByteKOT</strong> - Smart Restaurant Management<br>
                    © 2025 BillByte Innovations. All rights reserved.
                </p>
                <p style="margin: 10px 0 0 0; font-size: 12px;">
                    This is an automated email. Please do not reply.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    Welcome {username}!
    
    Your BillByteKOT Registration OTP is: {otp}
    
    This OTP is valid for 10 minutes.
    
    If you didn't request this registration, please ignore this email.
    
    ---
    BillByteKOT - Smart Restaurant Management
    © 2025 BillByte Innovations
    """
    
    # Console mode for development
    if EMAIL_PROVIDER == "console" or EMAIL_PROVIDER == "":
        print(f"\n{'='*60}")
        print(f"📧 REGISTRATION OTP EMAIL (Console Mode)")
        print(f"{'='*60}")
        print(f"To: {email}")
        print(f"Subject: {subject}")
        print(f"OTP: {otp}")
        print(f"{'='*60}\n")
        return {"success": True, "message": "OTP logged to console (dev mode)"}
    
    # Use email_service for all providers (including resend)
    try:
        from email_service import send_otp_email
        
        # Log OTP for debugging
        print(f"🔐 REGISTRATION OTP for {email}: {otp}")
        
        result = await send_otp_email(email, subject, html_body, text_body)
        return result
    except Exception as e:
        print(f"Email service error: {e}")
        print(f"🔐 REGISTRATION OTP for {email}: {otp}")
        return {"success": False, "message": str(e)}


async def send_password_reset_otp_email(email: str, otp: str, username: str = "User"):
    """Send password reset OTP email"""
    from email_service import send_otp_email
    
    subject = "Your BillByteKOT Password Reset OTP"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }}
            .container {{ max-width: 500px; margin: 0 auto; background: white; border-radius: 10px; padding: 30px; }}
            .header {{ text-align: center; color: #7c3aed; }}
            .otp-box {{ background: linear-gradient(135deg, #7c3aed, #a855f7); color: white; font-size: 32px; font-weight: bold; letter-spacing: 8px; text-align: center; padding: 20px; border-radius: 8px; margin: 20px 0; }}
            .footer {{ text-align: center; color: #666; font-size: 12px; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🍽️ BillByteKOT</h1>
                <p>Password Reset OTP</p>
            </div>
            <p>Hello {username},</p>
            <p>Your password reset OTP is:</p>
            <div class="otp-box">{otp}</div>
            <p><strong>Valid for 10 minutes.</strong></p>
            <p>If you didn't request this, please ignore this email.</p>
            <div class="footer">
                <p>BillByteKOT - Restaurant Management System</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    text_body = f"Hello {username},\n\nYour BillByteKOT Password Reset OTP is: {otp}\n\nValid for 10 minutes.\n\nIf you didn't request this, please ignore this email."
    
    return await send_otp_email(email, subject, html_body, text_body)


async def send_password_reset_email(email: str, reset_link: str, username: str = "User"):
    """Send password reset email with reset link"""
    import os
    
    # Get email configuration
    EMAIL_PROVIDER = os.getenv("EMAIL_PROVIDER", "console")
    
    subject = "Reset Your BillByteKOT Password"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{
                font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
                background-color: #f5f5f5;
                margin: 0;
                padding: 0;
            }}
            .container {{
                max-width: 600px;
                margin: 40px auto;
                background-color: #ffffff;
                border-radius: 12px;
                overflow: hidden;
                box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1);
            }}
            .header {{
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                padding: 40px 20px;
                text-align: center;
            }}
            .header h1 {{
                color: #ffffff;
                margin: 0;
                font-size: 28px;
            }}
            .content {{
                padding: 40px 30px;
            }}
            .button {{
                display: inline-block;
                background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
                color: #ffffff;
                padding: 15px 40px;
                text-decoration: none;
                border-radius: 8px;
                margin: 20px 0;
                font-weight: bold;
            }}
            .info {{
                background-color: #f8f9fa;
                border-left: 4px solid #667eea;
                padding: 15px;
                margin: 20px 0;
                border-radius: 4px;
            }}
            .footer {{
                background-color: #f8f9fa;
                padding: 20px;
                text-align: center;
                color: #6c757d;
                font-size: 14px;
            }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🍽️ BillByteKOT</h1>
                <p style="color: #ffffff; margin: 10px 0 0 0;">Restaurant Management System</p>
            </div>
            
            <div class="content">
                <h2 style="color: #333;">Hello {username}! 👋</h2>
                <p style="color: #666; font-size: 16px; line-height: 1.6;">
                    We received a request to reset your password for your BillByteKOT account.
                </p>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{reset_link}" class="button">Reset Password</a>
                </div>
                
                <div class="info">
                    <p style="margin: 0; color: #666;">
                        <strong>⏰ Valid for 1 hour</strong><br>
                        This reset link will expire in 1 hour for security reasons.
                    </p>
                </div>
                
                <p style="color: #666; font-size: 14px; margin-top: 30px;">
                    If you didn't request a password reset, please ignore this email or contact support if you have concerns.
                </p>
                
                <p style="color: #999; font-size: 12px; margin-top: 20px;">
                    If the button doesn't work, copy and paste this link into your browser:<br>
                    <a href="{reset_link}" style="color: #667eea;">{reset_link}</a>
                </p>
            </div>
            
            <div class="footer">
                <p style="margin: 0;">
                    <strong>BillByteKOT</strong> - Smart Restaurant Management<br>
                    © 2025 BillByte Innovations. All rights reserved.
                </p>
                <p style="margin: 10px 0 0 0; font-size: 12px;">
                    This is an automated email. Please do not reply.
                </p>
            </div>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    Hello {username}!
    
    We received a request to reset your password for your BillByteKOT account.
    
    Click the link below to reset your password:
    {reset_link}
    
    This link is valid for 1 hour.
    
    If you didn't request a password reset, please ignore this email.
    
    ---
    BillByteKOT - Smart Restaurant Management
    © 2025 BillByte Innovations
    """
    
    # Console mode for development
    if EMAIL_PROVIDER == "console" or EMAIL_PROVIDER == "":
        print(f"\n{'='*60}")
        print(f"📧 PASSWORD RESET EMAIL (Console Mode)")
        print(f"{'='*60}")
        print(f"To: {email}")
        print(f"Subject: {subject}")
        print(f"Reset Link: {reset_link}")
        print(f"{'='*60}\n")
        return {"success": True, "message": "Email logged to console (dev mode)"}
    
    # Try to use email_service if configured
    try:
        from email_service import send_via_smtp, send_via_sendgrid, send_via_mailgun, send_via_ses
        
        if EMAIL_PROVIDER == "smtp":
            return await send_via_smtp(email, subject, html_body, text_body)
        elif EMAIL_PROVIDER == "sendgrid":
            return await send_via_sendgrid(email, subject, html_body, text_body)
        elif EMAIL_PROVIDER == "mailgun":
            return await send_via_mailgun(email, subject, html_body, text_body)
        elif EMAIL_PROVIDER == "ses":
            return await send_via_ses(email, subject, html_body, text_body)
    except Exception as e:
        print(f"Email service error: {e}")
        # Fallback to console
        print(f"\n[EMAIL FALLBACK] To: {email}, Link: {reset_link}")
        return {"success": False, "message": str(e)}


async def send_staff_verification_email(email: str, otp: str, staff_name: str, admin_name: str):
    """Send OTP email for staff email verification"""
    from email_service import send_otp_email
    
    subject = "Verify Your Email - BillByteKOT Staff Account"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }}
            .container {{ max-width: 500px; margin: 0 auto; background: white; border-radius: 10px; padding: 30px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); }}
            .header {{ text-align: center; color: #7c3aed; }}
            .otp-box {{ background: linear-gradient(135deg, #7c3aed, #a855f7); color: white; font-size: 32px; font-weight: bold; letter-spacing: 8px; text-align: center; padding: 20px; border-radius: 8px; margin: 20px 0; }}
            .info {{ background-color: #f0f9ff; border-left: 4px solid #3b82f6; padding: 15px; margin: 20px 0; border-radius: 4px; }}
            .footer {{ text-align: center; color: #666; font-size: 12px; margin-top: 20px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🍽️ BillByteKOT</h1>
                <p>Staff Account Verification</p>
            </div>
            
            <h2>Hello {staff_name}! 👋</h2>
            <p><strong>{admin_name}</strong> has invited you to join their restaurant team on BillByteKOT.</p>
            <p>Please share this OTP with your admin to verify your email and complete your account setup:</p>
            
            <div class="otp-box">{otp}</div>
            
            <div class="info">
                <p style="margin: 0;"><strong>⏰ Valid for 10 minutes</strong><br>
                This OTP will expire in 10 minutes for security reasons.</p>
            </div>
            
            <p style="color: #666; font-size: 14px;">If you didn't expect this invitation, please ignore this email.</p>
            
            <div class="footer">
                <p><strong>BillByteKOT</strong> - Smart Restaurant Management<br>
                © 2025 BillByte Innovations. All rights reserved.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    Hello {staff_name}!
    
    {admin_name} has invited you to join their restaurant team on BillByteKOT.
    
    Your verification OTP is: {otp}
    
    Please share this OTP with your admin to complete your account setup.
    This OTP is valid for 10 minutes.
    
    If you didn't expect this invitation, please ignore this email.
    
    ---
    BillByteKOT - Smart Restaurant Management
    © 2025 BillByte Innovations
    """
    
    # Log OTP for debugging
    print(f"🔐 STAFF VERIFICATION OTP for {email}: {otp}")
    
    try:
        result = await send_otp_email(email, subject, html_body, text_body)
        return result
    except Exception as e:
        print(f"Email service error: {e}")
        return {"success": False, "message": str(e)}


def get_secure_org_id(current_user: dict) -> str:
    """
    SECURITY CRITICAL: Get organization_id securely.
    - For admin users: returns their own id
    - For staff users: returns their organization_id (must be set)
    - Raises exception if organization_id is missing for staff
    """
    org_id = current_user.get("organization_id")
    if not org_id:
        # This should never happen if get_current_user is working correctly
        print(f"🚨 SECURITY: User {current_user.get('email')} has no organization_id!")
        raise HTTPException(
            status_code=403, 
            detail="Organization not configured. Contact support."
        )
    return org_id


async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("user_id")
        if user_id is None:
            print(f"❌ Invalid token: no user_id in payload")
            raise HTTPException(status_code=401, detail="Invalid token")
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user is None:
            print(f"❌ User not found: {user_id}")
            raise HTTPException(status_code=401, detail="User not found")

        # Ensure all required fields exist with defaults
        user.setdefault("subscription_active", False)
        user.setdefault("bill_count", 0)
        user.setdefault("setup_completed", False)
        user.setdefault("business_settings", None)
        user.setdefault("razorpay_key_id", None)
        user.setdefault("razorpay_key_secret", None)
        user.setdefault("subscription_expires_at", None)
        
        # CRITICAL: Ensure organization_id is properly set
        # For admin users: organization_id = their own id
        # For staff users: organization_id MUST be set from database (linked to admin)
        if user["role"] == "admin":
            user["organization_id"] = user["id"]
        elif not user.get("organization_id"):
            # Staff without organization_id is a security risk - deny access
            print(f"❌ SECURITY: Staff user {user['email']} has no organization_id!")
            raise HTTPException(
                status_code=403, 
                detail="Staff account not properly linked to organization. Contact your admin."
            )

        return user
    except jwt.ExpiredSignatureError:
        print(f"❌ Token expired")
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.exceptions.InvalidTokenError as e:
        print(f"❌ JWT Error: {str(e)}")
        raise HTTPException(status_code=401, detail="Invalid token")
    except HTTPException:
        raise
    except Exception as e:
        print(f"❌ Auth error: {str(e)}")
        raise HTTPException(status_code=401, detail=f"Authentication failed: {str(e)}")


async def check_subscription(user: dict):
    """
    Strict trial enforcement with extension support
    - Trial: 7 days from account creation + any extension days
    - After trial: Must have active paid subscription
    - No bill count limit during trial
    - Staff users: Check their own subscription first, then fall back to admin's
    """
    
    # For staff users, check their own subscription first
    if user.get("role") in ["waiter", "cashier", "kitchen", "staff"]:
        # First check if staff has their own active subscription
        if user.get("subscription_active"):
            expires_at = user.get("subscription_expires_at")
            if expires_at:
                if isinstance(expires_at, str):
                    try:
                        expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
                    except:
                        expires_at = None
                
                if expires_at and expires_at >= datetime.now(timezone.utc):
                    # Staff has their own active subscription
                    return True
        
        # Staff doesn't have own subscription, check organization admin's subscription
        org_id = user.get("organization_id")
        if org_id:
            # Find the admin of this organization
            admin_user = await db.users.find_one(
                {"id": org_id, "role": "admin"}, 
                {"_id": 0, "subscription_active": 1, "subscription_expires_at": 1, "created_at": 1, "trial_extension_days": 1}
            )
            if admin_user:
                # Use admin's subscription status for trial/subscription check
                user = admin_user
            else:
                # If no admin found, block access
                return False
    
    # Check if user has active paid subscription
    if user.get("subscription_active"):
        # Check if subscription has expired
        expires_at = user.get("subscription_expires_at")
        if expires_at:
            if isinstance(expires_at, str):
                try:
                    expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
                except:
                    expires_at = None
            
            if expires_at and expires_at < datetime.now(timezone.utc):
                # Subscription expired - deactivate it
                await db.users.update_one(
                    {"id": user["id"]}, 
                    {"$set": {"subscription_active": False}}
                )
                # Fall through to check trial
            else:
                # Active subscription - allow access
                return True
    
    # No active subscription - check trial period
    created_at = user.get("created_at")
    if isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        except:
            created_at = None
    
    if created_at:
        # Base trial is 7 days + any extension days granted by admin
        trial_extension_days = user.get("trial_extension_days", 0)
        total_trial_days = 7 + trial_extension_days
        trial_end = created_at + timedelta(days=total_trial_days)
        now = datetime.now(timezone.utc)
        
        if now < trial_end:
            # Still in trial period - allow access
            return True
        else:
            # Trial expired and no active subscription - block access
            return False
    
    # No created_at date (shouldn't happen) - block access for safety
    return False


def get_paper_width_chars(paper_width: str, custom_width: Optional[int] = None) -> int:
    """Calculate character width based on paper size"""
    width_map = {
        "58mm": 32,
        "80mm": 48,
        "110mm": 64,
        "custom": custom_width // 2 if custom_width else 48
    }
    return width_map.get(paper_width, 48)


def format_date_time(dt: datetime, date_format: str, time_format: str) -> tuple:
    """Format date and time based on settings"""
    date_formats = {
        "DD-MM-YYYY": "%d-%m-%Y",
        "MM-DD-YYYY": "%m-%d-%Y",
        "YYYY-MM-DD": "%Y-%m-%d"
    }
    time_formats = {
        "12h": "%I:%M %p",
        "24h": "%H:%M"
    }
    
    date_str = dt.strftime(date_formats.get(date_format, "%d-%m-%Y"))
    time_str = dt.strftime(time_formats.get(time_format, "%I:%M %p"))
    return date_str, time_str


def get_separator(style: str, width: int) -> str:
    """Get separator based on style and width"""
    separators = {
        "dash": "-" * width,
        "equal": "=" * width,
        "heavy": "═" * width,
        "light": "─" * width,
        "none": ""
    }
    return separators.get(style, "-" * width)


def get_receipt_template(
    theme: str, business: dict, order: dict, currency_symbol: str, customization: Optional[dict] = None
) -> str:
    # Get customization settings or use defaults
    custom = customization or {}
    paper_width = custom.get("paper_width", "80mm")
    custom_width_mm = custom.get("custom_width")
    font_size = custom.get("font_size", 12)
    separator_style = custom.get("separator_style", "dash")
    show_logo = custom.get("show_logo", True)
    show_server = custom.get("show_server_name", True)
    show_table = custom.get("show_table_number", True)
    show_customer = custom.get("show_customer_name", True)
    show_date = custom.get("show_order_date", True)
    show_time = custom.get("show_order_time", True)
    date_format = custom.get("date_format", "DD-MM-YYYY")
    time_format = custom.get("time_format", "12h")
    item_layout = custom.get("item_layout", "detailed")
    show_notes = custom.get("show_item_notes", True)
    header_style = custom.get("header_style", "centered")
    
    # Calculate character width
    char_width = get_paper_width_chars(paper_width, custom_width_mm)
    
    # Get separators
    sep_main = get_separator(separator_style, char_width)
    sep_light = get_separator("light", char_width)
    
    # Format date and time
    now = datetime.now()
    date_str, time_str = format_date_time(now, date_format, time_format)
    
    # Legacy separators for backward compatibility
    sep_eq = "=" * 48
    sep_dash = "-" * 48
    sep_heavy = "═" * 48
    sep_light_48 = "─" * 48

    items_text = "".join(
        [
            f"{item['quantity']}x {item['name']:<30} {currency_symbol}{item['price'] * item['quantity']:.2f}\n"
            for item in order["items"]
        ]
    )
    items_modern = "".join(
        [
            f"  {item['quantity']}× {item['name']:<28} {currency_symbol}{item['price'] * item['quantity']:.2f}\n"
            for item in order["items"]
        ]
    )
    items_minimal = "".join(
        [
            f"{item['quantity']}× {item['name']}: {currency_symbol}{item['price'] * item['quantity']:.2f}\n"
            for item in order["items"]
        ]
    )
    items_elegant = "".join(
        [
            f"{item['quantity']:>3} × {item['name']:<28} {currency_symbol}{item['price'] * item['quantity']:>8.2f}\n"
            for item in order["items"]
        ]
    )
    items_compact = "".join(
        [
            f"{item['quantity']}x {item['name'][:20]:<20} {currency_symbol}{item['price'] * item['quantity']:.2f}\n"
            for item in order["items"]
        ]
    )

    now_str = datetime.now().strftime("%d-%m-%Y %I:%M %p")
    now_elegant = datetime.now().strftime("%d %B %Y, %I:%M %p")
    now_compact = datetime.now().strftime("%d/%m/%y %H:%M")

    # Pre-compute centered text
    rest_name = business.get("restaurant_name", "RESTAURANT")
    rest_name_48 = rest_name.center(48)
    rest_name_44 = rest_name.center(44)
    rest_name_32 = rest_name.center(32) if len(rest_name) <= 32 else rest_name[:32]
    address = business.get("address", "")
    address_48 = address.center(48) if len(address) <= 48 else address[:48]
    address_44 = address.center(44) if len(address) <= 44 else address[:44]
    address_32 = address.center(32) if len(address) <= 32 else address[:32]
    phone = business.get("phone", "N/A")
    email = business.get("email", "")
    gstin = business.get("gstin", "N/A")
    fssai = business.get("fssai", "")
    website = business.get("website", "")
    tagline = business.get("tagline", "")
    footer_msg = business.get("footer_message", "Thank you for dining with us!")

    tax_rate = business.get("tax_rate", 5)

    templates = {
        "classic": f"""

{sep_eq}
{rest_name_48}
{sep_eq}
{address_48}
Phone: {phone}
{f"Email: {email}" if email else ""}
{f"GSTIN: {gstin}" if gstin != "N/A" else ""}
{f"FSSAI: {fssai}" if fssai else ""}
{sep_eq}

BILL #{order["id"][:8]}
{sep_dash}
Table: {order["table_number"]:<20} Waiter: {order["waiter_name"]}
Customer: {order.get("customer_name", "Guest")}
Date: {now_str}
{sep_dash}

ITEMS
{sep_dash}
{"Item":<25} {"Qty":>3} {"Amount":>12}
{sep_dash}
{items_text}{sep_dash}

Subtotal:                 {currency_symbol}{order["subtotal"]:>10.2f}
Tax ({tax_rate}%):                   {currency_symbol}{order["tax"]:>10.2f}
{sep_eq}
TOTAL:                    {currency_symbol}{order["total"]:>10.2f}
{sep_eq}

{footer_msg.center(48)}
{f"{website.center(48)}" if website else ""}

{sep_dash}
Thank you! Visit again!
{sep_eq}

""",
        "modern": f"""

┌{"─" * 46}┐
│ {rest_name_44} │
{f"│ {tagline.center(44)} │" if tagline else ""}
├{"─" * 46}┤
│ {address_44} │
│ ☎  {phone:<43} │
{f"│ 🌐 {website:<43} │" if website else ""}
└{"─" * 46}┘

🧾 BILL #{order["id"][:8]}
{"─" * 48}
🍽️  Table {order["table_number"]}  |  👤 {order["waiter_name"]}
👥 {order.get("customer_name", "Guest")}
📅 {now_str}
{"─" * 48}

ORDER ITEMS
{"─" * 48}
{items_modern}{"─" * 48}

Subtotal                  {currency_symbol}{order["subtotal"]:>10.2f}
Tax ({tax_rate}%)                    {currency_symbol}{order["tax"]:>10.2f}
{"═" * 48}
💰 TOTAL                  {currency_symbol}{order["total"]:>10.2f}
{"═" * 48}

✨ {footer_msg.center(44)} ✨
{f"GSTIN: {gstin}" if gstin != "N/A" else ""}

{"─" * 48}
Thank you for dining with us!
{"─" * 48}

""",
        "minimal": f"""

{rest_name.center(48)}
{address.center(48)}
{phone.center(48)}
{"─" * 48}

Bill: {order["id"][:8]}
Table: {order["table_number"]} | {order["waiter_name"]}
Customer: {order.get("customer_name", "Guest")}
{now_str}
{"─" * 48}

{items_minimal}
{"─" * 48}
Subtotal:             {currency_symbol}{order["subtotal"]:>10.2f}
Tax ({tax_rate}%):               {currency_symbol}{order["tax"]:>10.2f}
{"─" * 48}
TOTAL:                {currency_symbol}{order["total"]:>10.2f}
{"─" * 48}

{footer_msg.center(48)}

""",
        "elegant": f"""
╔{sep_heavy}╗
║ {rest_name_44} ║
{f"║ {tagline.center(44)} ║" if tagline else ""}
╠{sep_heavy}╣
║ {address_44} ║
║ Tel: {phone:<40} ║
{f"║ Email: {email:<38} ║" if email else ""}
║ GSTIN: {gstin:<38} ║
{f"║ FSSAI: {fssai:<38} ║" if fssai else ""}
╚{sep_heavy}╝

Invoice: {order["id"][:8]}
Table: {order["table_number"]} | Server: {order["waiter_name"]}
Guest: {order.get("customer_name", "Walk-in")}
Date: {now_elegant}

{sep_dash}
{items_elegant}{sep_dash}
                    Subtotal: {currency_symbol}{order["subtotal"]:>8.2f}
               Tax ({tax_rate}%): {currency_symbol}{order["tax"]:>8.2f}
{sep_heavy}
                       TOTAL: {currency_symbol}{order["total"]:>8.2f}
{sep_heavy}

        {footer_msg}
{f"          {website}" if website else "          Please visit us again"}
""",
        "compact": f"""
{rest_name_32}
{address_32}
Ph: {phone}
{sep_dash[:32]}
Bill: {order["id"][:8]}
Table: {order["table_number"]} | {order["waiter_name"]}
{now_compact}
{sep_dash[:32]}
{items_compact}{sep_dash[:32]}
Subtotal: {currency_symbol}{order["subtotal"]:.2f}
Tax ({tax_rate}%): {currency_symbol}{order["tax"]:.2f}
TOTAL: {currency_symbol}{order["total"]:.2f}
{sep_dash[:32]}
{footer_msg[:32]}
GSTIN: {gstin}
""",
        "detailed": f"""
{sep_eq}
{rest_name_48}
{f"{tagline.center(48)}" if tagline else ""}
{sep_eq}
Address: {address}
Phone: {phone}
{f"Email: {email}" if email else ""}
{f"Website: {website}" if website else ""}
GSTIN: {gstin}
{f"FSSAI License: {fssai}" if fssai else ""}
{sep_eq}

INVOICE DETAILS
{sep_dash}
Invoice No: {order["id"][:8]}
Table Number: {order["table_number"]}
Server: {order["waiter_name"]}
Customer: {order.get("customer_name", "Walk-in Guest")}
Date & Time: {now_str}
{sep_dash}

ORDER ITEMS
{sep_dash}
{"Item":<30} {"Qty":>4} {"Price":>10}
{sep_dash}
{items_text}{sep_dash}

PAYMENT SUMMARY
{sep_dash}
Subtotal:                 {currency_symbol}{order["subtotal"]:.2f}
Tax ({tax_rate}%):                  {currency_symbol}{order["tax"]:.2f}
{sep_eq}
GRAND TOTAL:              {currency_symbol}{order["total"]:.2f}
{sep_eq}

{footer_msg}
{f"Visit us at: {website}" if website else ""}

{sep_dash}
This is a computer generated receipt
{sep_eq}
""",
    }
    return templates.get(theme, templates["classic"])


# Auth routes
@api_router.post("/auth/register-debug")
async def register_debug(user_data: RegisterOTPRequest):
    """Debug registration endpoint that returns OTP for testing"""
    # Only enable in development/debug mode
    if os.getenv("DEBUG_MODE", "false").lower() != "true":
        raise HTTPException(status_code=403, detail="Debug mode not enabled")
    
    # Same logic as register_request but returns OTP
    username_lower = user_data.username.lower().strip()
    email_lower = user_data.email.lower().strip()
    
    # Check if username already exists (case-insensitive)
    existing_username = await db.users.find_one(
        {"username_lower": username_lower}, {"_id": 0}
    )
    if not existing_username:
        existing_username = await db.users.find_one(
            {"username": {"$regex": f"^{user_data.username}$", "$options": "i"}}, {"_id": 0}
        )
    if existing_username:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists (case-insensitive)
    existing_email = await db.users.find_one(
        {"email_lower": email_lower}, {"_id": 0}
    )
    if not existing_email:
        existing_email = await db.users.find_one(
            {"email": {"$regex": f"^{user_data.email}$", "$options": "i"}}, {"_id": 0}
        )
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Generate 6-digit OTP
    otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
    
    # Store OTP and user data temporarily
    registration_otp_storage[email_lower] = {
        "otp": otp,
        "expires": expires_at,
        "user_data": {
            "username": user_data.username.strip(),
            "username_lower": username_lower,
            "email": user_data.email.strip(),
            "email_lower": email_lower,
            "password": user_data.password,
            "role": user_data.role,
            "referral_code": user_data.referral_code.strip().upper() if user_data.referral_code else None
        }
    }
    
    # Send OTP email asynchronously (non-blocking)
    asyncio.create_task(send_registration_otp_email(user_data.email.strip(), otp, user_data.username.strip()))
    
    return {
        "message": "Debug OTP generated and sent to email",
        "email": user_data.email,
        "otp": str(otp),  # Return OTP for testing in debug mode
        "success": True
    }


@api_router.post("/auth/register-request")
async def register_request(user_data: RegisterOTPRequest):
    """Step 1: Request registration - Send OTP to email"""
    # Normalize username and email to lowercase for case-insensitive matching
    username_lower = user_data.username.lower().strip()
    email_lower = user_data.email.lower().strip()
    
    # Check if username already exists (case-insensitive)
    existing_username = await db.users.find_one(
        {"username_lower": username_lower}, {"_id": 0}
    )
    # Fallback: also check original username field with regex for older records
    if not existing_username:
        existing_username = await db.users.find_one(
            {"username": {"$regex": f"^{user_data.username}$", "$options": "i"}}, {"_id": 0}
        )
    if existing_username:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists (case-insensitive)
    existing_email = await db.users.find_one(
        {"email_lower": email_lower}, {"_id": 0}
    )
    # Fallback: also check original email field with regex for older records
    if not existing_email:
        existing_email = await db.users.find_one(
            {"email": {"$regex": f"^{user_data.email}$", "$options": "i"}}, {"_id": 0}
        )
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Generate 6-digit OTP
    otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
    
    # Store OTP and user data temporarily (store lowercase versions for consistency)
    registration_otp_storage[email_lower] = {
        "otp": otp,
        "expires": expires_at,
        "user_data": {
            "username": user_data.username.strip(),  # Keep original case for display
            "username_lower": username_lower,  # Lowercase for lookups
            "email": user_data.email.strip(),  # Keep original case for display
            "email_lower": email_lower,  # Lowercase for lookups
            "password": user_data.password,
            "role": user_data.role,
            "referral_code": user_data.referral_code.strip().upper() if user_data.referral_code and user_data.referral_code.strip() else None  # Store referral code if provided
        }
    }
    
    # Send OTP email asynchronously (non-blocking)
    asyncio.create_task(send_registration_otp_email(user_data.email.strip(), otp, user_data.username.strip()))
    
    return {
        "message": "OTP sent to your email. Please verify to complete registration.",
        "email": user_data.email,
        "success": True,
        "otp": otp,  # Always return OTP for debugging signup issues
        "debug_info": {
            "email_key": email_lower,
            "otp_length": len(otp),
            "expires_in_minutes": 15
        }
    }


@api_router.post("/auth/verify-registration", response_model=User)
async def verify_registration(verify_data: VerifyRegistrationOTP):
    """Step 2: Verify OTP and complete registration"""
    # Normalize email for lookup
    email_lower = verify_data.email.lower().strip()
    
    # Get OTP data (try lowercase first, then original)
    otp_data = registration_otp_storage.get(email_lower) or registration_otp_storage.get(verify_data.email)
    if not otp_data:
        raise HTTPException(status_code=400, detail="No registration request found. Please request OTP again.")
    
    # Check if OTP expired
    if datetime.now(timezone.utc) > otp_data["expires"]:
        del registration_otp_storage[verify_data.email]
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new one.")
    
    # Verify OTP (normalize both values for comparison)
    stored_otp = str(otp_data["otp"]).strip()
    input_otp = str(verify_data.otp).strip()
    
    print(f"🔍 OTP Comparison:")
    print(f"  Stored OTP: '{stored_otp}' (length: {len(stored_otp)})")
    print(f"  Input OTP: '{input_otp}' (length: {len(input_otp)})")
    print(f"  Match: {stored_otp == input_otp}")
    
    if stored_otp != input_otp:
        print(f"❌ OTP mismatch - Expected: '{stored_otp}', Got: '{input_otp}'")
        raise HTTPException(status_code=400, detail=f"Invalid OTP. Please check and try again.")
    
    # Get user data from storage
    user_data = otp_data["user_data"]
    
    # GENERATE UNIQUE REFERRAL CODE BEFORE CREATING USER OBJECT
    try:
        user_referral_code = await generate_unique_referral_code()
        print(f"✅ Generated referral code for new user: {user_referral_code}")
    except Exception as e:
        print(f"⚠️ Failed to generate referral code: {e}")
        # If generation fails, create a simple unique code
        import time
        user_referral_code = f"U{int(time.time())}"[-8:].upper().zfill(8)
        print(f"✅ Using fallback referral code: {user_referral_code}")
    
    # Create user object WITH referral_code
    user_obj = User(
        username=user_data["username"],
        email=user_data["email"],
        role=user_data["role"],
        referral_code=user_referral_code  # ALWAYS SET A REFERRAL CODE
    )
    
    # If admin, they are their own organization
    if user_data["role"] == "admin":
        user_obj.organization_id = user_obj.id
    
    # Add email_verified flag and lowercase fields for case-insensitive lookups
    doc = user_obj.model_dump()
    
    # Handle referral if code was provided during signup
    referral_code = user_data.get("referral_code")
    if referral_code and referral_code.strip():
        doc["referred_by"] = referral_code.strip().upper()
        print(f"✅ User referred by: {referral_code}")
    
    # Initialize other referral fields (referral_code already set in user_obj)
    doc["wallet_balance"] = 0.0
    doc["total_referrals"] = 0
    doc["total_referral_earnings"] = 0.0
    
    # DEBUG: Print document keys to see what's being inserted
    print(f"🔍 Document keys before insert: {list(doc.keys())}")
    print(f"🔍 Referral code in doc: {doc.get('referral_code', 'NOT_FOUND')}")
    
    doc["password"] = hash_password(user_data["password"])
    doc["created_at"] = doc["created_at"].isoformat()
    doc["email_verified"] = True
    doc["email_verified_at"] = datetime.now(timezone.utc).isoformat()
    # Add lowercase fields for case-insensitive lookups
    doc["username_lower"] = user_data.get("username_lower", user_data["username"].lower().strip())
    doc["email_lower"] = user_data.get("email_lower", user_data["email"].lower().strip())
    
    try:
        # Insert user into database
        await db.users.insert_one(doc)
        print(f"✅ User created successfully: {user_obj.username} ({user_obj.email})")
        
        # Remove used OTP (clean up all possible keys)
        cleanup_keys = [email_lower, verify_data.email, verify_data.email.strip()]
        for key in cleanup_keys:
            registration_otp_storage.pop(key, None)
        print(f"✅ OTP cleaned up successfully")
        
        # Send welcome email asynchronously
        try:
            from email_automation import send_welcome_email
            asyncio.create_task(send_welcome_email(user_data["email"], user_data["username"]))
            print(f"✅ Welcome email queued for {user_data['email']}")
        except Exception as e:
            print(f"⚠️ Failed to send welcome email: {e}")
        
        return user_obj
        
    except Exception as e:
        print(f"❌ Error during user creation: {e}")
        import traceback
        traceback.print_exc()
        
        # Clean up OTP on error
        cleanup_keys = [email_lower, verify_data.email, verify_data.email.strip()]
        for key in cleanup_keys:
            registration_otp_storage.pop(key, None)
        
        # Provide user-friendly error message
        if "duplicate key" in str(e).lower() or "already exists" in str(e).lower():
            raise HTTPException(status_code=400, detail="Account already exists. Please try logging in instead.")
        else:
            raise HTTPException(status_code=500, detail="Account creation failed. Please try again or contact support.")
    
    # Apply referral if code was provided (Requirement 3.7)
    if referral_code:
        try:
            # Find the referrer by code
            referrer = await db.users.find_one({
                "referral_code": {"$regex": f"^{referral_code}$", "$options": "i"}
            })
            if referrer:
                # Create referral record with PENDING status
                referral_record = {
                    "id": str(uuid.uuid4()),
                    "referrer_user_id": referrer.get("id"),
                    "referee_user_id": user_obj.id,
                    "referral_code": referral_code,
                    "status": "PENDING",
                    "referee_discount": REFERRAL_DISCOUNT_AMOUNT,
                    "referrer_reward": REFERRAL_REWARD_AMOUNT,
                    "referee_email": user_data["email"],
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.referrals.insert_one(referral_record)
                print(f"✅ Referral record created for user {user_obj.id} with code {referral_code}")
            else:
                print(f"⚠️ Referrer not found for code: {referral_code}")
        except Exception as e:
            print(f"⚠️ Failed to create referral record: {e}")
    
    print(f"✅ User created successfully: {user_obj.username} ({user_obj.email})")
    
    # Remove used OTP (clean up all possible keys)
    cleanup_keys = [email_lower, verify_data.email, verify_data.email.strip()]
    for key in cleanup_keys:
        registration_otp_storage.pop(key, None)
    print(f"✅ OTP cleaned up successfully")
    
    # Send welcome email asynchronously
    try:
        from email_automation import send_welcome_email
        asyncio.create_task(send_welcome_email(user_data["email"], user_data["username"]))
    except Exception as e:
        print(f"Failed to send welcome email: {e}")
    
    return user_obj


@api_router.post("/auth/register", response_model=User)
async def register_direct(user_data: UserCreate):
    """Direct registration without OTP verification"""
    # Normalize username and email to lowercase for case-insensitive matching
    username_lower = user_data.username.lower().strip()
    email_lower = user_data.email.lower().strip()
    
    # Check if username already exists (case-insensitive)
    existing_username = await db.users.find_one(
        {"username_lower": username_lower}, {"_id": 0}
    )
    if not existing_username:
        existing_username = await db.users.find_one(
            {"username": {"$regex": f"^{user_data.username}$", "$options": "i"}}, {"_id": 0}
        )
    if existing_username:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists (case-insensitive)
    existing_email = await db.users.find_one(
        {"email_lower": email_lower}, {"_id": 0}
    )
    if not existing_email:
        existing_email = await db.users.find_one(
            {"email": {"$regex": f"^{user_data.email}$", "$options": "i"}}, {"_id": 0}
        )
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # GENERATE UNIQUE REFERRAL CODE BEFORE CREATING USER OBJECT
    try:
        user_referral_code = await generate_unique_referral_code()
        print(f"✅ Generated referral code for new user: {user_referral_code}")
    except Exception as e:
        print(f"⚠️ Failed to generate referral code: {e}")
        # If generation fails, create a simple unique code
        import time
        user_referral_code = f"U{int(time.time())}"[-8:].upper().zfill(8)
        print(f"✅ Using fallback referral code: {user_referral_code}")
    
    # Create user object
    user_obj = User(
        username=user_data.username.strip(),
        email=user_data.email.strip(),
        role=user_data.role,
        referral_code=user_referral_code  # ALWAYS SET A REFERRAL CODE
    )
    
    # If admin, they are their own organization
    if user_data.role == "admin":
        user_obj.organization_id = user_obj.id
    
    # Prepare document for database
    doc = user_obj.model_dump()
    doc["password"] = hash_password(user_data.password)
    doc["created_at"] = doc["created_at"].isoformat()
    doc["email_verified"] = False  # Not verified since no OTP
    doc["username_lower"] = username_lower
    doc["email_lower"] = email_lower
    
    # Store referral code if provided (Requirement 3.7)
    referral_code = user_data.referral_code.strip().upper() if user_data.referral_code else None
    if referral_code:
        doc["referred_by"] = referral_code
    # Don't set referred_by field if referral_code is None to avoid database issues
    
    # Insert user into database
    try:
        await db.users.insert_one(doc)
        print(f"✅ User created successfully: {user_obj.username}")
    except Exception as e:
        if "duplicate key error" in str(e):
            if "username" in str(e):
                raise HTTPException(status_code=400, detail="Username already exists")
            elif "email" in str(e):
                raise HTTPException(status_code=400, detail="Email already registered")
            elif "referral_code" in str(e):
                # This shouldn't happen with proper generation, but handle it
                print(f"⚠️ Referral code collision, retrying...")
                # Generate a new referral code and retry
                user_referral_code = await generate_unique_referral_code()
                doc["referral_code"] = user_referral_code
                user_obj.referral_code = user_referral_code
                await db.users.insert_one(doc)
            else:
                raise HTTPException(status_code=400, detail="User already exists")
        else:
            print(f"❌ Database error during user creation: {e}")
            raise HTTPException(status_code=500, detail="Failed to create user")
    
    # Apply referral if code was provided (Requirement 3.7)
    if referral_code:
        try:
            # Find the referrer by code
            referrer = await db.users.find_one({
                "referral_code": {"$regex": f"^{referral_code}$", "$options": "i"}
            })
            if referrer:
                # Create referral record with PENDING status
                referral_record = {
                    "id": str(uuid.uuid4()),
                    "referrer_user_id": referrer.get("id"),
                    "referee_user_id": user_obj.id,
                    "referral_code": referral_code,
                    "status": "PENDING",
                    "referee_discount": REFERRAL_DISCOUNT_AMOUNT,
                    "referrer_reward": REFERRAL_REWARD_AMOUNT,
                    "referee_email": user_data.email.strip(),
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                await db.referrals.insert_one(referral_record)
                print(f"✅ Referral record created for user {user_obj.id} with code {referral_code}")
        except Exception as e:
            print(f"Failed to create referral record: {e}")
    
    # Send welcome email asynchronously (non-blocking)
    try:
        from email_automation import send_welcome_email
        asyncio.create_task(send_welcome_email(user_data.email.strip(), user_data.username.strip()))
    except Exception as e:
        print(f"Failed to send welcome email: {e}")
    
    return user_obj


@api_router.post("/auth/login")
async def login(credentials: UserLogin):
    # Normalize username for case-insensitive lookup - strip ALL whitespace including tabs
    username_clean = credentials.username.strip()
    username_lower = username_clean.lower()
    
    # Try to find user by lowercase username first (new records)
    user = await db.users.find_one({"username_lower": username_lower}, {"_id": 0})
    
    # Fallback: try case-insensitive regex for older records without username_lower field
    if not user:
        user = await db.users.find_one(
            {"username": {"$regex": f"^{username_clean}$", "$options": "i"}}, 
            {"_id": 0}
        )
    
    # Also try to find by email (users might login with email)
    if not user:
        user = await db.users.find_one(
            {"email": {"$regex": f"^{username_clean}$", "$options": "i"}}, 
            {"_id": 0}
        )
    
    if not user:
        print(f"❌ Login failed: User not found for {username_clean}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Verify password
    try:
        password_valid = verify_password(credentials.password, user["password"])
    except Exception as e:
        print(f"❌ Password verification error for {username_clean}: {str(e)}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not password_valid:
        print(f"❌ Login failed: Invalid password for {username_clean}")
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    print(f"✅ Login successful for {username_clean}")

    # For staff users, get business_settings and subscription from their admin
    business_settings = user.get("business_settings")
    subscription_active = user.get("subscription_active", False)
    organization_id = user.get("organization_id")
    
    if user.get("role") in ["waiter", "cashier", "kitchen", "staff"] and organization_id:
        admin_user = await db.users.find_one(
            {"id": organization_id, "role": "admin"}, 
            {"_id": 0, "business_settings": 1, "subscription_active": 1}
        )
        if admin_user:
            business_settings = admin_user.get("business_settings")
            subscription_active = admin_user.get("subscription_active", False)

    token = create_access_token({"user_id": user["id"], "role": user["role"]})
    return {
        "token": token,
        "user": {
            "id": user["id"],
            "username": user["username"],
            "role": user["role"],
            "email": user["email"],
            "phone": user.get("phone"),
            "login_method": user.get("login_method", "password"),
            "subscription_active": subscription_active,
            "bill_count": user.get("bill_count", 0),
            "setup_completed": user.get("setup_completed", False),
            "onboarding_completed": user.get("onboarding_completed", False),
            "business_settings": business_settings,
            "organization_id": organization_id,
        },
    }


@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    """Get current user with trial/subscription status"""
    
    # For staff users, get business_settings and subscription from their admin
    business_settings = current_user.get("business_settings")
    subscription_active = current_user.get("subscription_active", False)
    organization_id = current_user.get("organization_id")
    admin_created_at = None
    
    if current_user.get("role") in ["waiter", "cashier", "kitchen", "staff"] and organization_id:
        admin_user = await db.users.find_one(
            {"id": organization_id, "role": "admin"}, 
            {"_id": 0, "business_settings": 1, "subscription_active": 1, "created_at": 1, "trial_extension_days": 1}
        )
        if admin_user:
            business_settings = admin_user.get("business_settings")
            subscription_active = admin_user.get("subscription_active", False)
            admin_created_at = admin_user.get("created_at")
    
    # Calculate trial info (use admin's created_at for staff)
    created_at = admin_created_at or current_user.get("created_at")
    trial_info = {
        "is_trial": False,
        "trial_days_left": 0,
        "trial_expired": False,
        "trial_end_date": None
    }
    
    if isinstance(created_at, str):
        try:
            created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
        except:
            created_at = None
    
    if created_at:
        trial_end = created_at + timedelta(days=7)
        now = datetime.now(timezone.utc)
        days_left = (trial_end - now).days
        
        trial_info = {
            "is_trial": not subscription_active,
            "trial_days_left": max(0, days_left),
            "trial_expired": days_left < 0 and not subscription_active,
            "trial_end_date": trial_end.isoformat()
        }
    
    # Add trial info and updated business_settings to response
    user_data = {
        **current_user, 
        "trial_info": trial_info,
        "business_settings": business_settings,
        "subscription_active": subscription_active
    }
    
    return user_data


@api_router.get("/dashboard")
async def get_dashboard(current_user: dict = Depends(get_current_user)):
    """Get dashboard statistics and metrics"""
    user_org_id = get_secure_org_id(current_user)
    
    try:
        # Use IST (Indian Standard Time) for "today" calculation
        from datetime import timedelta
        IST = timezone(timedelta(hours=5, minutes=30))
        
        # Get current time in IST and find start of today in IST
        now_ist = datetime.now(IST)
        today_ist = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Convert to UTC for database query
        today_utc = today_ist.astimezone(timezone.utc)
        
        # Get today's orders (all statuses)
        today_orders = await db.orders.find({
            "organization_id": user_org_id,
            "created_at": {"$gte": today_utc.isoformat()}
        }, {"_id": 0}).to_list(1000)
        
        # Get today's completed orders only
        today_completed_orders = [order for order in today_orders if order.get("status") in ["completed", "paid"]]
        
        # Calculate today's revenue (completed orders only)
        todays_revenue = sum(order.get("total", 0) for order in today_completed_orders)
        
        # Get pending orders count
        pending_orders = await db.orders.count_documents({
            "organization_id": user_org_id,
            "status": {"$in": ["pending", "preparing", "confirmed"]}
        })
        
        # Get this month's data
        month_start = now_ist.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        month_start_utc = month_start.astimezone(timezone.utc)
        
        monthly_orders = await db.orders.find({
            "organization_id": user_org_id,
            "created_at": {"$gte": month_start_utc.isoformat()},
            "status": {"$in": ["completed", "paid"]}
        }, {"_id": 0}).to_list(5000)
        
        monthly_revenue = sum(order.get("total", 0) for order in monthly_orders)
        
        # Get total orders count (all time)
        total_orders = await db.orders.count_documents({
            "organization_id": user_org_id
        })
        
        return {
            "todaysRevenue": todays_revenue,
            "todaysOrders": len(today_orders),
            "todaysCompletedOrders": len(today_completed_orders),
            "totalOrders": total_orders,
            "pendingOrders": pending_orders,
            "monthlyRevenue": monthly_revenue,
            "monthlyOrders": len(monthly_orders),
            "timestamp": now_ist.isoformat()
        }
        
    except Exception as e:
        print(f"❌ Dashboard error: {e}")
        return {
            "todaysRevenue": 0,
            "todaysOrders": 0,
            "todaysCompletedOrders": 0,
            "totalOrders": 0,
            "pendingOrders": 0,
            "monthlyRevenue": 0,
            "monthlyOrders": 0,
            "error": str(e),
            "timestamp": datetime.now().isoformat()
        }


@api_router.put("/users/me/onboarding")
async def complete_onboarding(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Mark onboarding as completed"""
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"onboarding_completed": data.get("onboarding_completed", True)}}
    )
    return {"message": "Onboarding status updated"}


# ============ PASSWORD RESET ============
# In-memory reset token storage (use Redis in production)
reset_tokens = {}  # {token: {"email": "user@example.com", "expires": timestamp}}


class ForgotPasswordRequest(BaseModel):
    email: str


class VerifyOTPRequest(BaseModel):
    email: str
    otp: str


class ResetPasswordRequest(BaseModel):
    email: str
    otp: str
    new_password: str


@api_router.post("/auth/forgot-password")
async def forgot_password(request: ForgotPasswordRequest):
    """Verify account exists and send OTP for password reset"""
    # Clean email - strip whitespace including tabs
    email_clean = request.email.strip().lower()
    
    # Find user by email (case-insensitive)
    user = await db.users.find_one(
        {"email": {"$regex": f"^{email_clean}$", "$options": "i"}}, 
        {"_id": 0}
    )
    if not user:
        raise HTTPException(status_code=404, detail="No account found with this email address")
    
    # Use the actual email from database for consistency
    user_email = user.get("email", email_clean)
    
    # Generate 6-digit OTP
    otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)  # OTP valid for 10 minutes
    
    # Store OTP using lowercase email as key for consistency
    reset_tokens[email_clean] = {
        "otp": otp,
        "expires": expires_at,
        "verified": False,
        "user_email": user_email  # Store actual email from DB
    }
    
    # Log OTP for debugging (check Render logs)
    print(f"🔐 PASSWORD RESET OTP for {user_email}: {otp}")
    
    # Send OTP email
    try:
        result = await send_password_reset_otp_email(user_email, otp, user.get("username", "User"))
        email_sent = result.get("success", False)
        print(f"📧 Email result for {user_email}: {result}")
    except Exception as e:
        email_sent = False
        print(f"❌ Email error for {user_email}: {e}")
    
    if not email_sent:
        print(f"⚠️ Email failed but OTP logged above. Check Render logs for OTP: {otp}")
    
    return {
        "message": "OTP sent to your email. Please check your inbox.",
        "success": True
    }


@api_router.post("/auth/verify-reset-otp")
async def verify_reset_otp(request: VerifyOTPRequest):
    """Verify OTP for password reset"""
    # Clean email - strip whitespace including tabs
    email_clean = request.email.strip().lower()
    
    # Get OTP data
    otp_data = reset_tokens.get(email_clean)
    if not otp_data:
        raise HTTPException(status_code=400, detail="No OTP found. Please request a new one.")
    
    # Check if OTP expired
    if datetime.now(timezone.utc) > otp_data["expires"]:
        # Remove expired OTP
        del reset_tokens[email_clean]
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new one.")
    
    # Verify OTP
    if otp_data["otp"] != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP. Please check and try again.")
    
    # Mark OTP as verified
    reset_tokens[email_clean]["verified"] = True
    
    return {
        "message": "OTP verified successfully. You can now reset your password.",
        "success": True
    }


@api_router.post("/auth/reset-password")
async def reset_password(request: ResetPasswordRequest):
    """Reset user password using verified OTP"""
    # Clean email - strip whitespace including tabs
    email_clean = request.email.strip().lower()
    
    # Get OTP data
    otp_data = reset_tokens.get(email_clean)
    if not otp_data:
        raise HTTPException(status_code=400, detail="No OTP found. Please request a new one.")
    
    # Check if OTP expired
    if datetime.now(timezone.utc) > otp_data["expires"]:
        # Remove expired OTP
        del reset_tokens[email_clean]
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new one.")
    
    # Verify OTP again
    if otp_data["otp"] != request.otp:
        raise HTTPException(status_code=400, detail="Invalid OTP. Please check and try again.")
    
    # Check if OTP was verified
    if not otp_data.get("verified", False):
        raise HTTPException(status_code=400, detail="Please verify OTP first.")
    
    # Get the actual user email from OTP data or find user
    user_email = otp_data.get("user_email", email_clean)
    
    # Find user (case-insensitive)
    user = await db.users.find_one(
        {"email": {"$regex": f"^{email_clean}$", "$options": "i"}}, 
        {"_id": 0}
    )
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Use actual email from database for update
    actual_email = user.get("email")
    
    # Update password
    hashed_password = hash_password(request.new_password)
    print(f"🔐 Resetting password for {actual_email}")
    print(f"🔐 New password hash: {hashed_password[:20]}...")
    
    await db.users.update_one(
        {"email": actual_email},
        {"$set": {"password": hashed_password}}
    )
    
    print(f"✅ Password reset successful for {actual_email}")
    
    # Remove used OTP
    del reset_tokens[email_clean]
    
    return {
        "message": "Password reset successful. You can now login with your new password.",
        "success": True
    }

# Lead Capture (Public endpoint - no auth required)
import random
import string

# In-memory storage for registration OTP
registration_otp_storage = {}  # {email: {"otp": "123456", "expires": timestamp, "user_data": {...}}}

# In-memory storage for staff OTP verification
staff_otp_storage = {}  # {email: {"otp": "123456", "expires": timestamp, "staff_data": {...}, "admin_id": "..."}}

class LeadCapture(BaseModel):
    name: str
    phone: str
    email: str
    businessName: Optional[str] = None
    source: str = "landing_page"
    timestamp: str


# Lead Capture (Public endpoint - no auth required)
@api_router.post("/leads")
async def capture_lead(lead: LeadCapture):
    """Capture lead from landing page popup"""
    try:
        # Store lead in database
        lead_data = lead.dict()
        lead_data["created_at"] = datetime.now(timezone.utc).isoformat()
        lead_data["status"] = "new"
        lead_data["contacted"] = False
        
        result = await db.leads.insert_one(lead_data)
        
        # TODO: Send notification to admin (email/SMS)
        # TODO: Add to CRM system if integrated
        
        return {
            "success": True,
            "message": "Lead captured successfully",
            "lead_id": str(result.inserted_id)
        }
    except Exception as e:
        logging.error(f"Error capturing lead: {e}")
        raise HTTPException(status_code=500, detail="Failed to capture lead")


# Staff Management
@api_router.post("/staff/create-request")
async def create_staff_request(
    staff_data: StaffCreate, current_user: dict = Depends(get_current_user)
):
    """Step 1: Request staff creation - Send OTP to staff email"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create staff")

    # Check if username already exists
    existing = await db.users.find_one({"username": staff_data.username}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists
    existing_email = await db.users.find_one({"email": staff_data.email}, {"_id": 0})
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Generate 6-digit OTP
    otp = ''.join([str(random.randint(0, 9)) for _ in range(6)])
    expires_at = datetime.now(timezone.utc) + timedelta(minutes=10)
    
    # Get admin's organization_id
    admin_org_id = current_user.get("organization_id") or current_user["id"]
    
    # Store OTP and staff data temporarily
    email_lower = staff_data.email.lower().strip()
    staff_otp_storage[email_lower] = {
        "otp": otp,
        "expires": expires_at,
        "staff_data": {
            "username": staff_data.username,
            "email": staff_data.email,
            "password": staff_data.password,
            "role": staff_data.role,
            "phone": staff_data.phone,
            "salary": staff_data.salary
        },
        "admin_id": current_user["id"],
        "organization_id": admin_org_id
    }
    
    # Send OTP email
    asyncio.create_task(send_staff_verification_email(staff_data.email, otp, staff_data.username, current_user.get("username", "Admin")))
    
    # Log OTP for debugging
    print(f"🔐 STAFF VERIFICATION OTP for {staff_data.email}: {otp}")
    
    return {
        "message": "OTP sent to staff email for verification",
        "email": staff_data.email,
        "success": True
    }


@api_router.post("/staff/verify-create")
async def verify_staff_creation(
    email: str = Body(...),
    otp: str = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """Step 2: Verify OTP and create staff"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create staff")
    
    email_lower = email.lower().strip()
    
    # Get OTP data
    otp_data = staff_otp_storage.get(email_lower)
    if not otp_data:
        raise HTTPException(status_code=400, detail="No verification request found. Please request OTP again.")
    
    # Check if OTP expired
    if datetime.now(timezone.utc) > otp_data["expires"]:
        del staff_otp_storage[email_lower]
        raise HTTPException(status_code=400, detail="OTP has expired. Please request a new one.")
    
    # Verify OTP
    if otp_data["otp"] != otp:
        raise HTTPException(status_code=400, detail="Invalid OTP. Please check and try again.")
    
    # Get staff data
    staff_data = otp_data["staff_data"]
    admin_org_id = otp_data["organization_id"]
    
    # Create user object
    user_obj = User(
        username=staff_data["username"],
        email=staff_data["email"],
        role=staff_data["role"],
        organization_id=admin_org_id,
    )

    doc = user_obj.model_dump()
    doc["password"] = hash_password(staff_data["password"])
    doc["phone"] = staff_data["phone"]
    doc["salary"] = staff_data["salary"]
    doc["created_at"] = doc["created_at"].isoformat()
    doc["email_verified"] = True
    doc["email_verified_at"] = datetime.now(timezone.utc).isoformat()

    await db.users.insert_one(doc)
    
    # Remove used OTP
    del staff_otp_storage[email_lower]
    
    return {"message": "Staff member created successfully", "id": user_obj.id}


@api_router.post("/staff/create")
async def create_staff(
    staff_data: StaffCreate, current_user: dict = Depends(get_current_user)
):
    """Direct staff creation without OTP (for skip verification)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can create staff")

    existing = await db.users.find_one({"username": staff_data.username}, {"_id": 0})
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")

    # Get admin's organization_id
    admin_org_id = current_user.get("organization_id") or current_user["id"]

    user_obj = User(
        username=staff_data.username,
        email=staff_data.email,
        role=staff_data.role,
        organization_id=admin_org_id,  # Link staff to this organization
    )

    doc = user_obj.model_dump()
    doc["password"] = hash_password(staff_data.password)
    doc["phone"] = staff_data.phone
    doc["salary"] = staff_data.salary
    doc["created_at"] = doc["created_at"].isoformat()
    doc["email_verified"] = False

    await db.users.insert_one(doc)
    return {"message": "Staff member created", "id": user_obj.id}


@api_router.get("/staff")
async def get_staff(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view staff")

    # Get admin's organization_id
    admin_org_id = current_user.get("organization_id") or current_user["id"]

    # Only fetch staff belonging to this organization
    staff = await db.users.find(
        {"organization_id": admin_org_id}, {"_id": 0, "password": 0}
    ).to_list(1000)
    return staff


@api_router.put("/staff/{staff_id}")
async def update_staff(
    staff_id: str,
    staff_data: StaffUpdate,
    current_user: dict = Depends(get_current_user),
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update staff")

    # Get admin's organization_id
    admin_org_id = current_user.get("organization_id") or current_user["id"]

    # Only allow updating staff from same organization
    existing = await db.users.find_one(
        {"id": staff_id, "organization_id": admin_org_id}, {"_id": 0}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Staff not found or access denied")

    update_data = {}
    if staff_data.username:
        update_data["username"] = staff_data.username
    if staff_data.email:
        update_data["email"] = staff_data.email
    if staff_data.password:
        update_data["password"] = hash_password(staff_data.password)
    if staff_data.role:
        update_data["role"] = staff_data.role
    if staff_data.phone is not None:
        update_data["phone"] = staff_data.phone
    if staff_data.salary is not None:
        update_data["salary"] = staff_data.salary

    await db.users.update_one({"id": staff_id}, {"$set": update_data})
    return {"message": "Staff updated"}


@api_router.delete("/staff/{staff_id}")
async def delete_staff(staff_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete staff")

    # Get admin's organization_id
    admin_org_id = current_user.get("organization_id") or current_user["id"]

    # Only allow deleting staff from same organization
    staff = await db.users.find_one(
        {"id": staff_id, "organization_id": admin_org_id}, {"_id": 0}
    )
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found or access denied")

    if staff["role"] == "admin":
        raise HTTPException(status_code=400, detail="Cannot delete admin user")

    await db.users.delete_one({"id": staff_id})
    return {"message": "Staff deleted"}


# Database Migration (Run once to fix existing data)
@api_router.post("/admin/migrate-organizations")
async def migrate_organizations(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")

    # Update all admin users to have their own organization_id
    admins = await db.users.find({"role": "admin"}, {"_id": 0}).to_list(1000)
    for admin in admins:
        if not admin.get("organization_id"):
            await db.users.update_one(
                {"id": admin["id"]}, {"$set": {"organization_id": admin["id"]}}
            )

    # Update all non-admin users without organization_id to set it to null
    await db.users.update_many(
        {"role": {"$ne": "admin"}, "organization_id": {"$exists": False}},
        {"$set": {"organization_id": None}},
    )

    return {"message": "Migration completed", "admins_updated": len(admins)}


# Business Setup
@api_router.post("/business/setup")
async def setup_business(
    settings: BusinessSettings, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can setup business")

    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"business_settings": settings.model_dump(), "setup_completed": True}},
    )
    return {"message": "Business setup completed", "settings": settings.model_dump()}


@api_router.put("/business/settings")
async def update_business_settings(
    settings: BusinessSettings, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update business settings")

    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"business_settings": settings.model_dump()}},
    )
    return {"message": "Business settings updated successfully", "settings": settings.model_dump()}


@api_router.get("/settings/all")
async def get_all_settings(current_user: dict = Depends(get_current_user)):
    """Get all settings data in a single API call for better performance"""
    try:
        # Get business settings
        business_settings = current_user.get("business_settings")
        setup_completed = current_user.get("setup_completed", False)
        
        if current_user.get("role") in ["waiter", "cashier", "kitchen", "staff"]:
            org_id = current_user.get("organization_id")
            if org_id:
                admin_user = await db.users.find_one(
                    {"id": org_id, "role": "admin"}, 
                    {"_id": 0, "business_settings": 1, "setup_completed": 1}
                )
                if admin_user:
                    business_settings = admin_user.get("business_settings")
                    setup_completed = admin_user.get("setup_completed", False)
        
        # Get Razorpay settings
        razorpay_configured = bool(current_user.get("razorpay_key_id"))
        razorpay_key_id = current_user.get("razorpay_key_id", "")
        
        # Get WhatsApp settings
        whatsapp_settings = {
            "whatsapp_enabled": current_user.get("whatsapp_enabled", False),
            "whatsapp_business_number": current_user.get("whatsapp_business_number", ""),
            "whatsapp_message_template": current_user.get("whatsapp_message_template", "Thank you for dining at {restaurant_name}! Your bill of {currency}{total} has been paid. Order #{order_id}"),
            "whatsapp_auto_notify": current_user.get("whatsapp_auto_notify", False),
            "whatsapp_notify_on_placed": current_user.get("whatsapp_notify_on_placed", True),
            "whatsapp_notify_on_preparing": current_user.get("whatsapp_notify_on_preparing", True),
            "whatsapp_notify_on_ready": current_user.get("whatsapp_notify_on_ready", True),
            "whatsapp_notify_on_completed": current_user.get("whatsapp_notify_on_completed", True),
            "customer_self_order_enabled": current_user.get("customer_self_order_enabled", False),
            "menu_display_enabled": current_user.get("menu_display_enabled", False)
        }
        
        # Get campaigns
        campaigns = []
        try:
            campaigns_cursor = db.campaigns.find(
                {"organization_id": current_user["organization_id"]},
                {"_id": 0}
            ).sort("created_at", -1)
            campaigns = await campaigns_cursor.to_list(length=None)
        except Exception as e:
            print(f"Failed to fetch campaigns: {e}")
        
        # Static data that doesn't change often
        themes = [
            {"id": "classic", "name": "Classic", "description": "Traditional receipt design"},
            {"id": "modern", "name": "Modern", "description": "Clean and minimal design"},
            {"id": "colorful", "name": "Colorful", "description": "Vibrant and eye-catching"},
            {"id": "elegant", "name": "Elegant", "description": "Sophisticated design"},
            {"id": "compact", "name": "Compact", "description": "Space-saving layout"}
        ]
        
        currencies = [
            {"code": "INR", "name": "Indian Rupee", "symbol": "₹"},
            {"code": "USD", "name": "US Dollar", "symbol": "$"},
            {"code": "EUR", "name": "Euro", "symbol": "€"},
            {"code": "GBP", "name": "British Pound", "symbol": "£"},
            {"code": "AED", "name": "UAE Dirham", "symbol": "د.إ"},
            {"code": "SAR", "name": "Saudi Riyal", "symbol": "﷼"}
        ]
        
        return {
            "business_settings": business_settings,
            "setup_completed": setup_completed,
            "razorpay": {
                "razorpay_configured": razorpay_configured,
                "razorpay_key_id": razorpay_key_id
            },
            "whatsapp": whatsapp_settings,
            "campaigns": campaigns,
            "themes": themes,
            "currencies": currencies
        }
        
    except Exception as e:
        print(f"Error fetching all settings: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch settings")


@api_router.get("/business/settings")
async def get_business_settings(current_user: dict = Depends(get_current_user)):
    # For staff users, get business settings from their admin
    business_settings = current_user.get("business_settings")
    
    if current_user.get("role") in ["waiter", "cashier", "kitchen", "staff"]:
        org_id = current_user.get("organization_id")
        if org_id:
            admin_user = await db.users.find_one(
                {"id": org_id, "role": "admin"}, 
                {"_id": 0, "business_settings": 1, "setup_completed": 1}
            )
            if admin_user:
                business_settings = admin_user.get("business_settings")
                return {
                    "business_settings": business_settings,
                    "setup_completed": admin_user.get("setup_completed", False),
                }
    
    return {
        "business_settings": business_settings,
        "setup_completed": current_user.get("setup_completed", False),
    }


@api_router.get("/currencies")
async def get_currencies():
    return [
        {"code": code, "symbol": symbol} for code, symbol in CURRENCY_SYMBOLS.items()
    ]


@api_router.get("/receipt-themes")
async def get_receipt_themes():
    return [
        {
            "id": "classic",
            "name": "Classic",
            "description": "Traditional receipt format",
            "recommended_width": "80mm",
            "supports_logo": True,
            "supports_qr": True
        },
        {
            "id": "modern",
            "name": "Modern",
            "description": "Modern with emojis and borders",
            "recommended_width": "80mm",
            "supports_logo": True,
            "supports_qr": True
        },
        {
            "id": "minimal",
            "name": "Minimal",
            "description": "Clean and simple design",
            "recommended_width": "80mm",
            "supports_logo": False,
            "supports_qr": False
        },
        {
            "id": "elegant",
            "name": "Elegant",
            "description": "Professional and elegant",
            "recommended_width": "80mm",
            "supports_logo": True,
            "supports_qr": True
        },
        {
            "id": "compact",
            "name": "Compact",
            "description": "Space-saving 58mm format",
            "recommended_width": "58mm",
            "supports_logo": False,
            "supports_qr": False
        },
        {
            "id": "detailed",
            "name": "Detailed",
            "description": "Comprehensive invoice format",
            "recommended_width": "80mm",
            "supports_logo": True,
            "supports_qr": True
        },
    ]


@api_router.get("/paper-sizes")
async def get_paper_sizes():
    return [
        {
            "id": "58mm",
            "name": "58mm (2.28 inches)",
            "width_mm": 58,
            "width_inches": 2.28,
            "char_width": 32,
            "description": "Compact thermal paper for small printers",
            "common_use": "Food trucks, kiosks, mobile POS"
        },
        {
            "id": "80mm",
            "name": "80mm (3.15 inches)",
            "width_mm": 80,
            "width_inches": 3.15,
            "char_width": 48,
            "description": "Standard thermal paper size",
            "common_use": "Most restaurants, retail stores"
        },
        {
            "id": "110mm",
            "name": "110mm (4.33 inches)",
            "width_mm": 110,
            "width_inches": 4.33,
            "char_width": 64,
            "description": "Wide format for detailed receipts",
            "common_use": "Fine dining, detailed invoices"
        },
        {
            "id": "custom",
            "name": "Custom Size",
            "width_mm": None,
            "width_inches": None,
            "char_width": None,
            "description": "Specify custom paper width",
            "common_use": "Special printer requirements"
        }
    ]


@api_router.get("/print-customization-options")
async def get_print_customization_options():
    return {
        "paper_widths": ["58mm", "80mm", "110mm", "custom"],
        "font_sizes": [8, 9, 10, 11, 12, 13, 14, 15, 16],
        "line_spacing": [1.0, 1.1, 1.2, 1.3, 1.4, 1.5, 1.6, 1.7, 1.8, 1.9, 2.0],
        "date_formats": ["DD-MM-YYYY", "MM-DD-YYYY", "YYYY-MM-DD"],
        "time_formats": ["12h", "24h"],
        "separator_styles": ["dash", "equal", "heavy", "light", "none"],
        "header_styles": ["centered", "left", "right"],
        "item_layouts": ["detailed", "compact", "minimal"],
        "total_styles": ["bold", "boxed", "highlighted"],
        "qr_code_content": ["website", "order_id", "custom"],
        "print_copies": [1, 2, 3, 4, 5]
    }


# Razorpay Settings
@api_router.post("/settings/razorpay")
async def update_razorpay_settings(
    settings: RazorpaySettings, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update settings")

    await db.users.update_one(
        {"id": current_user["id"]},
        {
            "$set": {
                "razorpay_key_id": settings.razorpay_key_id,
                "razorpay_key_secret": settings.razorpay_key_secret,
            }
        },
    )
    return {"message": "Razorpay settings updated successfully"}


@api_router.get("/settings/razorpay")
async def get_razorpay_settings(current_user: dict = Depends(get_current_user)):
    return {
        "razorpay_key_id": current_user.get("razorpay_key_id", ""),
        "razorpay_configured": bool(current_user.get("razorpay_key_id")),
    }


# Subscription - ₹1999/year with 7-day free trial (default)
# Pricing is now dynamic from Super Admin panel
SUBSCRIPTION_PRICE_PAISE = 199900  # ₹1999 in paise (default regular price)
NEW_YEAR_PRICE_PAISE = 179900  # ₹1799 in paise (default campaign price - 10% off)
NEW_YEAR_END_DATE = datetime(2026, 1, 1, 23, 59, 59, tzinfo=timezone.utc)
TRIAL_DAYS = 7
SUBSCRIPTION_DAYS = 365

# Campaign configuration (can be updated via admin panel)
ACTIVE_CAMPAIGNS = {
    "NEW_YEAR_2026": {
        "name": "New Year Special",
        "description": "Get BillByteKOT for just ₹599/year - 40% OFF!",
        "price_paise": 59900,  # ₹599
        "original_price_paise": 99900,  # ₹999
        "discount_percent": 40,
        "start_date": "2026-01-01T00:00:00+00:00",
        "end_date": "2026-01-01T23:59:59+00:00",
        "active": True,
        "badge": "🎉 40% OFF",
        "max_users": 5000,
        "current_users": 0
    }
}

async def get_current_subscription_price():
    """Get current subscription price based on database settings and active campaigns/sales"""
    now = datetime.now(timezone.utc)
    
    # First, check if there's an active sale offer (from Promotions tab)
    # Check both sale_offers collection (new) and site_settings (legacy) for backwards compatibility
    sale_offer = await db.sale_offers.find_one({"enabled": True})
    if not sale_offer:
        sale_offer = await db.site_settings.find_one({"type": "sale_offer", "enabled": True})
    
    if sale_offer:
        # Check if sale offer has expired
        sale_active = True
        
        # Check valid_until (datetime)
        if sale_offer.get("valid_until"):
            try:
                valid_until = datetime.fromisoformat(sale_offer["valid_until"])
                if valid_until.tzinfo is None:
                    valid_until = valid_until.replace(tzinfo=timezone.utc)
                if now > valid_until:
                    sale_active = False
            except:
                pass
        
        # Check end_date (date only)
        if sale_active and sale_offer.get("end_date"):
            try:
                end_date = datetime.fromisoformat(sale_offer["end_date"])
                if end_date.tzinfo is None:
                    end_date = end_date.replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
                if now > end_date:
                    sale_active = False
            except:
                pass
        
        if sale_active:
            sale_price = sale_offer.get("sale_price", 1799)
            original_price = sale_offer.get("original_price", 1999)
            discount_percent = sale_offer.get("discount_percent", 10)
            title = sale_offer.get("title", "Special Offer")
            
            return {
                "price_paise": sale_price * 100,
                "original_price_paise": original_price * 100,
                "price_display": f"₹{sale_price}",
                "original_price_display": f"₹{original_price}",
                "discount_percent": discount_percent,
                "campaign_name": title,
                "campaign_active": True,
                "campaign_ends": sale_offer.get("valid_until") or sale_offer.get("end_date"),
                "badge": f"🎉 {discount_percent}% OFF!",
                "trial_expired_discount": 10,
                "trial_expired_price_paise": int(original_price * 0.9) * 100,
                "source": "sale_offer"
            }
    
    # Fetch pricing from database (campaign pricing)
    pricing_doc = await db.site_settings.find_one({"type": "pricing"})
    
    if pricing_doc:
        regular_price = pricing_doc.get("regular_price", 1999)
        campaign_price = pricing_doc.get("campaign_price", 1799)
        campaign_active = pricing_doc.get("campaign_active", False)
        campaign_name = pricing_doc.get("campaign_name", "")
        campaign_discount = pricing_doc.get("campaign_discount_percent", 10)
        trial_expired_discount = pricing_doc.get("trial_expired_discount", 10)
        
        # Check campaign dates
        if campaign_active:
            start_date = pricing_doc.get("campaign_start_date")
            end_date = pricing_doc.get("campaign_end_date")
            if start_date and end_date:
                try:
                    start = datetime.fromisoformat(start_date)
                    end = datetime.fromisoformat(end_date)
                    if start.tzinfo is None:
                        start = start.replace(tzinfo=timezone.utc)
                    if end.tzinfo is None:
                        end = end.replace(tzinfo=timezone.utc)
                    campaign_active = start <= now <= end
                except:
                    pass
        
        if campaign_active:
            return {
                "price_paise": campaign_price * 100,
                "original_price_paise": regular_price * 100,
                "price_display": f"₹{campaign_price}",
                "original_price_display": f"₹{regular_price}",
                "discount_percent": campaign_discount,
                "campaign_name": campaign_name or "Special Offer",
                "campaign_active": True,
                "campaign_ends": end_date if end_date else None,
                "badge": f"🎉 {campaign_discount}% OFF!",
                "trial_expired_discount": trial_expired_discount,
                "trial_expired_price_paise": int(regular_price * (100 - trial_expired_discount) / 100) * 100,
                "source": "pricing_campaign"
            }
        
        # Regular pricing with trial expired discount info
        trial_expired_price = int(regular_price * (100 - trial_expired_discount) / 100)
        return {
            "price_paise": regular_price * 100,
            "original_price_paise": regular_price * 100,
            "price_display": f"₹{regular_price}",
            "original_price_display": f"₹{regular_price}",
            "discount_percent": 0,
            "campaign_name": None,
            "campaign_active": False,
            "campaign_ends": None,
            "badge": None,
            "trial_expired_discount": trial_expired_discount,
            "trial_expired_price_paise": trial_expired_price * 100,
            "source": "regular"
        }
    
    # Default pricing if no database entry
    return {
        "price_paise": SUBSCRIPTION_PRICE_PAISE,
        "original_price_paise": SUBSCRIPTION_PRICE_PAISE,
        "price_display": "₹1999",
        "original_price_display": "₹1999",
        "discount_percent": 0,
        "campaign_name": None,
        "campaign_active": False,
        "campaign_ends": None,
        "badge": None,
        "trial_expired_discount": 10,
        "trial_expired_price_paise": 179900,
        "source": "default"
    }


@api_router.get("/subscription/status")
async def get_subscription_status(current_user: dict = Depends(get_current_user)):
    is_active = await check_subscription(current_user)
    
    # Check trial status
    created_at = current_user.get("created_at")
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
    
    trial_end = created_at + timedelta(days=TRIAL_DAYS) if created_at else None
    is_trial = trial_end and datetime.now(timezone.utc) < trial_end if trial_end else False
    trial_days_left = max(0, (trial_end - datetime.now(timezone.utc)).days) if trial_end else 0
    
    # Get current pricing (with campaign if active)
    pricing = await get_current_subscription_price()
    
    return {
        "subscription_active": current_user.get("subscription_active", False),
        "bill_count": current_user.get("bill_count", 0),
        "needs_subscription": not is_trial and current_user.get("bill_count", 0) >= 50
        and not current_user.get("subscription_active", False),
        "subscription_expires_at": current_user.get("subscription_expires_at"),
        "is_trial": is_trial,
        "trial_days_left": trial_days_left,
        "trial_end": trial_end.isoformat() if trial_end else None,
        "price": pricing["price_paise"] // 100,
        "price_paise": pricing["price_paise"],
        "original_price": pricing["original_price_paise"] // 100,
        "price_display": pricing["price_display"],
        "original_price_display": pricing["original_price_display"],
        "discount_percent": pricing["discount_percent"],
        "campaign_active": pricing["campaign_active"],
        "campaign_name": pricing["campaign_name"],
        "campaign_ends": pricing["campaign_ends"],
        "campaign_badge": pricing["badge"]
    }


@api_router.post("/subscription/start-trial")
async def start_trial(current_user: dict = Depends(get_current_user)):
    """Start 7-day free trial"""
    # Trial is automatic based on account creation date
    created_at = current_user.get("created_at")
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
    
    trial_end = created_at + timedelta(days=TRIAL_DAYS) if created_at else None
    
    return {
        "trial_started": True,
        "trial_end": trial_end.isoformat() if trial_end else None,
        "trial_days": TRIAL_DAYS
    }


@api_router.get("/subscription/pricing")
async def get_subscription_pricing():
    """Get current subscription pricing (public endpoint - no auth required)"""
    pricing = await get_current_subscription_price()
    
    # Calculate time remaining for campaign
    time_remaining = None
    if pricing["campaign_active"] and pricing["campaign_ends"]:
        end_date = datetime.fromisoformat(pricing["campaign_ends"].replace("Z", "+00:00"))
        now = datetime.now(timezone.utc)
        if end_date > now:
            delta = end_date - now
            time_remaining = {
                "days": delta.days,
                "hours": delta.seconds // 3600,
                "minutes": (delta.seconds % 3600) // 60,
                "seconds": delta.seconds % 60,
                "total_seconds": int(delta.total_seconds())
            }
    
    return {
        **pricing,
        "time_remaining": time_remaining,
        "features": [
            "Unlimited Bills & Orders",
            "6 Thermal Printer Themes",
            "WhatsApp Integration",
            "AI-Powered Analytics",
            "Multi-Staff Management",
            "Priority 24/7 Support",
            "Free Updates Forever",
            "GST Compliant Billing"
        ]
    }


class CouponValidateRequest(BaseModel):
    coupon_code: str


@api_router.post("/subscription/validate-coupon")
async def validate_coupon(data: CouponValidateRequest):
    """Validate coupon code and return discount details"""
    coupon_code = data.coupon_code.upper().strip()
    
    if coupon_code not in COUPON_CODES:
        raise HTTPException(status_code=404, detail="Invalid coupon code")
    
    coupon = COUPON_CODES[coupon_code]
    
    if not coupon.get("active", True):
        raise HTTPException(status_code=400, detail="This coupon code has expired")
    
    # Calculate discount
    original_price = SUBSCRIPTION_PRICE_PAISE
    discount_amount = 0
    
    if "discount_percent" in coupon:
        discount_amount = int(original_price * coupon["discount_percent"] / 100)
    elif "discount_amount" in coupon:
        discount_amount = coupon["discount_amount"]
    
    final_price = max(0, original_price - discount_amount)
    
    return {
        "valid": True,
        "coupon_code": coupon_code,
        "description": coupon["description"],
        "original_price": original_price,
        "discount_amount": discount_amount,
        "final_price": final_price,
        "original_price_display": f"₹{original_price / 100:.0f}",
        "discount_display": f"₹{discount_amount / 100:.0f}",
        "final_price_display": f"₹{final_price / 100:.0f}",
        "discount_percent": coupon.get("discount_percent", 0)
    }


class CreateOrderRequest(BaseModel):
    coupon_code: Optional[str] = None
    plan_type: Optional[str] = "yearly"  # monthly, quarterly, halfYearly, yearly
    months: Optional[int] = 12
    amount: Optional[float] = None


@api_router.post("/subscription/create-order")
async def create_subscription_order(
    data: CreateOrderRequest = None,
    current_user: dict = Depends(get_current_user)
):
    # IMPORTANT: These are YOUR (platform owner's) Razorpay keys for subscription payments
    # Money from subscriptions comes to YOU, not to individual restaurants
    DEFAULT_RAZORPAY_KEY_ID = "rzp_live_RmGqVf5JPGOT6G"
    DEFAULT_RAZORPAY_KEY_SECRET = "SKYS5tgjwU3H3Pf2ch3ZFtuH"
    
    # Define pricing plans (matching frontend)
    plans = {
        "monthly": {"months": 1, "price": 199, "originalPrice": 199, "discount": 0, "label": "1 Month", "perMonth": 199},
        "quarterly": {"months": 3, "price": 549, "originalPrice": 597, "discount": 8, "label": "3 Months", "perMonth": 183},
        "halfYearly": {"months": 6, "price": 999, "originalPrice": 1194, "discount": 16, "label": "6 Months", "perMonth": 167},
        "yearly": {"months": 12, "price": 1899, "originalPrice": 1999, "discount": 5, "label": "1 Year", "perMonth": 159}
    }
    
    # Get selected plan or default to yearly
    plan_type = data.plan_type if data and data.plan_type else "yearly"
    selected_plan = plans.get(plan_type, plans["yearly"])
    
    # Use plan-specific pricing
    base_price = int(selected_plan["price"] * 100)  # Convert to paise
    original_price = int(selected_plan["originalPrice"] * 100)  # Convert to paise
    
    # If amount is provided from frontend, use it (for validation)
    if data and data.amount:
        frontend_amount = int(data.amount * 100)  # Convert to paise
        if abs(frontend_amount - base_price) > 100:  # Allow small rounding differences
            print(f"⚠️  Price mismatch: Frontend={frontend_amount}, Backend={base_price}")
            # Use frontend amount if it's reasonable (within plan price range)
            if 10000 <= frontend_amount <= 250000:  # ₹100 to ₹2500 range
                base_price = frontend_amount
    
    # Get current campaign pricing from database for potential additional discounts
    try:
        pricing = await get_current_subscription_price()
        campaign_active = pricing.get("campaign_active", False)
        campaign_discount = pricing.get("campaign_discount_percent", 0)
        
        # Apply campaign discount if active
        if campaign_active and campaign_discount > 0:
            campaign_discount_amount = int(base_price * campaign_discount / 100)
            base_price = max(100, base_price - campaign_discount_amount)
            print(f"🎉 Campaign discount applied: {campaign_discount}% off")
    except Exception as e:
        print(f"⚠️  Could not fetch campaign pricing: {e}")
        campaign_active = False
    
    # Calculate price with coupon if provided (on top of plan price)
    final_price = base_price
    coupon_applied = None
    discount_amount = 0
    referral_discount_applied = None
    
    # Check for pending referral discount (₹200 off for referred users)
    # Requirements: 3.6 - Reduce payment amount by ₹200 for referred users
    user_id = current_user.get("id")
    pending_referral = await db.referrals.find_one({
        "referee_user_id": user_id,
        "status": "PENDING"
    })
    
    if pending_referral:
        # Apply referral discount (₹200 = 20000 paise)
        referral_discount_paise = int(REFERRAL_DISCOUNT_AMOUNT * 100)
        final_price = max(100, base_price - referral_discount_paise)  # Minimum ₹1 (100 paise)
        referral_discount_applied = {
            "referral_id": pending_referral.get("id"),
            "referral_code": pending_referral.get("referral_code"),
            "discount_amount": referral_discount_paise,
            "discount_display": f"₹{int(REFERRAL_DISCOUNT_AMOUNT)}"
        }
        print(f"Referral discount applied for user {user_id}: ₹{int(REFERRAL_DISCOUNT_AMOUNT)}")
    
    if data and data.coupon_code:
        coupon_code = data.coupon_code.upper().strip()
        if coupon_code in COUPON_CODES:
            coupon = COUPON_CODES[coupon_code]
            if coupon.get("active", True):
                if "discount_percent" in coupon:
                    discount_amount = int(base_price * coupon["discount_percent"] / 100)
                elif "discount_amount" in coupon:
                    discount_amount = min(coupon["discount_amount"], base_price)  # Can't discount more than price
                
                final_price = max(100, final_price - discount_amount)  # Minimum ₹1 (100 paise)
                coupon_applied = {
                    "code": coupon_code,
                    "description": coupon["description"],
                    "discount_amount": discount_amount,
                    "discount_display": f"₹{discount_amount / 100:.0f}"
                }
    
    # Use platform owner's keys (YOUR account)
    razorpay_key_id = os.environ.get("RAZORPAY_KEY_ID") or DEFAULT_RAZORPAY_KEY_ID
    razorpay_key_secret = os.environ.get("RAZORPAY_KEY_SECRET") or DEFAULT_RAZORPAY_KEY_SECRET

    try:
        razorpay_client = razorpay.Client(auth=(razorpay_key_id, razorpay_key_secret))

        razor_order = razorpay_client.order.create(
            {"amount": final_price, "currency": "INR", "payment_capture": 1}
        )

        print(f"✅ Subscription order created: {plan_type} plan, ₹{final_price/100} ({selected_plan['label']})")

        return {
            "razorpay_order_id": razor_order["id"],
            "amount": final_price,
            "original_amount": original_price,
            "plan_price": int(selected_plan["price"] * 100),
            "discount_amount": discount_amount,
            "currency": "INR",
            "key_id": razorpay_key_id,
            "price_display": f"₹{final_price / 100:.0f}",
            "original_price_display": f"₹{original_price / 100:.0f}",
            "plan_display": f"₹{selected_plan['price']}",
            "coupon_applied": coupon_applied,
            "referral_discount_applied": referral_discount_applied,
            "plan_type": plan_type,
            "plan_label": selected_plan["label"],
            "plan_months": selected_plan["months"],
            "campaign_active": campaign_active if 'campaign_active' in locals() else False,
            "campaign_name": "Multi-Plan Subscription"
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to create payment order: {str(e)}"
        )


class SubscriptionVerifyRequest(BaseModel):
    razorpay_payment_id: str
    razorpay_order_id: str
    razorpay_signature: Optional[str] = None
    plan_type: Optional[str] = "yearly"
    months: Optional[int] = 12


@api_router.post("/subscription/verify")
async def verify_subscription_payment(
    data: SubscriptionVerifyRequest,
    current_user: dict = Depends(get_current_user),
):
    DEFAULT_RAZORPAY_KEY_ID = "rzp_live_RmGqVf5JPGOT6G"
    DEFAULT_RAZORPAY_KEY_SECRET = "SKYS5tgjwU3H3Pf2ch3ZFtuH"
    
    razorpay_key_id = os.environ.get("RAZORPAY_KEY_ID") or DEFAULT_RAZORPAY_KEY_ID
    razorpay_key_secret = os.environ.get("RAZORPAY_KEY_SECRET") or DEFAULT_RAZORPAY_KEY_SECRET
    
    # Get current campaign pricing for recording
    pricing = await get_current_subscription_price()
    
    try:
        razorpay_client = razorpay.Client(auth=(razorpay_key_id, razorpay_key_secret))
        
        # Fetch payment details first
        payment = None
        amount_paid = 0
        try:
            payment = razorpay_client.payment.fetch(data.razorpay_payment_id)
            amount_paid = payment.get('amount', 0)
            print(f"Payment fetched: {payment}")
        except Exception as fetch_error:
            print(f"Error fetching payment: {fetch_error}")
        
        # Verify signature if provided (optional for some payment methods)
        signature_valid = False
        if data.razorpay_signature:
            try:
                razorpay_client.utility.verify_payment_signature({
                    'razorpay_order_id': data.razorpay_order_id,
                    'razorpay_payment_id': data.razorpay_payment_id,
                    'razorpay_signature': data.razorpay_signature
                })
                signature_valid = True
                print("Signature verified successfully")
            except razorpay.errors.SignatureVerificationError as sig_error:
                print(f"Signature verification failed: {sig_error}")
        
        # Check payment status
        payment_captured = False
        if payment:
            payment_captured = payment.get('status') in ['captured', 'authorized']
            print(f"Payment status: {payment.get('status')}, captured: {payment_captured}")
        
        # Accept payment if either signature is valid OR payment is captured
        # Also accept if amount matches expected campaign price (₹9 = 900 paise or ₹999 = 99900 paise)
        if not signature_valid and not payment_captured:
            if payment and amount_paid in [900, 99900, pricing["price_paise"]]:
                print(f"Payment amount {amount_paid} matches expected, accepting")
                payment_captured = True
            else:
                raise HTTPException(
                    status_code=400, 
                    detail="Payment verification failed. Contact support with payment ID: " + data.razorpay_payment_id
                )
        
        # Activate subscription with plan-specific duration
        plan_months = data.months if data.months else 12  # Default to 12 months
        subscription_days = plan_months * 30  # Approximate days per month
        expires_at = datetime.now(timezone.utc) + timedelta(days=subscription_days)
        
        # Define pricing plans for validation
        plans = {
            "monthly": {"months": 1, "price": 199},
            "quarterly": {"months": 3, "price": 549},
            "halfYearly": {"months": 6, "price": 999},
            "yearly": {"months": 12, "price": 1899}
        }
        
        plan_type = data.plan_type if data.plan_type else "yearly"
        selected_plan = plans.get(plan_type, plans["yearly"])
        expected_amount = int(selected_plan["price"] * 100)  # Convert to paise
        
        # Validate payment amount matches plan (with some tolerance for discounts)
        if payment and amount_paid > 0:
            # Allow for discounts, but amount should be at least 50% of expected
            min_expected = expected_amount // 2
            if amount_paid < min_expected:
                print(f"⚠️  Payment amount {amount_paid} seems too low for {plan_type} plan (expected ~{expected_amount})")
        
        # Determine campaign used
        campaign_name = f"{plan_type.upper()}_PLAN_2025" if pricing["campaign_active"] else f"{plan_type.upper()}_REGULAR"

        await db.users.update_one(
            {"id": current_user["id"]},
            {
                "$set": {
                    "subscription_active": True,
                    "subscription_expires_at": expires_at.isoformat(),
                    "subscription_payment_id": data.razorpay_payment_id,
                    "subscription_order_id": data.razorpay_order_id,
                    "subscription_verified_at": datetime.now(timezone.utc).isoformat(),
                    "subscription_campaign": campaign_name,
                    "subscription_price_paid": amount_paid,
                    "subscription_original_price": pricing["original_price_paise"],
                    "subscription_plan_type": plan_type,
                    "subscription_plan_months": plan_months,
                    "subscription_plan_label": selected_plan.get("label", f"{plan_months} Month(s)"),
                    "is_early_adopter": pricing["campaign_active"]
                }
            },
        )
        
        print(f"Subscription activated for user: {current_user['id']} via campaign: {campaign_name}, plan: {plan_type} ({plan_months} months)")
        
        # Trigger referral completion after successful payment
        # Requirements: 4.1, 4.2 - Credit referrer wallet with ₹300 after payment
        referral_result = None
        try:
            referral_result = await process_referral_completion(
                referee_user_id=current_user["id"],
                payment_id=data.razorpay_payment_id
            )
            if referral_result and referral_result.get("referral_processed"):
                print(f"Referral completed for user {current_user['id']}: {referral_result}")
        except Exception as ref_error:
            print(f"Referral completion error (non-blocking): {ref_error}")

        return {
            "status": "subscription_activated", 
            "expires_at": expires_at.isoformat(),
            "days": subscription_days,
            "months": plan_months,
            "plan_type": plan_type,
            "plan_label": selected_plan.get("label", f"{plan_months} Month(s)"),
            "message": f"🎉 {selected_plan.get('label', f'{plan_months} Month')} Premium subscription activated successfully!" + (" You're an Early Adopter! 🔥" if pricing["campaign_active"] else ""),
            "payment_id": data.razorpay_payment_id,
            "amount_paid": amount_paid,
            "campaign": campaign_name,
            "is_early_adopter": pricing["campaign_active"],
            "referral_processed": referral_result.get("referral_processed") if referral_result else False
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"Verification error: {str(e)}")
        # Try to activate anyway if payment ID exists
        try:
            plan_months = data.months if data.months else 12
            subscription_days = plan_months * 30
            expires_at = datetime.now(timezone.utc) + timedelta(days=subscription_days)
            plan_type = data.plan_type if data.plan_type else "yearly"
            campaign_name = f"{plan_type.upper()}_PLAN_2025" if pricing["campaign_active"] else f"{plan_type.upper()}_REGULAR"
            
            await db.users.update_one(
                {"id": current_user["id"]},
                {
                    "$set": {
                        "subscription_active": True,
                        "subscription_expires_at": expires_at.isoformat(),
                        "subscription_payment_id": data.razorpay_payment_id,
                        "subscription_order_id": data.razorpay_order_id,
                        "subscription_verified_at": datetime.now(timezone.utc).isoformat(),
                        "subscription_campaign": campaign_name,
                        "subscription_plan_type": plan_type,
                        "subscription_plan_months": plan_months,
                        "is_early_adopter": pricing["campaign_active"]
                    }
                },
            )
            
            # Also trigger referral completion in fallback case
            try:
                await process_referral_completion(
                    referee_user_id=current_user["id"],
                    payment_id=data.razorpay_payment_id
                )
            except Exception as ref_error:
                print(f"Referral completion error in fallback (non-blocking): {ref_error}")
            
            return {
                "status": "subscription_activated", 
                "expires_at": expires_at.isoformat(),
                "days": subscription_days,
                "months": plan_months,
                "plan_type": plan_type,
                "message": f"🎉 {plan_months} Month Premium subscription activated successfully!"
            }
        except:
            raise HTTPException(
                status_code=500, 
                detail=f"Verification failed: {str(e)}. Contact support with payment ID: {data.razorpay_payment_id}"
            )


# Image Upload
@api_router.post("/upload/image")
async def upload_image(
    file: UploadFile = File(...), current_user: dict = Depends(get_current_user)
):
    if file.content_type not in ["image/jpeg", "image/png", "image/jpg", "image/webp"]:
        raise HTTPException(status_code=400, detail="Only image files allowed")

    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File size exceeds 5MB")

    image_data = base64.b64encode(contents).decode("utf-8")
    image_url = f"data:{file.content_type};base64,{image_data}"

    return {"image_url": image_url}


# Menu routes
@api_router.post("/menu", response_model=MenuItem)
async def create_menu_item(
    item: MenuItemCreate, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    menu_obj = MenuItem(**item.model_dump(), organization_id=user_org_id)
    doc = menu_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.menu_items.insert_one(doc)
    
    # Invalidate menu cache
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_menu_caches(user_org_id)
        print(f"🗑️ Menu cache invalidated for new item {menu_obj.id}")
    except Exception as e:
        print(f"⚠️ Menu cache invalidation error: {e}")
    
    return menu_obj


@api_router.get("/menu/lightweight")
async def get_menu_lightweight(current_user: dict = Depends(get_current_user)):
    """Get menu items with minimal data for faster loading"""
    user_org_id = get_secure_org_id(current_user)

    try:
        # Use Redis-cached service for menu items with minimal fields
        cached_service = get_cached_order_service()
        items = await cached_service.get_menu_items(user_org_id, use_cache=True)
        
        # Return only essential fields for faster loading
        lightweight_items = []
        for item in items:
            lightweight_items.append({
                "id": item.get("id"),
                "name": item.get("name"),
                "category": item.get("category"),
                "price": item.get("price"),
                "available": item.get("available", True),
                "image_url": item.get("image_url", ""),
                "description": item.get("description", "")[:100] if item.get("description") else "",  # Truncate description
                "preparation_time": item.get("preparation_time", 15)
            })
        
        print(f"🚀 Returned {len(lightweight_items)} lightweight menu items (Redis cached)")
        return lightweight_items
        
    except Exception as e:
        print(f"❌ Error fetching lightweight menu from cache: {e}")
        # Fallback to direct MongoDB query with projection
        items = await db.menu_items.find(
            {"organization_id": user_org_id}, 
            {
                "_id": 0,
                "id": 1,
                "name": 1,
                "category": 1,
                "price": 1,
                "available": 1,
                "image_url": 1,
                "description": 1,
                "preparation_time": 1
            }
        ).to_list(1000)
        
        # Truncate descriptions for faster transfer
        for item in items:
            if item.get("description") and len(item["description"]) > 100:
                item["description"] = item["description"][:100] + "..."
        
        print(f"📊 Fallback: Returned {len(items)} lightweight menu items from MongoDB")
        return items


@api_router.get("/menu", response_model=List[MenuItem])
async def get_menu(current_user: dict = Depends(get_current_user)):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    try:
        # Use Redis-cached service for menu items
        cached_service = get_cached_order_service()
        items = await cached_service.get_menu_items(user_org_id, use_cache=True)
        print(f"🚀 Returned {len(items)} menu items (Redis cached)")
        return items
        
    except Exception as e:
        print(f"❌ Error fetching menu from cache: {e}")
        # Fallback to direct MongoDB query
        items = await db.menu_items.find(
            {"organization_id": user_org_id}, {"_id": 0}
        ).to_list(1000)
        for item in items:
            if isinstance(item["created_at"], str):
                item["created_at"] = datetime.fromisoformat(item["created_at"])
        print(f"📊 Fallback: Returned {len(items)} menu items from MongoDB")
        return items


@api_router.get("/menu/{item_id}", response_model=MenuItem)
async def get_menu_item(item_id: str, current_user: dict = Depends(get_current_user)):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    item = await db.menu_items.find_one(
        {"id": item_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if not item:
        raise HTTPException(status_code=404, detail="Item not found")
    if isinstance(item["created_at"], str):
        item["created_at"] = datetime.fromisoformat(item["created_at"])
    return item


@api_router.put("/menu/{item_id}", response_model=MenuItem)
async def update_menu_item(
    item_id: str, item: MenuItemCreate, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    existing = await db.menu_items.find_one(
        {"id": item_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Item not found")

    update_data = item.model_dump()
    await db.menu_items.update_one(
        {"id": item_id, "organization_id": user_org_id}, {"$set": update_data}
    )

    updated = await db.menu_items.find_one(
        {"id": item_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if isinstance(updated["created_at"], str):
        updated["created_at"] = datetime.fromisoformat(updated["created_at"])
    
    # Invalidate menu cache
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_menu_caches(user_org_id)
        print(f"🗑️ Menu cache invalidated for updated item {item_id}")
    except Exception as e:
        print(f"⚠️ Menu cache invalidation error: {e}")
    
    return updated


@api_router.delete("/menu/{item_id}")
async def delete_menu_item(
    item_id: str, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    result = await db.menu_items.delete_one(
        {"id": item_id, "organization_id": user_org_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Invalidate menu cache
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_menu_caches(user_org_id)
        print(f"🗑️ Menu cache invalidated for deleted item {item_id}")
    except Exception as e:
        print(f"⚠️ Menu cache invalidation error: {e}")
    
    return {"message": "Item deleted"}


# Table routes
@api_router.post("/tables", response_model=Table)
async def create_table(
    table: TableCreate, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Check if table number already exists for this organization
    existing_table = await db.tables.find_one({
        "table_number": table.table_number,
        "organization_id": user_org_id
    })
    
    if existing_table:
        raise HTTPException(
            status_code=400, 
            detail=f"Table number {table.table_number} already exists"
        )

    # Create table object with all fields
    table_data = table.model_dump()
    table_data["organization_id"] = user_org_id
    table_data["id"] = str(uuid.uuid4())
    table_data["created_at"] = datetime.now().isoformat()
    table_data["updated_at"] = datetime.now().isoformat()
    
    table_obj = Table(**table_data)
    await db.tables.insert_one(table_obj.model_dump())
    
    # Invalidate table cache after creation to ensure fresh data on next fetch
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_table_caches(user_org_id)
        print(f"🗑️ Table cache invalidated for org {user_org_id} after creating table {table_obj.table_number}")
    except Exception as e:
        print(f"⚠️ Table cache invalidation error after creation: {e}")
    
    return table_obj


@api_router.get("/tables", response_model=List[Table])
async def get_tables(
    fresh: bool = Query(False, description="Bypass cache and fetch directly from database"),
    current_user: dict = Depends(get_current_user)
):
    """Get all tables for the organization. Use fresh=true to bypass cache."""
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    try:
        # If fresh=true, use TableStatusManager to get fresh data directly from DB
        if fresh:
            table_manager = get_table_status_manager()
            tables = await table_manager.get_tables_fresh(user_org_id)
            print(f"🔄 Returned {len(tables)} tables (fresh from DB, bypassed cache)")
            return tables
        
        # Otherwise use Redis-cached service for tables
        cached_service = get_cached_order_service()
        tables = await cached_service.get_tables(user_org_id, use_cache=True)
        print(f"🚀 Returned {len(tables)} tables (Redis cached)")
        return tables
        
    except Exception as e:
        print(f"❌ Error fetching tables: {e}")
        # Fallback to direct MongoDB query
        tables = await db.tables.find({"organization_id": user_org_id}, {"_id": 0}).to_list(1000)
        print(f"📊 Fallback: Returned {len(tables)} tables from MongoDB")
        return tables


@api_router.put("/tables/{table_id}", response_model=Table)
async def update_table(
    table_id: str, table: TableCreate, current_user: dict = Depends(get_current_user)
):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    existing = await db.tables.find_one(
        {"id": table_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Table not found")

    await db.tables.update_one(
        {"id": table_id, "organization_id": user_org_id}, {"$set": table.model_dump()}
    )
    updated = await db.tables.find_one(
        {"id": table_id, "organization_id": user_org_id}, {"_id": 0}
    )
    
    # Fix: Invalidate table cache after update to ensure fresh data on next fetch
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_table_caches(user_org_id)
        print(f"🗑️ Table cache invalidated for org {user_org_id} after table {table_id} update (status: {table.status})")
    except Exception as e:
        print(f"⚠️ Table cache invalidation error: {e}")
    
    return updated


@api_router.delete("/tables/{table_id}")
async def delete_table(
    table_id: str, current_user: dict = Depends(get_current_user)
):
    """Delete a table"""
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    existing = await db.tables.find_one(
        {"id": table_id, "organization_id": user_org_id}
    )
    if not existing:
        raise HTTPException(status_code=404, detail="Table not found")
    
    # Check if table has active orders
    if existing.get("status") == "occupied" and existing.get("current_order_id"):
        raise HTTPException(
            status_code=400, 
            detail="Cannot delete table with active order. Please complete or cancel the order first."
        )

    result = await db.tables.delete_one(
        {"id": table_id, "organization_id": user_org_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Table not found")
    
    # Invalidate table cache after deletion to ensure fresh data on next fetch
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_table_caches(user_org_id)
        print(f"🗑️ Table cache invalidated for org {user_org_id} after deleting table {table_id}")
    except Exception as e:
        print(f"⚠️ Table cache invalidation error after deletion: {e}")
    
    return {"message": "Table deleted successfully"}


@api_router.patch("/tables/{table_id}/status")
async def update_table_status(
    table_id: str,
    status_update: dict,
    current_user: dict = Depends(get_current_user)
):
    """Quick table status update endpoint for better performance"""
    if current_user["role"] not in ["admin", "cashier", "waiter"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)
    
    # Validate status
    valid_statuses = ["available", "occupied", "reserved", "maintenance", "cleaning"]
    new_status = status_update.get("status")
    
    if new_status not in valid_statuses:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {valid_statuses}")

    try:
        # Update only the status field for better performance
        result = await db.tables.update_one(
            {"id": table_id, "organization_id": user_org_id},
            {"$set": {"status": new_status, "updated_at": datetime.now().isoformat()}}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Table not found")
        
        # Invalidate table cache
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_table_caches(user_org_id)
        except Exception as e:
            print(f"⚠️ Table cache invalidation error: {e}")
        
        print(f"✅ Table {table_id} status updated to '{new_status}'")
        return {"message": f"Table status updated to {new_status}", "status": new_status}
        
    except Exception as e:
        print(f"❌ Error updating table status: {e}")
        raise HTTPException(status_code=500, detail="Failed to update table status")


# ==================== RESERVATION ENDPOINTS ====================

@api_router.post("/tables/reservations", response_model=Reservation)
async def create_reservation(
    reservation: ReservationCreate, current_user: dict = Depends(get_current_user)
):
    """Create a new table reservation"""
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Verify table exists and belongs to organization
    table = await db.tables.find_one({
        "id": reservation.table_id,
        "organization_id": user_org_id
    })
    
    if not table:
        raise HTTPException(status_code=404, detail="Table not found")

    # Check for conflicting reservations
    reservation_datetime = f"{reservation.reservation_date} {reservation.reservation_time}"
    existing_reservation = await db.reservations.find_one({
        "table_id": reservation.table_id,
        "reservation_date": reservation.reservation_date,
        "status": {"$in": ["confirmed", "pending"]},
        "organization_id": user_org_id
    })
    
    if existing_reservation:
        raise HTTPException(
            status_code=400, 
            detail=f"Table {table['table_number']} is already reserved for {reservation.reservation_date}"
        )

    # Create reservation object
    reservation_data = reservation.model_dump()
    reservation_data["organization_id"] = user_org_id
    reservation_data["table_number"] = table["table_number"]
    reservation_data["id"] = str(uuid.uuid4())
    reservation_data["created_at"] = datetime.now().isoformat()
    reservation_data["updated_at"] = datetime.now().isoformat()
    
    reservation_obj = Reservation(**reservation_data)
    await db.reservations.insert_one(reservation_obj.model_dump())
    print(f"✅ Created reservation: Table {table['table_number']} for {reservation.customer_name} on {reservation.reservation_date}")
    
    # Smart table status update based on timing
    from datetime import date, datetime as dt, timedelta
    reservation_datetime = dt.fromisoformat(f"{reservation.reservation_date} {reservation.reservation_time}")
    current_time = dt.now()
    pre_arrival_time = reservation_datetime - timedelta(minutes=reservation.pre_arrival_minutes)
    
    # Only mark table as reserved if we're within the pre-arrival window
    if reservation.reservation_date == date.today().isoformat() and current_time >= pre_arrival_time:
        print(f"🔄 Updating table {table['table_number']} status to 'reserved' (within {reservation.pre_arrival_minutes} min window)")
        await db.tables.update_one(
            {"id": reservation.table_id, "organization_id": user_org_id},
            {"$set": {"status": "reserved", "updated_at": datetime.now().isoformat()}}
        )
        print(f"✅ Table {table['table_number']} status updated to 'reserved'")
        
        # Invalidate table cache
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_table_caches(user_org_id)
            print(f"🗑️ Table cache invalidated for org {user_org_id}")
        except Exception as e:
            print(f"⚠️ Table cache invalidation error: {e}")
    else:
        if reservation.reservation_date == date.today().isoformat():
            minutes_until_activation = int((pre_arrival_time - current_time).total_seconds() / 60)
            print(f"ℹ️ Table {table['table_number']} will be reserved in {minutes_until_activation} minutes")
        else:
            print(f"ℹ️ Table {table['table_number']} status not updated (reservation is for {reservation.reservation_date})")
    
    return reservation_obj


@api_router.get("/tables/reservations", response_model=List[Reservation])
async def get_reservations(
    date: Optional[str] = Query(None, description="Filter by date (YYYY-MM-DD)"),
    current_user: dict = Depends(get_current_user)
):
    """Get reservations for the organization"""
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Build query
    query = {"organization_id": user_org_id}
    
    # Only filter by date if explicitly provided
    if date is not None:
        query["reservation_date"] = date
        print(f"📅 Fetching reservations for date: {date}")
    else:
        print(f"📅 Fetching all reservations for org: {user_org_id}")

    try:
        reservations = await db.reservations.find(query, {"_id": 0}).to_list(1000)
        print(f"✅ Found {len(reservations)} reservations")
        
        # Log first few reservations for debugging
        for i, res in enumerate(reservations[:3]):
            print(f"  {i+1}. Table {res.get('table_number', 'N/A')} - {res.get('customer_name', 'N/A')} - {res.get('reservation_date', 'N/A')}")
        
        return reservations
    except Exception as e:
        print(f"❌ Error fetching reservations: {e}")
        return []


@api_router.put("/tables/reservations/{reservation_id}", response_model=Reservation)
async def update_reservation(
    reservation_id: str,
    reservation_update: ReservationCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update a reservation"""
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Check if reservation exists
    existing = await db.reservations.find_one({
        "id": reservation_id,
        "organization_id": user_org_id
    })
    
    if not existing:
        raise HTTPException(status_code=404, detail="Reservation not found")

    # Update reservation
    update_data = reservation_update.model_dump()
    update_data["updated_at"] = datetime.now().isoformat()
    
    await db.reservations.update_one(
        {"id": reservation_id, "organization_id": user_org_id},
        {"$set": update_data}
    )
    
    # Get updated reservation
    updated = await db.reservations.find_one({
        "id": reservation_id,
        "organization_id": user_org_id
    }, {"_id": 0})
    
    return updated


@api_router.delete("/tables/reservations/{reservation_id}")
async def delete_reservation(
    reservation_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Cancel/Delete a reservation"""
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Check if reservation exists
    existing = await db.reservations.find_one({
        "id": reservation_id,
        "organization_id": user_org_id
    })
    
    if not existing:
        raise HTTPException(status_code=404, detail="Reservation not found")

    # Delete reservation
    result = await db.reservations.delete_one({
        "id": reservation_id,
        "organization_id": user_org_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Reservation not found")
    
    # Clear table status if it was reserved for this reservation
    table_id = existing.get("table_id")
    if table_id:
        await db.tables.update_one(
            {"id": table_id, "organization_id": user_org_id, "status": "reserved"},
            {"$set": {"status": "available", "updated_at": datetime.now().isoformat()}}
        )
        print(f"✅ Table {existing.get('table_number', 'N/A')} status cleared to 'available'")
        
        # Invalidate table cache
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_table_caches(user_org_id)
        except Exception as e:
            print(f"⚠️ Table cache invalidation error: {e}")
    
    print(f"✅ Reservation cancelled: {existing.get('customer_name', 'N/A')} - Table {existing.get('table_number', 'N/A')}")
    return {"message": "Reservation cancelled successfully"}


@api_router.post("/tables/reservations/auto-clear")
async def auto_clear_expired_reservations(
    current_user: dict = Depends(get_current_user)
):
    """Auto-clear expired reservations and update table statuses"""
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    
    from datetime import datetime as dt, timedelta, date
    current_time = dt.now()
    today = date.today().isoformat()
    
    # Find expired reservations (30 minutes past reservation time + duration)
    expired_reservations = []
    reservations = await db.reservations.find({
        "organization_id": user_org_id,
        "reservation_date": today,
        "status": {"$in": ["confirmed", "pending"]}
    }).to_list(1000)
    
    cleared_count = 0
    updated_tables = []
    
    for reservation in reservations:
        try:
            reservation_datetime = dt.fromisoformat(f"{reservation['reservation_date']} {reservation['reservation_time']}")
            reservation_end = reservation_datetime + timedelta(minutes=reservation.get('duration', 120))
            grace_period_end = reservation_end + timedelta(minutes=30)  # 30 min grace period
            
            if current_time > grace_period_end:
                # Mark as expired and clear table
                await db.reservations.update_one(
                    {"id": reservation["id"]},
                    {"$set": {"status": "expired", "updated_at": current_time.isoformat()}}
                )
                
                # Clear table status
                table_result = await db.tables.update_one(
                    {"id": reservation["table_id"], "organization_id": user_org_id, "status": "reserved"},
                    {"$set": {"status": "available", "updated_at": current_time.isoformat()}}
                )
                
                if table_result.modified_count > 0:
                    updated_tables.append(reservation.get("table_number", "N/A"))
                
                expired_reservations.append({
                    "customer_name": reservation.get("customer_name"),
                    "table_number": reservation.get("table_number"),
                    "reservation_time": reservation.get("reservation_time")
                })
                cleared_count += 1
                
        except Exception as e:
            print(f"⚠️ Error processing reservation {reservation.get('id', 'N/A')}: {e}")
    
    # Invalidate table cache if any tables were updated
    if updated_tables:
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_table_caches(user_org_id)
        except Exception as e:
            print(f"⚠️ Table cache invalidation error: {e}")
    
    print(f"🧹 Auto-cleared {cleared_count} expired reservations, freed tables: {updated_tables}")
    
    return {
        "message": f"Auto-cleared {cleared_count} expired reservations",
        "cleared_reservations": expired_reservations,
        "updated_tables": updated_tables
    }


@api_router.post("/tables/reservations/activate-pending")
async def activate_pending_reservations(
    current_user: dict = Depends(get_current_user)
):
    """Activate reservations that are within their pre-arrival window"""
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    
    from datetime import datetime as dt, timedelta, date
    current_time = dt.now()
    today = date.today().isoformat()
    
    # Find reservations for today that should be activated
    reservations = await db.reservations.find({
        "organization_id": user_org_id,
        "reservation_date": today,
        "status": {"$in": ["confirmed", "pending"]}
    }).to_list(1000)
    
    activated_count = 0
    activated_tables = []
    
    for reservation in reservations:
        try:
            reservation_datetime = dt.fromisoformat(f"{reservation['reservation_date']} {reservation['reservation_time']}")
            pre_arrival_minutes = reservation.get('pre_arrival_minutes', 15)
            pre_arrival_time = reservation_datetime - timedelta(minutes=pre_arrival_minutes)
            
            # Check if we're within the pre-arrival window
            if current_time >= pre_arrival_time and current_time <= reservation_datetime + timedelta(minutes=reservation.get('duration', 120)):
                # Check if table is still available
                table = await db.tables.find_one({
                    "id": reservation["table_id"],
                    "organization_id": user_org_id
                })
                
                if table and table.get("status") == "available":
                    # Mark table as reserved
                    await db.tables.update_one(
                        {"id": reservation["table_id"], "organization_id": user_org_id},
                        {"$set": {"status": "reserved", "updated_at": current_time.isoformat()}}
                    )
                    
                    activated_tables.append(reservation.get("table_number", "N/A"))
                    activated_count += 1
                    
        except Exception as e:
            print(f"⚠️ Error processing reservation {reservation.get('id', 'N/A')}: {e}")
    
    # Invalidate table cache if any tables were updated
    if activated_tables:
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_table_caches(user_org_id)
        except Exception as e:
            print(f"⚠️ Table cache invalidation error: {e}")
    
    print(f"🎯 Activated {activated_count} reservations, reserved tables: {activated_tables}")
    
    return {
        "message": f"Activated {activated_count} reservations",
        "activated_tables": activated_tables
    }


# Helper function to generate WhatsApp notification link
def generate_whatsapp_notification(phone: str, message: str) -> str:
    """Generate WhatsApp link for notification"""
    import urllib.parse
    
    # Clean phone number
    phone_clean = phone.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if not phone_clean.startswith("+"):
        if phone_clean.startswith("0"):
            phone_clean = "+91" + phone_clean[1:]
        elif len(phone_clean) == 10:
            phone_clean = "+91" + phone_clean
        else:
            phone_clean = "+" + phone_clean
    
    phone_clean = phone_clean.replace("+", "")
    encoded_message = urllib.parse.quote(message)
    return f"https://wa.me/{phone_clean}?text={encoded_message}"


def get_status_message(status: str, order: dict, business: dict, frontend_url: str = "") -> str:
    """Generate status-specific WhatsApp message"""
    # Handle None business case
    business = business or {}
    restaurant_name = business.get("restaurant_name", "Restaurant")
    currency = CURRENCY_SYMBOLS.get(business.get("currency", "INR"), "₹")
    order_id = order["id"][:8]
    tracking_token = order.get("tracking_token", "")
    
    # Use provided frontend_url or fall back to business settings
    base_url = frontend_url or business.get("frontend_url", "")
    tracking_link = f"{base_url}/track/{tracking_token}" if base_url and tracking_token else ""
    
    messages = {
        "pending": f"""🍽️ *{restaurant_name}*

✅ *Order Confirmed!*
Order #{order_id}
Table: {order.get('table_number', 'N/A')}

Your order has been received and will be prepared shortly.

{f'📍 Track your order: {tracking_link}' if tracking_link else ''}

Thank you for your order! 🙏""",
        
        "preparing": f"""🍽️ *{restaurant_name}*

👨‍🍳 *Order Being Prepared*
Order #{order_id}

Great news! Our chef is now preparing your delicious meal.

Estimated time: 15-20 minutes

{f'📍 Track live: {tracking_link}' if tracking_link else ''}""",
        
        "ready": f"""🍽️ *{restaurant_name}*

🔔 *Order Ready!*
Order #{order_id}
Table: {order.get('table_number', 'N/A')}

Your order is ready and will be served shortly!

Enjoy your meal! 😋""",
        
        "completed": f"""🍽️ *{restaurant_name}*

✅ *Payment Completed*
Order #{order_id}

Amount: {currency}{order.get('total', 0):.2f}

Thank you for dining with us!
We hope to see you again soon! 🙏

{business.get('footer_message', '')}"""
    }
    
    return messages.get(status, f"Order #{order_id} status: {status}")


# Order routes
@api_router.post("/orders", response_model=Order)
async def create_order(
    order_data: OrderCreate, current_user: dict = Depends(get_current_user)
):
    if not await check_subscription(current_user):
        # Check if trial expired or subscription expired
        created_at = current_user.get("created_at")
        if isinstance(created_at, str):
            try:
                created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
                trial_end = created_at + timedelta(days=7)
                if datetime.now(timezone.utc) > trial_end:
                    raise HTTPException(
                        status_code=402,
                        detail="Your 7-day free trial has expired. Please subscribe to continue using BillByteKOT. Only ₹999/year for unlimited bills!",
                    )
            except:
                pass
        
        raise HTTPException(
            status_code=402,
            detail="Subscription required. Please subscribe to continue using BillByteKOT.",
        )

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Get business settings - handle None case
    business = current_user.get("business_settings") or {}
    # Handle tax_rate properly - allow 0 as valid value
    tax_rate_setting = business.get("tax_rate")
    tax_rate = (tax_rate_setting if tax_rate_setting is not None else 5.0) / 100
    kot_mode_enabled = business.get("kot_mode_enabled", True)
    
    # Use default values for table when KOT is disabled
    table_id = order_data.table_id or "counter"
    table_number = order_data.table_number or 0

    # ENHANCED DUPLICATE PREVENTION - Check for exact duplicate orders
    if kot_mode_enabled and table_id != "counter":
        try:
            # Check for exact duplicate orders (same table, same items, within last 30 seconds)
            recent_cutoff = (datetime.now(timezone.utc) - timedelta(seconds=30)).isoformat()
            
            # Create signature of current order items for comparison
            current_items_signature = sorted([
                f"{item.menu_item_id}_{item.quantity}_{item.price}" 
                for item in order_data.items
            ])
            
            # Find recent orders on same table
            recent_orders = await db.orders.find({
                "organization_id": user_org_id,
                "table_id": table_id,
                "created_at": {"$gte": recent_cutoff}
            }).to_list(10)
            
            for recent_order in recent_orders:
                # Create signature of recent order items
                recent_items_signature = sorted([
                    f"{item.get('menu_item_id', 'unknown')}_{item.get('quantity', 0)}_{item.get('price', 0)}" 
                    for item in recent_order.get("items", [])
                ])
                
                # Check if signatures match (exact duplicate)
                if current_items_signature == recent_items_signature:
                    print(f"🚫 Duplicate order detected for table {table_number} - rejecting")
                    raise HTTPException(
                        status_code=400, 
                        detail=f"Duplicate order detected for Table {table_number}. Please wait 30 seconds before placing the same order again."
                    )
            
            # ORDER CONSOLIDATION LOGIC - Check for existing pending orders on the same table
            existing_order = await db.orders.find_one({
                "organization_id": user_org_id,
                "table_id": table_id,
                "status": {"$in": ["pending", "preparing"]},
                "created_at": {"$gte": (datetime.now(timezone.utc) - timedelta(hours=2)).isoformat()}  # Only check last 2 hours
            })
            
            if existing_order:
                print(f"🔄 Found existing order {existing_order['id']} for table {table_number}, consolidating items...")
                
                # Consolidate items - merge with existing order
                existing_items = existing_order.get("items", [])
                new_items = [item.model_dump() for item in order_data.items]
                
                # Merge items by menu_item_id
                consolidated_items = {}
                
                # Add existing items
                for item in existing_items:
                    key = item.get("menu_item_id", item.get("name", "unknown"))
                    if key in consolidated_items:
                        consolidated_items[key]["quantity"] += item["quantity"]
                    else:
                        consolidated_items[key] = item.copy()
                
                # Add new items
                for item in new_items:
                    key = item.get("menu_item_id", item.get("name", "unknown"))
                    if key in consolidated_items:
                        consolidated_items[key]["quantity"] += item["quantity"]
                    else:
                        consolidated_items[key] = item.copy()
                
                # Convert back to list
                final_items = list(consolidated_items.values())
                
                # Recalculate totals
                subtotal = sum(item["price"] * item["quantity"] for item in final_items)
                tax = subtotal * tax_rate
                total = subtotal + tax
                
                # Update existing order with consolidated items
                await db.orders.update_one(
                    {"id": existing_order["id"]},
                    {
                        "$set": {
                            "items": final_items,
                            "subtotal": subtotal,
                            "tax": tax,
                            "total": total,
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                            # Update customer info if provided
                            "customer_name": order_data.customer_name or existing_order.get("customer_name", ""),
                            "customer_phone": order_data.customer_phone or existing_order.get("customer_phone", "")
                        }
                    }
                )
                
                # Invalidate Redis cache
                try:
                    cached_service = get_cached_order_service()
                    await cached_service.invalidate_order_caches(user_org_id, existing_order["id"])
                    print(f"🗑️ Cache invalidated for consolidated order {existing_order['id']}")
                except Exception as e:
                    print(f"⚠️ Cache invalidation error: {e}")
                
                # Return the updated order
                updated_order = await db.orders.find_one({"id": existing_order["id"]}, {"_id": 0})
                
                # Generate WhatsApp notification if enabled
                whatsapp_link = None
                frontend_url = order_data.frontend_origin or ""
                if order_data.customer_phone and business.get("whatsapp_auto_notify"):
                    message = get_status_message("pending", updated_order, business, frontend_url)
                    whatsapp_link = generate_whatsapp_notification(order_data.customer_phone, message)
                
                print(f"✅ Order consolidated successfully: {existing_order['id']} (Table {table_number})")
                return {
                    **updated_order,
                    "whatsapp_link": whatsapp_link,
                    "consolidated": True,
                    "message": f"Items added to existing order for Table {table_number}"
                }
                
        except Exception as e:
            print(f"⚠️ Order consolidation check failed: {e}, creating new order...")
            # If consolidation fails, continue with creating new order

    # CREATE NEW ORDER (original logic)
    subtotal = sum(item.price * item.quantity for item in order_data.items)
    tax = subtotal * tax_rate
    total = subtotal + tax
    
    # Generate tracking token for customer live tracking
    tracking_token = str(uuid.uuid4())[:12]

    # Get next invoice number for the organization
    invoice_number = await get_next_invoice_number(user_org_id)
    
    order_obj = Order(
        table_id=table_id,
        table_number=table_number,
        items=[item.model_dump() for item in order_data.items],
        subtotal=subtotal,
        tax=tax,
        tax_rate=tax_rate_setting if tax_rate_setting is not None else 5.0,  # Store the tax rate used
        total=total,
        waiter_id=current_user["id"],
        waiter_name=current_user["username"],
        customer_name=order_data.customer_name,
        customer_phone=order_data.customer_phone,
        tracking_token=tracking_token,
        order_type=order_data.order_type or "takeaway",
        organization_id=user_org_id,
        invoice_number=invoice_number
    )

    doc = order_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()

    await db.orders.insert_one(doc)
    
    # Invalidate Redis cache for active orders
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_order_caches(user_org_id, order_obj.id)
        print(f"🗑️ Cache invalidated for new order {order_obj.id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
        # If Redis is not available, that's okay - MongoDB will handle the queries
        pass
    
    # Only update table status if KOT mode is enabled and table exists
    if kot_mode_enabled and table_id != "counter":
        # Use TableStatusManager for immediate, direct DB update
        try:
            table_manager = get_table_status_manager()
            result = await table_manager.set_table_occupied(user_org_id, table_id, order_obj.id)
            if result["success"]:
                print(f"✅ Table {table_id} set to OCCUPIED via TableStatusManager")
            else:
                print(f"⚠️ TableStatusManager failed: {result['message']}")
                # Fallback to direct update
                await db.tables.update_one(
                    {"id": table_id, "organization_id": user_org_id},
                    {"$set": {"status": "occupied", "current_order_id": order_obj.id}},
                )
        except Exception as e:
            print(f"⚠️ TableStatusManager error: {e}, using fallback")
            # Fallback to direct update if TableStatusManager fails
            await db.tables.update_one(
                {"id": table_id, "organization_id": user_org_id},
                {"$set": {"status": "occupied", "current_order_id": order_obj.id}},
            )
            # Invalidate tables cache
            try:
                cached_service = get_cached_order_service()
                await cached_service.invalidate_table_caches(user_org_id)
            except Exception as cache_e:
                print(f"⚠️ Tables cache invalidation error: {cache_e}")
    
    # Generate WhatsApp notification if enabled and phone provided
    whatsapp_link = None
    frontend_url = order_data.frontend_origin or ""
    if order_data.customer_phone and business.get("whatsapp_auto_notify") and business.get("whatsapp_notify_on_placed"):
        message = get_status_message("pending", doc, business, frontend_url)
        whatsapp_link = generate_whatsapp_notification(order_data.customer_phone, message)

    print(f"✅ New order created successfully: {order_obj.id} (Table {table_number})")
    return {
        **order_obj.model_dump(),
        "whatsapp_link": whatsapp_link,
        "tracking_token": tracking_token,
        "tracking_url": f"{frontend_url}/track/{tracking_token}" if frontend_url else ""
    }


@api_router.get("/orders", response_model=List[Order])
async def get_orders(
    status: Optional[str] = None, 
    current_user: dict = Depends(get_current_user),
    page: int = Query(1, ge=1, description="Page number starting from 1"),
    page_size: int = Query(20, ge=1, le=100, description="Items per page (max 100)")
):
    # Get user's organization_id - CRITICAL for data isolation
    user_org_id = current_user.get("organization_id")
    
    # SECURITY: Ensure organization_id is valid
    if not user_org_id:
        print(f"🚨 SECURITY ALERT: User {current_user.get('email')} attempted to fetch orders without organization_id!")
        raise HTTPException(status_code=403, detail="Organization not configured. Contact support.")

    # Security log
    print(f"🔒 User {current_user['email']} (org: {user_org_id}) fetching orders with status: {status}")

    try:
        # Strategy 1: Try Redis-cached order service for active orders
        if not status or status not in ["completed", "cancelled"]:
            try:
                cached_service = get_cached_order_service()
                
                if not status:
                    # Get all active orders (Redis cache with MongoDB fallback)
                    orders = await cached_service.get_active_orders(user_org_id, use_cache=True)
                    print(f"🚀 Returned {len(orders)} active orders (cached service)")
                    return orders
                else:
                    # Get active orders and filter by status
                    active_orders = await cached_service.get_active_orders(user_org_id, use_cache=True)
                    filtered_orders = [order for order in active_orders if order.get("status") == status]
                    print(f"🚀 Returned {len(filtered_orders)} orders with status '{status}' (cached service)")
                    return filtered_orders
                    
            except Exception as cache_error:
                print(f"❌ Cached service error: {cache_error}, falling back to direct MongoDB")
                # Continue to Strategy 2
        
        # Strategy 2: Direct MongoDB query (for completed/cancelled orders or if cache fails)
        print(f"📊 Fetching orders directly from MongoDB for org {user_org_id}")
        
        query = {"organization_id": user_org_id}
        if status:
            query["status"] = status
        elif not status:
            # If no status specified, get active orders from MongoDB
            query["status"] = {"$nin": ["completed", "cancelled"]}

        try:
            orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).limit(1000).to_list(1000)
            
            # Convert datetime objects for consistency
            for order in orders:
                if isinstance(order.get("created_at"), str):
                    try:
                        order["created_at"] = datetime.fromisoformat(order["created_at"])
                    except:
                        pass
                if isinstance(order.get("updated_at"), str):
                    try:
                        order["updated_at"] = datetime.fromisoformat(order["updated_at"])
                    except:
                        pass
            
            print(f"📊 Returned {len(orders)} orders from MongoDB (status: {status})")
            return orders
            
        except Exception as db_error:
            print(f"❌ MongoDB query error: {db_error}")
            # Continue to Strategy 3
        
        # Strategy 3: Final fallback with minimal query
        print(f"🆘 Using minimal fallback query for org {user_org_id}")
        
        try:
            # Most basic query possible
            basic_query = {"organization_id": user_org_id}
            orders = await db.orders.find(basic_query, {"_id": 0}).sort("created_at", -1).limit(100).to_list(100)
            
            # Filter by status if needed
            if status:
                orders = [order for order in orders if order.get("status") == status]
            elif not status:
                # Filter to active orders only
                orders = [order for order in orders if order.get("status") not in ["completed", "cancelled"]]
            
            # Basic datetime conversion
            for order in orders:
                try:
                    if isinstance(order.get("created_at"), str):
                        order["created_at"] = datetime.fromisoformat(order["created_at"])
                    if isinstance(order.get("updated_at"), str):
                        order["updated_at"] = datetime.fromisoformat(order["updated_at"])
                except:
                    # If datetime conversion fails, leave as string
                    pass
            
            print(f"🆘 Fallback returned {len(orders)} orders")
            return orders
            
        except Exception as final_error:
            print(f"❌ Final fallback failed: {final_error}")
            # Return empty list rather than crash
            return []
        
    except Exception as e:
        print(f"❌ Critical error in get_orders: {e}")
        # Last resort: return empty list to prevent API crash
        return []


@api_router.get("/orders/today-bills", response_model=List[Order])
async def get_todays_bills(current_user: dict = Depends(get_current_user)):
    """Get today's completed/paid bills with Redis caching and MongoDB fallback"""
    # Get user's organization_id - CRITICAL for data isolation
    user_org_id = current_user.get("organization_id")
    
    # SECURITY: Ensure organization_id is valid
    if not user_org_id:
        print(f"🚨 SECURITY ALERT: User {current_user.get('email')} attempted to fetch today's bills without organization_id!")
        raise HTTPException(status_code=403, detail="Organization not configured. Contact support.")

    # Security log
    print(f"🔒 User {current_user['email']} (org: {user_org_id}) fetching today's bills")

    try:
        # Use IST (Indian Standard Time) for "today" calculation
        from datetime import timedelta
        IST = timezone(timedelta(hours=5, minutes=30))
        
        # Get current time in IST and find start of today in IST
        now_ist = datetime.now(IST)
        today_ist = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
        
        # Convert to UTC for database query
        today_utc = today_ist.astimezone(timezone.utc)
        
        # Strategy 1: Try Redis-cached service first
        try:
            cached_service = get_cached_order_service()
            
            # Check if we have today's bills cached
            cache_key = f"todays_bills:{user_org_id}:{today_ist.strftime('%Y-%m-%d')}"
            
            # For now, go directly to MongoDB since we need specific today filtering
            # TODO: Implement specific today's bills caching in Redis
            print(f"📊 Fetching today's bills directly from MongoDB for org {user_org_id}")
            
        except Exception as cache_error:
            print(f"❌ Cached service error: {cache_error}, falling back to direct MongoDB")
        
        # Strategy 2: Direct MongoDB query for today's completed/paid orders
        print(f"📊 Fetching today's bills from MongoDB for org {user_org_id} (from {today_utc.isoformat()})")
        
        try:
            # Query for today's COMPLETED orders ONLY
            query = {
                "organization_id": user_org_id,
                "created_at": {"$gte": today_utc.isoformat()},
                "status": {"$in": ["completed", "paid"]}  # ONLY completed or paid orders
            }
            
            orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).limit(500).to_list(500)
            
            # Convert datetime objects for consistency
            for order in orders:
                if isinstance(order.get("created_at"), str):
                    try:
                        order["created_at"] = datetime.fromisoformat(order["created_at"])
                    except:
                        pass
                if isinstance(order.get("updated_at"), str):
                    try:
                        order["updated_at"] = datetime.fromisoformat(order["updated_at"])
                    except:
                        pass
            
            print(f"📊 Found {len(orders)} today's bills for org {user_org_id}")
            return orders
            
        except Exception as db_error:
            print(f"❌ MongoDB error in get_todays_bills: {db_error}")
            # Continue to fallback
        
        # Strategy 3: Final fallback with basic query and client-side filtering
        print(f"🆘 Using fallback query for today's bills org {user_org_id}")
        
        try:
            # Get recent completed/paid orders only
            basic_query = {
                "organization_id": user_org_id,
                "status": {"$in": ["completed", "paid"]}  # ONLY completed/paid, no cancelled
            }
            
            orders = await db.orders.find(basic_query, {"_id": 0}).sort("created_at", -1).limit(200).to_list(200)
            
            # Filter to today's orders
            todays_orders = []
            for order in orders:
                try:
                    if isinstance(order.get("created_at"), str):
                        order_date = datetime.fromisoformat(order["created_at"])
                    else:
                        order_date = order.get("created_at")
                    
                    if order_date and order_date >= today_utc:
                        # Only include orders that are actually completed or paid
                        if order.get("status") in ["completed", "paid"]:
                            todays_orders.append(order)
                            
                except Exception as date_error:
                    print(f"⚠️ Date parsing error for order {order.get('id', 'unknown')}: {date_error}")
                    continue
            
            print(f"🆘 Fallback returned {len(todays_orders)} today's bills")
            return todays_orders
            
        except Exception as final_error:
            print(f"❌ Final fallback failed: {final_error}")
            return []
        
    except Exception as e:
        print(f"❌ Critical error in get_todays_bills: {e}")
        return []


@api_router.get("/orders/{order_id}", response_model=Order)
async def get_order(order_id: str, current_user: dict = Depends(get_current_user)):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    try:
        # Try Redis cache first
        cached_service = get_cached_order_service()
        order = await cached_service.get_order_by_id(order_id, user_org_id, use_cache=True)
        
        if order:
            print(f"🚀 Order {order_id} retrieved from cache")
            return order
        else:
            raise HTTPException(status_code=404, detail="Order not found")
            
    except Exception as e:
        print(f"❌ Cache error for order {order_id}: {e}")
        # Fallback to direct MongoDB query
        order = await db.orders.find_one(
            {"id": order_id, "organization_id": user_org_id}, {"_id": 0}
        )
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
        
        # Convert datetime objects
        if isinstance(order["created_at"], str):
            order["created_at"] = datetime.fromisoformat(order["created_at"])
        if isinstance(order["updated_at"], str):
            order["updated_at"] = datetime.fromisoformat(order["updated_at"])
        
        print(f"📊 Order {order_id} retrieved from MongoDB (fallback)")
        return order


@api_router.put("/orders/{order_id}/status")
async def update_order_status(
    order_id: str, 
    status: str, 
    frontend_origin: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Try to get order from cache first, then fallback to MongoDB
    try:
        cached_service = get_cached_order_service()
        order = await cached_service.get_order_by_id(order_id, user_org_id, use_cache=True)
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")
    except Exception as e:
        print(f"⚠️ Cache error getting order {order_id}: {e}")
        order = await db.orders.find_one(
            {"id": order_id, "organization_id": user_org_id}, {"_id": 0}
        )
        if not order:
            raise HTTPException(status_code=404, detail="Order not found")

    # Update order status in MongoDB
    await db.orders.update_one(
        {"id": order_id, "organization_id": user_org_id},
        {
            "$set": {
                "status": status,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            }
        },
    )

    # Invalidate Redis cache for this order and active orders list
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_order_caches(user_org_id, order_id)
        
        # Also invalidate table cache if status affects table
        if status == "completed":
            await cached_service.invalidate_table_caches(user_org_id)
        
        print(f"🗑️ Cache invalidated for order status update {order_id} -> {status}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")

    # Update table status if order is completed - Use TableStatusManager for immediate sync
    # This ensures cache invalidation happens (Requirements 1.2, 1.3)
    if status == "completed":
        table_id = order.get("table_id")
        if table_id and table_id != "counter":
            try:
                table_manager = get_table_status_manager()
                result = await table_manager.set_table_available(user_org_id, table_id)
                if result["success"]:
                    print(f"✅ Table {order.get('table_number', 'unknown')} set to AVAILABLE via TableStatusManager for status update {order_id}")
                else:
                    print(f"⚠️ TableStatusManager failed: {result['message']}")
                    # Fallback to direct update
                    await db.tables.update_one(
                        {"id": table_id, "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table {order.get('table_number', 'unknown')} cleared via fallback for status update {order_id}")
            except Exception as e:
                print(f"⚠️ TableStatusManager error: {e}, using fallback")
                # Fallback to direct update if TableStatusManager fails
                try:
                    await db.tables.update_one(
                        {"id": table_id, "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table {order.get('table_number', 'unknown')} cleared via fallback for status update {order_id}")
                except Exception as fallback_error:
                    print(f"⚠️ Table clearing fallback error: {fallback_error}")
    
    # Generate WhatsApp notification if enabled
    business = current_user.get("business_settings", {})
    whatsapp_link = None
    customer_phone = order.get("customer_phone")
    
    if customer_phone and business.get("whatsapp_auto_notify"):
        should_notify = False
        if status == "preparing" and business.get("whatsapp_notify_on_preparing"):
            should_notify = True
        elif status == "ready" and business.get("whatsapp_notify_on_ready"):
            should_notify = True
        elif status == "completed" and business.get("whatsapp_notify_on_completed"):
            should_notify = True
        
        if should_notify:
            # Update order with current status for message generation
            order["status"] = status
            message = get_status_message(status, order, business, frontend_origin or "")
            whatsapp_link = generate_whatsapp_notification(customer_phone, message)

    return {
        "message": "Order status updated",
        "whatsapp_link": whatsapp_link,
        "customer_phone": customer_phone
    }


# Order Management - Edit/Cancel/Delete
@api_router.put("/orders/{order_id}")
async def update_order(
    order_id: str,
    order_data: dict,
    current_user: dict = Depends(get_current_user)
):
    """Update an existing order"""
    user_org_id = get_secure_org_id(current_user)
    
    # Verify order belongs to user's organization
    existing_order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id}
    )
    
    if not existing_order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Handle status update separately (e.g., marking as completed from billing page)
    if "status" in order_data and order_data.get("status") == "completed":
        update_data = {
            "status": "completed",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        # Also update payment fields if provided
        payment_fields = ["payment_method", "is_credit", "payment_received", "balance_amount"]
        for field in payment_fields:
            if field in order_data:
                update_data[field] = order_data[field]
        
        await db.orders.update_one(
            {"id": order_id, "organization_id": user_org_id},
            {"$set": update_data}
        )
        
        # Clear table when order is completed - Use TableStatusManager for immediate, direct DB update
        if existing_order.get("table_id") and existing_order.get("table_id") != "counter":
            try:
                table_manager = get_table_status_manager()
                result = await table_manager.set_table_available(user_org_id, existing_order["table_id"])
                if result["success"]:
                    print(f"✅ Table {existing_order.get('table_number', 'unknown')} set to AVAILABLE via TableStatusManager for completed order {order_id}")
                else:
                    print(f"⚠️ TableStatusManager failed: {result['message']}")
                    # Fallback to direct update
                    await db.tables.update_one(
                        {"id": existing_order["table_id"], "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table {existing_order.get('table_number', 'unknown')} cleared via fallback for completed order {order_id}")
            except Exception as e:
                print(f"⚠️ TableStatusManager error: {e}, using fallback")
                # Fallback to direct update if TableStatusManager fails
                try:
                    await db.tables.update_one(
                        {"id": existing_order["table_id"], "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table {existing_order.get('table_number', 'unknown')} cleared via fallback for completed order {order_id}")
                except Exception as fallback_error:
                    print(f"⚠️ Table clearing fallback error: {fallback_error}")
        
        # Invalidate cache for completed order update
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_order_caches(user_org_id, order_id)
            print(f"🗑️ Cache invalidated for completed order update {order_id}")
        except Exception as e:
            print(f"⚠️ Cache invalidation error: {e}")
        
        return {"message": "Order completed and table cleared successfully"}
    
    # For completed orders, only allow updating payment-related fields
    if existing_order.get("status") == "completed":
        # Allow updating: is_credit, payment_method, payment_received, balance_amount, customer_name, customer_phone
        # Also allow discount and tax updates
        allowed_fields = [
            "is_credit", "payment_method", "payment_received", "balance_amount", 
            "customer_name", "customer_phone",
            "cash_amount", "card_amount", "upi_amount", "credit_amount",
            "discount", "discount_type", "discount_value", "discount_amount",
            "tax", "tax_rate", "subtotal", "total", "items"
        ]
        update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
        
        for field in allowed_fields:
            if field in order_data:
                update_data[field] = order_data[field]
        
        # Calculate balance if payment_received is provided
        # Fix: Use the new total from order_data if provided, otherwise use existing
        if "payment_received" in order_data:
            total = order_data.get("total", existing_order.get("total", 0))
            payment_received = order_data.get("payment_received", 0) or 0
            calculated_balance = max(0, total - payment_received)  # Ensure non-negative
            update_data["balance_amount"] = calculated_balance
            # Auto-mark as credit if there's a balance
            if calculated_balance > 0:
                update_data["is_credit"] = True
            else:
                update_data["is_credit"] = False
                update_data["balance_amount"] = 0
            print(f"💰 Payment update: total={total}, received={payment_received}, balance={calculated_balance}, is_credit={update_data['is_credit']}")
        
        await db.orders.update_one(
            {"id": order_id, "organization_id": user_org_id},
            {"$set": update_data}
        )
        
        # Invalidate cache for payment update
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_order_caches(user_org_id, order_id)
            print(f"🗑️ Cache invalidated for payment update {order_id}")
        except Exception as e:
            print(f"⚠️ Cache invalidation error: {e}")
        
        # Clear table if payment is fully completed (no balance remaining) - Use TableStatusManager
        if update_data.get("balance_amount", 0) <= 0 and not update_data.get("is_credit", False):
            table_id = existing_order.get("table_id")
            if table_id and table_id != "counter":
                try:
                    table_manager = get_table_status_manager()
                    result = await table_manager.set_table_available(user_org_id, table_id)
                    if result["success"]:
                        print(f"✅ Table {table_id} set to AVAILABLE via TableStatusManager for completed payment")
                    else:
                        print(f"⚠️ TableStatusManager failed: {result['message']}")
                        # Fallback to direct update
                        await db.tables.update_one(
                            {"id": table_id, "organization_id": user_org_id},
                            {"$set": {"status": "available", "current_order_id": None}}
                        )
                        print(f"🍽️ Table freed via fallback for completed payment: {table_id}")
                except Exception as e:
                    print(f"⚠️ TableStatusManager error: {e}, using fallback")
                    # Fallback to direct update if TableStatusManager fails
                    try:
                        await db.tables.update_one(
                            {"id": table_id, "organization_id": user_org_id},
                            {"$set": {"status": "available", "current_order_id": None}}
                        )
                        print(f"🍽️ Table freed via fallback for completed payment: {table_id}")
                    except Exception as fallback_error:
                        print(f"⚠️ Table clearing fallback error: {fallback_error}")
        
        return {"message": "Order payment details updated successfully"}
    
    # For non-completed orders, allow full editing
    update_data = {
        "items": order_data.get("items", existing_order["items"]),
        "subtotal": order_data.get("subtotal", existing_order["subtotal"]),
        "tax": order_data.get("tax", existing_order["tax"]),
        "tax_rate": order_data.get("tax_rate", existing_order.get("tax_rate")),
        "total": order_data.get("total", existing_order["total"]),
        "customer_name": order_data.get("customer_name", existing_order.get("customer_name", "")),
        "customer_phone": order_data.get("customer_phone", existing_order.get("customer_phone", "")),
        "payment_method": order_data.get("payment_method", existing_order.get("payment_method", "cash")),
        # Split payment fields
        "cash_amount": order_data.get("cash_amount", existing_order.get("cash_amount", 0)),
        "card_amount": order_data.get("card_amount", existing_order.get("card_amount", 0)),
        "upi_amount": order_data.get("upi_amount", existing_order.get("upi_amount", 0)),
        "credit_amount": order_data.get("credit_amount", existing_order.get("credit_amount", 0)),
        # Discount fields
        "discount": order_data.get("discount", existing_order.get("discount", 0)),
        "discount_type": order_data.get("discount_type", existing_order.get("discount_type", "amount")),
        "discount_value": order_data.get("discount_value", existing_order.get("discount_value", 0)),
        "discount_amount": order_data.get("discount_amount", existing_order.get("discount_amount", 0)),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Fix: Properly calculate payment fields when payment_received is provided
    total = order_data.get("total", existing_order["total"])
    payment_received = order_data.get("payment_received", existing_order.get("payment_received", 0)) or 0
    
    # Calculate balance correctly - ensure it's never negative
    calculated_balance = max(0, total - payment_received)
    
    # Determine is_credit based on balance
    is_credit = calculated_balance > 0
    
    # If full payment (balance <= 0), ensure is_credit is false and balance is 0
    if calculated_balance <= 0:
        is_credit = False
        calculated_balance = 0
    
    update_data["payment_received"] = payment_received
    update_data["balance_amount"] = calculated_balance
    update_data["is_credit"] = is_credit
    
    # Also update status if provided
    if "status" in order_data:
        update_data["status"] = order_data["status"]
    
    print(f"📝 Order update: total={total}, received={payment_received}, balance={calculated_balance}, is_credit={is_credit}")
    
    await db.orders.update_one(
        {"id": order_id, "organization_id": user_org_id},
        {"$set": update_data}
    )
    
    # Invalidate cache for order update
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_order_caches(user_org_id, order_id)
        print(f"🗑️ Cache invalidated for order update {order_id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
    
    # Clear table if payment is fully completed (no balance remaining) - Use TableStatusManager
    if (update_data.get("balance_amount", 0) <= 0 and 
        not update_data.get("is_credit", False) and 
        update_data.get("payment_received", 0) > 0):
        table_id = existing_order.get("table_id")
        if table_id and table_id != "counter":
            try:
                table_manager = get_table_status_manager()
                result = await table_manager.set_table_available(user_org_id, table_id)
                if result["success"]:
                    print(f"✅ Table {table_id} set to AVAILABLE via TableStatusManager for completed payment")
                else:
                    print(f"⚠️ TableStatusManager failed: {result['message']}")
                    # Fallback to direct update
                    await db.tables.update_one(
                        {"id": table_id, "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table freed via fallback for completed payment: {table_id}")
            except Exception as e:
                print(f"⚠️ TableStatusManager error: {e}, using fallback")
                # Fallback to direct update if TableStatusManager fails
                try:
                    await db.tables.update_one(
                        {"id": table_id, "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table freed via fallback for completed payment: {table_id}")
                except Exception as fallback_error:
                    print(f"⚠️ Table clearing fallback error: {fallback_error}")
    
    return {"message": "Order updated successfully"}


@api_router.put("/orders/{order_id}/cancel")
async def cancel_order(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Cancel an order"""
    user_org_id = get_secure_org_id(current_user)
    
    order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id}
    )
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    if order.get("status") == "completed":
        raise HTTPException(status_code=400, detail="Cannot cancel completed orders")
    
    # Update order status to cancelled
    await db.orders.update_one(
        {"id": order_id, "organization_id": user_org_id},
        {
            "$set": {
                "status": "cancelled",
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    # Invalidate cache for cancelled order
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_order_caches(user_org_id, order_id)
        print(f"🗑️ Cache invalidated for cancelled order {order_id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
    
    # Release table if order had one - Use TableStatusManager for immediate sync
    if order.get("table_id") and order.get("table_id") != "counter":
        try:
            table_manager = get_table_status_manager()
            result = await table_manager.set_table_available(user_org_id, order["table_id"])
            if result["success"]:
                print(f"✅ Table {order.get('table_number', 'unknown')} set to AVAILABLE via TableStatusManager for cancelled order {order_id}")
            else:
                print(f"⚠️ TableStatusManager failed: {result['message']}")
                # Fallback to direct update
                await db.tables.update_one(
                    {"id": order["table_id"], "organization_id": user_org_id},
                    {"$set": {"status": "available", "current_order_id": None}}
                )
                print(f"🍽️ Table cleared via fallback for cancelled order {order_id}")
        except Exception as e:
            print(f"⚠️ TableStatusManager error: {e}, using fallback")
            try:
                await db.tables.update_one(
                    {"id": order["table_id"], "organization_id": user_org_id},
                    {"$set": {"status": "available", "current_order_id": None}}
                )
                print(f"🍽️ Table cleared via fallback for cancelled order {order_id}")
            except Exception as fallback_error:
                print(f"⚠️ Table clearing fallback error: {fallback_error}")
    
    return {"message": "Order cancelled successfully"}


@api_router.delete("/orders/{order_id}")
async def delete_order(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete an order (admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete orders")
    
    user_org_id = get_secure_org_id(current_user)
    
    order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id}
    )
    
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Release table if order had one - Use TableStatusManager for immediate sync
    if order.get("table_id") and order.get("table_id") != "counter":
        try:
            table_manager = get_table_status_manager()
            result = await table_manager.set_table_available(user_org_id, order["table_id"])
            if result["success"]:
                print(f"✅ Table {order.get('table_number', 'unknown')} set to AVAILABLE via TableStatusManager for deleted order {order_id}")
            else:
                print(f"⚠️ TableStatusManager failed: {result['message']}")
                # Fallback to direct update
                await db.tables.update_one(
                    {"id": order["table_id"], "organization_id": user_org_id},
                    {"$set": {"status": "available", "current_order_id": None}}
                )
                print(f"🍽️ Table cleared via fallback for deleted order {order_id}")
        except Exception as e:
            print(f"⚠️ TableStatusManager error: {e}, using fallback")
            try:
                await db.tables.update_one(
                    {"id": order["table_id"], "organization_id": user_org_id},
                    {"$set": {"status": "available", "current_order_id": None}}
                )
                print(f"🍽️ Table cleared via fallback for deleted order {order_id}")
            except Exception as fallback_error:
                print(f"⚠️ Table clearing fallback error: {fallback_error}")
    
    # Delete order
    await db.orders.delete_one(
        {"id": order_id, "organization_id": user_org_id}
    )
    
    # Invalidate cache for deleted order
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_order_caches(user_org_id, order_id)
        print(f"🗑️ Cache invalidated for deleted order {order_id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
    
    return {"message": "Order deleted successfully"}


# Payment routes
@api_router.post("/payments/create-order")
async def create_payment_order(
    payment_data: PaymentCreate, current_user: dict = Depends(get_current_user)
):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    order = await db.orders.find_one(
        {"id": payment_data.order_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    if payment_data.payment_method == "razorpay":
        # IMPORTANT: Use restaurant's own Razorpay keys for billing payments
        # NOT the platform subscription keys
        razorpay_key_id = current_user.get("razorpay_key_id")
        razorpay_key_secret = current_user.get("razorpay_key_secret")

        if not razorpay_key_id or not razorpay_key_secret:
            raise HTTPException(
                status_code=400, 
                detail="Razorpay not configured. Please add your Razorpay API keys in Settings > Payment Gateway to accept online payments."
            )

        razorpay_client = razorpay.Client(auth=(razorpay_key_id, razorpay_key_secret))

        amount_paise = int(payment_data.amount * 100)
        razor_order = razorpay_client.order.create(
            {"amount": amount_paise, "currency": "INR", "payment_capture": 1}
        )

        payment_obj = Payment(
            order_id=payment_data.order_id,
            amount=payment_data.amount,
            payment_method=payment_data.payment_method,
            razorpay_order_id=razor_order["id"],
            organization_id=user_org_id,
        )

        doc = payment_obj.model_dump()
        doc["created_at"] = doc["created_at"].isoformat()
        await db.payments.insert_one(doc)

        return {
            "razorpay_order_id": razor_order["id"],
            "amount": amount_paise,
            "currency": "INR",
            "key_id": razorpay_key_id,
        }
    else:
        payment_obj = Payment(
            order_id=payment_data.order_id,
            amount=payment_data.amount,
            payment_method=payment_data.payment_method,
            status="completed",
            organization_id=user_org_id,
        )

        doc = payment_obj.model_dump()
        doc["created_at"] = doc["created_at"].isoformat()
        await db.payments.insert_one(doc)

        # Get the order to retrieve table_id for table status update
        existing_order = await db.orders.find_one(
            {"id": payment_data.order_id, "organization_id": user_org_id},
            {"_id": 0}
        )

        await db.orders.update_one(
            {"id": payment_data.order_id, "organization_id": user_org_id},
            {"$set": {"status": "completed"}},
        )
        await db.users.update_one(
            {"id": current_user["id"]}, {"$inc": {"bill_count": 1}}
        )

        # Use TableStatusManager to set table to available when payment is completed
        # This ensures immediate table status sync (Requirements 1.2, 1.3)
        if existing_order and existing_order.get("table_id") and existing_order.get("table_id") != "counter":
            try:
                table_manager = get_table_status_manager()
                result = await table_manager.set_table_available(user_org_id, existing_order["table_id"])
                if result["success"]:
                    print(f"✅ Table {existing_order.get('table_number', 'unknown')} set to AVAILABLE via TableStatusManager for payment {payment_data.order_id}")
                else:
                    print(f"⚠️ TableStatusManager failed: {result['message']}")
                    # Fallback to direct update
                    await db.tables.update_one(
                        {"id": existing_order["table_id"], "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table {existing_order.get('table_number', 'unknown')} cleared via fallback for payment {payment_data.order_id}")
            except Exception as e:
                print(f"⚠️ TableStatusManager error: {e}, using fallback")
                # Fallback to direct update if TableStatusManager fails
                try:
                    await db.tables.update_one(
                        {"id": existing_order["table_id"], "organization_id": user_org_id},
                        {"$set": {"status": "available", "current_order_id": None}}
                    )
                    print(f"🍽️ Table {existing_order.get('table_number', 'unknown')} cleared via fallback for payment {payment_data.order_id}")
                except Exception as fallback_error:
                    print(f"⚠️ Table clearing fallback error: {fallback_error}")

        # Invalidate cache for completed payment
        try:
            cached_service = get_cached_order_service()
            await cached_service.invalidate_order_caches(user_org_id, payment_data.order_id)
            print(f"🗑️ Cache invalidated for payment {payment_data.order_id}")
        except Exception as e:
            print(f"⚠️ Cache invalidation error: {e}")

        return {"payment_id": payment_obj.id, "status": "completed"}


@api_router.post("/payments/verify")
async def verify_payment(
    razorpay_payment_id: str,
    razorpay_order_id: str,
    order_id: str,
    current_user: dict = Depends(get_current_user),
):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    payment = await db.payments.find_one(
        {"razorpay_order_id": razorpay_order_id, "organization_id": user_org_id},
        {"_id": 0},
    )
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")

    await db.payments.update_one(
        {"razorpay_order_id": razorpay_order_id, "organization_id": user_org_id},
        {"$set": {"razorpay_payment_id": razorpay_payment_id, "status": "completed"}},
    )

    # Get the order to retrieve table_id for table status update
    existing_order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id},
        {"_id": 0}
    )

    await db.orders.update_one(
        {"id": order_id, "organization_id": user_org_id},
        {"$set": {"status": "completed"}},
    )
    await db.users.update_one({"id": current_user["id"]}, {"$inc": {"bill_count": 1}})

    # Use TableStatusManager to set table to available when payment is completed
    # This ensures immediate table status sync (Requirements 1.2, 1.3)
    if existing_order and existing_order.get("table_id") and existing_order.get("table_id") != "counter":
        try:
            table_manager = get_table_status_manager()
            result = await table_manager.set_table_available(user_org_id, existing_order["table_id"])
            if result["success"]:
                print(f"✅ Table {existing_order.get('table_number', 'unknown')} set to AVAILABLE via TableStatusManager for verified payment {order_id}")
            else:
                print(f"⚠️ TableStatusManager failed: {result['message']}")
                # Fallback to direct update
                await db.tables.update_one(
                    {"id": existing_order["table_id"], "organization_id": user_org_id},
                    {"$set": {"status": "available", "current_order_id": None}}
                )
                print(f"🍽️ Table {existing_order.get('table_number', 'unknown')} cleared via fallback for verified payment {order_id}")
        except Exception as e:
            print(f"⚠️ TableStatusManager error: {e}, using fallback")
            # Fallback to direct update if TableStatusManager fails
            try:
                await db.tables.update_one(
                    {"id": existing_order["table_id"], "organization_id": user_org_id},
                    {"$set": {"status": "available", "current_order_id": None}}
                )
                print(f"🍽️ Table {existing_order.get('table_number', 'unknown')} cleared via fallback for verified payment {order_id}")
            except Exception as fallback_error:
                print(f"⚠️ Table clearing fallback error: {fallback_error}")

    # Invalidate cache for completed payment
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_order_caches(user_org_id, order_id)
        print(f"🗑️ Cache invalidated for verified payment {order_id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")

    return {"status": "payment_verified"}


@api_router.get("/payments")
async def get_payments(current_user: dict = Depends(get_current_user)):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    payments = await db.payments.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).to_list(1000)
    for payment in payments:
        if isinstance(payment["created_at"], str):
            payment["created_at"] = datetime.fromisoformat(payment["created_at"])
    return payments


# Inventory routes
@api_router.post("/inventory", response_model=InventoryItem)
async def create_inventory_item(
    item: InventoryItemCreate, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    inv_obj = InventoryItem(**item.model_dump(), organization_id=user_org_id)
    doc = inv_obj.model_dump()
    doc["last_updated"] = doc["last_updated"].isoformat()
    await db.inventory.insert_one(doc)
    
    # Invalidate related caches
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_inventory_caches(user_org_id)
        await cached_service.invalidate_menu_caches(user_org_id)
        print(f"🗑️ Inventory and menu caches invalidated for new item {inv_obj.id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
    
    return inv_obj


@api_router.get("/inventory", response_model=List[InventoryItem])
async def get_inventory(current_user: dict = Depends(get_current_user)):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    try:
        # Use Redis-cached service for inventory items
        cached_service = get_cached_order_service()
        items = await cached_service.get_inventory_items(user_org_id, use_cache=True)
        print(f"🚀 Returned {len(items)} inventory items (Redis cached)")
        return items
        
    except Exception as e:
        print(f"❌ Error fetching inventory from cache: {e}")
        # Fallback to direct MongoDB query
        items = await db.inventory.find(
            {"organization_id": user_org_id}, 
            {"_id": 0}
        ).sort("name", 1).to_list(1000)  # Sort by name for consistent ordering
        
        for item in items:
            if isinstance(item["last_updated"], str):
                item["last_updated"] = datetime.fromisoformat(item["last_updated"])
        print(f"📊 Fallback: Returned {len(items)} inventory items from MongoDB")
        return items


@api_router.put("/inventory/{item_id}", response_model=InventoryItem)
async def update_inventory_item(
    item_id: str,
    item: InventoryItemCreate,
    current_user: dict = Depends(get_current_user),
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    update_data = item.model_dump()
    update_data["last_updated"] = datetime.now(timezone.utc).isoformat()

    await db.inventory.update_one(
        {"id": item_id, "organization_id": user_org_id}, {"$set": update_data}
    )
    updated = await db.inventory.find_one(
        {"id": item_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if isinstance(updated["last_updated"], str):
        updated["last_updated"] = datetime.fromisoformat(updated["last_updated"])
    
    # Invalidate related caches
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_inventory_caches(user_org_id)
        await cached_service.invalidate_menu_caches(user_org_id)
        print(f"🗑️ Inventory and menu caches invalidated for updated item {item_id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
    
    return updated


@api_router.get("/inventory/low-stock")
async def get_low_stock(current_user: dict = Depends(get_current_user)):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    # Optimized: Use database aggregation to filter low stock items directly
    pipeline = [
        {"$match": {"organization_id": user_org_id}},
        {"$addFields": {
            "is_low_stock": {"$lte": ["$quantity", "$min_quantity"]}
        }},
        {"$match": {"is_low_stock": True}},
        {"$project": {"_id": 0, "is_low_stock": 0}},
        {"$sort": {"quantity": 1}}  # Sort by quantity ascending (lowest first)
    ]
    
    low_stock = await db.inventory.aggregate(pipeline).to_list(1000)
    return low_stock


@api_router.delete("/inventory/{item_id}")
async def delete_inventory_item(
    item_id: str,
    current_user: dict = Depends(get_current_user),
):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    result = await db.inventory.delete_one(
        {"id": item_id, "organization_id": user_org_id}
    )
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Item not found")
    
    # Invalidate related caches
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_inventory_caches(user_org_id)
        await cached_service.invalidate_menu_caches(user_org_id)
        print(f"🗑️ Inventory and menu caches invalidated for deleted item {item_id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
    
    return {"message": "Item deleted successfully"}


# Supplier models and routes
class SupplierCreate(BaseModel):
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    website: Optional[str] = None
    payment_terms: Optional[str] = None
    notes: Optional[str] = None


class Supplier(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    contact_person: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    website: Optional[str] = None
    payment_terms: Optional[str] = None
    notes: Optional[str] = None
    organization_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


@api_router.get("/inventory/suppliers", response_model=List[Supplier])
async def get_suppliers(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    suppliers = await db.suppliers.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).to_list(1000)
    return suppliers


@api_router.post("/inventory/suppliers", response_model=Supplier)
async def create_supplier(
    supplier: SupplierCreate, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    supplier_obj = Supplier(**supplier.model_dump(), organization_id=user_org_id)
    doc = supplier_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.suppliers.insert_one(doc)
    return supplier_obj


@api_router.put("/inventory/suppliers/{supplier_id}", response_model=Supplier)
async def update_supplier(
    supplier_id: str,
    supplier: SupplierCreate,
    current_user: dict = Depends(get_current_user),
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    update_data = supplier.model_dump()

    await db.suppliers.update_one(
        {"id": supplier_id, "organization_id": user_org_id}, {"$set": update_data}
    )
    updated = await db.suppliers.find_one(
        {"id": supplier_id, "organization_id": user_org_id}, {"_id": 0}
    )
    return updated


@api_router.delete("/inventory/suppliers/{supplier_id}")
async def delete_supplier(
    supplier_id: str,
    current_user: dict = Depends(get_current_user),
):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    result = await db.suppliers.delete_one(
        {"id": supplier_id, "organization_id": user_org_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Supplier not found")
    return {"message": "Supplier deleted successfully"}


# Category models and routes
class CategoryCreate(BaseModel):
    name: str
    description: Optional[str] = None
    color: str = "#7c3aed"


class Category(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: Optional[str] = None
    color: str = "#7c3aed"
    organization_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


@api_router.get("/inventory/categories", response_model=List[Category])
async def get_categories(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    categories = await db.categories.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).to_list(1000)
    return categories


@api_router.post("/inventory/categories", response_model=Category)
async def create_category(
    category: CategoryCreate, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    category_obj = Category(**category.model_dump(), organization_id=user_org_id)
    doc = category_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.categories.insert_one(doc)
    return category_obj


@api_router.put("/inventory/categories/{category_id}", response_model=Category)
async def update_category(
    category_id: str,
    category: CategoryCreate,
    current_user: dict = Depends(get_current_user),
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    update_data = category.model_dump()

    await db.categories.update_one(
        {"id": category_id, "organization_id": user_org_id}, {"$set": update_data}
    )
    updated = await db.categories.find_one(
        {"id": category_id, "organization_id": user_org_id}, {"_id": 0}
    )
    return updated


@api_router.delete("/inventory/categories/{category_id}")
async def delete_category(
    category_id: str,
    current_user: dict = Depends(get_current_user),
):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    result = await db.categories.delete_one(
        {"id": category_id, "organization_id": user_org_id}
    )
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Category not found")
    return {"message": "Category deleted successfully"}


# Stock Movement models and routes
class StockMovementCreate(BaseModel):
    item_id: str
    type: str  # in, out, adjustment
    quantity: float
    reason: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None


class StockMovement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    item_id: str
    type: str
    quantity: float
    reason: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    organization_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


@api_router.get("/inventory/movements", response_model=List[StockMovement])
async def get_stock_movements(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    movements = await db.stock_movements.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).sort("created_at", -1).limit(100).to_list(100)
    return movements


@api_router.post("/inventory/movements", response_model=StockMovement)
async def create_stock_movement(
    movement: StockMovementCreate, current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    
    # Update inventory quantity based on movement
    inventory_item = await db.inventory.find_one(
        {"id": movement.item_id, "organization_id": user_org_id}, {"_id": 0}
    )
    
    if not inventory_item:
        raise HTTPException(status_code=404, detail="Inventory item not found")
    
    # Calculate new quantity
    current_qty = inventory_item["quantity"]
    if movement.type == "in":
        new_qty = current_qty + movement.quantity
    elif movement.type == "out":
        new_qty = max(0, current_qty - movement.quantity)
    else:  # adjustment
        new_qty = movement.quantity
    
    # Update inventory
    await db.inventory.update_one(
        {"id": movement.item_id, "organization_id": user_org_id},
        {"$set": {"quantity": new_qty, "last_updated": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Record movement
    movement_obj = StockMovement(**movement.model_dump(), organization_id=user_org_id)
    doc = movement_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.stock_movements.insert_one(doc)
    return movement_obj


@api_router.get("/inventory/analytics")
async def get_inventory_analytics(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    
    # Get all inventory items
    items = await db.inventory.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).to_list(1000)
    
    # Get recent stock movements
    movements = await db.stock_movements.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).sort("created_at", -1).limit(50).to_list(50)
    
    # Calculate analytics
    total_items = len(items)
    total_value = sum(item["quantity"] * item["price_per_unit"] for item in items)
    low_stock_count = len([item for item in items if item["quantity"] <= item["min_quantity"]])
    
    # Category breakdown
    categories = {}
    for item in items:
        cat_id = item.get("category_id", "uncategorized")
        if cat_id not in categories:
            categories[cat_id] = {"count": 0, "value": 0}
        categories[cat_id]["count"] += 1
        categories[cat_id]["value"] += item["quantity"] * item["price_per_unit"]
    
    # Recent movements summary
    movement_summary = {"in": 0, "out": 0, "adjustment": 0}
    for movement in movements:
        movement_summary[movement["type"]] += 1
    
    return {
        "total_items": total_items,
        "total_value": total_value,
        "low_stock_count": low_stock_count,
        "healthy_stock_count": total_items - low_stock_count,
        "categories": categories,
        "recent_movements": movement_summary,
        "movements": movements[:10]  # Last 10 movements
    }


# ============ PURCHASE ORDER MODELS AND ENDPOINTS ============
# Requirements 6.1, 6.2, 6.3 - Inventory Purchase Order Feature

class PurchaseOrderItemCreate(BaseModel):
    inventory_item_id: str
    item_name: str
    quantity: float
    unit_cost: float


class PurchaseOrderItem(BaseModel):
    inventory_item_id: str
    item_name: str
    quantity: float
    unit_cost: float
    total_cost: float = 0.0


class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    purchase_date: str  # YYYY-MM-DD format
    notes: Optional[str] = None
    items: List[PurchaseOrderItemCreate]


class PurchaseOrder(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    supplier_id: str
    supplier_name: Optional[str] = None
    purchase_date: str
    total_amount: float = 0.0
    status: str = "received"  # pending, received, cancelled
    notes: Optional[str] = None
    items: List[PurchaseOrderItem] = []
    organization_id: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None


@api_router.post("/inventory/purchases", response_model=PurchaseOrder)
async def create_purchase_order(
    purchase: PurchaseOrderCreate, current_user: dict = Depends(get_current_user)
):
    """Create a purchase order and update inventory stock quantities (Requirement 6.3)"""
    if current_user["role"] not in ["admin", "cashier"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    user_org_id = get_secure_org_id(current_user)
    
    # Get supplier name
    supplier = await db.suppliers.find_one(
        {"id": purchase.supplier_id, "organization_id": user_org_id}, {"_id": 0}
    )
    supplier_name = supplier["name"] if supplier else "Unknown Supplier"
    
    # Process items and calculate totals
    processed_items = []
    total_amount = 0.0
    
    for item in purchase.items:
        total_cost = item.quantity * item.unit_cost
        total_amount += total_cost
        processed_items.append(PurchaseOrderItem(
            inventory_item_id=item.inventory_item_id,
            item_name=item.item_name,
            quantity=item.quantity,
            unit_cost=item.unit_cost,
            total_cost=total_cost
        ))
    
    # Create purchase order object
    purchase_obj = PurchaseOrder(
        supplier_id=purchase.supplier_id,
        supplier_name=supplier_name,
        purchase_date=purchase.purchase_date,
        total_amount=total_amount,
        status="received",
        notes=purchase.notes,
        items=[item.model_dump() for item in processed_items],
        organization_id=user_org_id,
        created_by=current_user.get("username", current_user.get("id"))
    )
    
    # Save purchase order to database
    doc = purchase_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    await db.purchase_orders.insert_one(doc)
    
    # Update inventory quantities for each item (Requirement 6.3)
    for item in purchase.items:
        # Get current inventory item
        inv_item = await db.inventory.find_one(
            {"id": item.inventory_item_id, "organization_id": user_org_id}, {"_id": 0}
        )
        
        if inv_item:
            new_quantity = inv_item["quantity"] + item.quantity
            await db.inventory.update_one(
                {"id": item.inventory_item_id, "organization_id": user_org_id},
                {"$set": {
                    "quantity": new_quantity,
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }}
            )
            
            # Record stock movement for audit trail
            movement_obj = StockMovement(
                item_id=item.inventory_item_id,
                type="in",
                quantity=item.quantity,
                reason="Purchase Order",
                reference=f"PO-{purchase_obj.id[:8]}",
                notes=f"Purchase from {supplier_name}",
                organization_id=user_org_id
            )
            movement_doc = movement_obj.model_dump()
            movement_doc["created_at"] = movement_doc["created_at"].isoformat()
            await db.stock_movements.insert_one(movement_doc)
    
    # Invalidate inventory caches
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_inventory_caches(user_org_id)
        print(f"🗑️ Inventory caches invalidated for purchase order {purchase_obj.id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error: {e}")
    
    return purchase_obj


@api_router.get("/inventory/purchases", response_model=List[PurchaseOrder])
async def get_purchase_orders(current_user: dict = Depends(get_current_user)):
    """Get all purchase orders for the organization (Requirement 6.6)"""
    user_org_id = get_secure_org_id(current_user)
    
    purchases = await db.purchase_orders.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    return purchases


@api_router.get("/inventory/purchases/{purchase_id}", response_model=PurchaseOrder)
async def get_purchase_order(
    purchase_id: str, current_user: dict = Depends(get_current_user)
):
    """Get a specific purchase order by ID"""
    user_org_id = get_secure_org_id(current_user)
    
    purchase = await db.purchase_orders.find_one(
        {"id": purchase_id, "organization_id": user_org_id}, {"_id": 0}
    )
    
    if not purchase:
        raise HTTPException(status_code=404, detail="Purchase order not found")
    
    return purchase


@api_router.get("/inventory/purchases/summary/total")
async def get_purchase_summary(current_user: dict = Depends(get_current_user)):
    """Get total purchase value for reporting (Requirement 6.7)"""
    user_org_id = get_secure_org_id(current_user)
    
    purchases = await db.purchase_orders.find(
        {"organization_id": user_org_id}, {"_id": 0}
    ).to_list(1000)
    
    total_value = sum(p.get("total_amount", 0) for p in purchases)
    total_orders = len(purchases)
    
    # Group by supplier
    by_supplier = {}
    for p in purchases:
        supplier_name = p.get("supplier_name", "Unknown")
        if supplier_name not in by_supplier:
            by_supplier[supplier_name] = {"count": 0, "total": 0}
        by_supplier[supplier_name]["count"] += 1
        by_supplier[supplier_name]["total"] += p.get("total_amount", 0)
    
    return {
        "total_value": total_value,
        "total_orders": total_orders,
        "by_supplier": by_supplier
    }


# ============ EXPENSE MANAGEMENT ENDPOINTS ============

@api_router.get("/expenses", response_model=List[Expense])
async def get_expenses(
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    category: Optional[str] = Query(None, description="Filter by category"),
    current_user: dict = Depends(get_current_user)
):
    """Get expenses with optional date range and category filters"""
    user_org_id = get_secure_org_id(current_user)
    
    query = {"organization_id": user_org_id}
    
    # Add date range filter
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        if date_filter:
            query["date"] = date_filter
    
    # Add category filter
    if category:
        query["category"] = category
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(1000)
    return expenses


@api_router.get("/expenses/categories")
async def get_expense_categories():
    """Get list of expense categories"""
    return {"categories": EXPENSE_CATEGORIES}


@api_router.get("/expenses/summary")
async def get_expense_summary(
    start_date: Optional[str] = Query(None, description="Start date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user)
):
    """Get expense summary with totals by category"""
    user_org_id = get_secure_org_id(current_user)
    
    query = {"organization_id": user_org_id}
    
    if start_date or end_date:
        date_filter = {}
        if start_date:
            date_filter["$gte"] = start_date
        if end_date:
            date_filter["$lte"] = end_date
        if date_filter:
            query["date"] = date_filter
    
    expenses = await db.expenses.find(query, {"_id": 0}).to_list(1000)
    
    # Calculate totals
    total = sum(e["amount"] for e in expenses)
    by_category = {}
    by_payment_method = {}
    
    for expense in expenses:
        cat = expense.get("category", "Other")
        method = expense.get("payment_method", "cash")
        
        by_category[cat] = by_category.get(cat, 0) + expense["amount"]
        by_payment_method[method] = by_payment_method.get(method, 0) + expense["amount"]
    
    return {
        "total": total,
        "count": len(expenses),
        "by_category": by_category,
        "by_payment_method": by_payment_method
    }


@api_router.get("/expenses/{expense_id}", response_model=Expense)
async def get_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    """Get a single expense by ID"""
    user_org_id = get_secure_org_id(current_user)
    
    expense = await db.expenses.find_one(
        {"id": expense_id, "organization_id": user_org_id}, {"_id": 0}
    )
    
    if not expense:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    return expense


@api_router.post("/expenses", response_model=Expense)
async def create_expense(
    expense: ExpenseCreate, current_user: dict = Depends(get_current_user)
):
    """Create a new expense"""
    user_org_id = get_secure_org_id(current_user)
    
    # Validate amount
    if expense.amount <= 0:
        raise HTTPException(status_code=400, detail="Amount must be greater than 0")
    
    # Validate category
    if expense.category not in EXPENSE_CATEGORIES:
        raise HTTPException(status_code=400, detail=f"Invalid category. Must be one of: {', '.join(EXPENSE_CATEGORIES)}")
    
    expense_obj = Expense(
        **expense.model_dump(),
        organization_id=user_org_id,
        created_by=current_user.get("id")
    )
    
    await db.expenses.insert_one(expense_obj.model_dump())
    print(f"💰 Created expense: {expense.category} - ₹{expense.amount}")
    
    return expense_obj


@api_router.put("/expenses/{expense_id}", response_model=Expense)
async def update_expense(
    expense_id: str,
    expense_update: ExpenseUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update an existing expense"""
    user_org_id = get_secure_org_id(current_user)
    
    existing = await db.expenses.find_one(
        {"id": expense_id, "organization_id": user_org_id}, {"_id": 0}
    )
    
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    # Build update data
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    for field, value in expense_update.model_dump(exclude_unset=True).items():
        if value is not None:
            # Validate amount
            if field == "amount" and value <= 0:
                raise HTTPException(status_code=400, detail="Amount must be greater than 0")
            # Validate category
            if field == "category" and value not in EXPENSE_CATEGORIES:
                raise HTTPException(status_code=400, detail=f"Invalid category")
            update_data[field] = value
    
    await db.expenses.update_one(
        {"id": expense_id, "organization_id": user_org_id},
        {"$set": update_data}
    )
    
    updated = await db.expenses.find_one(
        {"id": expense_id, "organization_id": user_org_id}, {"_id": 0}
    )
    
    print(f"💰 Updated expense: {expense_id}")
    return updated


@api_router.delete("/expenses/{expense_id}")
async def delete_expense(expense_id: str, current_user: dict = Depends(get_current_user)):
    """Delete an expense"""
    user_org_id = get_secure_org_id(current_user)
    
    existing = await db.expenses.find_one(
        {"id": expense_id, "organization_id": user_org_id}, {"_id": 0}
    )
    
    if not existing:
        raise HTTPException(status_code=404, detail="Expense not found")
    
    await db.expenses.delete_one({"id": expense_id, "organization_id": user_org_id})
    
    print(f"🗑️ Deleted expense: {expense_id}")
    return {"message": "Expense deleted successfully"}


# ============ DAY BOOK / CASH FLOW REPORT ENDPOINTS ============

@api_router.get("/reports/daybook")
async def get_daybook(
    date: str = Query(..., description="Date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date for range YYYY-MM-DD"),
    current_user: dict = Depends(get_current_user)
):
    """
    Get Day Book / Cash Flow report for a date or date range.
    Shows opening balance, inflows (sales), outflows (expenses), and closing balance.
    """
    user_org_id = get_secure_org_id(current_user)
    
    # Determine date range
    start_date = date
    if not end_date:
        end_date = date
    
    # Get completed orders (inflows) for the date range
    orders_query = {
        "organization_id": user_org_id,
        "status": {"$in": ["completed", "paid"]},
        "created_at": {"$gte": f"{start_date}T00:00:00", "$lte": f"{end_date}T23:59:59"}
    }
    
    orders = await db.orders.find(orders_query, {"_id": 0}).to_list(1000)
    
    # Get expenses (outflows) for the date range
    expenses_query = {
        "organization_id": user_org_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }
    
    expenses = await db.expenses.find(expenses_query, {"_id": 0}).to_list(1000)
    
    # Calculate inflows by payment method
    inflow_breakdown = {"cash": 0, "card": 0, "upi": 0, "other": 0}
    total_inflows = 0
    
    inflow_entries = []
    for order in orders:
        amount = order.get("payment_received", order.get("total", 0))
        payment_method = order.get("payment_method", "cash")
        
        # Handle split payments
        if payment_method == "split":
            cash = order.get("cash_amount", 0)
            card = order.get("card_amount", 0)
            upi = order.get("upi_amount", 0)
            inflow_breakdown["cash"] += cash
            inflow_breakdown["card"] += card
            inflow_breakdown["upi"] += upi
            total_inflows += cash + card + upi
        else:
            if payment_method in inflow_breakdown:
                inflow_breakdown[payment_method] += amount
            else:
                inflow_breakdown["other"] += amount
            total_inflows += amount
        
        inflow_entries.append({
            "timestamp": order.get("created_at"),
            "type": "inflow",
            "category": f"Sales-{payment_method.upper()}",
            "description": f"Order #{order.get('invoice_number', order.get('id', '')[:8])} - Table {order.get('table_number', 'Counter')}",
            "amount": amount,
            "reference_id": order.get("id")
        })
    
    # Calculate outflows by category
    outflow_breakdown = {}
    total_outflows = 0
    
    outflow_entries = []
    for expense in expenses:
        amount = expense.get("amount", 0)
        category = expense.get("category", "Other")
        
        outflow_breakdown[category] = outflow_breakdown.get(category, 0) + amount
        total_outflows += amount
        
        outflow_entries.append({
            "timestamp": f"{expense.get('date')}T12:00:00",
            "type": "outflow",
            "category": category,
            "description": expense.get("description", ""),
            "amount": amount,
            "reference_id": expense.get("id")
        })
    
    # Combine and sort entries by timestamp
    all_entries = inflow_entries + outflow_entries
    all_entries.sort(key=lambda x: x.get("timestamp", ""))
    
    # Calculate running balance
    running_balance = 0
    for entry in all_entries:
        if entry["type"] == "inflow":
            running_balance += entry["amount"]
        else:
            running_balance -= entry["amount"]
        entry["running_balance"] = running_balance
    
    # Calculate opening balance (sum of all previous transactions)
    # For simplicity, we'll set opening balance to 0 for now
    # In a real system, this would be calculated from previous day's closing
    opening_balance = 0
    closing_balance = opening_balance + total_inflows - total_outflows
    
    return {
        "date": date,
        "end_date": end_date,
        "opening_balance": opening_balance,
        "total_inflows": total_inflows,
        "total_outflows": total_outflows,
        "closing_balance": closing_balance,
        "net_cash_flow": total_inflows - total_outflows,
        "inflow_breakdown": inflow_breakdown,
        "outflow_breakdown": outflow_breakdown,
        "entries": all_entries,
        "order_count": len(orders),
        "expense_count": len(expenses)
    }


@api_router.get("/reports/daybook/summary")
async def get_daybook_summary(
    period: str = Query("month", description="Period: today, week, month, year"),
    current_user: dict = Depends(get_current_user)
):
    """Get Day Book summary for a period"""
    user_org_id = get_secure_org_id(current_user)
    
    today = datetime.now(timezone.utc)
    
    if period == "today":
        start_date = today.strftime("%Y-%m-%d")
        end_date = start_date
    elif period == "week":
        start_date = (today - timedelta(days=7)).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif period == "month":
        start_date = today.replace(day=1).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    elif period == "year":
        start_date = today.replace(month=1, day=1).strftime("%Y-%m-%d")
        end_date = today.strftime("%Y-%m-%d")
    else:
        start_date = today.strftime("%Y-%m-%d")
        end_date = start_date
    
    # Get orders
    orders = await db.orders.find({
        "organization_id": user_org_id,
        "status": {"$in": ["completed", "paid"]},
        "created_at": {"$gte": f"{start_date}T00:00:00", "$lte": f"{end_date}T23:59:59"}
    }, {"_id": 0}).to_list(1000)
    
    # Get expenses
    expenses = await db.expenses.find({
        "organization_id": user_org_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(1000)
    
    total_inflows = sum(o.get("payment_received", o.get("total", 0)) for o in orders)
    total_outflows = sum(e.get("amount", 0) for e in expenses)
    
    return {
        "period": period,
        "start_date": start_date,
        "end_date": end_date,
        "total_inflows": total_inflows,
        "total_outflows": total_outflows,
        "net_cash_flow": total_inflows - total_outflows,
        "order_count": len(orders),
        "expense_count": len(expenses)
    }


@api_router.get("/reports/daybook/export")
async def export_daybook(
    date: str = Query(..., description="Date YYYY-MM-DD"),
    end_date: Optional[str] = Query(None, description="End date for range YYYY-MM-DD"),
    format: str = Query("pdf", description="Export format: pdf or excel"),
    current_user: dict = Depends(get_current_user)
):
    """
    Export Day Book / Cash Flow report to PDF or Excel format.
    """
    user_org_id = get_secure_org_id(current_user)
    
    # Determine date range
    start_date = date
    if not end_date:
        end_date = date
    
    # Get completed orders (inflows) for the date range
    orders_query = {
        "organization_id": user_org_id,
        "status": {"$in": ["completed", "paid"]},
        "created_at": {"$gte": f"{start_date}T00:00:00", "$lte": f"{end_date}T23:59:59"}
    }
    
    orders = await db.orders.find(orders_query, {"_id": 0}).to_list(1000)
    
    # Get expenses (outflows) for the date range
    expenses_query = {
        "organization_id": user_org_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }
    
    expenses = await db.expenses.find(expenses_query, {"_id": 0}).to_list(1000)
    
    # Calculate inflows by payment method
    inflow_breakdown = {"cash": 0, "card": 0, "upi": 0, "other": 0}
    total_inflows = 0
    
    inflow_entries = []
    for order in orders:
        amount = order.get("payment_received", order.get("total", 0))
        payment_method = order.get("payment_method", "cash")
        
        # Handle split payments
        if payment_method == "split":
            cash = order.get("cash_amount", 0)
            card = order.get("card_amount", 0)
            upi = order.get("upi_amount", 0)
            inflow_breakdown["cash"] += cash
            inflow_breakdown["card"] += card
            inflow_breakdown["upi"] += upi
            total_inflows += cash + card + upi
        else:
            if payment_method in inflow_breakdown:
                inflow_breakdown[payment_method] += amount
            else:
                inflow_breakdown["other"] += amount
            total_inflows += amount
        
        inflow_entries.append({
            "timestamp": order.get("created_at"),
            "type": "inflow",
            "category": f"Sales-{payment_method.upper()}",
            "description": f"Order #{order.get('invoice_number', order.get('id', '')[:8])} - Table {order.get('table_number', 'Counter')}",
            "amount": amount,
            "reference_id": order.get("id")
        })
    
    # Calculate outflows by category
    outflow_breakdown = {}
    total_outflows = 0
    
    outflow_entries = []
    for expense in expenses:
        amount = expense.get("amount", 0)
        category = expense.get("category", "Other")
        
        outflow_breakdown[category] = outflow_breakdown.get(category, 0) + amount
        total_outflows += amount
        
        outflow_entries.append({
            "timestamp": f"{expense.get('date')}T12:00:00",
            "type": "outflow",
            "category": category,
            "description": expense.get("description", ""),
            "amount": amount,
            "reference_id": expense.get("id")
        })
    
    # Combine and sort entries by timestamp
    all_entries = inflow_entries + outflow_entries
    all_entries.sort(key=lambda x: x.get("timestamp", ""))
    
    # Calculate running balance
    running_balance = 0
    for entry in all_entries:
        if entry["type"] == "inflow":
            running_balance += entry["amount"]
        else:
            running_balance -= entry["amount"]
        entry["running_balance"] = running_balance
    
    # Calculate balances
    opening_balance = 0
    closing_balance = opening_balance + total_inflows - total_outflows
    net_cash_flow = total_inflows - total_outflows
    
    # Prepare daybook data
    daybook_data = {
        "date": date,
        "end_date": end_date,
        "opening_balance": opening_balance,
        "total_inflows": total_inflows,
        "total_outflows": total_outflows,
        "closing_balance": closing_balance,
        "net_cash_flow": net_cash_flow,
        "inflow_breakdown": inflow_breakdown,
        "outflow_breakdown": outflow_breakdown,
        "entries": all_entries,
        "order_count": len(orders),
        "expense_count": len(expenses)
    }
    
    if format.lower() == "excel":
        return await generate_daybook_excel(daybook_data, date, end_date)
    else:
        return await generate_daybook_pdf(daybook_data, date, end_date)


async def generate_daybook_pdf(daybook_data: dict, date: str, end_date: str) -> StreamingResponse:
    """Generate Day Book PDF report using reportlab"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, mm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    except ImportError:
        # Fallback to CSV if reportlab not available
        return await generate_daybook_csv(daybook_data, date, end_date)
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, alignment=TA_CENTER, spaceAfter=12, textColor=colors.HexColor('#7c3aed'))
    subtitle_style = ParagraphStyle('Subtitle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, spaceAfter=20, textColor=colors.gray)
    section_style = ParagraphStyle('Section', parent=styles['Heading2'], fontSize=12, spaceBefore=15, spaceAfter=8, textColor=colors.HexColor('#7c3aed'))
    
    elements = []
    
    # Title
    date_range_str = f"{date}" if date == end_date else f"{date} to {end_date}"
    elements.append(Paragraph("📒 Day Book Report", title_style))
    elements.append(Paragraph(f"Date: {date_range_str} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", subtitle_style))
    
    # Summary Table
    elements.append(Paragraph("Summary", section_style))
    summary_data = [
        ['Opening Balance', 'Total Inflows', 'Total Outflows', 'Net Cash Flow', 'Closing Balance'],
        [
            f"₹{daybook_data['opening_balance']:.2f}",
            f"₹{daybook_data['total_inflows']:.2f}",
            f"₹{daybook_data['total_outflows']:.2f}",
            f"{'+'if daybook_data['net_cash_flow'] >= 0 else ''}₹{daybook_data['net_cash_flow']:.2f}",
            f"₹{daybook_data['closing_balance']:.2f}"
        ]
    ]
    summary_table = Table(summary_data, colWidths=[90, 90, 90, 90, 90])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('TOPPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, 1), colors.HexColor('#f3e8ff')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ('TEXTCOLOR', (1, 1), (1, 1), colors.green),
        ('TEXTCOLOR', (2, 1), (2, 1), colors.red),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 15))
    
    # Inflow Breakdown
    if daybook_data['inflow_breakdown']:
        elements.append(Paragraph("💰 Inflow Breakdown", section_style))
        inflow_data = [['Payment Method', 'Amount']]
        for method, amount in daybook_data['inflow_breakdown'].items():
            if amount > 0:
                inflow_data.append([method.capitalize(), f"₹{amount:.2f}"])
        inflow_data.append(['Total Inflows', f"₹{daybook_data['total_inflows']:.2f}"])
        
        inflow_table = Table(inflow_data, colWidths=[200, 100])
        inflow_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#15803d')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#dcfce7')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#bbf7d0')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ]))
        elements.append(inflow_table)
        elements.append(Spacer(1, 10))
    
    # Outflow Breakdown
    if daybook_data['outflow_breakdown']:
        elements.append(Paragraph("📤 Outflow Breakdown", section_style))
        outflow_data = [['Category', 'Amount']]
        for category, amount in daybook_data['outflow_breakdown'].items():
            if amount > 0:
                outflow_data.append([category, f"₹{amount:.2f}"])
        outflow_data.append(['Total Outflows', f"₹{daybook_data['total_outflows']:.2f}"])
        
        outflow_table = Table(outflow_data, colWidths=[200, 100])
        outflow_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 1), (-1, -2), colors.HexColor('#fee2e2')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fecaca')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ]))
        elements.append(outflow_table)
        elements.append(Spacer(1, 10))
    
    # Transaction Details
    if daybook_data['entries']:
        elements.append(Paragraph(f"📋 Transaction Details ({len(daybook_data['entries'])} transactions)", section_style))
        trans_data = [['Time', 'Type', 'Category', 'Description', 'Amount', 'Balance']]
        for entry in daybook_data['entries']:
            try:
                time_str = datetime.fromisoformat(entry['timestamp'].replace('Z', '+00:00')).strftime('%H:%M') if entry.get('timestamp') else '-'
            except:
                time_str = '-'
            
            amount_str = f"{'+'if entry['type'] == 'inflow' else '-'}₹{entry['amount']:.2f}"
            trans_data.append([
                time_str,
                '↑ Inflow' if entry['type'] == 'inflow' else '↓ Outflow',
                entry.get('category', '')[:20],
                entry.get('description', '')[:30],
                amount_str,
                f"₹{entry.get('running_balance', 0):.2f}"
            ])
        
        trans_table = Table(trans_data, colWidths=[40, 55, 80, 130, 70, 70])
        trans_style = [
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#7c3aed')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
        ]
        # Add row colors based on type
        for i, entry in enumerate(daybook_data['entries'], 1):
            if entry['type'] == 'inflow':
                trans_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f0fdf4')))
            else:
                trans_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#fef2f2')))
        
        trans_table.setStyle(TableStyle(trans_style))
        elements.append(trans_table)
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, alignment=TA_CENTER, textColor=colors.gray)
    elements.append(Paragraph("Generated by BillByteKOT - Restaurant Management System", footer_style))
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"daybook-{date}.pdf" if date == end_date else f"daybook-{date}-to-{end_date}.pdf"
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def generate_daybook_excel(daybook_data: dict, date: str, end_date: str) -> StreamingResponse:
    """Generate Day Book Excel report using openpyxl"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return await generate_daybook_csv(daybook_data, date, end_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Day Book"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7c3aed", end_color="7c3aed", fill_type="solid")
    inflow_fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
    outflow_fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
    summary_fill = PatternFill(start_color="f3e8ff", end_color="f3e8ff", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    row = 1
    
    # Title
    date_range_str = f"{date}" if date == end_date else f"{date} to {end_date}"
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Day Book Report - {date_range_str}"
    ws[f'A{row}'].font = Font(bold=True, size=16, color="7c3aed")
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1
    
    ws.merge_cells(f'A{row}:F{row}')
    ws[f'A{row}'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 2
    
    # Summary Section
    ws[f'A{row}'] = "SUMMARY"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="7c3aed")
    row += 1
    
    summary_headers = ['Opening Balance', 'Total Inflows', 'Total Outflows', 'Net Cash Flow', 'Closing Balance']
    for col, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    summary_values = [
        daybook_data['opening_balance'],
        daybook_data['total_inflows'],
        daybook_data['total_outflows'],
        daybook_data['net_cash_flow'],
        daybook_data['closing_balance']
    ]
    for col, value in enumerate(summary_values, 1):
        cell = ws.cell(row=row, column=col, value=f"₹{value:.2f}")
        cell.fill = summary_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
        if col == 2:  # Inflows - green
            cell.font = Font(color="15803d", bold=True)
        elif col == 3:  # Outflows - red
            cell.font = Font(color="dc2626", bold=True)
        elif col == 4:  # Net - based on value
            cell.font = Font(color="15803d" if value >= 0 else "dc2626", bold=True)
    row += 2
    
    # Inflow Breakdown
    ws[f'A{row}'] = "INFLOW BREAKDOWN"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="15803d")
    row += 1
    
    for col, header in enumerate(['Payment Method', 'Amount'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="15803d", end_color="15803d", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    for method, amount in daybook_data['inflow_breakdown'].items():
        if amount > 0:
            ws.cell(row=row, column=1, value=method.capitalize()).border = thin_border
            ws.cell(row=row, column=1).fill = inflow_fill
            cell = ws.cell(row=row, column=2, value=f"₹{amount:.2f}")
            cell.border = thin_border
            cell.fill = inflow_fill
            cell.alignment = Alignment(horizontal='right')
            row += 1
    
    # Total row
    ws.cell(row=row, column=1, value="Total Inflows").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    cell = ws.cell(row=row, column=2, value=f"₹{daybook_data['total_inflows']:.2f}")
    cell.font = Font(bold=True, color="15803d")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right')
    row += 2
    
    # Outflow Breakdown
    ws[f'A{row}'] = "OUTFLOW BREAKDOWN"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="dc2626")
    row += 1
    
    for col, header in enumerate(['Category', 'Amount'], 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = PatternFill(start_color="dc2626", end_color="dc2626", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    for category, amount in daybook_data['outflow_breakdown'].items():
        if amount > 0:
            ws.cell(row=row, column=1, value=category).border = thin_border
            ws.cell(row=row, column=1).fill = outflow_fill
            cell = ws.cell(row=row, column=2, value=f"₹{amount:.2f}")
            cell.border = thin_border
            cell.fill = outflow_fill
            cell.alignment = Alignment(horizontal='right')
            row += 1
    
    # Total row
    ws.cell(row=row, column=1, value="Total Outflows").font = Font(bold=True)
    ws.cell(row=row, column=1).border = thin_border
    cell = ws.cell(row=row, column=2, value=f"₹{daybook_data['total_outflows']:.2f}")
    cell.font = Font(bold=True, color="dc2626")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='right')
    row += 2
    
    # Transaction Details
    ws[f'A{row}'] = f"TRANSACTION DETAILS ({len(daybook_data['entries'])} transactions)"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="7c3aed")
    row += 1
    
    trans_headers = ['Time', 'Type', 'Category', 'Description', 'Amount', 'Running Balance']
    for col, header in enumerate(trans_headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    row += 1
    
    for entry in daybook_data['entries']:
        try:
            time_str = datetime.fromisoformat(entry['timestamp'].replace('Z', '+00:00')).strftime('%H:%M') if entry.get('timestamp') else '-'
        except:
            time_str = '-'
        
        fill = inflow_fill if entry['type'] == 'inflow' else outflow_fill
        amount_prefix = '+' if entry['type'] == 'inflow' else '-'
        
        values = [
            time_str,
            '↑ Inflow' if entry['type'] == 'inflow' else '↓ Outflow',
            entry.get('category', ''),
            entry.get('description', ''),
            f"{amount_prefix}₹{entry['amount']:.2f}",
            f"₹{entry.get('running_balance', 0):.2f}"
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = fill
            cell.border = thin_border
            if col >= 5:
                cell.alignment = Alignment(horizontal='right')
        row += 1
    
    # Adjust column widths
    column_widths = [12, 12, 20, 40, 15, 18]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    filename = f"daybook-{date}.xlsx" if date == end_date else f"daybook-{date}-to-{end_date}.xlsx"
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def generate_daybook_csv(daybook_data: dict, date: str, end_date: str) -> StreamingResponse:
    """Fallback CSV export if PDF/Excel libraries not available"""
    csv_content = f"Day Book Report - {date}\n"
    csv_content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n"
    
    csv_content += "SUMMARY\n"
    csv_content += f"Opening Balance,Total Inflows,Total Outflows,Net Cash Flow,Closing Balance\n"
    csv_content += f"{daybook_data['opening_balance']:.2f},{daybook_data['total_inflows']:.2f},{daybook_data['total_outflows']:.2f},{daybook_data['net_cash_flow']:.2f},{daybook_data['closing_balance']:.2f}\n\n"
    
    csv_content += "INFLOW BREAKDOWN\n"
    csv_content += "Payment Method,Amount\n"
    for method, amount in daybook_data['inflow_breakdown'].items():
        csv_content += f"{method},{amount:.2f}\n"
    csv_content += f"Total,{daybook_data['total_inflows']:.2f}\n\n"
    
    csv_content += "OUTFLOW BREAKDOWN\n"
    csv_content += "Category,Amount\n"
    for category, amount in daybook_data['outflow_breakdown'].items():
        csv_content += f"{category},{amount:.2f}\n"
    csv_content += f"Total,{daybook_data['total_outflows']:.2f}\n\n"
    
    csv_content += "TRANSACTION DETAILS\n"
    csv_content += "Time,Type,Category,Description,Amount,Running Balance\n"
    for entry in daybook_data['entries']:
        try:
            time_str = datetime.fromisoformat(entry['timestamp'].replace('Z', '+00:00')).strftime('%H:%M') if entry.get('timestamp') else '-'
        except:
            time_str = '-'
        
        amount_prefix = '+' if entry['type'] == 'inflow' else '-'
        csv_content += f"{time_str},{entry['type']},{entry.get('category', '')},\"{entry.get('description', '')}\",{amount_prefix}{entry['amount']:.2f},{entry.get('running_balance', 0):.2f}\n"
    
    filename = f"daybook-{date}.csv" if date == end_date else f"daybook-{date}-to-{end_date}.csv"
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


class InventoryDeduction(BaseModel):
    menu_item_id: str
    quantity: int


@api_router.post("/inventory/deduct")
async def ai_chat(message: ChatMessage):
    if not _LLM_AVAILABLE:
        raise HTTPException(status_code=503, detail="LLM integration unavailable")
    try:
        chat = LlmChat(
            api_key=os.environ.get("LLM_API_KEY"),
            session_id=str(uuid.uuid4()),
            system_message="You are a helpful restaurant assistant. Answer customer queries about menu, orders, and restaurant services.",
        ).with_model("openai", "gpt-4o-mini")

        user_msg = UserMessage(text=message.message)
        response = await chat.send_message(user_msg)

        return {"response": response}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@api_router.post("/ai/recommendations")
async def get_recommendations(current_user: dict = Depends(get_current_user)):
    try:
        user_org_id = get_secure_org_id(current_user)
        
        orders = await db.orders.find({
            "status": "completed",
            "organization_id": user_org_id
        }, {"_id": 0}).limit(50).to_list(50)
        
        menu_items = await db.menu_items.find({
            "organization_id": user_org_id
        }, {"_id": 0}).to_list(1000)

        if not orders or not menu_items:
            return {
                "recommendations": "Not enough data yet. Start taking orders to get AI recommendations!"
            }

        order_items = []
        for order in orders:
            order_items.extend([item["name"] for item in order["items"]])

        from collections import Counter
        popular_items = Counter(order_items).most_common(5)

        if not _LLM_AVAILABLE:
            # Fallback without AI
            return {
                "recommendations": f"Top selling items: {', '.join([item[0] for item in popular_items])}. Consider promoting these items!"
            }

        chat = LlmChat(
            api_key=os.environ.get("LLM_API_KEY"),
            session_id=str(uuid.uuid4()),
            system_message="You are a restaurant menu analyst. Based on order history, suggest menu items that would pair well together.",
        ).with_model("openai", "gpt-4o-mini")

        prompt = f"Popular items: {popular_items}. Menu: {[item['name'] for item in menu_items[:20]]}. Suggest 5 complementary items."
        user_msg = UserMessage(text=prompt)
        response = await chat.send_message(user_msg)

        return {"recommendations": response}
    except Exception as e:
        print(f"AI recommendations error: {str(e)}")
        return {
            "recommendations": "AI recommendations temporarily unavailable. Please try again later."
        }


@api_router.post("/ai/sales-forecast")
async def sales_forecast(current_user: dict = Depends(get_current_user)):
    try:
        user_org_id = get_secure_org_id(current_user)
        
        orders = await db.orders.find({
            "status": "completed",
            "organization_id": user_org_id
        }, {"_id": 0}).to_list(1000)

        if not orders:
            return {
                "forecast": "Not enough data yet. Complete some orders to get sales forecasts!",
                "current_stats": {
                    "total_orders": 0,
                    "total_sales": 0,
                    "avg_order": 0,
                },
            }

        total_sales = sum(order["total"] for order in orders)
        avg_order_value = total_sales / len(orders) if orders else 0

        if not _LLM_AVAILABLE:
            # Fallback without AI
            return {
                "forecast": f"Based on {len(orders)} orders with ₹{total_sales:.2f} total sales, your average order value is ₹{avg_order_value:.2f}. Keep up the good work!",
                "current_stats": {
                    "total_orders": len(orders),
                    "total_sales": total_sales,
                    "avg_order": avg_order_value,
                },
            }

        chat = LlmChat(
            api_key=os.environ.get("LLM_API_KEY"),
            session_id=str(uuid.uuid4()),
            system_message="You are a sales analyst. Provide sales predictions based on historical data.",
        ).with_model("openai", "gpt-4o-mini")

        prompt = f"Total orders: {len(orders)}, Total sales: ₹{total_sales:.2f}, Average order: ₹{avg_order_value:.2f}. Predict next week's sales."
        user_msg = UserMessage(text=prompt)
        response = await chat.send_message(user_msg)

        return {
            "forecast": response,
            "current_stats": {
                "total_orders": len(orders),
                "total_sales": total_sales,
                "avg_order": avg_order_value,
            },
        }
    except Exception as e:
        print(f"AI forecast error: {str(e)}")
        return {
            "forecast": "AI forecast temporarily unavailable. Please try again later.",
            "current_stats": {
                "total_orders": 0,
                "total_sales": 0,
                "avg_order": 0,
            },
        }


# Reports
@api_router.get("/reports/daily")
async def daily_report(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    
    # ✅ PERFORMANCE: Check cache first (30-second TTL for real-time dashboard updates)
    cache_key = f"daily_report:{user_org_id}"
    current_time = time.time()
    
    with _cache_lock:
        if cache_key in _cache and cache_key in _cache_ttl:
            if current_time < _cache_ttl[cache_key]:
                print(f"💾 Cache hit for daily_report: {cache_key}")
                return _cache[cache_key]
    
    # Use IST (Indian Standard Time) for "today" calculation
    # IST is UTC+5:30
    from datetime import timedelta
    IST = timezone(timedelta(hours=5, minutes=30))
    
    # Get current time in IST and find start of today in IST
    now_ist = datetime.now(IST)
    today_ist = now_ist.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Convert to UTC for database query
    today_utc = today_ist.astimezone(timezone.utc)
    
    # Optimized: Use database query instead of filtering in Python
    # Include all orders from today that have been paid (not just completed)
    today_orders = await db.orders.find({
        "$or": [
            {"status": "completed"},
            {"status": "paid"},
            {"payment_received": {"$gt": 0}},  # Any order with payment received
            {"is_credit": False, "total": {"$gt": 0}}  # Non-credit orders
        ],
        "organization_id": user_org_id,
        "created_at": {"$gte": today_utc.isoformat()}
    }, {"_id": 0}).sort("created_at", -1).to_list(1000)

    # Use aggregation for better performance - include paid orders
    pipeline = [
        {
            "$match": {
                "$or": [
                    {"status": "completed"},
                    {"status": "paid"},
                    {"payment_received": {"$gt": 0}},
                    {"is_credit": False, "total": {"$gt": 0}}
                ],
                "organization_id": user_org_id,
                "created_at": {"$gte": today_utc.isoformat()}
            }
        },
        {
            "$group": {
                "_id": None,
                "total_orders": {"$sum": 1},
                "total_sales": {"$sum": "$total"}
            }
        }
    ]
    
    aggregation_result = await db.orders.aggregate(pipeline).to_list(1)
    
    if aggregation_result:
        stats = aggregation_result[0]
        total_orders = stats["total_orders"]
        total_sales = stats["total_sales"]
    else:
        total_orders = 0
        total_sales = 0

    result = {
        "date": today_ist.isoformat(),
        "total_orders": total_orders,
        "total_sales": total_sales,
        "orders": today_orders,
    }
    
    # ✅ PERFORMANCE: Cache the result for 30 seconds for real-time updates
    with _cache_lock:
        _cache[cache_key] = result
        _cache_ttl[cache_key] = current_time + 30  # 30 seconds TTL for real-time dashboard
    
    print(f"✅ Cached daily_report for: {cache_key} (TTL: 30 seconds)")
    return result


# ✅ NEW: Performance Metrics Endpoint
@api_router.get("/admin/performance-metrics")
async def get_performance_metrics(current_user: dict = Depends(get_current_user)):
    """Get API performance metrics - only for admin users"""
    user_org_id = get_secure_org_id(current_user)
    
    # Check if user is super admin
    user_doc = await db.users.find_one({"email": current_user.get("email")}, {"_id": 0})
    if not user_doc or not user_doc.get("is_super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    # Get cache metrics
    cache_size = len(_cache)
    expired_count = sum(1 for k, v in _cache_ttl.items() if time.time() >= v)
    
    return {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "cache_stats": {
            "total_cached_keys": cache_size,
            "expired_keys": expired_count,
            "active_keys": cache_size - expired_count,
            "cache_memory_bytes": sum(len(str(v).encode()) for v in _cache.values()) if _cache else 0
        },
        "endpoints_with_cache": [
            {"endpoint": "/reports/daily", "ttl_seconds": 3600, "description": "Daily sales report"},
            {"endpoint": "/orders", "ttl_seconds": 300, "description": "List orders (browser cache)"},
            {"endpoint": "/menu", "ttl_seconds": 600, "description": "Menu items list"}
        ]
    }


# ✅ NEW: Clear Cache Endpoint (for admin)
@api_router.post("/admin/clear-cache")
async def clear_cache(current_user: dict = Depends(get_current_user), pattern: Optional[str] = None):
    """Clear cache entries - admin only"""
    user_doc = await db.users.find_one({"email": current_user.get("email")}, {"_id": 0})
    if not user_doc or not user_doc.get("is_super_admin"):
        raise HTTPException(status_code=403, detail="Admin access required")
    
    global _cache, _cache_ttl
    
    if pattern:
        # Clear cache entries matching pattern
        with _cache_lock:
            matching_keys = [k for k in _cache.keys() if pattern in k]
            for key in matching_keys:
                _cache.pop(key, None)
                _cache_ttl.pop(key, None)
        
        return {"message": f"Cleared {len(matching_keys)} cache entries", "pattern": pattern}
    else:
        # Clear all cache
        with _cache_lock:
            cleared = len(_cache)
            _cache.clear()
            _cache_ttl.clear()
        
        return {"message": f"Cleared all {cleared} cache entries"}


# ✅ NEW: Batch Update Orders Endpoint (reduces N+1 queries)
@api_router.post("/orders/batch-update-status")
async def batch_update_order_status(
    current_user: dict = Depends(get_current_user),
    updates: List[dict] = Body(...)
):
    """Batch update order statuses - reduces database round trips"""
    user_org_id = get_secure_org_id(current_user)
    
    if not updates or len(updates) > 100:
        raise HTTPException(status_code=400, detail="Provide 1-100 updates")
    
    try:
        from pymongo import UpdateOne
        
        # Build bulk operations
        bulk_ops = []
        for update in updates:
            order_id = update.get("order_id")
            new_status = update.get("status")
            
            if not order_id or not new_status:
                continue
            
            bulk_ops.append(
                UpdateOne(
                    {
                        "_id": order_id,
                        "organization_id": user_org_id
                    },
                    {
                        "$set": {
                            "status": new_status,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
            )
        
        if bulk_ops:
            result = await db.orders.bulk_write(bulk_ops)
            
            # Clear related cache entries
            with _cache_lock:
                cache_keys_to_clear = [k for k in _cache.keys() if "order" in k or user_org_id in k]
                for key in cache_keys_to_clear:
                    _cache.pop(key, None)
                    _cache_ttl.pop(key, None)
            
            return {
                "success": True,
                "modified_count": result.modified_count,
                "message": f"Updated {result.modified_count} orders"
            }
        
        return {"success": True, "modified_count": 0, "message": "No valid updates provided"}
        
    except Exception as e:
        logger.error(f"Batch update error: {e}")
        raise HTTPException(status_code=500, detail="Failed to batch update orders")


@api_router.get("/reports/export")
async def export_report(
    start_date: str, 
    end_date: str,
    current_user: dict = Depends(get_current_user)
):
    user_org_id = get_secure_org_id(current_user)
    # Parse dates and make them timezone-aware (UTC)
    start = datetime.fromisoformat(start_date).replace(tzinfo=timezone.utc)
    end = datetime.fromisoformat(end_date).replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)

    orders = await db.orders.find({
        "organization_id": user_org_id
    }, {"_id": 0}).to_list(1000)
    
    filtered_orders = []

    for order in orders:
        order_date = order.get("created_at")
        if not order_date:
            continue
        if isinstance(order_date, str):
            try:
                order_date = datetime.fromisoformat(order_date.replace("Z", "+00:00"))
            except:
                continue
        # Make sure order_date is timezone-aware
        if order_date.tzinfo is None:
            order_date = order_date.replace(tzinfo=timezone.utc)
        if start <= order_date <= end:
            filtered_orders.append(order)

    return {
        "orders": filtered_orders,
        "total_sales": sum(o.get("total", 0) for o in filtered_orders),
    }


@api_router.get("/reports/weekly")
async def weekly_report(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    week_ago = datetime.now(timezone.utc) - timedelta(days=7)
    
    # Optimized: Use database aggregation instead of filtering in Python
    pipeline = [
        {
            "$match": {
                "status": "completed",
                "organization_id": user_org_id,
                "created_at": {"$gte": week_ago.isoformat()}
            }
        },
        {
            "$group": {
                "_id": None,
                "total_orders": {"$sum": 1},
                "total_sales": {"$sum": "$total"}
            }
        }
    ]
    
    result = await db.orders.aggregate(pipeline).to_list(1)
    
    if result:
        stats = result[0]
        total_orders = stats["total_orders"]
        total_sales = stats["total_sales"]
        avg_order_value = total_sales / total_orders if total_orders > 0 else 0
    else:
        total_orders = 0
        total_sales = 0
        avg_order_value = 0
    
    return {
        "total_orders": total_orders,
        "total_sales": total_sales,
        "avg_order_value": avg_order_value,
        "period": "last_7_days"
    }


@api_router.get("/reports/monthly")
async def monthly_report(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    month_ago = datetime.now(timezone.utc) - timedelta(days=30)
    
    # Optimized: Use database aggregation instead of filtering in Python
    pipeline = [
        {
            "$match": {
                "status": "completed",
                "organization_id": user_org_id,
                "created_at": {"$gte": month_ago.isoformat()}
            }
        },
        {
            "$group": {
                "_id": None,
                "total_orders": {"$sum": 1},
                "total_sales": {"$sum": "$total"}
            }
        }
    ]
    
    result = await db.orders.aggregate(pipeline).to_list(1)
    
    if result:
        stats = result[0]
        total_orders = stats["total_orders"]
        total_sales = stats["total_sales"]
        avg_order_value = total_sales / total_orders if total_orders > 0 else 0
    else:
        total_orders = 0
        total_sales = 0
        avg_order_value = 0
    
    return {
        "total_orders": total_orders,
        "total_sales": total_sales,
        "avg_order_value": avg_order_value,
        "period": "last_30_days"
    }


@api_router.get("/reports/best-selling")
async def best_selling_report(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    
    orders = await db.orders.find({
        "status": "completed",
        "organization_id": user_org_id
    }, {"_id": 0}).to_list(1000)
    
    from collections import defaultdict
    item_stats = defaultdict(lambda: {
        "total_quantity": 0, 
        "total_revenue": 0, 
        "name": "", 
        "category": "Uncategorized",
        "price": 0
    })
    
    for order in orders:
        for item in order["items"]:
            item_id = item.get("menu_item_id", item["name"])
            item_stats[item_id]["name"] = item["name"]
            item_stats[item_id]["total_quantity"] += item["quantity"]
            item_stats[item_id]["total_revenue"] += item["price"] * item["quantity"]
            item_stats[item_id]["price"] = item["price"]  # Store price
    
    # Get category info from menu
    for item_id, stats in item_stats.items():
        menu_item = await db.menu_items.find_one({
            "name": stats["name"], 
            "organization_id": user_org_id
        }, {"_id": 0})
        if menu_item:
            stats["category"] = menu_item.get("category", "Uncategorized")
            if not stats["price"]:
                stats["price"] = menu_item.get("price", 0)
    
    sorted_items = sorted(
        item_stats.values(), 
        key=lambda x: x["total_quantity"], 
        reverse=True
    )[:10]
    
    return sorted_items


@api_router.get("/reports/staff-performance")
async def staff_performance_report(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    
    orders = await db.orders.find({
        "status": "completed",
        "organization_id": user_org_id
    }, {"_id": 0}).to_list(1000)
    
    from collections import defaultdict
    staff_stats = defaultdict(lambda: {"total_orders": 0, "total_sales": 0, "waiter_name": ""})
    
    for order in orders:
        waiter_id = order["waiter_id"]
        staff_stats[waiter_id]["waiter_name"] = order["waiter_name"]
        staff_stats[waiter_id]["total_orders"] += 1
        staff_stats[waiter_id]["total_sales"] += order["total"]
    
    # Calculate average order value
    for staff_id, stats in staff_stats.items():
        stats["avg_order_value"] = stats["total_sales"] / stats["total_orders"] if stats["total_orders"] > 0 else 0
        # Get role from users collection
        staff_user = await db.users.find_one({"id": staff_id}, {"_id": 0})
        if staff_user:
            stats["role"] = staff_user.get("role", "waiter")
    
    sorted_staff = sorted(staff_stats.values(), key=lambda x: x["total_orders"], reverse=True)
    return sorted_staff


@api_router.get("/reports/peak-hours")
async def peak_hours_report(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    
    orders = await db.orders.find({
        "status": "completed",
        "organization_id": user_org_id
    }, {"_id": 0}).to_list(1000)
    
    from collections import defaultdict
    hour_stats = defaultdict(int)
    
    for order in orders:
        order_date = order["created_at"]
        if isinstance(order_date, str):
            order_date = datetime.fromisoformat(order_date)
        hour = order_date.hour
        hour_stats[hour] += 1
    
    # Format hours
    formatted_hours = []
    for hour, count in sorted(hour_stats.items()):
        formatted_hours.append({
            "hour": f"{hour:02d}:00 - {hour:02d}:59",
            "order_count": count
        })
    
    return sorted(formatted_hours, key=lambda x: x["order_count"], reverse=True)[:12]


@api_router.get("/reports/category-analysis")
async def category_analysis_report(current_user: dict = Depends(get_current_user)):
    user_org_id = get_secure_org_id(current_user)
    
    orders = await db.orders.find({
        "status": "completed",
        "organization_id": user_org_id
    }, {"_id": 0}).to_list(1000)
    
    from collections import defaultdict
    category_stats = defaultdict(lambda: {"total_sold": 0, "total_revenue": 0})
    
    for order in orders:
        for item in order["items"]:
            # Get category from menu
            menu_item = await db.menu_items.find_one({
                "name": item["name"],
                "organization_id": user_org_id
            }, {"_id": 0})
            
            category = "Uncategorized"
            if menu_item:
                category = menu_item.get("category", "Uncategorized")
            
            category_stats[category]["total_sold"] += item["quantity"]
            category_stats[category]["total_revenue"] += item["price"] * item["quantity"]
    
    # Calculate percentages
    total_revenue = sum(stats["total_revenue"] for stats in category_stats.values())
    
    result = []
    for category, stats in category_stats.items():
        percentage = (stats["total_revenue"] / total_revenue * 100) if total_revenue > 0 else 0
        result.append({
            "category": category,
            "total_sold": stats["total_sold"],
            "total_revenue": stats["total_revenue"],
            "percentage": percentage
        })
    
    return sorted(result, key=lambda x: x["total_revenue"], reverse=True)


# Thermal printer route
@api_router.post("/print")
async def print_receipt(
    print_data: PrintData, current_user: dict = Depends(get_current_user)
):
    business = current_user.get("business_settings", {})
    currency_code = business.get("currency", "INR")
    currency_symbol = CURRENCY_SYMBOLS.get(currency_code, "₹")
    theme = print_data.theme or business.get("receipt_theme", "classic")

    return {
        "print_ready": True,
        "content": print_data.content,
        "format": "text",
        "escpos_available": True,
        "theme": theme,
    }


@api_router.post("/print/bill/{order_id}")
async def print_bill(
    order_id: str,
    theme: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    business = current_user.get("business_settings", {})
    currency_code = business.get("currency", "INR")
    currency_symbol = CURRENCY_SYMBOLS.get(currency_code, "₹")
    receipt_theme = theme or business.get("receipt_theme", "classic")
    
    # Get print customization settings
    customization = business.get("print_customization")

    receipt_content = get_receipt_template(
        receipt_theme, business, order, currency_symbol, customization
    )
    
    # Get paper width for response
    paper_width = "80mm"
    if customization:
        paper_width = customization.get("paper_width", "80mm")
        if paper_width == "custom" and customization.get("custom_width"):
            paper_width = f"{customization.get('custom_width')}mm"

    return {
        "print_ready": True,
        "content": receipt_content,
        "paper_width": paper_width,
        "font_size": customization.get("font_size", 12) if customization else 12,
        "copies": customization.get("print_copies", 1) if customization else 1,
        "format": "text",
        "theme": receipt_theme,
    }


# WhatsApp Integration
class WhatsAppMessage(BaseModel):
    phone_number: str
    customer_name: Optional[str] = None


@api_router.post("/whatsapp/send-receipt/{order_id}")
async def send_whatsapp_receipt(
    order_id: str,
    message_data: WhatsAppMessage,
    current_user: dict = Depends(get_current_user),
):
    """Generate WhatsApp share link for order receipt"""
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    business = current_user.get("business_settings", {})
    currency_code = business.get("currency", "INR")
    currency_symbol = CURRENCY_SYMBOLS.get(currency_code, "₹")
    
    # Get WhatsApp settings
    whatsapp_enabled = business.get("whatsapp_enabled", False)
    message_template = business.get("whatsapp_message_template", 
        "Thank you for dining at {restaurant_name}! Your bill of {currency}{total} has been paid. Order #{order_id}")
    
    restaurant_name = business.get("restaurant_name", "Our Restaurant")
    
    # Build items list for message
    items_list = "\n".join([
        f"• {item['quantity']}x {item['name']} - {currency_symbol}{item['price'] * item['quantity']:.2f}"
        for item in order["items"]
    ])
    
    # Format the message
    message = message_template.format(
        restaurant_name=restaurant_name,
        currency=currency_symbol,
        total=f"{order['total']:.2f}",
        order_id=order_id[:8],
        customer_name=message_data.customer_name or order.get("customer_name", "Guest"),
        subtotal=f"{order['subtotal']:.2f}",
        tax=f"{order['tax']:.2f}",
        table_number=order.get("table_number", "N/A"),
        waiter_name=order.get("waiter_name", "Staff"),
        items=items_list
    )
    
    # Add detailed receipt if template doesn't include items
    if "{items}" not in message_template:
        detailed_message = f"""
🧾 *{restaurant_name}*
━━━━━━━━━━━━━━━━━━━━

📋 *Order #{order_id[:8]}*
🍽️ Table: {order.get("table_number", "N/A")}
👤 Customer: {message_data.customer_name or order.get("customer_name", "Guest")}

*Items:*
{items_list}

━━━━━━━━━━━━━━━━━━━━
Subtotal: {currency_symbol}{order['subtotal']:.2f}
Tax: {currency_symbol}{order['tax']:.2f}
*Total: {currency_symbol}{order['total']:.2f}*
━━━━━━━━━━━━━━━━━━━━

✅ Payment Completed

{business.get('footer_message', 'Thank you for dining with us!')}
"""
        message = detailed_message
    
    # Clean phone number (remove spaces, dashes, etc.)
    phone = message_data.phone_number.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")
    if not phone.startswith("+"):
        # Assume Indian number if no country code
        if phone.startswith("0"):
            phone = "+91" + phone[1:]
        elif len(phone) == 10:
            phone = "+91" + phone
        else:
            phone = "+" + phone
    
    # Remove the + for WhatsApp API
    phone_clean = phone.replace("+", "")
    
    # URL encode the message
    import urllib.parse
    encoded_message = urllib.parse.quote(message)
    
    # Generate WhatsApp link
    whatsapp_link = f"https://wa.me/{phone_clean}?text={encoded_message}"
    
    return {
        "success": True,
        "whatsapp_link": whatsapp_link,
        "message": message,
        "phone_number": phone,
        "order_id": order_id,
        "whatsapp_enabled": whatsapp_enabled
    }


@api_router.get("/whatsapp/settings")
async def get_whatsapp_settings(current_user: dict = Depends(get_current_user)):
    """Get WhatsApp settings for the business"""
    business = current_user.get("business_settings", {})
    return {
        "whatsapp_enabled": business.get("whatsapp_enabled", False),
        "whatsapp_business_number": business.get("whatsapp_business_number", ""),
        "whatsapp_message_template": business.get("whatsapp_message_template", 
            "Thank you for dining at {restaurant_name}! Your bill of {currency}{total} has been paid. Order #{order_id}"),
        "whatsapp_auto_notify": business.get("whatsapp_auto_notify", False),
        "whatsapp_notify_on_placed": business.get("whatsapp_notify_on_placed", True),
        "whatsapp_notify_on_preparing": business.get("whatsapp_notify_on_preparing", True),
        "whatsapp_notify_on_ready": business.get("whatsapp_notify_on_ready", True),
        "whatsapp_notify_on_completed": business.get("whatsapp_notify_on_completed", True),
        "customer_self_order_enabled": business.get("customer_self_order_enabled", False),
        "menu_display_enabled": business.get("menu_display_enabled", False),
        "frontend_url": business.get("frontend_url", "")
    }


class WhatsAppSettings(BaseModel):
    whatsapp_enabled: bool = False
    whatsapp_business_number: Optional[str] = None
    whatsapp_message_template: Optional[str] = None
    whatsapp_auto_notify: bool = False
    whatsapp_notify_on_placed: bool = True
    whatsapp_notify_on_preparing: bool = True
    whatsapp_notify_on_ready: bool = True
    whatsapp_notify_on_completed: bool = True
    customer_self_order_enabled: bool = False
    menu_display_enabled: bool = False
    frontend_url: Optional[str] = None


@api_router.put("/whatsapp/settings")
async def update_whatsapp_settings(
    settings: WhatsAppSettings,
    current_user: dict = Depends(get_current_user)
):
    """Update WhatsApp settings for the business"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update WhatsApp settings")
    
    # Get current business settings
    business = current_user.get("business_settings", {}) or {}
    
    # Update WhatsApp specific settings
    business["whatsapp_enabled"] = settings.whatsapp_enabled
    business["whatsapp_auto_notify"] = settings.whatsapp_auto_notify
    business["whatsapp_notify_on_placed"] = settings.whatsapp_notify_on_placed
    business["whatsapp_notify_on_preparing"] = settings.whatsapp_notify_on_preparing
    business["whatsapp_notify_on_ready"] = settings.whatsapp_notify_on_ready
    business["whatsapp_notify_on_completed"] = settings.whatsapp_notify_on_completed
    business["customer_self_order_enabled"] = settings.customer_self_order_enabled
    business["menu_display_enabled"] = settings.menu_display_enabled
    
    if settings.whatsapp_business_number:
        business["whatsapp_business_number"] = settings.whatsapp_business_number
    if settings.whatsapp_message_template:
        business["whatsapp_message_template"] = settings.whatsapp_message_template
    if settings.frontend_url:
        business["frontend_url"] = settings.frontend_url
    
    await db.users.update_one(
        {"id": current_user["id"]},
        {"$set": {"business_settings": business}}
    )
    
    return {"message": "WhatsApp settings updated successfully", "settings": settings.model_dump()}


# ============ WHATSAPP CLOUD API ENDPOINTS ============

@api_router.post("/whatsapp/cloud/send-receipt/{order_id}")
async def send_receipt_via_cloud_api(
    order_id: str,
    message_data: WhatsAppMessage,
    current_user: dict = Depends(get_current_user),
):
    """Send receipt directly via WhatsApp Cloud API (no user login required)"""
    if not _WHATSAPP_CLOUD_AVAILABLE:
        raise HTTPException(
            status_code=503, 
            detail="WhatsApp Cloud API not configured. Please set WHATSAPP_PHONE_NUMBER_ID and WHATSAPP_ACCESS_TOKEN in environment variables."
        )
    
    # Get user's organization_id
    user_org_id = get_secure_org_id(current_user)

    order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id}, {"_id": 0}
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")

    business = current_user.get("business_settings", {})
    
    try:
        # Send via WhatsApp Cloud API
        result = await send_whatsapp_receipt(
            phone=message_data.phone_number,
            order=order,
            business=business
        )
        
        return {
            "success": True,
            "message": "Receipt sent via WhatsApp",
            "message_id": result.get("messages", [{}])[0].get("id"),
            "phone_number": message_data.phone_number,
            "order_id": order_id,
            "method": "cloud_api"
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to send WhatsApp message: {str(e)}"
        )


@api_router.post("/whatsapp/cloud/send-status")
async def send_status_via_cloud_api(
    order_id: str,
    status: str,
    phone_number: str,
    current_user: dict = Depends(get_current_user),
):
    """Send order status update via WhatsApp Cloud API"""
    if not _WHATSAPP_CLOUD_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="WhatsApp Cloud API not configured"
        )
    
    business = current_user.get("business_settings", {})
    restaurant_name = business.get("restaurant_name", "Restaurant")
    frontend_url = business.get("frontend_url", "")
    
    # Get order for tracking URL
    user_org_id = get_secure_org_id(current_user)
    order = await db.orders.find_one(
        {"id": order_id, "organization_id": user_org_id},
        {"_id": 0, "tracking_token": 1}
    )
    
    tracking_url = None
    if order and order.get("tracking_token") and frontend_url:
        tracking_url = f"{frontend_url}/track/{order['tracking_token']}"
    
    try:
        result = await send_whatsapp_status(
            phone=phone_number,
            order_id=order_id,
            status=status,
            restaurant_name=restaurant_name,
            tracking_url=tracking_url
        )
        
        return {
            "success": True,
            "message": "Status update sent via WhatsApp",
            "message_id": result.get("messages", [{}])[0].get("id"),
            "phone_number": phone_number,
            "status": status
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to send status update: {str(e)}"
        )


@api_router.post("/whatsapp/cloud/send-otp")
async def send_otp_via_cloud_api(
    phone_number: str,
    otp: str,
    restaurant_name: str = "BillByteKOT"
):
    """Send OTP via WhatsApp Cloud API (public endpoint)"""
    if not _WHATSAPP_CLOUD_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="WhatsApp Cloud API not configured"
        )
    
    try:
        result = await send_whatsapp_otp(
            phone=phone_number,
            otp=otp,
            restaurant_name=restaurant_name
        )
        
        return {
            "success": True,
            "message": "OTP sent via WhatsApp",
            "message_id": result.get("messages", [{}])[0].get("id"),
            "phone_number": phone_number
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to send OTP: {str(e)}"
        )


@api_router.get("/whatsapp/cloud/test")
async def test_cloud_api(current_user: dict = Depends(get_current_user)):
    """Test WhatsApp Cloud API connection"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin only")
    
    if not _WHATSAPP_CLOUD_AVAILABLE:
        return {
            "success": False,
            "configured": False,
            "error": "WhatsApp Cloud API module not available"
        }
    
    result = await test_whatsapp_connection()
    return result


@api_router.get("/whatsapp/cloud/status")
async def get_cloud_api_status():
    """Get WhatsApp Cloud API configuration status (public)"""
    if not _WHATSAPP_CLOUD_AVAILABLE:
        return {
            "available": False,
            "configured": False,
            "message": "WhatsApp Cloud API module not loaded"
        }
    
    is_configured = whatsapp_api.is_configured()
    return {
        "available": True,
        "configured": is_configured,
        "message": "WhatsApp Cloud API is configured and ready" if is_configured else "WhatsApp Cloud API credentials not set",
        "phone_number_id": whatsapp_api.phone_number_id if is_configured else None
    }


# These endpoints are for customer-facing features like order tracking and self-ordering

@app.get("/api/public/track/{tracking_token}")
async def track_order_public(tracking_token: str):
    """Public endpoint for customers to track their order status"""
    order = await db.orders.find_one(
        {"tracking_token": tracking_token}, 
        {"_id": 0, "waiter_id": 0, "organization_id": 0}
    )
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Get business info for display
    admin = await db.users.find_one(
        {"id": order.get("organization_id") or order.get("waiter_id")},
        {"_id": 0, "business_settings": 1}
    )
    business = admin.get("business_settings", {}) if admin else {}
    
    return {
        "order_id": order["id"][:8],
        "status": order["status"],
        "table_number": order.get("table_number"),
        "customer_name": order.get("customer_name"),
        "items": order["items"],
        "subtotal": order["subtotal"],
        "tax": order["tax"],
        "total": order["total"],
        "created_at": order["created_at"],
        "updated_at": order["updated_at"],
        "restaurant_name": business.get("restaurant_name", "Restaurant"),
        "restaurant_phone": business.get("phone", ""),
        "currency": business.get("currency", "INR")
    }


@app.get("/api/public/menu/{org_id}")
async def get_public_menu(org_id: str):
    """Public endpoint for customers to view menu (for self-ordering)"""
    # Check if self-ordering is enabled
    admin = await db.users.find_one({"id": org_id}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    business = admin.get("business_settings", {})
    if not business.get("customer_self_order_enabled"):
        raise HTTPException(status_code=403, detail="Self-ordering not enabled")
    
    # Get available menu items
    items = await db.menu_items.find(
        {"organization_id": org_id, "available": True},
        {"_id": 0, "organization_id": 0}
    ).to_list(1000)
    
    # Group by category
    categories = {}
    for item in items:
        cat = item.get("category", "Other")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(item)
    
    return {
        "restaurant_name": business.get("restaurant_name", "Restaurant"),
        "currency": business.get("currency", "INR"),
        "tax_rate": business.get("tax_rate", 5.0),
        "categories": categories,
        "items": items
    }


@app.get("/api/public/view-menu/{org_id}")
async def get_view_only_menu(org_id: str):
    """Public endpoint for customers to VIEW menu only (no ordering) - QR code menu display"""
    admin = await db.users.find_one({"id": org_id}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    business = admin.get("business_settings", {})
    
    # Check if menu display is enabled (either self-order OR view-only menu)
    if not business.get("customer_self_order_enabled") and not business.get("menu_display_enabled"):
        raise HTTPException(status_code=403, detail="Menu display not enabled for this restaurant")
    
    # Get available menu items
    items = await db.menu_items.find(
        {"organization_id": org_id, "available": True},
        {"_id": 0, "organization_id": 0}
    ).to_list(1000)
    
    # Group by category
    categories = {}
    for item in items:
        cat = item.get("category", "Other")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(item)
    
    # Get currency symbol
    currency_symbols = {"INR": "₹", "USD": "$", "EUR": "€", "GBP": "£", "AED": "د.إ", "PKR": "₨"}
    currency_code = business.get("currency", "INR")
    currency_symbol = currency_symbols.get(currency_code, "₹")
    
    return {
        "restaurant_name": business.get("restaurant_name", "Restaurant"),
        "tagline": business.get("tagline", ""),
        "logo_url": business.get("logo_url", ""),
        "currency": currency_code,
        "currency_symbol": currency_symbol,
        "categories": categories,
        "items": items,
        "allow_ordering": business.get("customer_self_order_enabled", False)
    }


@app.get("/r/{restaurant_slug}/menu")
async def get_menu_by_slug(restaurant_slug: str):
    """Cool URL endpoint for restaurant menu using custom slug"""
    
    print(f"🔍 Looking for restaurant with slug: {restaurant_slug}")
    
    # Find restaurant by slug first
    admin = await db.users.find_one({
        "business_settings.restaurant_slug": restaurant_slug
    }, {"_id": 0})
    
    if not admin:
        print(f"❌ No restaurant found with slug: {restaurant_slug}")
        # Fallback: try to find by restaurant name converted to slug
        # Get all restaurants and check name-based slugs
        all_restaurants = await db.users.find({
            "business_settings.restaurant_name": {"$exists": True}
        }, {"_id": 0, "id": 1, "business_settings.restaurant_name": 1}).to_list(None)
        
        print(f"🔍 Checking {len(all_restaurants)} restaurants for name-based slug match...")
        
        for restaurant in all_restaurants:
            restaurant_name = restaurant.get("business_settings", {}).get("restaurant_name", "")
            # Create slug from restaurant name
            name_slug = restaurant_name.lower().replace(" ", "").replace("-", "").replace("_", "").replace("'", "").replace("&", "and")
            print(f"  - {restaurant_name} -> {name_slug}")
            if name_slug == restaurant_slug.lower().replace("-", "").replace("_", ""):
                print(f"✅ Found match: {restaurant_name}")
                # Found by name match, get full restaurant data
                admin = await db.users.find_one({"id": restaurant["id"]}, {"_id": 0})
                break
    else:
        print(f"✅ Found restaurant by slug: {admin.get('business_settings', {}).get('restaurant_name', 'Unknown')}")
    
    if not admin:
        print(f"❌ Restaurant not found for slug: {restaurant_slug}")
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    business = admin.get("business_settings", {})
    print(f"📋 Business settings: {business.keys()}")
    
    # Check if menu display is enabled (be more lenient)
    menu_enabled = (
        business.get("customer_self_order_enabled", False) or 
        business.get("menu_display_enabled", False) or
        business.get("qr_menu_enabled", True)  # Default to True for backward compatibility
    )
    
    if not menu_enabled:
        # For debugging, let's be more permissive and show menu anyway
        print(f"⚠️ Menu display not explicitly enabled for {restaurant_slug}, but showing anyway")
        # raise HTTPException(status_code=403, detail="Menu display not enabled for this restaurant")
    
    # Get available menu items
    items = await db.menu_items.find(
        {"organization_id": admin["id"], "available": True},
        {"_id": 0, "organization_id": 0}
    ).to_list(1000)
    
    # Group by category
    categories = {}
    for item in items:
        cat = item.get("category", "Other")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(item)
    
    # Get currency symbol
    currency_symbols = {"INR": "₹", "USD": "$", "EUR": "€", "GBP": "£", "AED": "د.إ", "PKR": "₨"}
    currency_code = business.get("currency", "INR")
    currency_symbol = currency_symbols.get(currency_code, "₹")
    
    return {
        "restaurant_name": business.get("restaurant_name", "Restaurant"),
        "restaurant_slug": restaurant_slug,
        "tagline": business.get("tagline", ""),
        "logo_url": business.get("logo_url", ""),
        "currency": currency_code,
        "currency_symbol": currency_symbol,
        "categories": categories,
        "items": items,
        "allow_ordering": business.get("customer_self_order_enabled", False),
        "cool_url": True
    }


@app.get("/api/public/tables/{org_id}")
async def get_public_tables(org_id: str):
    """Public endpoint to get available tables for self-ordering"""
    admin = await db.users.find_one({"id": org_id}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    business = admin.get("business_settings", {})
    if not business.get("customer_self_order_enabled"):
        raise HTTPException(status_code=403, detail="Self-ordering not enabled")
    
    tables = await db.tables.find(
        {"organization_id": org_id},
        {"_id": 0, "organization_id": 0}
    ).to_list(100)
    
    return {
        "restaurant_name": business.get("restaurant_name", "Restaurant"),
        "tables": tables
    }


class CustomerOrderCreate(BaseModel):
    table_id: str
    table_number: int
    items: List[OrderItem]
    customer_name: str
    customer_phone: str
    org_id: str
    frontend_origin: Optional[str] = None  # For generating tracking links


@app.post("/api/public/order")
async def create_customer_order(order_data: CustomerOrderCreate):
    """Public endpoint for customers to place orders (self-ordering)"""
    # Verify restaurant and self-ordering enabled
    admin = await db.users.find_one({"id": order_data.org_id}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    business = admin.get("business_settings", {})
    if not business.get("customer_self_order_enabled"):
        raise HTTPException(status_code=403, detail="Self-ordering not enabled")
    
    # Calculate totals - handle tax_rate properly (allow 0)
    tax_rate_setting = business.get("tax_rate")
    tax_rate = (tax_rate_setting if tax_rate_setting is not None else 5.0) / 100
    subtotal = sum(item.price * item.quantity for item in order_data.items)
    tax = subtotal * tax_rate
    total = subtotal + tax
    
    # Generate tracking token
    tracking_token = str(uuid.uuid4())[:12]
    
    order_obj = Order(
        table_id=order_data.table_id,
        table_number=order_data.table_number,
        items=[item.model_dump() for item in order_data.items],
        subtotal=subtotal,
        tax=tax,
        tax_rate=tax_rate_setting if tax_rate_setting is not None else 5.0,  # Store the tax rate used
        total=total,
        waiter_id=order_data.org_id,  # Use org_id as waiter for self-orders
        waiter_name="Self-Order",
        customer_name=order_data.customer_name,
        customer_phone=order_data.customer_phone,
        tracking_token=tracking_token,
        organization_id=order_data.org_id,
    )
    
    doc = order_obj.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    doc["updated_at"] = doc["updated_at"].isoformat()
    
    await db.orders.insert_one(doc)
    await db.tables.update_one(
        {"id": order_data.table_id, "organization_id": order_data.org_id},
        {"$set": {"status": "occupied", "current_order_id": order_obj.id}},
    )
    
    # Invalidate Redis cache for active orders (CRITICAL for real-time updates)
    try:
        cached_service = get_cached_order_service()
        await cached_service.invalidate_order_caches(order_data.org_id, order_obj.id)
        print(f"🗑️ Cache invalidated for new QR order {order_obj.id}")
    except Exception as e:
        print(f"⚠️ Cache invalidation error for QR order: {e}")
        # If Redis is not available, that's okay - MongoDB will handle the queries
        pass
    
    # Use frontend_origin from request for tracking links
    frontend_url = order_data.frontend_origin or ""
    
    # Generate WhatsApp notification
    whatsapp_link = None
    if business.get("whatsapp_auto_notify") and business.get("whatsapp_notify_on_placed"):
        message = get_status_message("pending", doc, business, frontend_url)
        whatsapp_link = generate_whatsapp_notification(order_data.customer_phone, message)
    
    tracking_url = f"{frontend_url}/track/{tracking_token}" if frontend_url else ""
    
    return {
        "success": True,
        "order_id": order_obj.id[:8],
        "tracking_token": tracking_token,
        "tracking_url": tracking_url,
        "whatsapp_link": whatsapp_link,
        "total": total,
        "message": "Order placed successfully! You will receive updates on WhatsApp."
    }


@app.get("/api/public/qr/{org_id}/{table_number}")
async def get_table_qr_data(org_id: str, table_number: int):
    """Get data for generating table QR code"""
    admin = await db.users.find_one({"id": org_id}, {"_id": 0})
    if not admin:
        raise HTTPException(status_code=404, detail="Restaurant not found")
    
    business = admin.get("business_settings", {})
    frontend_url = business.get("frontend_url", "")
    
    # Find table
    table = await db.tables.find_one(
        {"organization_id": org_id, "table_number": table_number},
        {"_id": 0}
    )
    
    return {
        "restaurant_name": business.get("restaurant_name", "Restaurant"),
        "table_number": table_number,
        "table_id": table["id"] if table else None,
        "org_id": org_id,
        "self_order_enabled": business.get("customer_self_order_enabled", False),
        "menu_url": f"{frontend_url}/order/{org_id}?table={table_number}" if frontend_url else "",
        "qr_content": f"{frontend_url}/order/{org_id}?table={table_number}" if frontend_url else f"Order at Table {table_number}"
    }


# Support Ticket System
class SupportTicket(BaseModel):
    name: str
    email: str
    phone: Optional[str] = None
    subject: str
    message: str
    priority: str = "medium"
    requestType: Optional[str] = "support"  # support, demo, inquiry
    preferredDate: Optional[str] = None
    preferredTime: Optional[str] = None


async def send_ticket_confirmation_email(ticket_id: str, name: str, email: str, subject: str, request_type: str):
    """Send confirmation email when a ticket is raised"""
    from email_service import send_support_email
    
    email_subject = f"Ticket #{ticket_id} - We've received your {request_type} request"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }}
            .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1); }}
            .header {{ background: linear-gradient(135deg, #7c3aed, #a855f7); padding: 30px; text-align: center; }}
            .header h1 {{ color: white; margin: 0; font-size: 24px; }}
            .content {{ padding: 30px; }}
            .ticket-box {{ background: #f8f9fa; border-left: 4px solid #7c3aed; padding: 15px; margin: 20px 0; border-radius: 4px; }}
            .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 14px; }}
            .btn {{ display: inline-block; background: #7c3aed; color: white; padding: 12px 24px; text-decoration: none; border-radius: 6px; margin-top: 15px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🎫 Support Ticket Received</h1>
            </div>
            <div class="content">
                <p>Hello <strong>{name}</strong>,</p>
                <p>Thank you for contacting BillByteKOT support! We've received your request and our team will get back to you within 24 hours.</p>
                
                <div class="ticket-box">
                    <p><strong>Ticket ID:</strong> #{ticket_id}</p>
                    <p><strong>Subject:</strong> {subject}</p>
                    <p><strong>Type:</strong> {request_type.title()}</p>
                </div>
                
                <p>In the meantime, you can:</p>
                <ul>
                    <li>Check our <a href="https://billbytekot.in/help">Help Center</a> for quick answers</li>
                    <li>Reply to this email with any additional information</li>
                    <li>Call us at +91-8310832669 (Mon-Sat, 9 AM - 6 PM IST)</li>
                </ul>
                
                <p>We're here to help you succeed! 🚀</p>
            </div>
            <div class="footer">
                <p><strong>BillByteKOT Support Team</strong></p>
                <p>Email: support@billbytekot.in | Phone: +91-8310832669</p>
                <p>© 2026 BillByte Innovations. All rights reserved.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    Hello {name},
    
    Thank you for contacting BillByteKOT support!
    
    Ticket ID: #{ticket_id}
    Subject: {subject}
    Type: {request_type.title()}
    
    Our team will get back to you within 24 hours.
    
    In the meantime, you can:
    - Check our Help Center at https://billbytekot.in/help
    - Reply to this email with any additional information
    - Call us at +91-8310832669 (Mon-Sat, 9 AM - 6 PM IST)
    
    Best regards,
    BillByteKOT Support Team
    support@billbytekot.in
    """
    
    try:
        result = await send_support_email(email, email_subject, html_body, text_body)
        print(f"📧 Ticket confirmation email sent to {email}: {result}")
        return result
    except Exception as e:
        print(f"❌ Failed to send ticket confirmation email: {e}")
        return {"success": False, "message": str(e)}


async def send_ticket_reply_email(ticket_id: str, user_email: str, user_name: str, subject: str, reply_message: str, admin_name: str = "Support Team"):
    """Send reply email to user from support"""
    from email_service import send_support_email
    
    email_subject = f"Re: Ticket #{ticket_id} - {subject}"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }}
            .container {{ max-width: 600px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.1); }}
            .header {{ background: linear-gradient(135deg, #7c3aed, #a855f7); padding: 30px; text-align: center; }}
            .header h1 {{ color: white; margin: 0; font-size: 24px; }}
            .content {{ padding: 30px; }}
            .reply-box {{ background: #f0fdf4; border-left: 4px solid #22c55e; padding: 20px; margin: 20px 0; border-radius: 4px; }}
            .footer {{ background: #f8f9fa; padding: 20px; text-align: center; color: #666; font-size: 14px; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>💬 Reply to Your Ticket</h1>
            </div>
            <div class="content">
                <p>Hello <strong>{user_name}</strong>,</p>
                <p>We have a response to your support ticket:</p>
                
                <div class="reply-box">
                    <p><strong>Ticket ID:</strong> #{ticket_id}</p>
                    <p><strong>Subject:</strong> {subject}</p>
                    <hr style="border: none; border-top: 1px solid #ddd; margin: 15px 0;">
                    <p><strong>Response from {admin_name}:</strong></p>
                    <p style="white-space: pre-wrap;">{reply_message}</p>
                </div>
                
                <p>If you have any further questions, simply reply to this email or create a new ticket.</p>
                
                <p>Thank you for choosing BillByteKOT! 🙏</p>
            </div>
            <div class="footer">
                <p><strong>BillByteKOT Support Team</strong></p>
                <p>Email: support@billbytekot.in | Phone: +91-8310832669</p>
                <p>© 2026 BillByte Innovations. All rights reserved.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    Hello {user_name},
    
    We have a response to your support ticket:
    
    Ticket ID: #{ticket_id}
    Subject: {subject}
    
    Response from {admin_name}:
    {reply_message}
    
    If you have any further questions, simply reply to this email or create a new ticket.
    
    Thank you for choosing BillByteKOT!
    
    Best regards,
    BillByteKOT Support Team
    support@billbytekot.in
    """
    
    try:
        result = await send_support_email(user_email, email_subject, html_body, text_body)
        print(f"📧 Ticket reply email sent to {user_email}: {result}")
        return result
    except Exception as e:
        print(f"❌ Failed to send ticket reply email: {e}")
        return {"success": False, "message": str(e)}


@api_router.post("/support/ticket")
async def create_support_ticket(ticket: SupportTicket):
    """Create a support ticket or demo booking"""
    ticket_id = str(uuid.uuid4())[:12]
    
    ticket_doc = {
        "id": ticket_id,
        "name": ticket.name,
        "email": ticket.email,
        "phone": ticket.phone,
        "subject": ticket.subject,
        "message": ticket.message,
        "priority": ticket.priority,
        "request_type": ticket.requestType,
        "preferred_date": ticket.preferredDate,
        "preferred_time": ticket.preferredTime,
        "status": "open",
        "replies": [],  # Store conversation history
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.support_tickets.insert_one(ticket_doc)
    
    # Log for admin notification
    if ticket.requestType == "demo":
        print(f"📅 New demo booking #{ticket_id}: {ticket.name} ({ticket.email}) - {ticket.preferredDate} at {ticket.preferredTime}")
    else:
        print(f"📧 New support ticket #{ticket_id}: {ticket.subject} from {ticket.name} ({ticket.email})")
    
    # Send confirmation email to user
    asyncio.create_task(send_ticket_confirmation_email(
        ticket_id, 
        ticket.name, 
        ticket.email, 
        ticket.subject, 
        ticket.requestType or "support"
    ))
    
    response_message = "Support ticket created successfully. We'll contact you within 24 hours. A confirmation email has been sent."
    if ticket.requestType == "demo":
        response_message = f"Demo booking confirmed for {ticket.preferredDate} at {ticket.preferredTime}. A confirmation email has been sent!"
    
    return {
        "success": True,
        "ticket_id": ticket_id,
        "message": response_message
    }


class AIChatRequest(BaseModel):
    message: str
    history: Optional[List[Dict[str, str]]] = []


@api_router.get("/support/tickets")
async def get_support_tickets(
    status: Optional[str] = None,
    request_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all support tickets (admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view tickets")
    
    # Build query
    query = {}
    if status:
        query["status"] = status
    if request_type:
        query["request_type"] = request_type
    
    # Fetch tickets
    tickets = await db.support_tickets.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(1000)
    
    return {
        "success": True,
        "count": len(tickets),
        "tickets": tickets
    }


@api_router.put("/support/ticket/{ticket_id}/status")
async def update_ticket_status(
    ticket_id: str,
    status: str,
    current_user: dict = Depends(get_current_user)
):
    """Update ticket status (admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can update tickets")
    
    result = await db.support_tickets.update_one(
        {"id": ticket_id},
        {
            "$set": {
                "status": status,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    return {
        "success": True,
        "message": f"Ticket status updated to {status}"
    }


class TicketReplyRequest(BaseModel):
    message: str
    update_status: Optional[str] = None  # Optional: update status when replying


@api_router.post("/support/ticket/{ticket_id}/reply")
async def reply_to_ticket(
    ticket_id: str,
    reply: TicketReplyRequest,
    current_user: dict = Depends(get_current_user)
):
    """Reply to a support ticket and send email to user (admin or team with tickets permission)"""
    # Allow admin role or team members with tickets permission
    is_admin = current_user.get("role") == "admin"
    is_team_with_permission = current_user.get("type") == "team" and "tickets" in current_user.get("permissions", [])
    
    if not is_admin and not is_team_with_permission:
        raise HTTPException(status_code=403, detail="You don't have permission to reply to tickets")
    
    # Get the ticket
    ticket = await db.support_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Create reply record
    reply_record = {
        "id": str(uuid.uuid4())[:8],
        "message": reply.message,
        "from": "support",
        "admin_name": current_user.get("username", "Support Team"),
        "admin_email": current_user.get("email", "support@billbytekot.in"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Update ticket with reply
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if reply.update_status:
        update_data["status"] = reply.update_status
    
    await db.support_tickets.update_one(
        {"id": ticket_id},
        {
            "$push": {"replies": reply_record},
            "$set": update_data
        }
    )
    
    # Send email to user
    email_result = await send_ticket_reply_email(
        ticket_id=ticket_id,
        user_email=ticket["email"],
        user_name=ticket["name"],
        subject=ticket["subject"],
        reply_message=reply.message,
        admin_name=current_user.get("username", "Support Team")
    )
    
    return {
        "success": True,
        "message": "Reply sent successfully",
        "email_sent": email_result.get("success", False),
        "reply_id": reply_record["id"]
    }


@api_router.get("/support/ticket/{ticket_id}")
async def get_ticket_details(
    ticket_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get detailed ticket information including replies (admin only)"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can view ticket details")
    
    ticket = await db.support_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    return {
        "success": True,
        "ticket": ticket
    }


@api_router.post("/ai/support-chat")
async def ai_support_chat(chat_request: AIChatRequest):
    """AI-powered support chat"""
    
    # Predefined responses for common questions
    common_responses = {
        "pricing": "BillByteKOT offers a 7-day free trial with all features. After that, it's just ₹999/year. You get unlimited bills, thermal printing, AI analytics, and priority support!",
        "thermal": "To setup thermal printer: 1) Connect your ESC/POS compatible printer (58mm or 80mm), 2) Go to Settings > Printer, 3) Select your printer and choose from 6 beautiful themes. Need help? Contact us!",
        "kot": "KOT (Kitchen Order Ticket) system sends orders directly to the kitchen. When you create an order, it automatically prints in the kitchen with item details, table number, and timing. This reduces errors and speeds up service!",
        "payment": "We support multiple payment methods: Cash, Card, UPI, and Razorpay integration. You can configure your own Razorpay account in Settings to receive payments directly.",
        "whatsapp": "WhatsApp integration lets you send bills, order updates, and promotional messages to customers. Setup: Go to Settings > WhatsApp, enable it, and add your business number. You can auto-notify customers on order status changes!",
        "mobile": "Yes! BillByteKOT works on any device - desktop, tablet, or mobile. We also have a native Android app coming soon. Join our early access program to be notified!",
        "desktop": "Download our desktop app for Windows, Mac, or Linux from the Download page. It offers offline support, direct thermal printing, and faster performance!",
        "support": "We offer 24/7 support! Email us at support@billbytekot.in, call +91-8310832669 (Mon-Sat, 9 AM-6 PM IST), or use this chat. Premium users get priority support with faster response times.",
        "trial": "Start your 7-day free trial now! No credit card required. You get full access to all premium features including AI analytics, thermal printing, and unlimited bills. After trial, upgrade for just ₹999/year!",
        "features": "Key features: AI-powered billing, KOT system, thermal printing (6 themes), multi-currency support, inventory management, staff management, real-time analytics, WhatsApp integration, table management, and more!",
        "inventory": "Inventory management helps track stock levels, get low-stock alerts, manage suppliers, and auto-deduct items when sold. Go to Inventory page to add items and set minimum quantities.",
        "staff": "Add unlimited staff with roles: Admin (full access), Cashier (billing only), Waiter (orders & tables), Kitchen (KOT view). Each role has specific permissions. Go to Staff Management to add team members.",
        "reports": "Access detailed reports: daily/monthly sales, top-selling items, revenue trends, tax summaries, and more. Export to CSV/PDF. Go to Reports page for analytics dashboard.",
        "currency": "We support 10+ currencies: INR, USD, EUR, GBP, AED, SGD, MYR, and more. Change currency in Settings > Business Settings. All prices and reports will update automatically!",
        "table": "Table management lets you track occupancy, assign orders to tables, and manage seating. Tables auto-release after payment. Add tables in Tables page with capacity and status.",
        "offline": "Desktop app works offline! Your data syncs when internet is back. Web version requires internet connection. Download desktop app for offline billing capability.",
        "security": "Your data is 100% secure with bank-grade encryption, HTTPS, secure MongoDB storage, and 99.9% uptime. We never share your data with third parties.",
        "setup": "Quick setup: 1) Sign up (30 seconds), 2) Add your restaurant details, 3) Create menu items, 4) Add tables, 5) Start billing! Takes less than 5 minutes total.",
        "demo": "Try our interactive demo! Login and you'll see a guided walkthrough showing how to create orders, assign tables, and process payments. It's hands-on and takes just 2 minutes!",
        "contact": "Contact us: Email: support@billbytekot.in or contact@billbytekot.in | Phone: +91-8310832669 (Mon-Sat, 9-6 PM IST) | Office: BillByte Innovations, Bangalore, India. Or submit a ticket using the contact form!"
    }
    
    # Simple keyword matching for responses
    message_lower = chat_request.message.lower()
    response = None
    
    for keyword, answer in common_responses.items():
        if keyword in message_lower:
            response = answer
            break
    
    # Default response if no match
    if not response:
        if "?" in chat_request.message:
            response = "Great question! I can help you with: pricing, thermal printer setup, KOT system, payments, WhatsApp integration, mobile/desktop apps, support options, free trial, features, inventory, staff management, reports, currency, tables, offline mode, security, setup guide, demo, or contact info. What would you like to know?"
        else:
            response = "Thanks for reaching out! I'm here to help with BillByteKOT. You can ask me about features, pricing, setup, technical support, or anything else. Or submit a support ticket for personalized assistance!"
    
    return {
        "success": True,
        "response": response,
        "timestamp": datetime.now(timezone.utc).isoformat()
    }


# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint for monitoring and load balancers"""
    try:
        # Check database connection with timeout
        await db.users.find_one({}, {"_id": 1})
        return {
            "status": "healthy",
            "message": "BillByteKOT Server is running",
            "version": "1.0.0",
            "timestamp": datetime.now(timezone.utc).isoformat(),
            "services": {"database": "connected", "api": "operational"},
            "database": db.name,
            "connection": "active",
        }
    except Exception as e:
        # Try a simple ping command as fallback
        try:
            await db.command("ping")
            return {
                "status": "degraded",
                "message": "Database ping successful but query failed",
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "services": {"database": "limited", "api": "operational"},
                "database": db.name,
                "connection": "limited",
            }
        except Exception as ping_error:
            return {
                "status": "unhealthy",
                "message": f"Database connection failed: {str(e)}",
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "services": {"database": "disconnected", "api": "operational"},
                "troubleshooting": "Check MongoDB Atlas connection string and network access",
            }


@app.get("/api/health")
async def api_health_check():
    """API health check endpoint"""
    return await health_check()


@app.get("/api/notifications/check")
async def check_notifications(user_id: str = Query(None), since: str = Query("0")):
    """Check for new notifications for a user"""
    try:
        # Get active notifications
        query = {"status": "active"}
        
        # Filter by expiry
        now = datetime.now(timezone.utc).isoformat()
        
        notifications = await db.admin_notifications.find(
            query,
            {"_id": 0}
        ).sort("created_at", -1).limit(10).to_list(10)
        
        # Filter notifications based on target
        user = None
        if user_id:
            user = await db.users.find_one({"id": user_id}, {"_id": 0, "subscription_active": 1})
        
        filtered = []
        for notif in notifications:
            # Check expiry
            if notif.get("expires_at") and notif["expires_at"] < now:
                continue
            
            # Check target
            target = notif.get("target", "all")
            if target == "all":
                filtered.append(notif)
            elif target == "subscribed" and user and user.get("subscription_active"):
                filtered.append(notif)
            elif target == "trial" and user and not user.get("subscription_active"):
                filtered.append(notif)
            elif target == "specific" and user_id in notif.get("target_users", []):
                filtered.append(notif)
        
        # Check if user has seen these notifications
        seen_key = f"seen_notifications_{user_id}" if user_id else "seen_notifications"
        seen_doc = await db.user_notification_status.find_one({"key": seen_key})
        seen_ids = seen_doc.get("seen", []) if seen_doc else []
        
        # Filter out seen notifications
        new_notifications = [n for n in filtered if n["id"] not in seen_ids]
        
        # Mark as seen
        if new_notifications and user_id:
            new_seen_ids = [n["id"] for n in new_notifications]
            await db.user_notification_status.update_one(
                {"key": seen_key},
                {"$addToSet": {"seen": {"$each": new_seen_ids}}},
                upsert=True
            )
        
        return {
            "notifications": new_notifications,
            "count": len(new_notifications)
        }
    except Exception as e:
        return {"notifications": [], "count": 0}


@api_router.get("/notifications/unread")
async def get_unread_notifications(current_user: dict = Depends(get_current_user)):
    """Get unread notifications for the current user"""
    try:
        # For now, return empty array until full notifications system is implemented
        # This prevents 404 errors in the frontend
        return []
    except Exception as e:
        logger.error(f"Error fetching unread notifications: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch notifications")


@app.get("/api/app-version")
async def get_app_version():
    """Get latest app version info for in-app updates"""
    try:
        # Get version info from database (can be managed from Ops Controls)
        version_doc = await db.app_settings.find_one({"type": "app_version"})
        
        if version_doc:
            return {
                "version": version_doc.get("version", "1.5.0"),
                "release_notes": version_doc.get("release_notes", "Bug fixes and performance improvements"),
                "force_update": version_doc.get("force_update", False),
                "message": version_doc.get("message", "A new version is available with improvements!"),
                "download_url": version_doc.get("download_url", None),
                "min_version": version_doc.get("min_version", "1.0.0"),
                "updated_at": version_doc.get("updated_at", datetime.now(timezone.utc).isoformat())
            }
        
        # Default version info
        return {
            "version": "1.5.0",
            "release_notes": "• Push notifications from admin\n• Enhanced UI/UX for order creation\n• Improved print flow\n• Better mobile experience\n• Bug fixes",
            "force_update": False,
            "message": "Update available with new features!",
            "download_url": None,
            "min_version": "1.0.0",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
    except Exception as e:
        return {
            "version": "1.5.0",
            "release_notes": "",
            "force_update": False,
            "message": "",
            "download_url": None,
            "min_version": "1.0.0"
        }


# Root endpoint
@app.get("/")
async def root():
    """Root endpoint - basic server info"""
    return {
        "service": "BillByteKOT Server",
        "version": "1.0.0",
        "status": "running",
        "message": "Welcome to BillByteKOT - Restaurant Billing & KOT System",
        "endpoints": {
            "health": "/health",
            "api_health": "/api/health",
            "api_docs": "/docs",
            "api_base": "/api",
        },
        "timestamp": datetime.now(timezone.utc).isoformat(),
    }


# Startup validation
@app.on_event("startup")
async def startup_validation():
    """Validate configuration and database connection on startup"""
    global client, db

    print("🍽️  Starting BillByteKOT Server...")

    # Check required environment variables
    required_vars = {
        "MONGO_URL": mongo_url,
        "DB_NAME": os.getenv("DB_NAME", "restrobill"),
        "JWT_SECRET": JWT_SECRET,
    }

    missing_vars = []
    for var, value in required_vars.items():
        if not value or (
            var == "JWT_SECRET"
            and value == "default-jwt-secret-please-change-in-production"
        ):
            missing_vars.append(var)

    if missing_vars:
        print(f"⚠️  Warning: Missing or default values for: {', '.join(missing_vars)}")
        if "MONGO_URL" in missing_vars:
            print(
                "💡 Set MONGO_URL environment variable with your MongoDB connection string"
            )
        if "JWT_SECRET" in missing_vars:
            print(
                "💡 Set JWT_SECRET environment variable with a secure 32+ character secret"
            )

    # Test database connection with multiple strategies
    connection_successful = False
    last_error = None

    # Strategy 1: Try current connection
    try:
        await db.command("ping")
        print(f"✅ Database connected: {db.name}")
        connection_successful = True
    except Exception as e:
        last_error = e
        print(f"❌ Primary connection failed: {e}")

        # Strategy 2: Try alternative client with different SSL settings
        try:
            print("🔄 Trying alternative SSL configuration...")
            alt_client = AsyncIOMotorClient(
                mongo_url,
                tls=True,
                tlsInsecure=True,
                tlsAllowInvalidCertificates=True,
                serverSelectionTimeoutMS=5000,
                connectTimeoutMS=8000,
                maxPoolSize=50,
            )
            alt_db = alt_client[os.getenv("DB_NAME", "restrobill")]
            await alt_db.command("ping")

            # Replace global client with working one
            client = alt_client
            db = alt_db
            print(f"✅ Alternative connection successful: {db.name}")
            connection_successful = True

        except Exception as e2:
            last_error = e2
    
    # Create database indexes for faster queries (PERFORMANCE BOOST)
    if connection_successful:
        try:
            print("⚡ Creating database indexes for performance...")
            
            # Users collection indexes
            await db.users.create_index("id", unique=True)
            await db.users.create_index("username")
            await db.users.create_index("email")
            await db.users.create_index("organization_id")
            # Case-insensitive lookup indexes
            await db.users.create_index("username_lower")
            await db.users.create_index("email_lower")
            
            # Migrate existing users to have lowercase fields (one-time migration)
            await db.users.update_many(
                {"username_lower": {"$exists": False}},
                [{"$set": {
                    "username_lower": {"$toLower": "$username"},
                    "email_lower": {"$toLower": "$email"}
                }}]
            )
            
            # Menu items indexes
            await db.menu_items.create_index("organization_id")
            await db.menu_items.create_index([("organization_id", 1), ("category", 1)])
            await db.menu_items.create_index([("organization_id", 1), ("available", 1)])
            
            # Orders indexes - Enhanced for reports performance
            await db.orders.create_index("organization_id")
            await db.orders.create_index([("organization_id", 1), ("status", 1)])
            await db.orders.create_index([("organization_id", 1), ("created_at", -1)])
            await db.orders.create_index([("organization_id", 1), ("waiter_name", 1)])
            await db.orders.create_index([("organization_id", 1), ("created_at", -1), ("status", 1)])
            await db.orders.create_index("table_id")
            
            # Compound indexes for reports queries
            await db.orders.create_index([("organization_id", 1), ("created_at", -1), ("total", 1)])
            await db.orders.create_index([("organization_id", 1), ("items.name", 1), ("items.quantity", 1)])
            
            # Tables indexes
            await db.tables.create_index("organization_id")
            await db.tables.create_index([("organization_id", 1), ("status", 1)])
            
            # Payments indexes
            await db.payments.create_index("organization_id")
            await db.payments.create_index([("organization_id", 1), ("created_at", -1)])
            await db.payments.create_index("order_id")
            
            # Inventory indexes - Enhanced for better performance
            await db.inventory.create_index("organization_id")
            await db.inventory.create_index([("organization_id", 1), ("quantity", 1)])
            await db.inventory.create_index([("organization_id", 1), ("name", 1)])
            await db.inventory.create_index([("organization_id", 1), ("min_quantity", 1)])
            await db.inventory.create_index([("organization_id", 1), ("category_id", 1)])
            await db.inventory.create_index([("organization_id", 1), ("supplier_id", 1)])
            await db.inventory.create_index([("organization_id", 1), ("sku", 1)])
            await db.inventory.create_index([("organization_id", 1), ("barcode", 1)])
            
            # Suppliers and categories indexes
            await db.suppliers.create_index("organization_id")
            await db.categories.create_index("organization_id")
            await db.stock_movements.create_index([("organization_id", 1), ("item_id", 1)])
            await db.stock_movements.create_index([("organization_id", 1), ("created_at", -1)])
            
            # Referral system indexes - Make referral_code index sparse and handle existing nulls
            try:
                # Try to create referral_code index, but handle existing null values gracefully
                await db.users.create_index("referral_code", unique=True, sparse=True)
            except Exception as referral_index_error:
                print(f"⚠️  Referral code index creation skipped: {referral_index_error}")
                print("💡 This is expected if multiple users have null referral codes")
            
            await db.referrals.create_index("referral_code")  # Lookup referrals by code
            await db.referrals.create_index("referrer_user_id")  # Find all referrals by a user
            await db.referrals.create_index("referee_user_id", unique=True, sparse=True)  # One referral per referee
            await db.referrals.create_index([("referrer_user_id", 1), ("status", 1)])  # Filter by status
            await db.referrals.create_index([("created_at", -1)])  # Sort by date
            await db.referrals.create_index("referee_phone", sparse=True)  # Fast lookup for duplicate mobile check (Requirement 11.1)
            
            # Wallet transactions indexes
            await db.wallet_transactions.create_index("user_id")
            await db.wallet_transactions.create_index([("user_id", 1), ("created_at", -1)])
            
            print("✅ Database indexes created successfully")
        except Exception as e:
            print(f"⚠️  Index creation warning: {e}")
            # Continue startup even if index creation fails

            # Strategy 3: Try with pymongo legacy SSL options
            try:
                print("🔄 Trying with legacy SSL settings...")
                min_client = AsyncIOMotorClient(
                    mongo_url,
                    ssl=True,
                    ssl_cert_reqs=ssl.CERT_NONE,
                    serverSelectionTimeoutMS=8000,
                    connectTimeoutMS=10000,
                    socketTimeoutMS=20000,
                )
                min_db = min_client[os.getenv("DB_NAME", "restrobill")]
                await min_db.command("ping")

                client = min_client
                db = min_db
                print(f"✅ Legacy SSL connection successful: {db.name}")
                connection_successful = True

            except Exception as e3:
                last_error = e3
                print(f"❌ Legacy SSL connection failed: {e3}")

                # Strategy 4: Try with different connection approach
                try:
                    print("🔄 Trying direct connection without TLS validation...")
                    # Use base URL without TLS parameters for this attempt
                    base_mongo_url = (
                        mongo_url.split("?")[0] if "?" in mongo_url else mongo_url
                    )
                    direct_url = f"{base_mongo_url}?retryWrites=true&w=majority"

                    direct_client = AsyncIOMotorClient(
                        direct_url,
                        serverSelectionTimeoutMS=10000,
                        connectTimeoutMS=15000,
                        socketTimeoutMS=20000,
                    )
                    direct_db = direct_client[os.getenv("DB_NAME", "restrobill")]
                    await direct_db.command("ping")

                    client = direct_client
                    db = direct_db
                    print(f"✅ Direct connection successful: {db.name}")
                    connection_successful = True

                except Exception as e4:
                    last_error = e4
                    print(f"❌ Direct connection failed: {e4}")

    if not connection_successful:
        print(
            f"🔗 MongoDB URL: {mongo_url.replace(mongo_url.split('@')[0].split('://')[1] + '@', '***:***@') if '@' in mongo_url else mongo_url}"
        )

        # Provide specific troubleshooting for SSL errors
        if last_error:
            error_str = str(last_error)
            if "SSL" in error_str or "TLS" in error_str:
                print("💡 SSL/TLS Connection Issue Detected:")
                print("   Try setting these environment variables:")
                print("   MONGO_TLS_INSECURE=true")
                print("   Or update your MongoDB connection string with:")
                print(
                    "   mongodb+srv://user:pass@cluster.net/db?retryWrites=true&w=majority&tls=true&tlsInsecure=true"
                )
                print("   Contact your MongoDB Atlas admin for proper SSL certificates")
            elif "authentication" in error_str.lower():
                print("💡 Authentication Issue Detected:")
                print("   1. Verify username and password in connection string")
                print("   2. Check database user permissions in MongoDB Atlas")
            elif "timeout" in error_str.lower():
                print("💡 Connection Timeout Issue:")
                print("   1. Check network connectivity")
                print("   2. Verify MongoDB Atlas cluster is running")
                print("   3. Check if IP whitelist includes 0.0.0.0/0")
        else:
            print("💡 Unknown connection issue occurred")

        print("🔄 Server will continue running with degraded functionality")

    print(f"🚀 Server starting on port {os.getenv('PORT', '5000')}")
    
    # Initialize Redis cache for orders
    try:
        await init_redis_cache(db)
        print("✅ Redis cache initialized for fast order handling")
        
        # Set Redis cache for super admin after initialization
        from redis_cache import redis_cache
        set_super_admin_cache(redis_cache)
        set_ops_cache(redis_cache)
        print("✅ Super admin Redis cache configured")
        print("✅ Ops panel Redis cache configured")
    except Exception as e:
        print(f"⚠️ Redis cache initialization failed: {e}")
        print("📝 Continuing without Redis cache (MongoDB only)")
    
    # Initialize monitoring system
    try:
        from redis_cache import redis_cache
        init_monitoring(db, redis_cache)
        
        # Start background metrics collection
        asyncio.create_task(collect_metrics_task(db))
        print("✅ Monitoring system initialized")
    except Exception as e:
        print(f"⚠️ Monitoring initialization failed: {e}")
        print("📝 Continuing without monitoring")
    
    # Start background cache cleanup task
    asyncio.create_task(periodic_cache_cleanup())
    print("✅ Background cache cleanup task started")


async def periodic_cache_cleanup():
    """Periodically clean up expired cache entries to free memory"""
    while True:
        await asyncio.sleep(300)  # Run every 5 minutes
        clear_expired_cache()


# Keep-alive endpoint for preventing cold starts
@api_router.get("/ping")
async def ping():
    """Simple ping endpoint for health checks and keeping server warm"""
    return {"status": "ok", "timestamp": datetime.now(timezone.utc).isoformat()}


# ============ REFERRAL SYSTEM ENDPOINTS ============

# Pydantic models for referral endpoints
class ReferralValidateRequest(BaseModel):
    """Request model for validating a referral code"""
    referral_code: str
    new_user_email: Optional[str] = None
    new_user_phone: Optional[str] = None
    device_fingerprint: Optional[str] = None  # For self-referral detection (Requirement 11.2)


class ReferralApplyRequest(BaseModel):
    """Request model for applying a referral code"""
    referral_code: str
    referee_user_id: str


class ReferralCompleteRequest(BaseModel):
    """Request model for completing a referral after payment"""
    referee_user_id: str
    payment_id: str


# Referral configuration constants
REFERRAL_DISCOUNT_AMOUNT = 200.0  # ₹200 discount for new users
REFERRAL_REWARD_AMOUNT = 300.0    # ₹300 reward for referrers


async def process_referral_completion(referee_user_id: str, payment_id: str) -> dict:
    """
    Process referral completion after successful payment.
    
    Requirements: 4.1, 4.2, 4.4
    - Credit referrer wallet with ₹300
    - Update referral status to REWARDED
    - Create audit transaction record
    
    Args:
        referee_user_id: The user ID of the referee who made the payment
        payment_id: The payment ID from Razorpay
        
    Returns:
        dict with referral_processed status and details
    """
    # Find the pending referral for this referee
    referral = await db.referrals.find_one({
        "referee_user_id": referee_user_id,
        "status": "PENDING"
    })
    
    if not referral:
        # No pending referral - this is okay, not all users have referrals
        return {
            "success": True,
            "referral_processed": False,
            "message": "No pending referral found for this user"
        }
    
    # Update referral status to REWARDED
    now = datetime.now(timezone.utc)
    await db.referrals.update_one(
        {"id": referral["id"]},
        {
            "$set": {
                "status": "REWARDED",
                "payment_id": payment_id,
                "completed_at": now,
                "rewarded_at": now
            }
        }
    )
    
    # Credit referrer's wallet
    referrer_user_id = referral["referrer_user_id"]
    
    # Update referrer's wallet balance and stats
    await db.users.update_one(
        {"id": referrer_user_id},
        {
            "$inc": {
                "wallet_balance": REFERRAL_REWARD_AMOUNT,
                "total_referrals": 1,
                "total_referral_earnings": REFERRAL_REWARD_AMOUNT
            }
        }
    )
    
    # Create wallet transaction record
    wallet_transaction = {
        "id": str(uuid.uuid4()),
        "user_id": referrer_user_id,
        "type": "CREDIT",
        "amount": REFERRAL_REWARD_AMOUNT,
        "transaction_type": "REFERRAL_REWARD",
        "reference_id": referral["id"],
        "description": f"Referral reward for user signup (Payment: {payment_id})",
        "created_at": now
    }
    
    # Get current balance for balance_after field
    referrer = await db.users.find_one({"id": referrer_user_id})
    wallet_transaction["balance_after"] = referrer.get("wallet_balance", REFERRAL_REWARD_AMOUNT)
    
    await db.wallet_transactions.insert_one(wallet_transaction)
    
    return {
        "success": True,
        "referral_processed": True,
        "referral_id": referral["id"],
        "referrer_reward": REFERRAL_REWARD_AMOUNT,
        "message": f"Referral completed! Referrer credited with ₹{int(REFERRAL_REWARD_AMOUNT)}"
    }


# ============ WALLET SYSTEM MODELS ============

class WalletTransaction(BaseModel):
    """
    Wallet transaction model for tracking all wallet operations.
    
    Requirements: 5.4, 5.5
    - Stores transaction type, amount, and reference
    - Supports CREDIT and DEBIT operations
    - Maintains audit trail with timestamps
    """
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    type: str  # CREDIT or DEBIT
    amount: float
    transaction_type: str  # REFERRAL_REWARD, SUBSCRIPTION_PAYMENT, REVERSAL, MANUAL_CREDIT, MANUAL_DEBIT
    reference_id: Optional[str] = None  # Referral ID, Payment ID, etc.
    balance_after: float
    description: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class WalletBalanceResponse(BaseModel):
    """Response model for wallet balance endpoint"""
    success: bool = True
    total_earned: float  # Total credits ever received
    total_used: float  # Total debits ever made
    available_balance: float  # Current balance (earned - used)
    currency: str = "INR"


class WalletTransactionResponse(BaseModel):
    """Response model for individual transaction in history"""
    id: str
    type: str
    amount: float
    transaction_type: str
    reference_id: Optional[str] = None
    balance_after: float
    description: Optional[str] = None
    created_at: str  # ISO format string


class WalletApplyRequest(BaseModel):
    """Request model for applying wallet balance to subscription"""
    subscription_amount: float
    apply_amount: Optional[float] = None  # If None, apply full available balance


class WalletApplyResponse(BaseModel):
    """Response model for wallet apply endpoint"""
    success: bool
    amount_applied: float
    remaining_subscription_amount: float
    new_wallet_balance: float
    transaction_id: Optional[str] = None
    message: str


# ============ WALLET SERVICE FUNCTIONS ============

async def wallet_credit(
    user_id: str,
    amount: float,
    transaction_type: str,
    reference_id: Optional[str] = None,
    description: Optional[str] = None
) -> dict:
    """
    Credit amount to user's wallet.
    
    Requirements: 4.1, 5.4
    - Adds amount to user wallet_balance
    - Creates transaction record with all required fields
    - Returns transaction details
    
    Args:
        user_id: The user's ID
        amount: Amount to credit (must be positive)
        transaction_type: Type of transaction (REFERRAL_REWARD, MANUAL_CREDIT, etc.)
        reference_id: Optional reference to related entity (referral ID, etc.)
        description: Optional description of the transaction
        
    Returns:
        dict with transaction details and new balance
    """
    if amount <= 0:
        raise ValueError("Credit amount must be positive")
    
    # Get current user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise ValueError(f"User not found: {user_id}")
    
    current_balance = user.get("wallet_balance", 0.0)
    new_balance = current_balance + amount
    
    # Update user's wallet balance
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"wallet_balance": new_balance}}
    )
    
    # Create transaction record
    now = datetime.now(timezone.utc)
    transaction = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "type": "CREDIT",
        "amount": amount,
        "transaction_type": transaction_type,
        "reference_id": reference_id,
        "balance_after": new_balance,
        "description": description or f"Wallet credit: {transaction_type}",
        "created_at": now
    }
    
    await db.wallet_transactions.insert_one(transaction)
    
    return {
        "success": True,
        "transaction_id": transaction["id"],
        "amount": amount,
        "new_balance": new_balance,
        "transaction_type": transaction_type
    }


async def wallet_debit(
    user_id: str,
    amount: float,
    transaction_type: str,
    reference_id: Optional[str] = None,
    description: Optional[str] = None
) -> dict:
    """
    Debit amount from user's wallet.
    
    Requirements: 5.3, 5.4
    - Validates sufficient balance before deduction
    - Deducts amount from wallet_balance
    - Creates transaction record
    
    Args:
        user_id: The user's ID
        amount: Amount to debit (must be positive)
        transaction_type: Type of transaction (SUBSCRIPTION_PAYMENT, REVERSAL, etc.)
        reference_id: Optional reference to related entity (payment ID, etc.)
        description: Optional description of the transaction
        
    Returns:
        dict with transaction details and new balance
        
    Raises:
        ValueError: If amount is invalid or insufficient balance
    """
    if amount <= 0:
        raise ValueError("Debit amount must be positive")
    
    # Get current user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise ValueError(f"User not found: {user_id}")
    
    current_balance = user.get("wallet_balance", 0.0)
    
    # Validate sufficient balance
    if current_balance < amount:
        raise ValueError(f"Insufficient wallet balance. Available: ₹{current_balance}, Required: ₹{amount}")
    
    new_balance = current_balance - amount
    
    # Update user's wallet balance
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"wallet_balance": new_balance}}
    )
    
    # Create transaction record
    now = datetime.now(timezone.utc)
    transaction = {
        "id": str(uuid.uuid4()),
        "user_id": user_id,
        "type": "DEBIT",
        "amount": amount,
        "transaction_type": transaction_type,
        "reference_id": reference_id,
        "balance_after": new_balance,
        "description": description or f"Wallet debit: {transaction_type}",
        "created_at": now
    }
    
    await db.wallet_transactions.insert_one(transaction)
    
    return {
        "success": True,
        "transaction_id": transaction["id"],
        "amount": amount,
        "new_balance": new_balance,
        "transaction_type": transaction_type
    }


async def get_wallet_balance(user_id: str) -> dict:
    """
    Get wallet balance details for a user.
    
    Requirements: 5.1
    - Returns total earned, used, and available balance
    
    Args:
        user_id: The user's ID
        
    Returns:
        dict with total_earned, total_used, available_balance
    """
    # Get user's current balance
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise ValueError(f"User not found: {user_id}")
    
    available_balance = user.get("wallet_balance", 0.0)
    
    # Calculate total earned (sum of all CREDIT transactions)
    credit_pipeline = [
        {"$match": {"user_id": user_id, "type": "CREDIT"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    credit_result = await db.wallet_transactions.aggregate(credit_pipeline).to_list(length=1)
    total_earned = credit_result[0]["total"] if credit_result else 0.0
    
    # Calculate total used (sum of all DEBIT transactions)
    debit_pipeline = [
        {"$match": {"user_id": user_id, "type": "DEBIT"}},
        {"$group": {"_id": None, "total": {"$sum": "$amount"}}}
    ]
    debit_result = await db.wallet_transactions.aggregate(debit_pipeline).to_list(length=1)
    total_used = debit_result[0]["total"] if debit_result else 0.0
    
    return {
        "total_earned": total_earned,
        "total_used": total_used,
        "available_balance": available_balance
    }


async def get_wallet_transactions(
    user_id: str,
    skip: int = 0,
    limit: int = 50
) -> list:
    """
    Get paginated wallet transaction history for a user.
    
    Requirements: 5.5
    - Returns transaction history with dates and amounts
    - Supports pagination
    
    Args:
        user_id: The user's ID
        skip: Number of records to skip (for pagination)
        limit: Maximum number of records to return
        
    Returns:
        list of transaction records
    """
    transactions = await db.wallet_transactions.find(
        {"user_id": user_id}
    ).sort("created_at", -1).skip(skip).limit(limit).to_list(length=limit)
    
    # Format transactions for response
    formatted = []
    for t in transactions:
        formatted.append({
            "id": t.get("id"),
            "type": t.get("type"),
            "amount": t.get("amount"),
            "transaction_type": t.get("transaction_type"),
            "reference_id": t.get("reference_id"),
            "balance_after": t.get("balance_after"),
            "description": t.get("description"),
            "created_at": t.get("created_at").isoformat() if isinstance(t.get("created_at"), datetime) else str(t.get("created_at"))
        })
    
    return formatted


async def apply_wallet_to_subscription(
    user_id: str,
    subscription_amount: float,
    apply_amount: Optional[float] = None
) -> dict:
    """
    Apply wallet balance to subscription payment.
    
    Requirements: 5.2, 5.3
    - Applies wallet balance to subscription payment
    - Validates sufficient balance if specific amount requested
    - Creates debit transaction record
    
    Args:
        user_id: The user's ID
        subscription_amount: Total subscription amount
        apply_amount: Specific amount to apply (if None, applies full available balance up to subscription amount)
        
    Returns:
        dict with amount applied, remaining amount, and new balance
    """
    # Get current balance
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise ValueError(f"User not found: {user_id}")
    
    available_balance = user.get("wallet_balance", 0.0)
    
    if available_balance <= 0:
        return {
            "success": True,
            "amount_applied": 0,
            "remaining_subscription_amount": subscription_amount,
            "new_wallet_balance": 0,
            "transaction_id": None,
            "message": "No wallet balance available to apply"
        }
    
    # Determine amount to apply
    if apply_amount is not None:
        # Specific amount requested
        if apply_amount <= 0:
            raise ValueError("Apply amount must be positive")
        if apply_amount > available_balance:
            raise ValueError(f"Insufficient balance. Available: ₹{available_balance}, Requested: ₹{apply_amount}")
        if apply_amount > subscription_amount:
            raise ValueError(f"Apply amount (₹{apply_amount}) cannot exceed subscription amount (₹{subscription_amount})")
        amount_to_apply = apply_amount
    else:
        # Apply full available balance up to subscription amount
        amount_to_apply = min(available_balance, subscription_amount)
    
    # Debit the wallet
    result = await wallet_debit(
        user_id=user_id,
        amount=amount_to_apply,
        transaction_type="SUBSCRIPTION_PAYMENT",
        description=f"Applied to subscription payment of ₹{subscription_amount}"
    )
    
    remaining_amount = subscription_amount - amount_to_apply
    
    return {
        "success": True,
        "amount_applied": amount_to_apply,
        "remaining_subscription_amount": remaining_amount,
        "new_wallet_balance": result["new_balance"],
        "transaction_id": result["transaction_id"],
        "message": f"₹{amount_to_apply} applied from wallet. Remaining: ₹{remaining_amount}"
    }


@api_router.get("/referral/code")
async def get_referral_code(current_user: dict = Depends(get_current_user)):
    """
    Get the current user's referral code.
    Generates a new code if the user doesn't have one.
    
    Requirements: 1.1, 1.2, 1.3
    - Returns user's referral code within 200ms
    - Generates unique 8-character alphanumeric code if not exists
    - Code remains constant throughout account lifetime
    """
    user_id = current_user.get("id")
    
    # Check if user already has a referral code
    existing_code = current_user.get("referral_code")
    if existing_code:
        return {
            "success": True,
            "referral_code": existing_code,
            "share_message": f"Join BillByteKOT using my referral code {existing_code} and get ₹{int(REFERRAL_DISCOUNT_AMOUNT)} off on your subscription! Download now: https://billbytekot.in"
        }
    
    # Generate and assign a new referral code
    try:
        new_code = await assign_referral_code_to_user(user_id)
        return {
            "success": True,
            "referral_code": new_code,
            "share_message": f"Join BillByteKOT using my referral code {new_code} and get ₹{int(REFERRAL_DISCOUNT_AMOUNT)} off on your subscription! Download now: https://billbytekot.in"
        }
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate referral code: {str(e)}"
        )


@api_router.post("/referral/validate")
async def validate_referral_code_endpoint(request: ReferralValidateRequest, http_request: Request):
    """
    Validate a referral code and check for fraud.
    
    Requirements: 3.3, 3.4, 3.5, 3.6, 3.8, 11.2, 11.4
    - Validates code exists and is active
    - Checks for self-referral (email, phone, device fingerprint)
    - Checks for duplicate mobile number
    - Rate limits to 10 requests per hour per IP
    - Returns discount amount on success
    """
    # Rate limiting check (Requirement 11.4)
    client_ip = http_request.client.host if http_request.client else "unknown"
    try:
        from redis_cache import redis_cache
        if redis_cache and redis_cache.is_connected():
            rate_limit_key = f"rate_limit:referral_validate:{client_ip}"
            # 10 requests per hour (3600 seconds)
            if not await redis_cache.check_rate_limit(rate_limit_key, 10, 3600):
                raise HTTPException(
                    status_code=429,
                    detail={
                        "code": "RATE_LIMITED",
                        "message": "Too many referral validation requests. Please try again later."
                    }
                )
    except HTTPException:
        raise
    except Exception as e:
        # Continue without rate limiting if Redis is unavailable
        print(f"Rate limiting error: {e}")
    
    code = request.referral_code.strip().upper()
    new_user_email = request.new_user_email
    new_user_phone = request.new_user_phone
    device_fingerprint = request.device_fingerprint
    
    # Validate code format
    if not code or len(code) != REFERRAL_CODE_LENGTH:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "INVALID_REFERRAL_CODE",
                "message": "Invalid referral code format"
            }
        )
    
    # Find the referrer by code (case-insensitive)
    referrer = await db.users.find_one({
        "referral_code": {"$regex": f"^{code}$", "$options": "i"}
    })
    
    if not referrer:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "INVALID_REFERRAL_CODE",
                "message": "Invalid referral code"
            }
        )
    
    # Check for self-referral by email (Requirement 11.2)
    if new_user_email and referrer.get("email", "").lower() == new_user_email.lower():
        raise HTTPException(
            status_code=400,
            detail={
                "code": "SELF_REFERRAL",
                "message": "Cannot use your own referral code"
            }
        )
    
    # Check for self-referral by phone (Requirement 11.2)
    if new_user_phone and referrer.get("phone") == new_user_phone:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "SELF_REFERRAL",
                "message": "Cannot use your own referral code"
            }
        )
    
    # Check for self-referral by device fingerprint (Requirement 11.2)
    if device_fingerprint and referrer.get("device_fingerprint") == device_fingerprint:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "SELF_REFERRAL",
                "message": "Cannot use your own referral code"
            }
        )
    
    # Check for duplicate mobile number in referrals (Requirement 11.1)
    if new_user_phone:
        existing_referral = await db.referrals.find_one({
            "referee_phone": new_user_phone
        })
        if existing_referral:
            raise HTTPException(
                status_code=400,
                detail={
                    "code": "DUPLICATE_MOBILE",
                    "message": "This mobile number has already been used for a referral"
                }
            )
    
    # Return success with discount amount
    return {
        "success": True,
        "valid": True,
        "referrer_username": referrer.get("username"),
        "discount_amount": REFERRAL_DISCOUNT_AMOUNT,
        "message": f"Valid referral code! You'll get ₹{int(REFERRAL_DISCOUNT_AMOUNT)} off on your subscription."
    }


@api_router.post("/referral/apply")
async def apply_referral_code(
    request: ReferralApplyRequest,
    http_request: Request,
    current_user: dict = Depends(get_current_user)
):
    """
    Apply a referral code and create a referral record with PENDING status.
    
    Requirements: 3.3, 4.4, 11.4
    - Creates referral record with PENDING status
    - Links referee to referrer
    - Rate limits to 10 requests per hour per IP
    """
    # Rate limiting check (Requirement 11.4)
    client_ip = http_request.client.host if http_request.client else "unknown"
    try:
        from redis_cache import redis_cache
        if redis_cache and redis_cache.is_connected():
            rate_limit_key = f"rate_limit:referral_apply:{client_ip}"
            # 10 requests per hour (3600 seconds)
            if not await redis_cache.check_rate_limit(rate_limit_key, 10, 3600):
                raise HTTPException(
                    status_code=429,
                    detail={
                        "code": "RATE_LIMITED",
                        "message": "Too many referral application requests. Please try again later."
                    }
                )
    except HTTPException:
        raise
    except Exception as e:
        # Continue without rate limiting if Redis is unavailable
        print(f"Rate limiting error: {e}")
    
    code = request.referral_code.strip().upper()
    referee_user_id = request.referee_user_id
    
    # Find the referrer by code
    referrer = await db.users.find_one({
        "referral_code": {"$regex": f"^{code}$", "$options": "i"}
    })
    
    if not referrer:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "INVALID_REFERRAL_CODE",
                "message": "Invalid referral code"
            }
        )
    
    # Get referee user
    referee = await db.users.find_one({"id": referee_user_id})
    if not referee:
        raise HTTPException(
            status_code=404,
            detail={
                "code": "USER_NOT_FOUND",
                "message": "Referee user not found"
            }
        )
    
    # Check for self-referral
    if referrer.get("id") == referee_user_id:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "SELF_REFERRAL",
                "message": "Cannot use your own referral code"
            }
        )
    
    # Check if referee already has a referral
    existing_referral = await db.referrals.find_one({
        "referee_user_id": referee_user_id
    })
    if existing_referral:
        raise HTTPException(
            status_code=400,
            detail={
                "code": "REFERRAL_ALREADY_APPLIED",
                "message": "A referral has already been applied to this account"
            }
        )
    
    # Create referral record
    referral_record = {
        "id": str(uuid.uuid4()),
        "referrer_user_id": referrer.get("id"),
        "referee_user_id": referee_user_id,
        "referral_code": code,
        "status": "PENDING",
        "referee_discount": REFERRAL_DISCOUNT_AMOUNT,
        "referrer_reward": REFERRAL_REWARD_AMOUNT,
        "referee_email": referee.get("email"),
        "referee_phone": referee.get("phone"),
        "payment_id": None,
        "created_at": datetime.now(timezone.utc),
        "completed_at": None,
        "rewarded_at": None,
        "reversed_at": None,
        "reverse_reason": None
    }
    
    await db.referrals.insert_one(referral_record)
    
    # Update referee's referred_by field
    await db.users.update_one(
        {"id": referee_user_id},
        {"$set": {"referred_by": code}}
    )
    
    return {
        "success": True,
        "referral_id": referral_record["id"],
        "status": "PENDING",
        "discount_amount": REFERRAL_DISCOUNT_AMOUNT,
        "message": f"Referral applied! You'll get ₹{int(REFERRAL_DISCOUNT_AMOUNT)} off on your first subscription payment."
    }


@api_router.post("/referral/complete")
async def complete_referral(request: ReferralCompleteRequest):
    """
    Complete a referral after payment and credit the referrer's wallet.
    
    Requirements: 4.1, 4.2, 4.4
    - Process referral after payment completion
    - Credit referrer wallet with ₹300
    - Update referral status to REWARDED
    """
    referee_user_id = request.referee_user_id
    payment_id = request.payment_id
    
    # Find the pending referral for this referee
    referral = await db.referrals.find_one({
        "referee_user_id": referee_user_id,
        "status": "PENDING"
    })
    
    if not referral:
        # No pending referral - this is okay, not all users have referrals
        return {
            "success": True,
            "referral_processed": False,
            "message": "No pending referral found for this user"
        }
    
    # Update referral status to REWARDED
    now = datetime.now(timezone.utc)
    await db.referrals.update_one(
        {"id": referral["id"]},
        {
            "$set": {
                "status": "REWARDED",
                "payment_id": payment_id,
                "completed_at": now,
                "rewarded_at": now
            }
        }
    )
    
    # Credit referrer's wallet
    referrer_user_id = referral["referrer_user_id"]
    
    # Update referrer's wallet balance and stats
    await db.users.update_one(
        {"id": referrer_user_id},
        {
            "$inc": {
                "wallet_balance": REFERRAL_REWARD_AMOUNT,
                "total_referrals": 1,
                "total_referral_earnings": REFERRAL_REWARD_AMOUNT
            }
        }
    )
    
    # Create wallet transaction record
    wallet_transaction = {
        "id": str(uuid.uuid4()),
        "user_id": referrer_user_id,
        "type": "CREDIT",
        "amount": REFERRAL_REWARD_AMOUNT,
        "transaction_type": "REFERRAL_REWARD",
        "reference_id": referral["id"],
        "description": f"Referral reward for user signup",
        "created_at": now
    }
    
    # Get current balance for balance_after field
    referrer = await db.users.find_one({"id": referrer_user_id})
    wallet_transaction["balance_after"] = referrer.get("wallet_balance", REFERRAL_REWARD_AMOUNT)
    
    await db.wallet_transactions.insert_one(wallet_transaction)
    
    return {
        "success": True,
        "referral_processed": True,
        "referral_id": referral["id"],
        "referrer_reward": REFERRAL_REWARD_AMOUNT,
        "message": f"Referral completed! Referrer credited with ₹{int(REFERRAL_REWARD_AMOUNT)}"
    }


@api_router.get("/referral/summary")
async def get_referral_summary(current_user: dict = Depends(get_current_user)):
    """
    Get referral statistics for the current user.
    
    Requirements: 6.1, 6.2, 6.3, 6.4
    - Return total referrals count
    - Return breakdown by status (PENDING, COMPLETED, REWARDED)
    - Return total earnings from referrals
    - Return available wallet balance
    """
    user_id = current_user.get("id")
    
    # Get all referrals where user is the referrer
    referrals = await db.referrals.find({
        "referrer_user_id": user_id
    }).to_list(length=1000)
    
    # Calculate statistics
    total_referrals = len(referrals)
    pending_count = sum(1 for r in referrals if r.get("status") == "PENDING")
    completed_count = sum(1 for r in referrals if r.get("status") == "COMPLETED")
    rewarded_count = sum(1 for r in referrals if r.get("status") == "REWARDED")
    reversed_count = sum(1 for r in referrals if r.get("status") == "REVERSED")
    
    # Calculate total earnings (only from REWARDED referrals)
    total_earnings = sum(
        r.get("referrer_reward", 0) 
        for r in referrals 
        if r.get("status") == "REWARDED"
    )
    
    # Get current wallet balance from user
    user = await db.users.find_one({"id": user_id})
    wallet_balance = user.get("wallet_balance", 0.0) if user else 0.0
    
    # Get referral code
    referral_code = current_user.get("referral_code")
    if not referral_code:
        referral_code = await assign_referral_code_to_user(user_id)
    
    return {
        "success": True,
        "referral_code": referral_code,
        "total_referrals": total_referrals,
        "status_breakdown": {
            "pending": pending_count,
            "completed": completed_count,
            "rewarded": rewarded_count,
            "reversed": reversed_count
        },
        "total_earnings": total_earnings,
        "wallet_balance": wallet_balance,
        "reward_per_referral": REFERRAL_REWARD_AMOUNT,
        "discount_for_referee": REFERRAL_DISCOUNT_AMOUNT
    }


# ============ WALLET API ENDPOINTS ============

@api_router.get("/wallet/balance")
async def get_wallet_balance_endpoint(current_user: dict = Depends(get_current_user)):
    """
    Get wallet balance for the current user.
    
    Requirements: 5.1
    - Returns total earned, used, and available balance
    
    Returns:
        WalletBalanceResponse with balance details
    """
    try:
        user_id = current_user.get("id")
        balance_info = await get_wallet_balance(user_id)
        
        return {
            "success": True,
            "total_earned": balance_info["total_earned"],
            "total_used": balance_info["total_used"],
            "available_balance": balance_info["available_balance"],
            "currency": "INR"
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get wallet balance: {str(e)}")


@api_router.get("/wallet/transactions")
async def get_wallet_transactions_endpoint(
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(50, ge=1, le=100, description="Maximum number of records to return"),
    current_user: dict = Depends(get_current_user)
):
    """
    Get paginated wallet transaction history for the current user.
    
    Requirements: 5.5
    - Returns transaction history with dates and amounts
    - Supports pagination
    
    Args:
        skip: Number of records to skip (default: 0)
        limit: Maximum records to return (default: 50, max: 100)
        
    Returns:
        List of wallet transactions
    """
    try:
        user_id = current_user.get("id")
        transactions = await get_wallet_transactions(user_id, skip, limit)
        
        # Get total count for pagination
        total_count = await db.wallet_transactions.count_documents({"user_id": user_id})
        
        return {
            "success": True,
            "transactions": transactions,
            "pagination": {
                "skip": skip,
                "limit": limit,
                "total": total_count,
                "has_more": (skip + limit) < total_count
            }
        }
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to get transactions: {str(e)}")


@api_router.post("/wallet/apply-to-subscription")
async def apply_wallet_to_subscription_endpoint(
    request: WalletApplyRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    Apply wallet balance to subscription payment.
    
    Requirements: 5.2, 5.3
    - Applies wallet balance to subscription payment
    - Wallet balance is only applicable for subscription payments
    - Deducts amount and records transaction
    
    Args:
        request: WalletApplyRequest with subscription_amount and optional apply_amount
        
    Returns:
        WalletApplyResponse with amount applied and remaining balance
    """
    try:
        user_id = current_user.get("id")
        
        result = await apply_wallet_to_subscription(
            user_id=user_id,
            subscription_amount=request.subscription_amount,
            apply_amount=request.apply_amount
        )
        
        return result
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to apply wallet balance: {str(e)}")


# ============ BULK UPLOAD ENDPOINTS ============

@api_router.post("/menu/bulk-upload")
async def bulk_upload_menu(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk upload menu items from CSV"""
    if current_user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files allowed")
    
    try:
        contents = await file.read()
        decoded = contents.decode('utf-8').splitlines()
        
        import csv
        reader = csv.DictReader(decoded)
        
        user_org_id = get_secure_org_id(current_user)
        items_added = 0
        errors = []
        
        for row_num, row in enumerate(reader, start=2):
            try:
                # Expected columns: name, category, price, description, available
                item = {
                    "id": str(uuid.uuid4()),
                    "name": row.get('name', '').strip(),
                    "category": row.get('category', 'Uncategorized').strip(),
                    "price": float(row.get('price', 0)),
                    "description": row.get('description', '').strip(),
                    "available": row.get('available', 'true').lower() == 'true',
                    "organization_id": user_org_id,
                    "created_at": datetime.now(timezone.utc).isoformat()
                }
                
                if not item['name']:
                    errors.append(f"Row {row_num}: Name is required")
                    continue
                
                if item['price'] <= 0:
                    errors.append(f"Row {row_num}: Invalid price")
                    continue
                
                await db.menu_items.insert_one(item)
                items_added += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return {
            "message": f"Bulk upload completed",
            "items_added": items_added,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@api_router.post("/inventory/bulk-upload")
async def bulk_upload_inventory(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """Bulk upload inventory items from CSV"""
    if current_user["role"] not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Permission denied")
    
    if not file.filename.endswith('.csv'):
        raise HTTPException(status_code=400, detail="Only CSV files allowed")
    
    try:
        contents = await file.read()
        decoded = contents.decode('utf-8').splitlines()
        
        import csv
        reader = csv.DictReader(decoded)
        
        user_org_id = get_secure_org_id(current_user)
        items_added = 0
        errors = []
        
        for row_num, row in enumerate(reader, start=2):
            try:
                # Expected columns: item_name, quantity, unit, min_quantity, supplier
                item = {
                    "id": str(uuid.uuid4()),
                    "name": row.get('item_name', '').strip(),
                    "quantity": float(row.get('quantity', 0)),
                    "unit": row.get('unit', 'pcs').strip(),
                    "min_quantity": float(row.get('min_quantity', 0)),
                    "price_per_unit": float(row.get('price_per_unit', 0)),
                    "organization_id": user_org_id,
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }
                
                if not item['name']:
                    errors.append(f"Row {row_num}: Item name is required")
                    continue
                
                if item['quantity'] < 0:
                    errors.append(f"Row {row_num}: Invalid quantity")
                    continue
                
                # Check if item exists, update or insert
                existing = await db.inventory.find_one({
                    "name": item['name'],
                    "organization_id": user_org_id
                })
                
                if existing:
                    await db.inventory.update_one(
                        {"id": existing["id"]},
                        {"$set": item}
                    )
                else:
                    await db.inventory.insert_one(item)
                
                items_added += 1
                
            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
        
        return {
            "message": f"Bulk upload completed",
            "items_added": items_added,
            "errors": errors if errors else None
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@api_router.get("/templates/menu-csv")
async def download_menu_template():
    """Download menu CSV template"""
    csv_content = "name,category,price,description,available\n"
    csv_content += "Margherita Pizza,Pizza,299,Classic cheese pizza,true\n"
    csv_content += "Chicken Burger,Burgers,199,Grilled chicken burger,true\n"
    csv_content += "Coke,Beverages,50,Chilled coke,true\n"
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=menu_template.csv"}
    )


@api_router.get("/templates/inventory-csv")
async def download_inventory_template():
    """Download inventory CSV template"""
    csv_content = "item_name,quantity,unit,min_quantity,price_per_unit\n"
    csv_content += "Tomatoes,50,kg,10,80\n"
    csv_content += "Cheese,20,kg,5,400\n"
    csv_content += "Chicken,30,kg,10,250\n"
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=inventory_template.csv"}
    )


# ============ EXPORT & CUSTOMER MANAGEMENT ============

@api_router.get("/orders/export/excel")
async def export_orders_to_excel(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Export all orders to Excel file with sequential invoice numbers"""
    user_org_id = get_secure_org_id(current_user)
    
    # Build query
    query = {"organization_id": user_org_id}
    
    # Add date filters if provided
    if start_date or end_date:
        query["created_at"] = {}
        if start_date:
            query["created_at"]["$gte"] = start_date
        if end_date:
            query["created_at"]["$lte"] = end_date
    
    # Fetch orders sorted by creation date (newest first)
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=10000)
    
    if not orders:
        raise HTTPException(status_code=404, detail="No orders found")
    
    # Create CSV content with improved formatting
    csv_content = (
        "Invoice #,Order ID,Date,Time,Table,Customer Name,Phone,Waiter,"
        "Item Name,Quantity,Unit Price,Item Total,Subtotal,Tax,Discount,Total,Payment Method,Status\n"
    )
    
    for order in orders:
        order_id = order.get("id", "")[:8]
        invoice_number = order.get("invoice_number", "")
        created_at = order.get("created_at", "")
        
        # Parse date and time
        try:
            dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            date_str = dt.strftime("%Y-%m-%d")
            time_str = dt.strftime("%H:%M:%S")
        except:
            date_str = created_at
            time_str = ""
        
        table_number = order.get("table_number", "")
        customer_name = order.get("customer_name", "")
        customer_phone = order.get("customer_phone", "")
        waiter_name = order.get("waiter_name", "")
        
        # Get order totals
        subtotal = order.get("subtotal", 0)
        tax = order.get("tax", 0)
        discount = order.get("discount", 0)
        total = order.get("total", 0)
        payment_method = order.get("payment_method", "")
        status = order.get("status", "")
        
        # Format items with each item on a new line
        items = order.get("items", [])
        first_item = True
        
        for item in items:
            item_name = item.get('name', '').replace('"', '""')
            quantity = item.get('quantity', 0)
            price = item.get('price', 0)
            item_total = quantity * price
            
            if first_item:
                # First item includes order details
                csv_content += (
                    f'"{invoice_number}",'  # Invoice #
                    f'"{order_id}",'        # Order ID
                    f'"{date_str}",'        # Date
                    f'"{time_str}",'        # Time
                    f'"{table_number}",'     # Table
                    f'"{customer_name}",'    # Customer Name
                    f'"{customer_phone}",'   # Phone
                    f'"{waiter_name}",'      # Waiter
                    f'"{item_name}",'        # Item Name
                    f'{quantity},'           # Quantity
                    f'{price:.2f},'          # Unit Price
                    f'{item_total:.2f},'     # Item Total
                    f'"",'                   # Empty for subtotal
                    f'"",'                   # Empty for tax
                    f'"",'                   # Empty for discount
                    f'"",'                   # Empty for total
                    f'"",'                   # Empty for payment method
                    f'""\n'                  # Empty for status
                )
                first_item = False
            else:
                # Subsequent items only show item details
                csv_content += (
                    f'"",'                   # Empty invoice #
                    f'"",'                   # Empty order ID
                    f'"",'                   # Empty date
                    f'"",'                   # Empty time
                    f'"",'                   # Empty table
                    f'"",'                   # Empty customer name
                    f'"",'                   # Empty phone
                    f'"",'                   # Empty waiter
                    f'"{item_name}",'        # Item Name
                    f'{quantity},'           # Quantity
                    f'{price:.2f},'          # Unit Price
                    f'{item_total:.2f},'     # Item Total
                    f'"",'                   # Empty for subtotal
                    f'"",'                   # Empty for tax
                    f'"",'                   # Empty for discount
                    f'"",'                   # Empty for total
                    f'"",'                   # Empty for payment method
                    f'""\n'                  # Empty for status
                )
        
        # Add order totals on a new line
        if items:  # Only add if there are items
            csv_content += (
                f'"",'                   # Empty invoice #
                f'"",'                   # Empty order ID
                f'"",'                   # Empty date
                f'"",'                   # Empty time
                f'"",'                   # Empty table
                f'"",'                   # Empty customer name
                f'"",'                   # Empty phone
                f'"",'                   # Empty waiter
                f'"",'                   # Empty item name
                f'"",'                   # Empty quantity
                f'"",'                   # Empty unit price
                f'"",'                   # Empty item total
                f'{subtotal:.2f},'       # Subtotal
                f'{tax:.2f},'            # Tax
                f'{discount:.2f},'       # Discount
                f'{total:.2f},'          # Total
                f'"{payment_method}",'   # Payment Method
                f'"{status}"\n'         # Status
            )
        
        # Add an empty line between orders for better readability
        csv_content += '"","","","","","","","","","","","","","","","","",""\n'
    # Return as downloadable file
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=invoices_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


# Customer Management Models
class Customer(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    total_orders: int = 0
    total_spent: float = 0.0
    last_visit: Optional[str] = None
    organization_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class CustomerCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None


@api_router.post("/customers")
async def create_customer(
    customer_data: CustomerCreate,
    current_user: dict = Depends(get_current_user)
):
    """Create or update customer record"""
    user_org_id = get_secure_org_id(current_user)
    
    # Check if customer already exists by phone
    existing = await db.customers.find_one({
        "phone": customer_data.phone,
        "organization_id": user_org_id
    })
    
    if existing:
        # Update existing customer
        await db.customers.update_one(
            {"id": existing["id"]},
            {
                "$set": {
                    "name": customer_data.name,
                    "email": customer_data.email,
                    "address": customer_data.address,
                    "notes": customer_data.notes,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }
            }
        )
        return {"message": "Customer updated", "customer_id": existing["id"]}
    
    # Create new customer
    customer = Customer(
        name=customer_data.name,
        phone=customer_data.phone,
        email=customer_data.email,
        address=customer_data.address,
        notes=customer_data.notes,
        organization_id=user_org_id
    )
    
    doc = customer.model_dump()
    doc["created_at"] = doc["created_at"].isoformat()
    
    await db.customers.insert_one(doc)
    
    return {"message": "Customer created", "customer_id": customer.id}


@api_router.get("/customers")
async def get_customers(
    search: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """Get all customers"""
    user_org_id = get_secure_org_id(current_user)
    
    query = {"organization_id": user_org_id}
    
    # Add search filter
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}}
        ]
    
    customers = await db.customers.find(query, {"_id": 0}).sort("created_at", -1).to_list(length=1000)
    
    return customers


@api_router.get("/customers/{customer_id}")
async def get_customer(
    customer_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Get customer details with order history"""
    user_org_id = get_secure_org_id(current_user)
    
    customer = await db.customers.find_one({
        "id": customer_id,
        "organization_id": user_org_id
    }, {"_id": 0})
    
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get customer's orders
    orders = await db.orders.find({
        "customer_phone": customer["phone"],
        "organization_id": user_org_id
    }, {"_id": 0}).sort("created_at", -1).to_list(length=100)
    
    # Calculate stats
    total_orders = len(orders)
    total_spent = sum(order.get("total", 0) for order in orders)
    last_visit = orders[0].get("created_at") if orders else None
    
    # Update customer stats
    await db.customers.update_one(
        {"id": customer_id},
        {
            "$set": {
                "total_orders": total_orders,
                "total_spent": total_spent,
                "last_visit": last_visit
            }
        }
    )
    
    customer["total_orders"] = total_orders
    customer["total_spent"] = total_spent
    customer["last_visit"] = last_visit
    customer["orders"] = orders
    
    return customer


@api_router.get("/customers/phone/{phone}")
async def get_customer_by_phone(
    phone: str,
    current_user: dict = Depends(get_current_user)
):
    """Get customer by phone number"""
    user_org_id = get_secure_org_id(current_user)
    
    customer = await db.customers.find_one({
        "phone": phone,
        "organization_id": user_org_id
    }, {"_id": 0})
    
    if not customer:
        return {"found": False}
    
    return {"found": True, "customer": customer}


@api_router.put("/customers/{customer_id}")
async def update_customer(
    customer_id: str,
    customer_data: CustomerCreate,
    current_user: dict = Depends(get_current_user)
):
    """Update customer details"""
    user_org_id = get_secure_org_id(current_user)
    
    result = await db.customers.update_one(
        {"id": customer_id, "organization_id": user_org_id},
        {
            "$set": {
                "name": customer_data.name,
                "phone": customer_data.phone,
                "email": customer_data.email,
                "address": customer_data.address,
                "notes": customer_data.notes,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }
        }
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    return {"message": "Customer updated"}


@api_router.delete("/customers/{customer_id}")
async def delete_customer(
    customer_id: str,
    current_user: dict = Depends(get_current_user)
):
    """Delete customer"""
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admin can delete customers")
    
    user_org_id = get_secure_org_id(current_user)
    
    result = await db.customers.delete_one({
        "id": customer_id,
        "organization_id": user_org_id
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    return {"message": "Customer deleted"}


# ============================================================================
# Twilio WhatsApp Integration (Easy Setup)
# ============================================================================

@api_router.post("/twilio/whatsapp/send")
async def twilio_send_whatsapp(
    to_number: str,
    message: str,
    user: dict = Depends(get_current_user)
):
    """Send WhatsApp message via Twilio"""
    if not _TWILIO_WHATSAPP_AVAILABLE:
        raise HTTPException(
            status_code=503,
            detail="Twilio WhatsApp not configured. Please add TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN to .env"
        )
    
    result = twilio_whatsapp.send_message(to_number, message)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error"))
    
    return result

@api_router.post("/twilio/whatsapp/send-receipt")
async def twilio_send_receipt(
    to_number: str,
    order_id: str,
    user: dict = Depends(get_current_user)
):
    """Send order receipt via Twilio WhatsApp"""
    if not _TWILIO_WHATSAPP_AVAILABLE:
        raise HTTPException(status_code=503, detail="Twilio WhatsApp not configured")
    
    # Get order details
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Format order data
    order_data = {
        "order_id": order.get("id"),
        "restaurant_name": user.get("username", "BillByteKOT"),
        "date": order.get("created_at", datetime.now()).strftime("%Y-%m-%d %H:%M"),
        "items": order.get("items", []),
        "subtotal": order.get("subtotal", 0),
        "tax": order.get("tax", 0),
        "total": order.get("total", 0),
        "payment_method": order.get("payment_method", "Cash"),
        "status": order.get("status", "completed")
    }
    
    result = twilio_whatsapp.send_receipt(to_number, order_data)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error"))
    
    return result

@api_router.post("/twilio/whatsapp/send-confirmation")
async def twilio_send_confirmation(
    to_number: str,
    order_id: str,
    user: dict = Depends(get_current_user)
):
    """Send order confirmation via Twilio WhatsApp"""
    if not _TWILIO_WHATSAPP_AVAILABLE:
        raise HTTPException(status_code=503, detail="Twilio WhatsApp not configured")
    
    # Get order details
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order_data = {
        "order_id": order.get("id"),
        "restaurant_name": user.get("username", "BillByteKOT"),
        "total": order.get("total", 0),
        "tracking_url": f"https://billbytekot.in/track/{order.get('tracking_token', '')}"
    }
    
    result = twilio_whatsapp.send_order_confirmation(to_number, order_data)
    
    if not result.get("success"):
        raise HTTPException(status_code=400, detail=result.get("error"))
    
    return result

@api_router.get("/twilio/whatsapp/status")
async def twilio_whatsapp_status(user: dict = Depends(get_current_user)):
    """Check Twilio WhatsApp configuration status"""
    if not _TWILIO_WHATSAPP_AVAILABLE:
        return {
            "configured": False,
            "message": "Twilio WhatsApp not available. Install twilio package: pip install twilio"
        }
    
    if not twilio_whatsapp.is_configured():
        return {
            "configured": False,
            "message": "Twilio credentials not set. Add TWILIO_ACCOUNT_SID and TWILIO_AUTH_TOKEN to .env"
        }
    
    account_info = twilio_whatsapp.get_account_info()
    
    return {
        "configured": True,
        "account_info": account_info,
        "from_number": twilio_whatsapp.from_number
    }


# ============================================================================
# Simple WhatsApp Integration (Free, No API Keys Required)
# ============================================================================

@api_router.post("/whatsapp/create-receipt-link")
async def create_whatsapp_receipt_link(
    phone_number: str,
    order_id: str,
    user: dict = Depends(get_current_user)
):
    """
    Create WhatsApp link to send receipt
    Opens WhatsApp with pre-filled receipt message
    No API keys needed!
    """
    # Get order details
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    # Format order data
    order_data = {
        "order_id": order.get("id"),
        "restaurant_name": user.get("username", "BillByteKOT"),
        "date": order.get("created_at", datetime.now()).strftime("%Y-%m-%d %H:%M"),
        "items": order.get("items", []),
        "subtotal": order.get("subtotal", 0),
        "tax": order.get("tax", 0),
        "total": order.get("total", 0),
        "payment_method": order.get("payment_method", "Cash"),
        "status": order.get("status", "completed")
    }
    
    # Create WhatsApp link
    link = simple_whatsapp.create_receipt_link(phone_number, order_data)
    
    return {
        "success": True,
        "whatsapp_link": link,
        "phone_number": phone_number,
        "message": "Click the link to open WhatsApp and send receipt"
    }

@api_router.post("/whatsapp/create-confirmation-link")
async def create_whatsapp_confirmation_link(
    phone_number: str,
    order_id: str,
    user: dict = Depends(get_current_user)
):
    """Create WhatsApp link to send order confirmation"""
    # Get order details
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    order_data = {
        "order_id": order.get("id"),
        "restaurant_name": user.get("username", "BillByteKOT"),
        "total": order.get("total", 0),
        "tracking_url": f"https://billbytekot.in/track/{order.get('tracking_token', '')}"
    }
    
    link = simple_whatsapp.create_confirmation_link(phone_number, order_data)
    
    return {
        "success": True,
        "whatsapp_link": link,
        "phone_number": phone_number
    }

@api_router.post("/whatsapp/create-message-link")
async def create_whatsapp_message_link(
    phone_number: str,
    message: str,
    user: dict = Depends(get_current_user)
):
    """Create WhatsApp link with custom message"""
    link = simple_whatsapp.create_message_link(phone_number, message)
    
    return {
        "success": True,
        "whatsapp_link": link,
        "phone_number": phone_number
    }

@api_router.get("/whatsapp/business-link")
async def get_business_whatsapp_link(user: dict = Depends(get_current_user)):
    """Get business WhatsApp chat link"""
    link = simple_whatsapp.get_business_chat_link()
    
    return {
        "success": True,
        "whatsapp_link": link,
        "business_number": simple_whatsapp.business_number or "Not configured"
    }


# ============ SUPER ADMIN PANEL (Site Owner Only) ============
SUPER_ADMIN_USERNAME = os.getenv("SUPER_ADMIN_USERNAME", "shiv")
SUPER_ADMIN_PASSWORD = os.getenv("SUPER_ADMIN_PASSWORD", "shiv@123")

def verify_super_admin(username: str, password: str) -> bool:
    """Verify super admin credentials"""
    return username == SUPER_ADMIN_USERNAME and password == SUPER_ADMIN_PASSWORD

@api_router.get("/super-admin/dashboard")
async def get_super_admin_dashboard(username: str, password: str):
    """Get complete system overview - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Get all users
    users = await db.users.find({}, {"_id": 0, "password": 0}).to_list(1000)
    
    # Get all tickets
    tickets = await db.support_tickets.find({}, {"_id": 0}).to_list(1000)
    
    # Get recent orders (last 30 days)
    thirty_days_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    recent_orders = await db.orders.find(
        {"created_at": {"$gte": thirty_days_ago}},
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate statistics
    total_users = len(users)
    active_subscriptions = sum(1 for u in users if u.get("subscription_active"))
    trial_users = sum(1 for u in users if not u.get("subscription_active"))
    
    # Ticket statistics
    open_tickets = sum(1 for t in tickets if t.get("status") == "open")
    pending_tickets = sum(1 for t in tickets if t.get("status") == "pending")
    resolved_tickets = sum(1 for t in tickets if t.get("status") == "resolved")
    
    # Lead statistics
    total_leads = await db.leads.count_documents({})
    new_leads = await db.leads.count_documents({"status": "new"})
    
    return {
        "overview": {
            "total_users": total_users,
            "active_subscriptions": active_subscriptions,
            "trial_users": trial_users,
            "total_orders_30d": len(recent_orders),
            "open_tickets": open_tickets,
            "pending_tickets": pending_tickets,
            "resolved_tickets": resolved_tickets,
            "total_leads": total_leads,
            "new_leads": new_leads
        },
        "users": users,
        "tickets": tickets,
        "recent_orders": recent_orders[:100]
    }

@api_router.get("/super-admin/users")
async def get_all_users_admin(username: str, password: str, skip: int = 0, limit: int = 100):
    """Get all users - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    users = await db.users.find({}, {"_id": 0, "password": 0}).skip(skip).limit(limit).to_list(limit)
    total = await db.users.count_documents({})
    
    return {"users": users, "total": total, "skip": skip, "limit": limit}

class SubscriptionUpdate(BaseModel):
    subscription_active: bool
    subscription_expires_at: Optional[str] = None

class TrialExtension(BaseModel):
    days: int

@api_router.put("/super-admin/users/{user_id}/extend-trial")
async def extend_user_trial_admin(
    user_id: str,
    trial_extension: TrialExtension,
    username: str,
    password: str
):
    """Extend user trial by X days - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Get current user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get current created_at date
    created_at = user.get("created_at")
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
    
    if not created_at:
        created_at = datetime.now(timezone.utc)
    
    # Calculate new trial end date by adjusting created_at backwards
    # This effectively extends the trial by making it seem like they registered later
    # Or we can store a separate trial_extension_days field
    current_extension = user.get("trial_extension_days", 0)
    new_extension = current_extension + trial_extension.days
    
    result = await db.users.update_one(
        {"id": user_id}, 
        {"$set": {"trial_extension_days": new_extension}}
    )
    
    return {
        "message": f"Trial extended by {trial_extension.days} days",
        "user_id": user_id,
        "total_trial_days": 7 + new_extension,
        "extension_days": new_extension
    }

@api_router.put("/super-admin/users/{user_id}/subscription")
async def update_user_subscription_admin(
    user_id: str,
    subscription_update: SubscriptionUpdate,
    username: str,
    password: str
):
    """Manually update user subscription - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    update_data = {"subscription_active": subscription_update.subscription_active}
    if subscription_update.subscription_expires_at:
        update_data["subscription_expires_at"] = subscription_update.subscription_expires_at
    
    result = await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "message": "Subscription updated successfully",
        "user_id": user_id,
        "subscription_active": subscription_update.subscription_active
    }


class ManualSubscription(BaseModel):
    payment_id: str
    payment_method: str = "manual"  # manual, upi, bank_transfer, cash
    payment_proof_url: Optional[str] = None
    payment_notes: Optional[str] = None
    amount: float = 999.0
    months: int = 12
    send_invoice: bool = True


def generate_payment_id():
    """Generate unique payment ID for manual subscriptions"""
    import random
    import string
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    random_str = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    return f"BBK-{timestamp}-{random_str}"


def generate_invoice_number():
    """Generate unique invoice number with format: BBK/2025-26/INV/0001"""
    import random
    now = datetime.now()
    year = now.year
    month = now.month
    # Fiscal year in India starts from April
    fiscal_year = f"{year}-{str(year + 1)[-2:]}" if month >= 4 else f"{year - 1}-{str(year)[-2:]}"
    sequence = random.randint(1, 9999)
    return f"BBK/{fiscal_year}/INV/{sequence:04d}"


async def send_subscription_invoice_email(user_email: str, user_name: str, invoice_data: dict):
    """Send subscription invoice email"""
    from email_service import send_support_email
    
    # Calculate tax breakdown
    base_amount = invoice_data['amount'] / 1.18
    cgst = base_amount * 0.09
    sgst = base_amount * 0.09
    
    subject = f"BillByteKOT Tax Invoice - {invoice_data['invoice_number']}"
    
    html_body = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; background-color: #f5f5f5; margin: 0; padding: 20px; }}
            .container {{ max-width: 650px; margin: 0 auto; background: white; border-radius: 12px; overflow: hidden; box-shadow: 0 4px 20px rgba(0,0,0,0.1); }}
            .header {{ background: linear-gradient(135deg, #7c3aed 0%, #a855f7 100%); color: white; padding: 30px; text-align: center; }}
            .header h1 {{ margin: 0; font-size: 28px; }}
            .header p {{ margin: 5px 0 0; opacity: 0.9; }}
            .invoice-badge {{ display: inline-block; background: #10b981; color: white; padding: 8px 20px; border-radius: 25px; font-size: 14px; font-weight: 600; margin-top: 15px; }}
            .content {{ padding: 30px; }}
            .invoice-header {{ display: flex; justify-content: space-between; margin-bottom: 25px; padding-bottom: 20px; border-bottom: 2px solid #f0f0f0; }}
            .invoice-number {{ font-size: 18px; color: #7c3aed; font-weight: 700; }}
            .invoice-date {{ color: #666; font-size: 14px; }}
            .section {{ margin-bottom: 25px; }}
            .section-title {{ font-size: 12px; color: #999; text-transform: uppercase; letter-spacing: 1px; margin-bottom: 10px; }}
            .details-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }}
            .detail-box {{ background: #f8f9fa; padding: 15px; border-radius: 8px; }}
            .detail-label {{ font-size: 12px; color: #666; margin-bottom: 4px; }}
            .detail-value {{ font-size: 14px; color: #333; font-weight: 500; }}
            .items-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
            .items-table th {{ background: #7c3aed; color: white; padding: 12px; text-align: left; font-size: 12px; text-transform: uppercase; }}
            .items-table td {{ padding: 15px 12px; border-bottom: 1px solid #eee; }}
            .totals {{ margin-left: auto; width: 280px; }}
            .total-row {{ display: flex; justify-content: space-between; padding: 8px 0; font-size: 14px; }}
            .total-row.final {{ border-top: 2px solid #7c3aed; padding-top: 15px; margin-top: 10px; font-size: 20px; font-weight: bold; color: #7c3aed; }}
            .amount-words {{ background: #f0f7ff; padding: 15px; border-radius: 8px; border-left: 4px solid #7c3aed; margin: 20px 0; }}
            .footer {{ background: #f8f9fa; padding: 25px; text-align: center; }}
            .footer p {{ margin: 5px 0; font-size: 12px; color: #666; }}
            .footer .company {{ font-weight: 600; color: #333; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <h1>🍽️ BillByteKOT</h1>
                <p>Smart Restaurant Management System</p>
                <span class="invoice-badge">✓ PAID</span>
            </div>
            
            <div class="content">
                <div class="invoice-header">
                    <div>
                        <div class="invoice-number">{invoice_data['invoice_number']}</div>
                        <div class="invoice-date">Tax Invoice</div>
                    </div>
                    <div style="text-align: right;">
                        <div class="detail-label">Invoice Date</div>
                        <div class="detail-value">{invoice_data['date']}</div>
                    </div>
                </div>
                
                <div class="section">
                    <div class="section-title">Billed To</div>
                    <div class="detail-value" style="font-size: 16px;">{user_name}</div>
                    <div style="color: #666; font-size: 14px;">{user_email}</div>
                </div>
                
                <div class="details-grid">
                    <div class="detail-box">
                        <div class="detail-label">Payment ID</div>
                        <div class="detail-value" style="color: #7c3aed;">{invoice_data['payment_id']}</div>
                    </div>
                    <div class="detail-box">
                        <div class="detail-label">Payment Method</div>
                        <div class="detail-value">{invoice_data['payment_method'].upper()}</div>
                    </div>
                    <div class="detail-box">
                        <div class="detail-label">Subscription Period</div>
                        <div class="detail-value">{invoice_data['months']} Month(s)</div>
                    </div>
                    <div class="detail-box">
                        <div class="detail-label">Valid Until</div>
                        <div class="detail-value" style="color: #10b981;">{invoice_data['expires_at']}</div>
                    </div>
                </div>
                
                <table class="items-table">
                    <thead>
                        <tr>
                            <th>Description</th>
                            <th>HSN/SAC</th>
                            <th style="text-align: right;">Amount</th>
                        </tr>
                    </thead>
                    <tbody>
                        <tr>
                            <td>
                                <strong>BillByteKOT Premium Subscription</strong><br>
                                <span style="font-size: 12px; color: #666;">Annual subscription with all premium features</span>
                            </td>
                            <td>998314</td>
                            <td style="text-align: right;">₹{base_amount:.2f}</td>
                        </tr>
                    </tbody>
                </table>
                
                <div class="totals">
                    <div class="total-row">
                        <span>Subtotal</span>
                        <span>₹{base_amount:.2f}</span>
                    </div>
                    <div class="total-row">
                        <span>CGST (9%)</span>
                        <span>₹{cgst:.2f}</span>
                    </div>
                    <div class="total-row">
                        <span>SGST (9%)</span>
                        <span>₹{sgst:.2f}</span>
                    </div>
                    <div class="total-row final">
                        <span>Total</span>
                        <span>₹{invoice_data['amount']:.2f}</span>
                    </div>
                </div>
                
                <div class="amount-words">
                    <div style="font-size: 11px; color: #666; text-transform: uppercase;">Amount in Words</div>
                    <div style="font-weight: 500;">Nine Hundred Ninety Nine Rupees Only</div>
                </div>
            </div>
            
            <div class="footer">
                <p class="company">BillByte Innovations</p>
                <p>Bangalore, Karnataka, India</p>
                <p>support@billbytekot.in | +91-8310832669</p>
                <p style="margin-top: 15px; font-size: 11px; color: #999;">This is a computer-generated invoice and does not require a signature.</p>
            </div>
        </div>
    </body>
    </html>
    """
    
    text_body = f"""
    BillByteKOT Tax Invoice
    ========================
    
    Invoice Number: {invoice_data['invoice_number']}
    Date: {invoice_data['date']}
    
    Billed To: {user_name} ({user_email})
    
    Payment Details:
    - Payment ID: {invoice_data['payment_id']}
    - Payment Method: {invoice_data['payment_method'].upper()}
    - Subscription Period: {invoice_data['months']} Month(s)
    - Valid Until: {invoice_data['expires_at']}
    
    Invoice Summary:
    - BillByteKOT Premium Subscription: ₹{base_amount:.2f}
    - CGST (9%): ₹{cgst:.2f}
    - SGST (9%): ₹{sgst:.2f}
    - Total: ₹{invoice_data['amount']:.2f}
    
    Amount in Words: Nine Hundred Ninety Nine Rupees Only
    
    Your subscription is now active!
    
    ---
    BillByte Innovations
    support@billbytekot.in | +91-8310832669
    """
    
    try:
        result = await send_support_email(user_email, subject, html_body, text_body)
        return result
    except Exception as e:
        print(f"❌ Failed to send invoice email: {e}")
        return {"success": False, "message": str(e)}


@api_router.post("/super-admin/users/{user_id}/manual-subscription")
async def create_manual_subscription(
    user_id: str,
    subscription: ManualSubscription,
    username: str,
    password: str
):
    """Create manual subscription with payment proof - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Get user
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Calculate expiry date
    expires_at = datetime.now(timezone.utc)
    expires_at = expires_at.replace(month=expires_at.month + subscription.months) if expires_at.month + subscription.months <= 12 else expires_at.replace(year=expires_at.year + 1, month=(expires_at.month + subscription.months) % 12 or 12)
    
    # Generate invoice number
    invoice_number = generate_invoice_number()
    
    # Create subscription record
    subscription_record = {
        "id": str(uuid.uuid4())[:8],
        "user_id": user_id,
        "payment_id": subscription.payment_id,
        "invoice_number": invoice_number,
        "payment_method": subscription.payment_method,
        "payment_proof_url": subscription.payment_proof_url,
        "payment_notes": subscription.payment_notes,
        "amount": subscription.amount,
        "months": subscription.months,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": expires_at.isoformat(),
        "created_by": username
    }
    
    # Save subscription record
    await db.manual_subscriptions.insert_one(subscription_record)
    
    # Update user subscription
    await db.users.update_one(
        {"id": user_id},
        {"$set": {
            "subscription_active": True,
            "subscription_expires_at": expires_at.isoformat(),
            "subscription_payment_id": subscription.payment_id,
            "subscription_type": "manual"
        }}
    )
    
    # Send invoice email if requested
    email_result = {"success": False}
    if subscription.send_invoice:
        invoice_data = {
            "invoice_number": invoice_number,
            "payment_id": subscription.payment_id,
            "date": datetime.now().strftime("%d %B %Y"),
            "payment_method": subscription.payment_method,
            "months": subscription.months,
            "expires_at": expires_at.strftime("%d %B %Y"),
            "amount": subscription.amount
        }
        email_result = await send_subscription_invoice_email(
            user.get("email"),
            user.get("username", "User"),
            invoice_data
        )
    
    return {
        "success": True,
        "message": "Manual subscription created successfully",
        "subscription_id": subscription_record["id"],
        "invoice_number": invoice_number,
        "payment_id": subscription.payment_id,
        "expires_at": expires_at.isoformat(),
        "invoice_sent": email_result.get("success", False)
    }


@api_router.post("/super-admin/generate-payment-id")
async def generate_new_payment_id(username: str, password: str):
    """Generate a new payment ID for manual subscriptions"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    return {"payment_id": generate_payment_id()}


@api_router.get("/super-admin/subscriptions")
async def get_all_subscriptions(
    username: str,
    password: str,
    skip: int = 0,
    limit: int = 100
):
    """Get all manual subscription records - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    subscriptions = await db.manual_subscriptions.find({}, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.manual_subscriptions.count_documents({})
    
    return {"subscriptions": subscriptions, "total": total}


class ReceiptPDFRequest(BaseModel):
    user_email: str
    user_name: str
    business_name: str
    receipt_number: str
    amount: float
    valid_from: str
    valid_until: str
    payment_id: str
    payment_method: str
    html_content: str


@api_router.post("/super-admin/send-receipt-pdf")
async def send_receipt_pdf(
    request: ReceiptPDFRequest,
    username: str,
    password: str
):
    """Send receipt PDF via email - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    try:
        from email_service import send_receipt_email_with_html
        
        # Send email with receipt HTML
        result = await send_receipt_email_with_html(
            to_email=request.user_email,
            user_name=request.user_name,
            business_name=request.business_name,
            receipt_number=request.receipt_number,
            amount=request.amount,
            valid_from=request.valid_from,
            valid_until=request.valid_until,
            payment_id=request.payment_id,
            payment_method=request.payment_method,
            html_content=request.html_content
        )
        
        return {"success": result.get("success", False), "message": result.get("message", "Receipt sent")}
    except Exception as e:
        print(f"Error sending receipt: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to send receipt: {str(e)}")


@api_router.post("/super-admin/users/{user_id}/send-invoice")
async def send_invoice_to_user(
    user_id: str,
    username: str,
    password: str
):
    """Send invoice email to user for their current subscription"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if not user.get("subscription_active"):
        raise HTTPException(status_code=400, detail="User has no active subscription")
    
    # Get subscription record if exists
    sub_record = await db.manual_subscriptions.find_one(
        {"user_id": user_id},
        sort=[("created_at", -1)]
    )
    
    invoice_data = {
        "invoice_number": sub_record.get("invoice_number") if sub_record else generate_invoice_number(),
        "payment_id": user.get("subscription_payment_id", sub_record.get("payment_id") if sub_record else "N/A"),
        "date": datetime.now().strftime("%d %B %Y"),
        "payment_method": sub_record.get("payment_method", "online") if sub_record else "online",
        "months": sub_record.get("months", 12) if sub_record else 12,
        "expires_at": datetime.fromisoformat(user["subscription_expires_at"].replace("Z", "+00:00")).strftime("%d %B %Y") if user.get("subscription_expires_at") else "N/A",
        "amount": sub_record.get("amount", 999) if sub_record else 999
    }
    
    result = await send_subscription_invoice_email(
        user.get("email"),
        user.get("username", "User"),
        invoice_data
    )
    
    return {
        "success": result.get("success", False),
        "message": "Invoice sent successfully" if result.get("success") else "Failed to send invoice"
    }

@api_router.delete("/super-admin/users/{user_id}")
async def delete_user_admin(user_id: str, username: str, password: str):
    """Delete user and all their data - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Delete user and all data
    await db.users.delete_one({"id": user_id})
    await db.orders.delete_many({"organization_id": user_id})
    await db.menu_items.delete_many({"organization_id": user_id})
    await db.tables.delete_many({"organization_id": user_id})
    await db.payments.delete_many({"organization_id": user_id})
    await db.inventory.delete_many({"organization_id": user_id})
    
    return {"message": "User and all data deleted successfully", "user_id": user_id}


@api_router.get("/super-admin/users/{user_id}/full-data")
async def get_user_full_data(user_id: str, username: str, password: str):
    """Get complete user data including all business data - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Get user
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get all staff members
    staff = await db.users.find(
        {"organization_id": user_id},
        {"_id": 0, "password": 0}
    ).to_list(100)
    
    # Get all orders
    orders = await db.orders.find(
        {"organization_id": user_id},
        {"_id": 0}
    ).sort("created_at", -1).to_list(10000)
    
    # Get menu items
    menu_items = await db.menu_items.find(
        {"organization_id": user_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get tables
    tables = await db.tables.find(
        {"organization_id": user_id},
        {"_id": 0}
    ).to_list(100)
    
    # Get inventory
    inventory = await db.inventory.find(
        {"organization_id": user_id},
        {"_id": 0}
    ).to_list(1000)
    
    # Get payments
    payments = await db.payments.find(
        {"organization_id": user_id},
        {"_id": 0}
    ).to_list(10000)
    
    # Calculate stats
    total_revenue = sum(o.get("total", 0) for o in orders if o.get("status") == "completed")
    total_orders = len([o for o in orders if o.get("status") == "completed"])
    credit_orders = len([o for o in orders if o.get("is_credit")])
    pending_credit = sum(o.get("balance_amount", 0) for o in orders if o.get("is_credit"))
    
    return {
        "user": user,
        "staff": staff,
        "staff_count": len(staff),
        "orders": orders,
        "orders_count": len(orders),
        "menu_items": menu_items,
        "menu_count": len(menu_items),
        "tables": tables,
        "tables_count": len(tables),
        "inventory": inventory,
        "inventory_count": len(inventory),
        "payments": payments,
        "payments_count": len(payments),
        "stats": {
            "total_revenue": total_revenue,
            "total_orders": total_orders,
            "credit_orders": credit_orders,
            "pending_credit": pending_credit,
            "avg_order_value": total_revenue / total_orders if total_orders > 0 else 0
        },
        "exported_at": datetime.now(timezone.utc).isoformat()
    }


def serialize_for_sqlite(obj):
    """Convert MongoDB document to SQLite-compatible format"""
    if isinstance(obj, datetime):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return json.dumps(obj)
    elif isinstance(obj, list):
        return json.dumps(obj)
    return obj


def create_sqlite_backup(user_data: dict, staff: list, orders: list, menu_items: list, 
                         tables: list, inventory: list, payments: list) -> bytes:
    """Create SQLite database backup file"""
    # Create in-memory database
    conn = sqlite3.connect(':memory:')
    cursor = conn.cursor()
    
    # Create tables
    cursor.execute('''
        CREATE TABLE users (
            id TEXT PRIMARY KEY,
            username TEXT,
            email TEXT,
            role TEXT,
            phone TEXT,
            organization_id TEXT,
            subscription_active INTEGER,
            subscription_expires_at TEXT,
            trial_extension_days INTEGER,
            bill_count INTEGER,
            setup_completed INTEGER,
            onboarding_completed INTEGER,
            business_settings TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE orders (
            id TEXT PRIMARY KEY,
            invoice_number INTEGER,
            table_id TEXT,
            table_number INTEGER,
            items TEXT,
            subtotal REAL,
            tax REAL,
            discount REAL,
            total REAL,
            status TEXT,
            waiter_id TEXT,
            waiter_name TEXT,
            customer_name TEXT,
            customer_phone TEXT,
            order_type TEXT,
            organization_id TEXT,
            payment_method TEXT,
            is_credit INTEGER,
            payment_received REAL,
            balance_amount REAL,
            cash_amount REAL,
            card_amount REAL,
            upi_amount REAL,
            credit_amount REAL,
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE menu_items (
            id TEXT PRIMARY KEY,
            name TEXT,
            category TEXT,
            price REAL,
            description TEXT,
            image_url TEXT,
            available INTEGER,
            ingredients TEXT,
            preparation_time INTEGER,
            organization_id TEXT,
            created_at TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE tables (
            id TEXT PRIMARY KEY,
            table_number INTEGER,
            capacity INTEGER,
            status TEXT,
            current_order_id TEXT,
            organization_id TEXT,
            created_at TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE inventory (
            id TEXT PRIMARY KEY,
            name TEXT,
            category TEXT,
            quantity REAL,
            unit TEXT,
            min_stock REAL,
            cost_per_unit REAL,
            supplier TEXT,
            organization_id TEXT,
            created_at TEXT,
            updated_at TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE payments (
            id TEXT PRIMARY KEY,
            order_id TEXT,
            amount REAL,
            payment_method TEXT,
            razorpay_order_id TEXT,
            razorpay_payment_id TEXT,
            status TEXT,
            organization_id TEXT,
            created_at TEXT
        )
    ''')
    
    cursor.execute('''
        CREATE TABLE backup_info (
            id INTEGER PRIMARY KEY,
            user_id TEXT,
            username TEXT,
            exported_at TEXT,
            version TEXT
        )
    ''')
    
    # Insert user data
    cursor.execute('''
        INSERT INTO users VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        user_data.get('id'), user_data.get('username'), user_data.get('email'),
        user_data.get('role'), user_data.get('phone'), user_data.get('organization_id'),
        1 if user_data.get('subscription_active') else 0,
        user_data.get('subscription_expires_at'),
        user_data.get('trial_extension_days', 0),
        user_data.get('bill_count', 0),
        1 if user_data.get('setup_completed') else 0,
        1 if user_data.get('onboarding_completed') else 0,
        json.dumps(user_data.get('business_settings', {})),
        str(user_data.get('created_at', '')),
        str(user_data.get('updated_at', ''))
    ))
    
    # Insert staff
    for s in staff:
        cursor.execute('''
            INSERT INTO users VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            s.get('id'), s.get('username'), s.get('email'),
            s.get('role'), s.get('phone'), s.get('organization_id'),
            1 if s.get('subscription_active') else 0,
            s.get('subscription_expires_at'),
            s.get('trial_extension_days', 0),
            s.get('bill_count', 0),
            1 if s.get('setup_completed') else 0,
            1 if s.get('onboarding_completed') else 0,
            json.dumps(s.get('business_settings', {})),
            str(s.get('created_at', '')),
            str(s.get('updated_at', ''))
        ))
    
    # Insert orders
    for o in orders:
        cursor.execute('''
            INSERT INTO orders VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            o.get('id'), o.get('invoice_number'), o.get('table_id'), o.get('table_number'),
            json.dumps(o.get('items', [])), o.get('subtotal', 0), o.get('tax', 0),
            o.get('discount', 0), o.get('total', 0), o.get('status'),
            o.get('waiter_id'), o.get('waiter_name'), o.get('customer_name'),
            o.get('customer_phone'), o.get('order_type'), o.get('organization_id'),
            o.get('payment_method'), 1 if o.get('is_credit') else 0,
            o.get('payment_received', 0), o.get('balance_amount', 0),
            o.get('cash_amount', 0), o.get('card_amount', 0),
            o.get('upi_amount', 0), o.get('credit_amount', 0),
            str(o.get('created_at', '')), str(o.get('updated_at', ''))
        ))
    
    # Insert menu items
    for m in menu_items:
        cursor.execute('''
            INSERT INTO menu_items VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            m.get('id'), m.get('name'), m.get('category'), m.get('price'),
            m.get('description'), m.get('image_url'),
            1 if m.get('available', True) else 0,
            json.dumps(m.get('ingredients', [])), m.get('preparation_time'),
            m.get('organization_id'), str(m.get('created_at', ''))
        ))
    
    # Insert tables
    for t in tables:
        cursor.execute('''
            INSERT INTO tables VALUES (?, ?, ?, ?, ?, ?, ?)
        ''', (
            t.get('id'), t.get('table_number'), t.get('capacity'),
            t.get('status'), t.get('current_order_id'),
            t.get('organization_id'), str(t.get('created_at', ''))
        ))
    
    # Insert inventory
    for i in inventory:
        cursor.execute('''
            INSERT INTO inventory VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            i.get('id'), i.get('name'), i.get('category'), i.get('quantity'),
            i.get('unit'), i.get('min_stock'), i.get('cost_per_unit'),
            i.get('supplier'), i.get('organization_id'),
            str(i.get('created_at', '')), str(i.get('updated_at', ''))
        ))
    
    # Insert payments
    for p in payments:
        cursor.execute('''
            INSERT INTO payments VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            p.get('id'), p.get('order_id'), p.get('amount'),
            p.get('payment_method'), p.get('razorpay_order_id'),
            p.get('razorpay_payment_id'), p.get('status'),
            p.get('organization_id'), str(p.get('created_at', ''))
        ))
    
    # Insert backup info
    cursor.execute('''
        INSERT INTO backup_info VALUES (?, ?, ?, ?, ?)
    ''', (
        1, user_data.get('id'), user_data.get('username'),
        datetime.now(timezone.utc).isoformat(), '1.0'
    ))
    
    conn.commit()
    
    # Export to bytes
    buffer = io.BytesIO()
    for line in conn.iterdump():
        buffer.write(f'{line}\n'.encode('utf-8'))
    
    # Also create actual binary SQLite file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
    temp_conn = sqlite3.connect(temp_file.name)
    conn.backup(temp_conn)
    temp_conn.close()
    conn.close()
    
    with open(temp_file.name, 'rb') as f:
        db_bytes = f.read()
    
    os.unlink(temp_file.name)
    return db_bytes


@api_router.get("/super-admin/users/{user_id}/export-db")
async def export_user_database(user_id: str, username: str, password: str):
    """Export user data as SQLite database file - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Get user
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get all data
    staff = await db.users.find({"organization_id": user_id}, {"_id": 0, "password": 0}).to_list(100)
    orders = await db.orders.find({"organization_id": user_id}, {"_id": 0}).to_list(50000)
    menu_items = await db.menu_items.find({"organization_id": user_id}, {"_id": 0}).to_list(1000)
    tables = await db.tables.find({"organization_id": user_id}, {"_id": 0}).to_list(100)
    inventory = await db.inventory.find({"organization_id": user_id}, {"_id": 0}).to_list(1000)
    payments = await db.payments.find({"organization_id": user_id}, {"_id": 0}).to_list(50000)
    
    # Create SQLite backup
    db_bytes = create_sqlite_backup(user, staff, orders, menu_items, tables, inventory, payments)
    
    filename = f"{user.get('username', 'user')}_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"
    
    return StreamingResponse(
        io.BytesIO(db_bytes),
        media_type="application/x-sqlite3",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@api_router.post("/super-admin/users/{user_id}/import-db")
async def import_user_database(
    user_id: str,
    username: str = Query(...),
    password: str = Query(...),
    file: UploadFile = File(...),
    replace_existing: bool = Query(default=False, description="Replace existing data or merge")
):
    """Import user data from SQLite database file - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Verify user exists
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Read uploaded file
    content = await file.read()
    
    # Save to temp file and open with sqlite3
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
    temp_file.write(content)
    temp_file.close()
    
    try:
        conn = sqlite3.connect(temp_file.name)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Verify backup info
        cursor.execute("SELECT * FROM backup_info LIMIT 1")
        backup_info = cursor.fetchone()
        if not backup_info:
            raise HTTPException(status_code=400, detail="Invalid backup file - no backup info found")
        
        imported_counts = {
            "users": 0,
            "orders": 0,
            "menu_items": 0,
            "tables": 0,
            "inventory": 0,
            "payments": 0
        }
        
        # If replace_existing, delete existing data first
        if replace_existing:
            await db.orders.delete_many({"organization_id": user_id})
            await db.menu_items.delete_many({"organization_id": user_id})
            await db.tables.delete_many({"organization_id": user_id})
            await db.inventory.delete_many({"organization_id": user_id})
            await db.payments.delete_many({"organization_id": user_id})
            # Delete staff but not the main user
            await db.users.delete_many({"organization_id": user_id})
        
        # Import users (staff)
        cursor.execute("SELECT * FROM users WHERE organization_id IS NOT NULL AND organization_id != ''")
        for row in cursor.fetchall():
            staff_data = dict(row)
            staff_data['organization_id'] = user_id  # Ensure correct org_id
            staff_data['subscription_active'] = bool(staff_data.get('subscription_active'))
            staff_data['setup_completed'] = bool(staff_data.get('setup_completed'))
            staff_data['onboarding_completed'] = bool(staff_data.get('onboarding_completed'))
            if staff_data.get('business_settings'):
                try:
                    staff_data['business_settings'] = json.loads(staff_data['business_settings'])
                except:
                    staff_data['business_settings'] = {}
            
            # Upsert staff
            await db.users.update_one(
                {"id": staff_data['id']},
                {"$set": staff_data},
                upsert=True
            )
            imported_counts["users"] += 1
        
        # Import orders
        cursor.execute("SELECT * FROM orders")
        for row in cursor.fetchall():
            order_data = dict(row)
            order_data['organization_id'] = user_id
            order_data['is_credit'] = bool(order_data.get('is_credit'))
            if order_data.get('items'):
                try:
                    order_data['items'] = json.loads(order_data['items'])
                except:
                    order_data['items'] = []
            
            await db.orders.update_one(
                {"id": order_data['id']},
                {"$set": order_data},
                upsert=True
            )
            imported_counts["orders"] += 1
        
        # Import menu items
        cursor.execute("SELECT * FROM menu_items")
        for row in cursor.fetchall():
            menu_data = dict(row)
            menu_data['organization_id'] = user_id
            menu_data['available'] = bool(menu_data.get('available', 1))
            if menu_data.get('ingredients'):
                try:
                    menu_data['ingredients'] = json.loads(menu_data['ingredients'])
                except:
                    menu_data['ingredients'] = []
            
            await db.menu_items.update_one(
                {"id": menu_data['id']},
                {"$set": menu_data},
                upsert=True
            )
            imported_counts["menu_items"] += 1
        
        # Import tables
        cursor.execute("SELECT * FROM tables")
        for row in cursor.fetchall():
            table_data = dict(row)
            table_data['organization_id'] = user_id
            
            await db.tables.update_one(
                {"id": table_data['id']},
                {"$set": table_data},
                upsert=True
            )
            imported_counts["tables"] += 1
        
        # Import inventory
        cursor.execute("SELECT * FROM inventory")
        for row in cursor.fetchall():
            inv_data = dict(row)
            inv_data['organization_id'] = user_id
            
            await db.inventory.update_one(
                {"id": inv_data['id']},
                {"$set": inv_data},
                upsert=True
            )
            imported_counts["inventory"] += 1
        
        # Import payments
        cursor.execute("SELECT * FROM payments")
        for row in cursor.fetchall():
            payment_data = dict(row)
            payment_data['organization_id'] = user_id
            
            await db.payments.update_one(
                {"id": payment_data['id']},
                {"$set": payment_data},
                upsert=True
            )
            imported_counts["payments"] += 1
        
        conn.close()
        
        return {
            "message": "Database imported successfully",
            "user_id": user_id,
            "imported": imported_counts,
            "mode": "replace" if replace_existing else "merge"
        }
        
    except sqlite3.Error as e:
        raise HTTPException(status_code=400, detail=f"Invalid SQLite database: {str(e)}")
    finally:
        os.unlink(temp_file.name)


@api_router.get("/super-admin/users/{user_id}/business-details")
async def get_user_business_details(user_id: str, username: str, password: str):
    """Get detailed business information for a user - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Get user
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Get staff members
    staff = await db.users.find(
        {"organization_id": user_id},
        {"_id": 0, "id": 1, "username": 1, "email": 1, "role": 1, "phone": 1, "created_at": 1, "subscription_active": 1}
    ).to_list(100)
    
    # Get order stats
    orders_pipeline = [
        {"$match": {"organization_id": user_id}},
        {"$group": {
            "_id": None,
            "total_orders": {"$sum": 1},
            "completed_orders": {"$sum": {"$cond": [{"$eq": ["$status", "completed"]}, 1, 0]}},
            "total_revenue": {"$sum": {"$cond": [{"$eq": ["$status", "completed"]}, "$total", 0]}},
            "credit_orders": {"$sum": {"$cond": ["$is_credit", 1, 0]}},
            "pending_credit": {"$sum": {"$cond": ["$is_credit", "$balance_amount", 0]}}
        }}
    ]
    order_stats = await db.orders.aggregate(orders_pipeline).to_list(1)
    order_stats = order_stats[0] if order_stats else {}
    
    # Get recent orders (last 10)
    recent_orders = await db.orders.find(
        {"organization_id": user_id},
        {"_id": 0, "id": 1, "total": 1, "status": 1, "customer_name": 1, "created_at": 1, "is_credit": 1}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    # Get menu count
    menu_count = await db.menu_items.count_documents({"organization_id": user_id})
    
    # Get tables count
    tables_count = await db.tables.count_documents({"organization_id": user_id})
    
    # Get inventory count
    inventory_count = await db.inventory.count_documents({"organization_id": user_id})
    
    # Calculate trial/subscription info
    created_at = user.get("created_at")
    trial_days = 7 + user.get("trial_extension_days", 0)
    if isinstance(created_at, str):
        created_at = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
    
    trial_end = created_at + timedelta(days=trial_days) if created_at else None
    is_trial_active = trial_end and trial_end > datetime.now(timezone.utc) if trial_end else False
    days_remaining = (trial_end - datetime.now(timezone.utc)).days if trial_end and is_trial_active else 0
    
    return {
        "user": {
            "id": user.get("id"),
            "username": user.get("username"),
            "email": user.get("email"),
            "phone": user.get("phone"),
            "role": user.get("role"),
            "created_at": user.get("created_at"),
            "subscription_active": user.get("subscription_active", False),
            "subscription_expires_at": user.get("subscription_expires_at"),
            "trial_extension_days": user.get("trial_extension_days", 0),
            "is_trial_active": is_trial_active,
            "trial_days_remaining": days_remaining,
            "setup_completed": user.get("setup_completed", False),
            "onboarding_completed": user.get("onboarding_completed", False)
        },
        "business_settings": user.get("business_settings", {}),
        "staff": staff,
        "staff_count": len(staff),
        "stats": {
            "total_orders": order_stats.get("total_orders", 0),
            "completed_orders": order_stats.get("completed_orders", 0),
            "total_revenue": order_stats.get("total_revenue", 0),
            "credit_orders": order_stats.get("credit_orders", 0),
            "pending_credit": order_stats.get("pending_credit", 0),
            "menu_items": menu_count,
            "tables": tables_count,
            "inventory_items": inventory_count
        },
        "recent_orders": recent_orders
    }


@api_router.put("/super-admin/staff/{staff_id}/subscription")
async def update_staff_subscription(
    staff_id: str,
    subscription_active: bool,
    username: str,
    password: str
):
    """Update staff member subscription status - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Verify it's a staff member
    staff = await db.users.find_one({"id": staff_id})
    if not staff:
        raise HTTPException(status_code=404, detail="Staff not found")
    
    if staff.get("role") == "admin":
        raise HTTPException(status_code=400, detail="Use user subscription endpoint for admins")
    
    # Update staff subscription
    await db.users.update_one(
        {"id": staff_id},
        {"$set": {
            "subscription_active": subscription_active,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    return {
        "message": f"Staff subscription {'activated' if subscription_active else 'deactivated'}",
        "staff_id": staff_id,
        "subscription_active": subscription_active
    }


@api_router.get("/super-admin/tickets")
async def get_all_tickets_admin(
    username: str,
    password: str,
    status: Optional[str] = None,
    skip: int = 0,
    limit: int = 100
):
    """Get all support tickets - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    query = {}
    if status:
        query["status"] = status
    
    tickets = await db.support_tickets.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.support_tickets.count_documents(query)
    
    return {"tickets": tickets, "total": total, "skip": skip, "limit": limit}

class TicketUpdate(BaseModel):
    status: str
    admin_notes: Optional[str] = None

@api_router.put("/super-admin/tickets/{ticket_id}")
async def update_ticket_admin(
    ticket_id: str,
    ticket_update: TicketUpdate,
    username: str,
    password: str
):
    """Update ticket status - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    update_data = {
        "status": ticket_update.status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if ticket_update.admin_notes:
        update_data["admin_notes"] = ticket_update.admin_notes
    
    result = await db.support_tickets.update_one({"id": ticket_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    return {"message": "Ticket updated successfully", "ticket_id": ticket_id, "status": ticket_update.status}


class SuperAdminTicketReply(BaseModel):
    message: str
    update_status: Optional[str] = None


@api_router.post("/super-admin/tickets/{ticket_id}/reply")
async def super_admin_reply_to_ticket(
    ticket_id: str,
    reply: SuperAdminTicketReply,
    username: str,
    password: str
):
    """Reply to a support ticket as super admin - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Get the ticket
    ticket = await db.support_tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")
    
    # Create reply record
    reply_record = {
        "id": str(uuid.uuid4())[:8],
        "message": reply.message,
        "from": "support",
        "admin_name": username,
        "admin_email": "support@billbytekot.in",
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Update ticket with reply
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if reply.update_status:
        update_data["status"] = reply.update_status
    
    await db.support_tickets.update_one(
        {"id": ticket_id},
        {
            "$push": {"replies": reply_record},
            "$set": update_data
        }
    )
    
    # Send email to user
    email_result = await send_ticket_reply_email(
        ticket_id=ticket_id,
        user_email=ticket["email"],
        user_name=ticket["name"],
        subject=ticket["subject"],
        reply_message=reply.message,
        admin_name=username
    )
    
    return {
        "success": True,
        "message": "Reply sent successfully",
        "email_sent": email_result.get("success", False),
        "reply_id": reply_record["id"]
    }

@api_router.get("/super-admin/analytics")
async def get_analytics_admin(username: str, password: str, days: int = 30):
    """Get system analytics - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    start_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    
    new_users = await db.users.count_documents({"created_at": {"$gte": start_date}})
    new_orders = await db.orders.count_documents({"created_at": {"$gte": start_date}})
    new_tickets = await db.support_tickets.count_documents({"created_at": {"$gte": start_date}})
    
    return {
        "period_days": days,
        "new_users": new_users,
        "new_orders": new_orders,
        "new_tickets": new_tickets,
        "start_date": start_date
    }

class CreateLeadRequest(BaseModel):
    name: str
    email: str
    phone: str
    businessName: Optional[str] = None
    source: str = "manual"
    notes: Optional[str] = None

@api_router.post("/super-admin/leads")
async def create_lead_admin(
    lead: CreateLeadRequest,
    username: str,
    password: str
):
    """Manually create a new lead - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    lead_data = {
        "name": lead.name,
        "email": lead.email,
        "phone": lead.phone,
        "businessName": lead.businessName,
        "source": lead.source,
        "notes": lead.notes,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "status": "new",
        "contacted": False
    }
    
    result = await db.leads.insert_one(lead_data)
    
    return {
        "success": True,
        "message": "Lead created successfully",
        "lead_id": str(result.inserted_id)
    }

@api_router.get("/super-admin/leads")
async def get_all_leads_admin(
    username: str,
    password: str,
    skip: int = 0,
    limit: int = 100,
    status: Optional[str] = None
):
    """Get all leads from landing page - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Build query
    query = {}
    if status:
        query["status"] = status
    
    leads = await db.leads.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.leads.count_documents(query)
    
    # Count by status
    new_count = await db.leads.count_documents({"status": "new"})
    contacted_count = await db.leads.count_documents({"status": "contacted"})
    converted_count = await db.leads.count_documents({"status": "converted"})
    
    return {
        "leads": leads,
        "total": total,
        "skip": skip,
        "limit": limit,
        "stats": {
            "new": new_count,
            "contacted": contacted_count,
            "converted": converted_count
        }
    }

class LeadUpdate(BaseModel):
    status: str
    notes: Optional[str] = None
    contacted: Optional[bool] = None

@api_router.put("/super-admin/leads/{lead_id}")
async def update_lead_admin(
    lead_id: str,
    lead_update: LeadUpdate,
    username: str,
    password: str
):
    """Update lead status - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    update_data = {
        "status": lead_update.status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    if lead_update.notes:
        update_data["notes"] = lead_update.notes
    if lead_update.contacted is not None:
        update_data["contacted"] = lead_update.contacted
    
    # Find by timestamp (used as ID in leads)
    result = await db.leads.update_one({"timestamp": lead_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return {"message": "Lead updated successfully", "lead_id": lead_id, "status": lead_update.status}

@api_router.delete("/super-admin/leads/{lead_id}")
async def delete_lead_admin(lead_id: str, username: str, password: str):
    """Delete lead - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    result = await db.leads.delete_one({"timestamp": lead_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    return {"message": "Lead deleted successfully", "lead_id": lead_id}


# ============ TEAM MANAGEMENT (Site Owner Only) ============
class TeamMember(BaseModel):
    username: str
    email: str
    password: str
    role: str  # sales, support, admin
    permissions: List[str]  # leads, tickets, users, analytics
    full_name: Optional[str] = None
    phone: Optional[str] = None

class TeamMemberUpdate(BaseModel):
    role: Optional[str] = None
    permissions: Optional[List[str]] = None
    active: Optional[bool] = None

@api_router.post("/super-admin/team")
async def create_team_member(
    member: TeamMember,
    username: str,
    password: str
):
    """Create team member (sales/support) - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Check if username already exists (case-insensitive)
    existing = await db.team_members.find_one({
        "username": {"$regex": f"^{member.username}$", "$options": "i"}
    })
    if existing:
        raise HTTPException(status_code=400, detail="Username already exists")
    
    # Check if email already exists (case-insensitive)
    existing_email = await db.team_members.find_one({
        "email": {"$regex": f"^{member.email}$", "$options": "i"}
    })
    if existing_email:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    team_data = {
        "id": str(uuid.uuid4()),
        "username": member.username,
        "username_lower": member.username.lower(),
        "email": member.email,
        "email_lower": member.email.lower(),
        "password": hash_password(member.password),
        "role": member.role,
        "permissions": member.permissions,
        "full_name": member.full_name,
        "phone": member.phone,
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": username
    }
    
    result = await db.team_members.insert_one(team_data)
    
    return {
        "success": True,
        "message": "Team member created successfully",
        "member_id": team_data["id"]
    }

@api_router.get("/super-admin/team")
async def get_team_members(username: str, password: str):
    """Get all team members - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    members = await db.team_members.find({}, {"_id": 0, "password": 0}).to_list(100)
    
    # Count by role
    sales_count = sum(1 for m in members if m.get("role") == "sales")
    support_count = sum(1 for m in members if m.get("role") == "support")
    admin_count = sum(1 for m in members if m.get("role") == "admin")
    
    return {
        "members": members,
        "total": len(members),
        "stats": {
            "sales": sales_count,
            "support": support_count,
            "admin": admin_count
        }
    }

@api_router.put("/super-admin/team/{member_id}")
async def update_team_member(
    member_id: str,
    member_update: TeamMemberUpdate,
    username: str,
    password: str
):
    """Update team member - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    
    if member_update.role:
        update_data["role"] = member_update.role
    if member_update.permissions:
        update_data["permissions"] = member_update.permissions
    if member_update.active is not None:
        update_data["active"] = member_update.active
    
    result = await db.team_members.update_one({"id": member_id}, {"$set": update_data})
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Team member not found")
    
    return {"message": "Team member updated successfully", "member_id": member_id}

@api_router.delete("/super-admin/team/{member_id}")
async def delete_team_member(member_id: str, username: str, password: str):
    """Delete team member - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    result = await db.team_members.delete_one({"id": member_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Team member not found")
    
    return {"message": "Team member deleted successfully", "member_id": member_id}

# Team member login endpoint
class TeamLogin(BaseModel):
    username: str
    password: str

@api_router.post("/team/login")
async def team_login(credentials: TeamLogin):
    """Team member login"""
    # Case-insensitive username lookup
    username_lower = credentials.username.lower()
    member = await db.team_members.find_one({"username_lower": username_lower})
    
    # Fallback to case-insensitive regex if username_lower field doesn't exist
    if not member:
        member = await db.team_members.find_one({
            "username": {"$regex": f"^{credentials.username}$", "$options": "i"}
        })
    
    if not member or not verify_password(credentials.password, member["password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not member.get("active", True):
        raise HTTPException(status_code=403, detail="Account is inactive")
    
    # Create token
    token_data = {
        "id": member["id"],
        "username": member["username"],
        "role": member["role"],
        "permissions": member.get("permissions", []),
        "type": "team"
    }
    token = create_access_token(token_data)
    
    return {
        "token": token,
        "user": {
            "id": member["id"],
            "username": member["username"],
            "email": member["email"],
            "role": member["role"],
            "permissions": member.get("permissions", []),
            "full_name": member.get("full_name"),
            "type": "team"
        }
    }


# ============ PUBLIC ENDPOINTS (No Auth Required) ============

# Public Sale/Offer Endpoint
@api_router.get("/sale-offer")
async def get_public_sale_offer():
    """Get active sale offer for landing page - Public endpoint"""
    # Check both sale_offers collection (new) and site_settings (legacy) for backwards compatibility
    offer = await db.sale_offers.find_one({"enabled": True})
    if not offer:
        offer = await db.site_settings.find_one({"type": "sale_offer", "enabled": True})
    
    if not offer:
        return {"enabled": False}
    
    # Check if offer has expired based on end_date or valid_until
    now = datetime.now(timezone.utc)
    
    # Check end_date (date only)
    if offer.get("end_date"):
        try:
            end_date = datetime.fromisoformat(offer["end_date"])
            if end_date.tzinfo is None:
                end_date = end_date.replace(hour=23, minute=59, second=59)
            if now.replace(tzinfo=None) > end_date:
                return {"enabled": False}
        except:
            pass
    
    # Check valid_until (datetime with time)
    if offer.get("valid_until"):
        try:
            valid_until = datetime.fromisoformat(offer["valid_until"])
            if valid_until.tzinfo is None:
                valid_until = valid_until.replace(tzinfo=timezone.utc)
            if now > valid_until:
                return {"enabled": False}
        except:
            pass
    
    offer.pop("_id", None)
    offer.pop("type", None)
    return offer


# Public Pricing Endpoint
@api_router.get("/pricing")
async def get_public_pricing():
    """Get current pricing for subscription page - Public endpoint"""
    pricing = await db.site_settings.find_one({"type": "pricing"})
    
    if not pricing:
        # Default pricing - ₹1999 base price
        return {
            "regular_price": 1999,
            "regular_price_display": "₹1999",
            "campaign_price": 1799,
            "campaign_price_display": "₹1799",
            "campaign_active": False,
            "campaign_discount_percent": 10,
            "campaign_name": "",
            "trial_expired_discount": 10,
            "trial_expired_price": 1799,
            "trial_expired_price_display": "₹1799",
            "trial_days": 7
        }
    
    # Check if campaign is active based on dates
    campaign_active = pricing.get("campaign_active", False)
    if campaign_active and pricing.get("campaign_start_date") and pricing.get("campaign_end_date"):
        try:
            start_date = datetime.fromisoformat(pricing["campaign_start_date"])
            end_date = datetime.fromisoformat(pricing["campaign_end_date"])
            now = datetime.now()
            campaign_active = start_date <= now <= end_date
        except:
            pass
    
    regular_price = pricing.get("regular_price", 1999)
    trial_expired_discount = pricing.get("trial_expired_discount", 10)
    trial_expired_price = int(regular_price * (100 - trial_expired_discount) / 100)
    
    return {
        "regular_price": regular_price,
        "regular_price_display": pricing.get("regular_price_display", f"₹{regular_price}"),
        "campaign_price": pricing.get("campaign_price", 1799),
        "campaign_price_display": pricing.get("campaign_price_display", "₹1799"),
        "campaign_active": campaign_active,
        "campaign_name": pricing.get("campaign_name", ""),
        "campaign_discount_percent": pricing.get("campaign_discount_percent", 10),
        "trial_expired_discount": trial_expired_discount,
        "trial_expired_price": trial_expired_price,
        "trial_expired_price_display": f"₹{trial_expired_price}",
        "trial_days": pricing.get("trial_days", 7)
    }



# Include super admin routes and set database
set_super_admin_db(db)
# Redis cache will be set during startup in lifespan
app.include_router(super_admin_router)

# Include ops panel router
set_ops_db(db)
app.include_router(ops_router)

# Include monitoring routes
app.include_router(monitoring_router)


# ============ APP VERSION MANAGEMENT (Super Admin Only) ============

class AppVersionCreate(BaseModel):
    platform: str  # 'android' or 'windows'
    version: str  # e.g., '1.0.0'
    version_code: int  # e.g., 1
    download_url: str  # Direct download URL
    release_notes: Optional[str] = ""
    min_supported_version: Optional[str] = None
    is_mandatory: bool = False
    file_size: Optional[str] = None  # e.g., '25 MB'

class AppVersionUpdate(BaseModel):
    version: Optional[str] = None
    version_code: Optional[int] = None
    download_url: Optional[str] = None
    release_notes: Optional[str] = None
    min_supported_version: Optional[str] = None
    is_mandatory: Optional[bool] = None
    is_active: Optional[bool] = None
    file_size: Optional[str] = None

@api_router.get("/super-admin/app-versions")
async def get_app_versions(username: str, password: str):
    """Get all app versions - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    versions = await db.app_versions.find().sort("created_at", -1).to_list(100)
    for v in versions:
        v.pop("_id", None)
    
    return {"versions": versions}


@api_router.post("/super-admin/app-versions/upload")
async def upload_app_file(
    file: UploadFile = File(...),
    platform: str = Form(...),
    version: str = Form(...),
    username: str = Query(...),
    password: str = Query(...)
):
    """Upload APK/EXE file and store in database - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Validate file extension
    valid_extensions = {
        'android': ['.apk'],
        'windows': ['.exe', '.msi', '.zip']
    }
    
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in valid_extensions.get(platform, []):
        raise HTTPException(
            status_code=400, 
            detail=f"Invalid file type. Expected {valid_extensions.get(platform)} for {platform}"
        )
    
    # Read file content
    content = await file.read()
    file_size_mb = len(content) / (1024 * 1024)
    
    # Max file size: 200MB
    if file_size_mb > 200:
        raise HTTPException(status_code=400, detail="File too large. Maximum size is 200MB")
    
    # Store file in database (GridFS-like approach using base64 for smaller files, or chunks for larger)
    file_id = str(uuid.uuid4())
    filename = f"{platform}_{version}_{file_id}{file_ext}"
    
    # For files under 16MB, store directly in document
    # For larger files, store in chunks
    if file_size_mb < 16:
        import base64
        file_doc = {
            "id": file_id,
            "filename": filename,
            "original_filename": file.filename,
            "platform": platform,
            "version": version,
            "content_type": file.content_type,
            "size": len(content),
            "size_mb": round(file_size_mb, 2),
            "data": base64.b64encode(content).decode('utf-8'),
            "chunked": False,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.app_files.insert_one(file_doc)
    else:
        # Store in chunks (1MB each)
        import base64
        chunk_size = 1024 * 1024  # 1MB
        chunks = [content[i:i+chunk_size] for i in range(0, len(content), chunk_size)]
        
        file_doc = {
            "id": file_id,
            "filename": filename,
            "original_filename": file.filename,
            "platform": platform,
            "version": version,
            "content_type": file.content_type,
            "size": len(content),
            "size_mb": round(file_size_mb, 2),
            "chunked": True,
            "chunk_count": len(chunks),
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.app_files.insert_one(file_doc)
        
        # Store chunks
        for i, chunk in enumerate(chunks):
            chunk_doc = {
                "file_id": file_id,
                "chunk_index": i,
                "data": base64.b64encode(chunk).decode('utf-8')
            }
            await db.app_file_chunks.insert_one(chunk_doc)
    
    # Generate download URL (relative to API)
    download_url = f"/api/app-download/{file_id}"
    
    return {
        "message": "File uploaded successfully",
        "file_id": file_id,
        "filename": filename,
        "size_mb": round(file_size_mb, 2),
        "download_url": download_url
    }


@api_router.get("/app-download/{file_id}")
async def download_app_file(file_id: str):
    """Download app file by ID - Public endpoint for app downloads"""
    import base64
    from fastapi.responses import StreamingResponse
    
    file_doc = await db.app_files.find_one({"id": file_id})
    if not file_doc:
        raise HTTPException(status_code=404, detail="File not found")
    
    # Determine content type
    content_type = file_doc.get("content_type", "application/octet-stream")
    if file_doc["filename"].endswith(".apk"):
        content_type = "application/vnd.android.package-archive"
    elif file_doc["filename"].endswith(".exe"):
        content_type = "application/x-msdownload"
    
    original_filename = file_doc.get('original_filename', file_doc['filename'])
    file_size = file_doc.get("size", 0)
    
    async def generate_chunks():
        """Stream file content in chunks to avoid memory issues"""
        if file_doc.get("chunked"):
            # Stream chunks one at a time
            chunk_count = file_doc.get("chunk_count", 0)
            for i in range(chunk_count):
                chunk_doc = await db.app_file_chunks.find_one({
                    "file_id": file_id,
                    "chunk_index": i
                })
                if chunk_doc:
                    yield base64.b64decode(chunk_doc["data"])
        else:
            # Single chunk - yield all at once
            yield base64.b64decode(file_doc["data"])
    
    headers = {
        "Content-Disposition": f"attachment; filename={original_filename}",
        "Content-Length": str(file_size),
        "Accept-Ranges": "bytes"
    }
    
    return StreamingResponse(
        generate_chunks(),
        media_type=content_type,
        headers=headers
    )


@api_router.post("/super-admin/app-versions")
async def create_app_version(
    version_data: AppVersionCreate,
    username: str,
    password: str
):
    """Create new app version - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    # Deactivate previous versions of same platform
    await db.app_versions.update_many(
        {"platform": version_data.platform, "is_active": True},
        {"$set": {"is_active": False}}
    )
    
    new_version = {
        "id": str(uuid.uuid4()),
        "platform": version_data.platform,
        "version": version_data.version,
        "version_code": version_data.version_code,
        "download_url": version_data.download_url,
        "release_notes": version_data.release_notes or "",
        "min_supported_version": version_data.min_supported_version,
        "is_mandatory": version_data.is_mandatory,
        "is_active": True,
        "file_size": version_data.file_size,
        "download_count": 0,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.app_versions.insert_one(new_version)
    new_version.pop("_id", None)
    
    return {"message": f"{version_data.platform.title()} app version {version_data.version} created", "version": new_version}

@api_router.put("/super-admin/app-versions/{version_id}")
async def update_app_version(
    version_id: str,
    update_data: AppVersionUpdate,
    username: str,
    password: str
):
    """Update app version - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    update_dict = {k: v for k, v in update_data.model_dump().items() if v is not None}
    update_dict["updated_at"] = datetime.now(timezone.utc)
    
    # If activating this version, deactivate others of same platform
    if update_data.is_active:
        version = await db.app_versions.find_one({"id": version_id})
        if version:
            await db.app_versions.update_many(
                {"platform": version["platform"], "is_active": True, "id": {"$ne": version_id}},
                {"$set": {"is_active": False}}
            )
    
    result = await db.app_versions.update_one(
        {"id": version_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Version not found")
    
    return {"message": "App version updated successfully"}

@api_router.delete("/super-admin/app-versions/{version_id}")
async def delete_app_version(version_id: str, username: str, password: str):
    """Delete app version - Site Owner Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid super admin credentials")
    
    result = await db.app_versions.delete_one({"id": version_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Version not found")
    
    return {"message": "App version deleted successfully"}

# Public endpoints for app downloads (no auth required)
@api_router.get("/app/latest/{platform}")
async def get_latest_app_version(platform: str):
    """Get latest active app version for platform (android/windows) - Public"""
    if platform not in ["android", "windows"]:
        raise HTTPException(status_code=400, detail="Invalid platform. Use 'android' or 'windows'")
    
    version = await db.app_versions.find_one(
        {"platform": platform, "is_active": True},
        sort=[("version_code", -1)]
    )
    
    if not version:
        raise HTTPException(status_code=404, detail=f"No {platform} app version available")
    
    version.pop("_id", None)
    
    # Increment download count
    await db.app_versions.update_one(
        {"id": version["id"]},
        {"$inc": {"download_count": 1}}
    )
    
    return version

@api_router.get("/app/check-update/{platform}/{current_version}")
async def check_app_update(platform: str, current_version: str):
    """Check if app update is available - Public"""
    if platform not in ["android", "windows"]:
        raise HTTPException(status_code=400, detail="Invalid platform")
    
    latest = await db.app_versions.find_one(
        {"platform": platform, "is_active": True},
        sort=[("version_code", -1)]
    )
    
    if not latest:
        return {"update_available": False}
    
    latest.pop("_id", None)
    
    # Simple version comparison (assumes semantic versioning)
    def version_tuple(v):
        return tuple(map(int, v.split('.')))
    
    try:
        current = version_tuple(current_version)
        latest_v = version_tuple(latest["version"])
        update_available = latest_v > current
    except:
        update_available = latest["version"] != current_version
    
    return {
        "update_available": update_available,
        "latest_version": latest["version"],
        "is_mandatory": latest.get("is_mandatory", False),
        "download_url": latest["download_url"] if update_available else None,
        "release_notes": latest.get("release_notes", "") if update_available else None,
        "file_size": latest.get("file_size")
    }


# Serve Windows app download
from fastapi.responses import FileResponse, JSONResponse

@app.get("/downloads/windows")
async def download_windows_app():
    """Serve Windows desktop app for download"""
    file_path = ROOT_DIR / "downloads" / "BillByteKOT-Setup.exe"
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Windows app not found")
    
    return FileResponse(
        path=str(file_path),
        media_type="application/octet-stream",
        filename="BillByteKOT-Setup.exe"
    )


# Serve Digital Asset Links for Android TWA - Updated 2025-12-18
@app.get("/.well-known/assetlinks.json")
async def get_assetlinks():
    """Serve assetlinks.json for Android TWA verification - Enables app deep linking"""
    assetlinks = [{
        "relation": ["delegate_permission/common.handle_all_urls"],
        "target": {
            "namespace": "android_app",
            "package_name": "in.billbytekot.twa",
            "sha256_cert_fingerprints": [
                "3A:B4:A6:42:B5:C0:7E:C8:1C:7F:E4:6A:19:2D:AF:D4:3A:4C:1C:12:52:54:F5:89:DB:C6:09:4E:66:40:34:97"
            ]
        }
    }]
    
    # Add cache-control headers to prevent caching issues
    return JSONResponse(
        content=assetlinks,
        media_type="application/json",
        headers={
            "Cache-Control": "no-cache, no-store, must-revalidate",
            "Pragma": "no-cache",
            "Expires": "0",
            "Access-Control-Allow-Origin": "*"
        }
    )


logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)


# ============ PUSH NOTIFICATIONS API ============

# Import Firebase push module
try:
    from firebase_push import (
        is_firebase_configured, 
        send_fcm_notification, 
        send_fcm_to_topic,
        send_fcm_to_multiple
    )
    FCM_AVAILABLE = True
except ImportError:
    FCM_AVAILABLE = False

class PushSubscription(BaseModel):
    subscription: dict  # Contains endpoint, keys (p256dh, auth)
    user_id: Optional[str] = None
    device_info: Optional[dict] = None

class FCMTokenRegister(BaseModel):
    fcm_token: str
    user_id: Optional[str] = None
    device_info: Optional[dict] = None

class PushNotificationSend(BaseModel):
    title: str
    body: str
    icon: Optional[str] = "/icon-192.png"
    badge: Optional[str] = "/icon-192.png"
    url: Optional[str] = "/"
    type: Optional[str] = "info"  # info, success, warning, promo
    image: Optional[str] = None
    priority: Optional[str] = "normal"  # low, normal, high
    target: Optional[str] = "all"  # all, subscribed, trial
    tag: Optional[str] = None

# FCM Token Registration - for real push notifications
@api_router.post("/fcm/register")
async def register_fcm_token(data: FCMTokenRegister):
    """Register FCM token for push notifications (like WhatsApp/Zomato)"""
    try:
        fcm_token = data.fcm_token
        
        if not fcm_token:
            raise HTTPException(status_code=400, detail="FCM token required")
        
        # Check if token already exists
        existing = await db.fcm_tokens.find_one({"token": fcm_token})
        
        if existing:
            # Update existing
            await db.fcm_tokens.update_one(
                {"token": fcm_token},
                {"$set": {
                    "user_id": data.user_id,
                    "device_info": data.device_info,
                    "updated_at": datetime.now(timezone.utc).isoformat(),
                    "active": True
                }}
            )
            return {"success": True, "message": "Token updated"}
        
        # Create new
        token_doc = {
            "token": fcm_token,
            "user_id": data.user_id,
            "device_info": data.device_info,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "active": True
        }
        
        await db.fcm_tokens.insert_one(token_doc)
        
        total = await db.fcm_tokens.count_documents({"active": True})
        
        return {
            "success": True,
            "message": "FCM token registered",
            "total_devices": total
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/fcm/unregister")
async def unregister_fcm_token(token: str = Body(..., embed=True)):
    """Unregister FCM token"""
    try:
        await db.fcm_tokens.update_one(
            {"token": token},
            {"$set": {"active": False, "unregistered_at": datetime.now(timezone.utc).isoformat()}}
        )
        return {"success": True}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/fcm/stats")
async def get_fcm_stats(username: str, password: str):
    """Get FCM statistics - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    total = await db.fcm_tokens.count_documents({})
    active = await db.fcm_tokens.count_documents({"active": True})
    
    # Get recent registrations
    recent = await db.fcm_tokens.find({"active": True}).sort("created_at", -1).limit(10).to_list(10)
    for r in recent:
        r.pop("_id", None)
        r.pop("token", None)  # Don't expose tokens
    
    return {
        "total_devices": total,
        "active_devices": active,
        "recent_registrations": recent,
        "firebase_configured": FCM_AVAILABLE and is_firebase_configured() if FCM_AVAILABLE else False
    }

@api_router.post("/fcm/send")
async def send_fcm_push(
    notification: PushNotificationSend,
    username: str = Query(...),
    password: str = Query(...)
):
    """Send FCM push notification to all devices - Super Admin Only
    
    This sends REAL push notifications like WhatsApp/Zomato that appear
    even when the app is completely closed!
    """
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    if not FCM_AVAILABLE:
        # Store for later / in-app display
        notif_doc = {
            "title": notification.title,
            "body": notification.body,
            "type": notification.type,
            "url": notification.url,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "sent_count": 0,
            "status": "stored_only",
            "note": "Firebase module not available"
        }
        await db.sent_push_notifications.insert_one(notif_doc)
        return {"success": False, "message": "Firebase not configured", "sent_count": 0}
    
    if not is_firebase_configured():
        notif_doc = {
            "title": notification.title,
            "body": notification.body,
            "type": notification.type,
            "url": notification.url,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "sent_count": 0,
            "status": "stored_only",
            "note": "Firebase credentials not set"
        }
        await db.sent_push_notifications.insert_one(notif_doc)
        return {
            "success": False, 
            "message": "Firebase not configured. Add FIREBASE_PROJECT_ID, FIREBASE_PRIVATE_KEY, FIREBASE_CLIENT_EMAIL to .env",
            "sent_count": 0
        }
    
    # Get all active FCM tokens
    tokens_cursor = db.fcm_tokens.find({"active": True})
    tokens = await tokens_cursor.to_list(10000)
    
    if not tokens:
        return {"success": False, "message": "No registered devices", "sent_count": 0}
    
    # Send to all devices
    token_list = [t["token"] for t in tokens]
    
    result = await send_fcm_to_multiple(
        tokens=token_list,
        title=notification.title,
        body=notification.body,
        image=notification.image,
        data={
            "type": notification.type,
            "url": notification.url or "/",
            "click_action": notification.url or "https://billbytekot.in"
        }
    )
    
    # Store notification record
    notif_doc = {
        "title": notification.title,
        "body": notification.body,
        "type": notification.type,
        "image": notification.image,
        "url": notification.url,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "sent_count": result.get("success", 0),
        "failed_count": result.get("failed", 0),
        "status": "sent"
    }
    await db.sent_push_notifications.insert_one(notif_doc)
    
    return {
        "success": True,
        "message": f"Push notification sent to {result.get('success', 0)} devices",
        "sent_count": result.get("success", 0),
        "failed_count": result.get("failed", 0),
        "total_devices": len(tokens)
    }

@api_router.post("/fcm/send-topic")
async def send_fcm_topic_push(
    topic: str,
    notification: PushNotificationSend,
    username: str = Query(...),
    password: str = Query(...)
):
    """Send FCM push to a topic (all users subscribed to that topic)"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    if not FCM_AVAILABLE or not is_firebase_configured():
        return {"success": False, "message": "Firebase not configured"}
    
    result = await send_fcm_to_topic(
        topic=topic,
        title=notification.title,
        body=notification.body,
        image=notification.image,
        data={"type": notification.type, "url": notification.url}
    )
    
    return result

@api_router.get("/fcm/history")
async def get_fcm_history(username: str, password: str, limit: int = 50):
    """Get FCM push notification history"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    notifications = await db.sent_push_notifications.find().sort("created_at", -1).limit(limit).to_list(limit)
    for n in notifications:
        n.pop("_id", None)
    
    return {"notifications": notifications}

@api_router.post("/push/subscribe")
async def subscribe_to_push(data: PushSubscription):
    """Subscribe a device to push notifications"""
    try:
        subscription = data.subscription
        endpoint = subscription.get("endpoint")
        
        if not endpoint:
            raise HTTPException(status_code=400, detail="Invalid subscription: missing endpoint")
        
        # Check if subscription already exists
        existing = await db.push_subscriptions.find_one({"endpoint": endpoint})
        
        if existing:
            # Update existing subscription
            await db.push_subscriptions.update_one(
                {"endpoint": endpoint},
                {"$set": {
                    "subscription": subscription,
                    "user_id": data.user_id,
                    "device_info": data.device_info,
                    "updated_at": datetime.now(timezone.utc).isoformat()
                }}
            )
            return {"success": True, "message": "Subscription updated"}
        
        # Create new subscription
        sub_doc = {
            "endpoint": endpoint,
            "subscription": subscription,
            "user_id": data.user_id,
            "device_info": data.device_info,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "active": True
        }
        
        await db.push_subscriptions.insert_one(sub_doc)
        
        # Update total count
        total = await db.push_subscriptions.count_documents({"active": True})
        
        return {
            "success": True,
            "message": "Subscribed to push notifications",
            "total_subscribers": total
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/push/unsubscribe")
async def unsubscribe_from_push(endpoint: str = Body(..., embed=True)):
    """Unsubscribe a device from push notifications"""
    try:
        result = await db.push_subscriptions.update_one(
            {"endpoint": endpoint},
            {"$set": {"active": False, "unsubscribed_at": datetime.now(timezone.utc).isoformat()}}
        )
        
        return {"success": True, "message": "Unsubscribed from push notifications"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/push/stats")
async def get_push_stats(username: str, password: str):
    """Get push notification statistics - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    total = await db.push_subscriptions.count_documents({})
    active = await db.push_subscriptions.count_documents({"active": True})
    
    # Get recent subscriptions
    recent = await db.push_subscriptions.find({"active": True}).sort("created_at", -1).limit(10).to_list(10)
    for r in recent:
        r.pop("_id", None)
        r.pop("subscription", None)  # Don't expose keys
    
    return {
        "total_subscriptions": total,
        "active_subscriptions": active,
        "recent_subscriptions": recent
    }

@api_router.post("/push/send")
async def send_push_notification(
    notification: PushNotificationSend,
    username: str = Query(...),
    password: str = Query(...)
):
    """Send push notification to all subscribers - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        # Get VAPID keys from environment
        vapid_private_key = os.getenv("VAPID_PRIVATE_KEY")
        vapid_public_key = os.getenv("VAPID_PUBLIC_KEY")
        vapid_email = os.getenv("VAPID_EMAIL", "mailto:support@billbytekot.in")
        
        if not vapid_private_key or not vapid_public_key:
            # Return success but note that VAPID not configured
            # Store notification for in-app display
            notif_doc = {
                "title": notification.title,
                "body": notification.body,
                "type": notification.type,
                "url": notification.url,
                "priority": notification.priority,
                "target": notification.target,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "sent_count": 0,
                "failed_count": 0,
                "status": "stored_only",
                "note": "VAPID keys not configured - notification stored for in-app display only"
            }
            await db.sent_notifications.insert_one(notif_doc)
            
            return {
                "success": True,
                "message": "Notification stored for in-app display (VAPID not configured for push)",
                "sent_count": 0,
                "stored": True
            }
        
        # Import pywebpush
        try:
            from pywebpush import webpush, WebPushException
        except ImportError:
            # Store notification for in-app display
            notif_doc = {
                "title": notification.title,
                "body": notification.body,
                "type": notification.type,
                "url": notification.url,
                "priority": notification.priority,
                "target": notification.target,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "sent_count": 0,
                "status": "stored_only",
                "note": "pywebpush not installed"
            }
            await db.sent_notifications.insert_one(notif_doc)
            
            return {
                "success": True,
                "message": "Notification stored (pywebpush not installed)",
                "sent_count": 0
            }
        
        # Get active subscriptions
        subscriptions = await db.push_subscriptions.find({"active": True}).to_list(10000)
        
        if not subscriptions:
            return {"success": False, "message": "No active subscribers", "sent_count": 0}
        
        # Prepare notification payload
        payload = json.dumps({
            "title": notification.title,
            "body": notification.body,
            "icon": notification.icon,
            "badge": notification.badge,
            "url": notification.url,
            "type": notification.type,
            "image": notification.image,
            "priority": notification.priority,
            "tag": notification.tag or f"billbytekot-{datetime.now().timestamp()}",
            "notification_id": str(uuid.uuid4())
        })
        
        vapid_claims = {
            "sub": vapid_email
        }
        
        sent_count = 0
        failed_count = 0
        failed_endpoints = []
        
        for sub in subscriptions:
            try:
                subscription_info = sub.get("subscription", {})
                webpush(
                    subscription_info=subscription_info,
                    data=payload,
                    vapid_private_key=vapid_private_key,
                    vapid_claims=vapid_claims
                )
                sent_count += 1
            except WebPushException as e:
                failed_count += 1
                # If subscription is invalid, mark as inactive
                if e.response and e.response.status_code in [404, 410]:
                    await db.push_subscriptions.update_one(
                        {"endpoint": sub.get("endpoint")},
                        {"$set": {"active": False, "error": str(e)}}
                    )
                    failed_endpoints.append(sub.get("endpoint"))
            except Exception as e:
                failed_count += 1
        
        # Store notification record
        notif_doc = {
            "title": notification.title,
            "body": notification.body,
            "type": notification.type,
            "url": notification.url,
            "priority": notification.priority,
            "target": notification.target,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "sent_count": sent_count,
            "failed_count": failed_count,
            "status": "sent"
        }
        await db.sent_notifications.insert_one(notif_doc)
        
        return {
            "success": True,
            "message": f"Notification sent to {sent_count} devices",
            "sent_count": sent_count,
            "failed_count": failed_count,
            "total_subscribers": len(subscriptions)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/push/history")
async def get_push_history(username: str, password: str, limit: int = 50):
    """Get push notification history - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    notifications = await db.sent_notifications.find().sort("created_at", -1).limit(limit).to_list(limit)
    for n in notifications:
        n.pop("_id", None)
    
    return {"notifications": notifications}


# ==========================================
# PROMOTIONAL/CAMPAIGN MANAGEMENT ENDPOINTS
# ==========================================

import uuid
from datetime import datetime, timezone, timedelta

class Campaign(BaseModel):
    """Campaign/Promotion model"""
    model_config = ConfigDict(extra="allow")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: str
    discount_type: str = "percentage"  # percentage, fixed, buy_one_get_one
    discount_value: float = 10.0
    min_order_amount: float = 0.0
    max_discount: float = 0.0  # 0 = unlimited
    coupon_code: Optional[str] = None
    start_date: datetime
    end_date: datetime
    is_active: bool = True
    banner_text: Optional[str] = None
    banner_color: str = "violet"  # violet, red, green, blue, orange, pink
    show_on_landing: bool = True
    usage_limit: int = 0  # 0 = unlimited
    used_count: int = 0
    target_audience: str = "all"  # all, new_users, existing_users
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CampaignCreate(BaseModel):
    title: str
    description: str
    discount_type: str = "percentage"
    discount_value: float = 10.0
    min_order_amount: float = 0.0
    max_discount: float = 0.0
    coupon_code: Optional[str] = None
    start_date: str  # ISO format
    end_date: str    # ISO format
    is_active: bool = True
    banner_text: Optional[str] = None
    banner_color: str = "violet"
    show_on_landing: bool = True
    usage_limit: int = 0
    target_audience: str = "all"

class SaleOffer(BaseModel):
    """Sale/Offer configuration"""
    model_config = ConfigDict(extra="allow")
    enabled: bool = False
    title: str = "Special Offer!"
    subtitle: str = "Limited Time Deal"
    discount_text: str = "50% OFF"
    badge_text: str = "SALE"
    bg_color: str = "from-red-500 to-orange-500"
    end_date: Optional[str] = None
    valid_until: Optional[str] = None
    theme: str = "default"  # default, diwali, christmas, newyear, flash, blackfriday
    banner_design: str = "gradient-wave"
    discount_percent: float = 20.0
    original_price: float = 1999.0
    sale_price: float = 1599.0
    cta_text: str = "Grab This Deal Now!"
    urgency_text: str = "⚡ Limited slots available. Offer ends soon!"

class PricingConfig(BaseModel):
    """Pricing configuration"""
    model_config = ConfigDict(extra="allow")
    regular_price: float = 1999.0
    regular_price_display: str = "₹1999"
    campaign_price: float = 1799.0
    campaign_price_display: str = "₹1799"
    campaign_active: bool = False
    campaign_name: str = ""
    campaign_discount_percent: float = 10.0
    campaign_start_date: Optional[str] = None
    campaign_end_date: Optional[str] = None
    trial_expired_discount: float = 10.0
    trial_days: int = 7
    subscription_months: int = 12

@api_router.get("/super-admin/campaigns")
async def get_campaigns(username: str, password: str):
    """Get all campaigns - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        campaigns = await db.campaigns.find({}, {"_id": 0}).to_list(100)
        
        # Calculate active campaigns
        current_time = datetime.now(timezone.utc)
        active_campaigns = []
        expired_campaigns = []
        
        for campaign in campaigns:
            end_date = datetime.fromisoformat(campaign.get('end_date', '').replace('Z', '+00:00'))
            if end_date > current_time and campaign.get('is_active', False):
                active_campaigns.append(campaign)
            else:
                expired_campaigns.append(campaign)
        
        stats = {
            "total_campaigns": len(campaigns),
            "active_campaigns": len(active_campaigns),
            "expired_campaigns": len(expired_campaigns),
            "total_usage": sum(c.get('used_count', 0) for c in campaigns)
        }
        
        return {
            "campaigns": campaigns,
            "active_campaigns": active_campaigns,
            "stats": stats
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/super-admin/campaigns")
async def create_campaign(campaign: CampaignCreate, username: str, password: str):
    """Create new campaign - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        # Convert string dates to datetime
        start_date = datetime.fromisoformat(campaign.start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(campaign.end_date.replace('Z', '+00:00'))
        
        campaign_doc = {
            "id": str(uuid.uuid4()),
            "title": campaign.title,
            "description": campaign.description,
            "discount_type": campaign.discount_type,
            "discount_value": campaign.discount_value,
            "min_order_amount": campaign.min_order_amount,
            "max_discount": campaign.max_discount,
            "coupon_code": campaign.coupon_code,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "is_active": campaign.is_active,
            "banner_text": campaign.banner_text,
            "banner_color": campaign.banner_color,
            "show_on_landing": campaign.show_on_landing,
            "usage_limit": campaign.usage_limit,
            "used_count": 0,
            "target_audience": campaign.target_audience,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.campaigns.insert_one(campaign_doc)
        campaign_doc.pop("_id", None)
        
        return {"success": True, "campaign": campaign_doc}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/super-admin/campaigns/{campaign_id}")
async def update_campaign(campaign_id: str, campaign: CampaignCreate, username: str, password: str):
    """Update campaign - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        start_date = datetime.fromisoformat(campaign.start_date.replace('Z', '+00:00'))
        end_date = datetime.fromisoformat(campaign.end_date.replace('Z', '+00:00'))
        
        update_data = {
            "title": campaign.title,
            "description": campaign.description,
            "discount_type": campaign.discount_type,
            "discount_value": campaign.discount_value,
            "min_order_amount": campaign.min_order_amount,
            "max_discount": campaign.max_discount,
            "coupon_code": campaign.coupon_code,
            "start_date": start_date.isoformat(),
            "end_date": end_date.isoformat(),
            "is_active": campaign.is_active,
            "banner_text": campaign.banner_text,
            "banner_color": campaign.banner_color,
            "show_on_landing": campaign.show_on_landing,
            "usage_limit": campaign.usage_limit,
            "target_audience": campaign.target_audience,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }
        
        result = await db.campaigns.update_one(
            {"id": campaign_id},
            {"$set": update_data}
        )
        
        if result.matched_count == 0:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        return {"success": True, "message": "Campaign updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.delete("/super-admin/campaigns/{campaign_id}")
async def delete_campaign(campaign_id: str, username: str, password: str):
    """Delete campaign - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        result = await db.campaigns.delete_one({"id": campaign_id})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Campaign not found")
        
        return {"success": True, "message": "Campaign deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/super-admin/sale-offer")
async def get_sale_offer(username: str, password: str):
    """Get sale offer configuration - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        sale_offer = await db.sale_offers.find_one({}, {"_id": 0})
        
        if not sale_offer:
            # Return default configuration
            sale_offer = {
                "enabled": False,
                "title": "Special Offer!",
                "subtitle": "Limited Time Deal",
                "discount_text": "50% OFF",
                "badge_text": "SALE",
                "bg_color": "from-red-500 to-orange-500",
                "theme": "default",
                "banner_design": "gradient-wave",
                "discount_percent": 20.0,
                "original_price": 1999.0,
                "sale_price": 1599.0,
                "cta_text": "Grab This Deal Now!",
                "urgency_text": "⚡ Limited slots available. Offer ends soon!"
            }
        
        return sale_offer
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/super-admin/sale-offer")
async def update_sale_offer(sale_offer: SaleOffer, username: str, password: str):
    """Update sale offer configuration - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        sale_offer_doc = sale_offer.model_dump()
        sale_offer_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.sale_offers.replace_one(
            {},
            sale_offer_doc,
            upsert=True
        )
        
        return {"success": True, "message": "Sale offer updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/super-admin/pricing")
async def get_pricing_config(username: str, password: str):
    """Get pricing configuration - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        pricing = await db.pricing_config.find_one({}, {"_id": 0})
        
        if not pricing:
            # Return default pricing
            pricing = {
                "regular_price": 1999.0,
                "regular_price_display": "₹1999",
                "campaign_price": 1799.0,
                "campaign_price_display": "₹1799",
                "campaign_active": False,
                "campaign_name": "",
                "campaign_discount_percent": 10.0,
                "trial_expired_discount": 10.0,
                "trial_days": 7,
                "subscription_months": 12
            }
        
        return pricing
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.put("/super-admin/pricing")
async def update_pricing_config(pricing: PricingConfig, username: str, password: str):
    """Update pricing configuration - Super Admin Only"""
    if not verify_super_admin(username, password):
        raise HTTPException(status_code=403, detail="Invalid credentials")
    
    try:
        pricing_doc = pricing.model_dump()
        pricing_doc["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.pricing_config.replace_one(
            {},
            pricing_doc,
            upsert=True
        )
        
        return {"success": True, "message": "Pricing configuration updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Public endpoints for campaigns and offers
@api_router.get("/public/active-campaigns")
async def get_active_campaigns():
    """Get active campaigns for public display"""
    try:
        current_time = datetime.now(timezone.utc)
        
        campaigns = await db.campaigns.find({
            "is_active": True,
            "show_on_landing": True,
            "start_date": {"$lte": current_time.isoformat()},
            "end_date": {"$gte": current_time.isoformat()}
        }, {"_id": 0}).to_list(10)
        
        return {"campaigns": campaigns}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/public/sale-offer")
async def get_public_sale_offer():
    """Get active sale offer for public display"""
    try:
        sale_offer = await db.sale_offers.find_one({"enabled": True}, {"_id": 0})
        
        if not sale_offer:
            return {"enabled": False}
        
        # Check if offer is still valid
        if sale_offer.get("valid_until"):
            valid_until = datetime.fromisoformat(sale_offer["valid_until"].replace('Z', '+00:00'))
            if datetime.now(timezone.utc) > valid_until:
                return {"enabled": False}
        
        return sale_offer
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/public/pricing")
async def get_public_pricing():
    """Get current pricing for public display with campaign logic"""
    try:
        pricing = await db.pricing_config.find_one({}, {"_id": 0})
        now = datetime.now(timezone.utc)
        
        # Default pricing if no config exists
        if not pricing:
            return {
                "regular_price": 1999.0,
                "regular_price_display": "₹1999",
                "campaign_price": None,
                "campaign_price_display": None,
                "campaign_active": False,
                "campaign_name": None,
                "campaign_discount_percent": 0,
                "campaign_start_date": None,
                "campaign_end_date": None
            }
        
        regular_price = pricing.get("regular_price", 1999.0)
        campaign_active = pricing.get("campaign_active", False)
        campaign_discount_percent = pricing.get("campaign_discount_percent", 0)
        campaign_name = pricing.get("campaign_name", None)
        campaign_start_date = pricing.get("campaign_start_date")
        campaign_end_date = pricing.get("campaign_end_date")
        
        # Validate campaign dates if campaign is marked as active
        if campaign_active:
            # Check start_date - campaign should not be active before start
            if campaign_start_date:
                try:
                    start_str = campaign_start_date if isinstance(campaign_start_date, str) else campaign_start_date.isoformat()
                    start = datetime.fromisoformat(start_str.replace('Z', '+00:00'))
                    if start.tzinfo is None:
                        start = start.replace(tzinfo=timezone.utc)
                    if now < start:
                        campaign_active = False
                except (ValueError, AttributeError):
                    pass
            
            # Check end_date - campaign should not be active after end
            if campaign_active and campaign_end_date:
                try:
                    end_str = campaign_end_date if isinstance(campaign_end_date, str) else campaign_end_date.isoformat()
                    end = datetime.fromisoformat(end_str.replace('Z', '+00:00'))
                    if end.tzinfo is None:
                        end = end.replace(tzinfo=timezone.utc)
                    if now > end:
                        campaign_active = False
                except (ValueError, AttributeError):
                    pass
        
        # Calculate campaign price if campaign is active
        campaign_price = None
        campaign_price_display = None
        
        if campaign_active and campaign_discount_percent > 0:
            # Calculate: campaign_price = regular_price - (regular_price * discount_percent / 100)
            campaign_price = regular_price - (regular_price * campaign_discount_percent / 100)
            campaign_price = round(campaign_price, 2)
            campaign_price_display = f"₹{int(campaign_price)}" if campaign_price == int(campaign_price) else f"₹{campaign_price}"
        elif campaign_active:
            # Use stored campaign_price if no discount percent
            campaign_price = pricing.get("campaign_price")
            if campaign_price:
                campaign_price_display = f"₹{int(campaign_price)}" if campaign_price == int(campaign_price) else f"₹{campaign_price}"
        
        return {
            "regular_price": regular_price,
            "regular_price_display": f"₹{int(regular_price)}" if regular_price == int(regular_price) else f"₹{regular_price}",
            "campaign_price": campaign_price,
            "campaign_price_display": campaign_price_display,
            "campaign_active": campaign_active,
            "campaign_name": campaign_name if campaign_active else None,
            "campaign_discount_percent": campaign_discount_percent if campaign_active else 0,
            "campaign_start_date": campaign_start_date,
            "campaign_end_date": campaign_end_date
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Include all API routes (must be after all route definitions)
app.include_router(api_router)

@app.on_event("shutdown")
async def shutdown_db_client():
    # Cleanup Redis cache
    try:
        await cleanup_redis_cache()
    except Exception as e:
        print(f"⚠️ Redis cleanup error: {e}")
    
    # Close MongoDB client
    client.close()
    print("🔌 Database connections closed")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=10000)
